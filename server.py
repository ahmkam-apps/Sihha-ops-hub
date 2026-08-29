import os
import hashlib
import secrets
import sqlite3
import uuid
import logging
from datetime import datetime, timedelta
from functools import wraps
from flask import Flask, request, jsonify, send_from_directory, g
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from werkzeug.exceptions import HTTPException
from werkzeug.middleware.proxy_fix import ProxyFix

# ── Config ────────────────────────────────────────────────────────────────────

app = Flask(__name__, static_folder='public')
# gzip/brotli-compress responses (audit P2): index.html is ~267KB of inline HTML/JS
# served no-store, so it re-downloads every admin visit. Compression cuts that ~4-5×
# with a single line. flask-compress only touches compressible content-types above a
# size threshold; JSON API responses benefit too. Degrades gracefully if not installed.
try:
    from flask_compress import Compress
    Compress(app)
except ImportError:
    import logging as _logging
    _logging.getLogger(__name__).warning('flask-compress not installed — responses uncompressed')
# Railway terminates TLS at a single reverse proxy that appends the real client IP
# to X-Forwarded-For. Trust exactly ONE hop (x_for=1) so request.remote_addr is the
# real client and cannot be spoofed by a client-supplied X-Forwarded-For (audit P1.5:
# login-throttle bypass). Do NOT raise x_for above 1 — that would re-open the spoof.
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1)
# CORS restricted to known origins (audit 2.3) — override with CORS_ORIGINS env var
# (comma-separated) if domains change. Same-origin SPA/iframe traffic is unaffected
# by CORS; this only blocks third-party sites from scripting the API.
_CORS_ORIGINS = [o.strip() for o in os.environ.get(
    'CORS_ORIGINS',
    'https://sihha-ops-hub-production.up.railway.app,https://ops.sihha.org,'
    'https://sihha.org,https://www.sihha.org'
).split(',') if o.strip()]
CORS(app, origins=_CORS_ORIGINS)

@app.after_request
def add_security_headers(response):
    """Low-risk baseline headers; frame policy/CSP need a separate widget/inline-JS migration."""
    response.headers.setdefault('X-Content-Type-Options', 'nosniff')
    response.headers.setdefault('Referrer-Policy', 'same-origin')
    response.headers.setdefault(
        'Permissions-Policy', 'camera=(self), geolocation=(), microphone=(), payment=()'
    )
    if (request.path.startswith('/confirm/')
            or request.path.startswith('/api/family/confirm/')
            or request.path == '/activate'
            or request.path.startswith('/api/auth/access-invitation')):
        response.headers['Referrer-Policy'] = 'no-referrer'
        response.headers['Cache-Control'] = 'no-store'
    if request.is_secure:
        response.headers.setdefault('Strict-Transport-Security', 'max-age=31536000')
    return response
# Cap request body size (receipt photos incl. HEIC) — prevents disk-fill DoS on the
# Railway volume that also holds the DB. Flask returns 413 automatically when exceeded.
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16 MB

DB_PATH         = os.environ.get('DB_PATH', 'data/sihaa.db')
UPLOAD_FOLDER   = os.environ.get('UPLOAD_FOLDER', 'data/uploads')
SESSION_HOURS   = int(os.environ.get('SESSION_EXPIRY_HOURS', 24))
PORT            = int(os.environ.get('PORT', 5000))
ALLOWED_EXT     = {'png', 'jpg', 'jpeg', 'gif', 'pdf', 'heic'}
MAX_UPLOAD_BYTES = int(os.environ.get('MAX_UPLOAD_BYTES', 12 * 1024 * 1024))
MAX_IMAGE_PIXELS = int(os.environ.get('MAX_IMAGE_PIXELS', 40_000_000))
UPLOAD_FILES_PER_DAY = int(os.environ.get('UPLOAD_FILES_PER_DAY', 20))
UPLOAD_BYTES_PER_DAY = int(os.environ.get('UPLOAD_BYTES_PER_DAY', 64 * 1024 * 1024))
UPLOAD_TOTAL_BYTES = int(os.environ.get('UPLOAD_TOTAL_BYTES', 2 * 1024 * 1024 * 1024))
CONFIRMATION_TOKEN_HOURS = int(os.environ.get('CONFIRMATION_TOKEN_HOURS', 168))
ACCOUNT_INVITATION_MINUTES = 60
ACCOUNT_INVITATION_URL = 'https://ops.sihha.org/activate'
EMAIL_PROVIDER = os.environ.get('EMAIL_PROVIDER', 'sendgrid').strip().lower() or 'sendgrid'
SENDGRID_API_KEY  = os.environ.get('SENDGRID_API_KEY', '').strip()
NOTIFY_FROM_EMAIL = os.environ.get('NOTIFY_FROM_EMAIL', 'ops@sihha.org')
TWILIO_EMAIL_API_KEY_SID = os.environ.get('TWILIO_EMAIL_API_KEY_SID', '').strip()
TWILIO_EMAIL_API_KEY_SECRET = os.environ.get('TWILIO_EMAIL_API_KEY_SECRET', '').strip()
TWILIO_EMAIL_ENDPOINT = 'https://comms.twilio.com/v1/Emails'
TWILIO_EMAIL_MAX_REQUEST_BYTES = 9_500_000
EMAIL_ATTACHMENT_MAX_BYTES = 7_000_000

def _env_flag(name, default=False):
    """Parse a conventional boolean environment variable."""
    raw = os.environ.get(name)
    if raw is None:
        return default
    return raw.strip().lower() in ('1', 'true', 'yes', 'on')

# Production sets REQUIRE_EXISTING_DB=1. Staging intentionally leaves it off so
# its synthetic database can be rebuilt. This prevents a missing Railway volume
# from silently bootstrapping an empty production database that still looks healthy.
REQUIRE_EXISTING_DB = _env_flag('REQUIRE_EXISTING_DB', False)

# ── Receipt vision-parsing (Phase A) ──────────────────────────────────────────
# Auto-extract store/date/total/line-items from an uploaded receipt photo via the
# Anthropic vision API. Fully OPTIONAL: if the key is unset or the flag is off, the
# app behaves exactly as before (manual entry) — nothing breaks. Parsed values only
# ever pre-fill the form; a treasurer/admin still approves before any money moves.
ANTHROPIC_API_KEY       = os.environ.get('ANTHROPIC_API_KEY', '')
RECEIPT_PARSE_MODEL     = os.environ.get('RECEIPT_PARSE_MODEL', 'claude-haiku-4-5-20251001')
# Auto-read is ON by default whenever an API key is present — you only set
# ENABLE_RECEIPT_PARSING to a falsey value (0/false/no/off) to turn it OFF. This
# avoids the footgun of setting the key but forgetting a separate enable flag.
ENABLE_RECEIPT_PARSING  = os.environ.get('ENABLE_RECEIPT_PARSING', '').strip().lower() not in ('0', 'false', 'no', 'off')
# Only actually call the API when parsing is enabled AND a key is present.
RECEIPT_PARSING_ACTIVE  = ENABLE_RECEIPT_PARSING and bool(ANTHROPIC_API_KEY)

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
_early_log = logging.getLogger(__name__)
_early_log.info(
    'Email provider=%s configured=%s notify_from=%r',
    EMAIL_PROVIDER,
    bool(SENDGRID_API_KEY) if EMAIL_PROVIDER == 'sendgrid' else bool(
        TWILIO_EMAIL_API_KEY_SID and TWILIO_EMAIL_API_KEY_SECRET
    ),
    NOTIFY_FROM_EMAIL,
)

os.makedirs(os.path.dirname(DB_PATH) if os.path.dirname(DB_PATH) else '.', exist_ok=True)
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
log = logging.getLogger(__name__)

# ── DB Helpers ────────────────────────────────────────────────────────────────

def make_conn():
    """Connection factory for non-request contexts (scheduler jobs, scripts).
    Applies the same pragmas as get_db() — without this, background jobs ran
    with foreign_keys OFF and no busy_timeout (audit 2.5)."""
    conn = sqlite3.connect(DB_PATH, detect_types=sqlite3.PARSE_DECLTYPES)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA journal_mode=WAL')
    conn.execute('PRAGMA foreign_keys=ON')
    conn.execute('PRAGMA busy_timeout=5000')
    return conn

def get_db():
    if 'db' not in g:
        g.db = make_conn()
    return g.db

@app.teardown_appcontext
def close_db(error):
    db = g.pop('db', None)
    if db is not None:
        db.close()

def now():
    return datetime.utcnow().isoformat()

ADMIN_PASSWORD_ENV_STATE_KEY = 'admin_password_env_hash'

def _sync_admin_password_from_env(conn, admin_pw):
    """Apply ADMIN_PASSWORD only when that environment value actually changes.

    The environment variable remains the break-glass recovery mechanism, while a
    password changed inside the app survives ordinary restarts and deployments.
    A slow password hash is stored as the environment-state verifier so the
    database never contains the raw environment password or a fast reusable digest.
    """
    admin_row = conn.execute(
        "SELECT id, password_hash FROM users WHERE username='admin'"
    ).fetchone()
    if not admin_row:
        return False

    state = conn.execute(
        "SELECT value FROM app_settings WHERE key=?",
        (ADMIN_PASSWORD_ENV_STATE_KEY,)
    ).fetchone()
    env_value_changed = bool(
        state and not check_password_hash(state['value'], admin_pw)
    )
    first_tracking_boot = state is None
    password_needs_sync = not check_password_hash(admin_row['password_hash'], admin_pw)

    # On the migration boot, preserve the historical contract by reconciling a
    # mismatch once. After that, only a changed env value is authoritative.
    should_sync = password_needs_sync and (first_tracking_boot or env_value_changed)
    if should_sync:
        changed_at = now()
        conn.execute(
            '''UPDATE users SET password_hash=?, must_change_password=0,
               password_changed_at=? WHERE id=?''',
            (generate_password_hash(admin_pw), changed_at, admin_row['id'])
        )
        conn.execute("DELETE FROM sessions WHERE user_id=?", (admin_row['id'],))
        conn.execute(
            '''UPDATE account_invitations SET invalidated_at=?
               WHERE user_id=? AND used_at IS NULL AND invalidated_at IS NULL''',
            (changed_at, admin_row['id'])
        )
        conn.execute(
            '''INSERT INTO account_access_events
               (id,user_id,event_type,actor_user_id,detail,created_at)
               VALUES (?,?,?,?,?,?)''',
            (str(uuid.uuid4()), admin_row['id'],
             'admin_password_reset_from_environment', None,
             'ADMIN_PASSWORD value changed', changed_at)
        )
        log.info('Admin password changed from ADMIN_PASSWORD; existing sessions revoked.')

    if first_tracking_boot or env_value_changed:
        conn.execute(
            '''INSERT INTO app_settings (key,value,updated_at) VALUES (?,?,?)
               ON CONFLICT(key) DO UPDATE SET value=excluded.value,
                                              updated_at=excluded.updated_at''',
            (ADMIN_PASSWORD_ENV_STATE_KEY, generate_password_hash(admin_pw), now())
        )
    return should_sync

def _bundle_letter(family_size):
    """Return S / M / L based on household size."""
    size = int(family_size or 1)
    if size <= 2:   return 'S'
    elif size <= 5: return 'M'
    else:           return 'L'

def _normalize_phone(phone):
    """Strip all non-digit characters. '555-123-4567' → '5551234567'."""
    return ''.join(c for c in (phone or '') if c.isdigit())

VALID_VOLUNTEER_ROLES = {'shopper', 'delivery', 'both', 'general'}

def _normalize_volunteer_role(role):
    """Normalize the old public-form label while keeping DB values canonical."""
    value = (role or '').strip().lower()
    return 'delivery' if value == 'driver' else value

def _make_family_code(phone, family_size, db_conn=None, exclude_id=None):
    """Generate unique human-readable family reference.
    Format: last 6 digits of phone + '-' + bundle size letter.
    Example: phone=5855551234, family_size=4 → '551234-M'
    If a collision exists, appends -2, -3, … until unique.
    Pass db_conn to enforce uniqueness against the families table.
    """
    digits = ''.join(c for c in (phone or '') if c.isdigit())
    phone_part = digits[-6:].zfill(6) if digits else '000000'
    bundle = _bundle_letter(family_size)
    base = f'{phone_part}-{bundle}'

    if db_conn is None:
        return base

    # Enforce uniqueness
    candidate = base
    suffix = 2
    while True:
        row = db_conn.execute(
            "SELECT id FROM families WHERE family_code=?" +
            (" AND id!=?" if exclude_id else ""),
            (candidate, exclude_id) if exclude_id else (candidate,)
        ).fetchone()
        if not row:
            return candidate
        candidate = f'{base}-{suffix}'
        suffix += 1

def bootstrap_db():
    # ── Diagnostics: log exactly where the DB lives and whether it's fresh ────
    abs_db = os.path.abspath(DB_PATH)
    db_dir = os.path.dirname(abs_db)
    log.info(f'DB_PATH={DB_PATH}  →  absolute={abs_db}')
    log.info(f'Working directory: {os.getcwd()}')

    # A production volume outage must stop the release. Without this guard,
    # sqlite3.connect() creates a new empty file and the health check can report
    # success even though all operational data is absent.
    if REQUIRE_EXISTING_DB:
        valid_existing_db = False
        try:
            with open(abs_db, 'rb') as existing_db:
                valid_existing_db = existing_db.read(16) == b'SQLite format 3\x00'
        except OSError:
            pass
        if not valid_existing_db:
            raise RuntimeError(
                'REQUIRE_EXISTING_DB is enabled but DB_PATH is missing or is not a valid '
                'SQLite database. Refusing to create an empty production database; verify '
                'the Railway volume mount and DB_PATH.'
            )

    if os.path.exists(abs_db):
        size_kb = os.path.getsize(abs_db) / 1024
        log.info(f'DB EXISTS ({size_kb:.1f} KB) — data will be preserved ✓')
    else:
        log.warning(f'DB NOT FOUND — creating fresh database. '
                    f'If this happens every deploy, the Railway Volume is not mounted at: {db_dir}')

    os.makedirs(db_dir, exist_ok=True)
    conn = sqlite3.connect(abs_db)
    conn.execute('PRAGMA journal_mode=WAL')
    conn.execute('PRAGMA foreign_keys=ON')
    conn.execute('PRAGMA busy_timeout=5000')  # 2 workers bootstrap concurrently
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    c.executescript('''
        CREATE TABLE IF NOT EXISTS users (
            id                   TEXT PRIMARY KEY,
            username             TEXT UNIQUE NOT NULL,
            password_hash        TEXT NOT NULL,
            name                 TEXT,
            role                 TEXT NOT NULL DEFAULT 'viewer'
                                 CHECK(role IN ('admin','volunteer','finance','treasurer','viewer','family')),
            email                TEXT,
            wa_phone             TEXT,
            wa_apikey            TEXT,
            active               INTEGER NOT NULL DEFAULT 1,
            linked_id            TEXT,
            linked_type          TEXT,
            must_change_password INTEGER NOT NULL DEFAULT 1,
            password_changed_at  TEXT,
            last_login_at        TEXT,
            created_at           TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS sessions (
            token       TEXT PRIMARY KEY,
            user_id     TEXT NOT NULL,
            expires_at  TEXT NOT NULL,
            created_at  TEXT NOT NULL,
            FOREIGN KEY (user_id) REFERENCES users(id)
        );

        -- Account invitations never store the bearer token itself. Only its
        -- SHA-256 digest is retained, and a token becomes unusable after one
        -- successful password creation, expiry, or explicit invalidation.
        CREATE TABLE IF NOT EXISTS account_invitations (
            id               TEXT PRIMARY KEY,
            user_id          TEXT NOT NULL,
            token_hash       TEXT UNIQUE NOT NULL,
            delivery_email   TEXT NOT NULL,
            created_by       TEXT,
            created_at       TEXT NOT NULL,
            expires_at       TEXT NOT NULL,
            email_sent_at    TEXT,
            used_at          TEXT,
            invalidated_at   TEXT,
            FOREIGN KEY (user_id) REFERENCES users(id),
            FOREIGN KEY (created_by) REFERENCES users(id)
        );

        -- Append-only security audit trail for invitation and credential events.
        -- It intentionally excludes raw tokens, passwords, email addresses, and IPs.
        CREATE TABLE IF NOT EXISTS account_access_events (
            id               TEXT PRIMARY KEY,
            user_id          TEXT NOT NULL,
            invitation_id    TEXT,
            event_type       TEXT NOT NULL,
            actor_user_id    TEXT,
            detail           TEXT,
            created_at       TEXT NOT NULL
        );

        -- Internal operational state. Values must never contain raw secrets.
        CREATE TABLE IF NOT EXISTS app_settings (
            key              TEXT PRIMARY KEY,
            value            TEXT NOT NULL,
            updated_at       TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS families (
            id              TEXT PRIMARY KEY,
            name            TEXT NOT NULL,
            phone           TEXT,
            address         TEXT,
            city            TEXT,
            family_size     INTEGER,
            children_count  INTEGER,
            dietary_notes   TEXT,
            frequency       TEXT,
            income_range    TEXT,
            status          TEXT NOT NULL DEFAULT 'pending'
                            CHECK(status IN ('pending','active','inactive','paused')),
            notes           TEXT,
            source          TEXT DEFAULT 'admin',
            created_at      TEXT NOT NULL,
            updated_at      TEXT
        );

        CREATE TABLE IF NOT EXISTS volunteers (
            id                  TEXT PRIMARY KEY,
            name                TEXT NOT NULL,
            phone               TEXT,
            email               TEXT,
            role                TEXT DEFAULT 'shopper'
                                CHECK(role IN ('shopper','delivery','both','general')),
            availability        TEXT,
            service_area        TEXT,
            contact_preference  TEXT,
            volunteer_areas     TEXT,
            comfort_level       INTEGER,
            skills              TEXT,
            other_info          TEXT,
            status              TEXT NOT NULL DEFAULT 'pending'
                                CHECK(status IN ('pending','active','inactive')),
            notes               TEXT,
            source              TEXT DEFAULT 'admin',
            created_at          TEXT NOT NULL,
            updated_at          TEXT
        );

        CREATE TABLE IF NOT EXISTS assignments (
            id              TEXT PRIMARY KEY,
            family_id       TEXT NOT NULL,
            volunteer_id    TEXT,
            task_type       TEXT CHECK(task_type IN ('shopping','delivery','both')),
            due_date        TEXT,
            status          TEXT NOT NULL DEFAULT 'pending'
                            CHECK(status IN ('pending','assigned','in_progress','completed','cancelled')),
            notes           TEXT,
            created_by      TEXT,
            created_at      TEXT NOT NULL,
            updated_at      TEXT,
            FOREIGN KEY (family_id) REFERENCES families(id),
            FOREIGN KEY (volunteer_id) REFERENCES volunteers(id)
        );

        CREATE TABLE IF NOT EXISTS receipts (
            id              TEXT PRIMARY KEY,
            assignment_id   TEXT,
            volunteer_id    TEXT,
            family_id       TEXT,
            store           TEXT,
            purchase_date   TEXT,
            amount          REAL,
            file_url        TEXT,
            status          TEXT NOT NULL DEFAULT 'pending'
                            CHECK(status IN ('pending','approved','rejected')),
            notes           TEXT,
            created_at      TEXT NOT NULL,
            updated_at      TEXT,
            FOREIGN KEY (assignment_id) REFERENCES assignments(id),
            FOREIGN KEY (volunteer_id) REFERENCES volunteers(id),
            FOREIGN KEY (family_id) REFERENCES families(id)
        );

        CREATE TABLE IF NOT EXISTS reimbursements (
            id              TEXT PRIMARY KEY,
            receipt_id      TEXT NOT NULL,
            volunteer_id    TEXT,
            amount          REAL,
            status          TEXT NOT NULL DEFAULT 'pending'
                            CHECK(status IN ('pending','approved','paid','rejected')),
            payment_method  TEXT CHECK(payment_method IN ('venmo','zelle','check','cash','bank_transfer','cheque','other')),
            payment_ref     TEXT,
            paid_date       TEXT,
            approved_by     TEXT,
            notes           TEXT,
            created_at      TEXT NOT NULL,
            updated_at      TEXT,
            FOREIGN KEY (receipt_id) REFERENCES receipts(id)
        );

        CREATE TABLE IF NOT EXISTS donations (
            id          TEXT PRIMARY KEY,
            donor_name  TEXT,
            donor_email TEXT,
            amount      REAL,
            type        TEXT DEFAULT 'online'
                        CHECK(type IN ('online','cash','check','bank')),
            date        TEXT,
            source      TEXT DEFAULT 'manual',
            reference_id TEXT,
            cycle_id    TEXT,
            notes       TEXT,
            created_at  TEXT NOT NULL
        );

        -- ── Food Catalog ──────────────────────────────────────────────────────

        CREATE TABLE IF NOT EXISTS food_categories (
            id            TEXT PRIMARY KEY,
            name          TEXT NOT NULL,
            display_order INTEGER NOT NULL DEFAULT 0,
            is_active     INTEGER NOT NULL DEFAULT 1,
            created_at    TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS food_items (
            id            TEXT PRIMARY KEY,
            category_id   TEXT NOT NULL,
            name          TEXT NOT NULL,
            unit          TEXT NOT NULL DEFAULT 'each',
            is_active     INTEGER NOT NULL DEFAULT 1,
            display_order INTEGER NOT NULL DEFAULT 0,
            created_at    TEXT NOT NULL,
            FOREIGN KEY (category_id) REFERENCES food_categories(id)
        );

        CREATE TABLE IF NOT EXISTS bundle_quantities (
            id           TEXT PRIMARY KEY,
            food_item_id TEXT NOT NULL,
            bundle_size  TEXT NOT NULL CHECK(bundle_size IN ('S','M','L')),
            quantity     TEXT NOT NULL,
            UNIQUE(food_item_id, bundle_size),
            FOREIGN KEY (food_item_id) REFERENCES food_items(id)
        );

        CREATE TABLE IF NOT EXISTS bundle_size_rules (
            id            TEXT PRIMARY KEY,
            bundle_size   TEXT NOT NULL UNIQUE CHECK(bundle_size IN ('S','M','L')),
            label         TEXT NOT NULL,
            min_household INTEGER NOT NULL,
            max_household INTEGER
        );

        -- ── Delivery Cycles ───────────────────────────────────────────────────

        CREATE TABLE IF NOT EXISTS delivery_cycles (
            id                  TEXT PRIMARY KEY,
            title               TEXT NOT NULL,
            delivery_date_start TEXT NOT NULL,
            delivery_date_end   TEXT NOT NULL,
            request_open_at     TEXT NOT NULL,
            request_close_at    TEXT NOT NULL,
            status              TEXT NOT NULL DEFAULT 'draft'
                                CHECK(status IN ('draft','open','closed','shopping','delivered')),
            notes               TEXT,
            created_by          TEXT,
            created_at          TEXT NOT NULL,
            updated_at          TEXT
        );

        -- ── Food Orders ───────────────────────────────────────────────────────

        CREATE TABLE IF NOT EXISTS food_requests (
            id                  TEXT PRIMARY KEY,
            cycle_id            TEXT NOT NULL,
            family_id           TEXT NOT NULL,
            bundle_size         TEXT NOT NULL CHECK(bundle_size IN ('S','M','L')),
            submitted_at        TEXT NOT NULL,
            status              TEXT NOT NULL DEFAULT 'submitted'
                                CHECK(status IN ('submitted','assigned','delivered','cancelled')),
            assigned_volunteer_id TEXT,
            delivered_at        TEXT,
            notes               TEXT,
            confirmation_expires_at TEXT,
            UNIQUE(cycle_id, family_id),
            FOREIGN KEY (cycle_id) REFERENCES delivery_cycles(id),
            FOREIGN KEY (family_id) REFERENCES families(id)
        );

        CREATE TABLE IF NOT EXISTS food_request_items (
            id           TEXT PRIMARY KEY,
            request_id   TEXT NOT NULL,
            food_item_id TEXT NOT NULL,
            selected     INTEGER NOT NULL DEFAULT 0,
            UNIQUE(request_id, food_item_id),
            FOREIGN KEY (request_id) REFERENCES food_requests(id),
            FOREIGN KEY (food_item_id) REFERENCES food_items(id)
        );

        -- ── Volunteer Cycle Assignments ────────────────────────────────────────

        CREATE TABLE IF NOT EXISTS cycle_assignments (
            id           TEXT PRIMARY KEY,
            cycle_id     TEXT NOT NULL,
            volunteer_id TEXT NOT NULL,
            family_id    TEXT,
            task_type    TEXT NOT NULL CHECK(task_type IN ('shopping','delivery')),
            task_date    TEXT,
            task_time    TEXT,
            status       TEXT NOT NULL DEFAULT 'pending'
                         CHECK(status IN ('pending','confirmed','completed','cancelled')),
            notes        TEXT,
            created_at   TEXT NOT NULL,
            updated_at   TEXT,
            FOREIGN KEY (cycle_id) REFERENCES delivery_cycles(id),
            FOREIGN KEY (volunteer_id) REFERENCES volunteers(id),
            FOREIGN KEY (family_id) REFERENCES families(id)
        );

        -- ── Financial Reconciliation ──────────────────────────────────────────

        CREATE TABLE IF NOT EXISTS bank_transactions (
            id                  TEXT PRIMARY KEY,
            transaction_date    TEXT NOT NULL,
            description         TEXT,
            amount              REAL NOT NULL,
            matched_donation_id TEXT,
            reconcile_status    TEXT NOT NULL DEFAULT 'unmatched'
                                CHECK(reconcile_status IN ('matched','unmatched','ignored')),
            imported_at         TEXT NOT NULL,
            FOREIGN KEY (matched_donation_id) REFERENCES donations(id)
        );

        CREATE TABLE IF NOT EXISTS reconciliation_runs (
            id                      TEXT PRIMARY KEY,
            run_date                TEXT NOT NULL,
            period_start            TEXT,
            period_end              TEXT,
            total_online_donations  REAL DEFAULT 0,
            total_bank_deposits     REAL DEFAULT 0,
            variance                REAL DEFAULT 0,
            notes                   TEXT,
            run_by                  TEXT,
            created_at              TEXT NOT NULL
        );

        -- ── Phase 4D: Stripe / Wix / Bank reconciliation ─────────────────────

        CREATE TABLE IF NOT EXISTS stripe_transactions (
            id               TEXT PRIMARY KEY,
            stripe_charge_id TEXT UNIQUE,
            stripe_payout_id TEXT,
            donor_name       TEXT,
            donor_email      TEXT,
            amount           REAL,
            fee              REAL,
            net              REAL,
            charge_date      TEXT,
            payout_date      TEXT,
            description      TEXT,
            synced_at        TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS wix_donations (
            id              TEXT PRIMARY KEY,
            wix_order_id    TEXT UNIQUE,
            donor_name      TEXT,
            donor_email     TEXT,
            amount          REAL,
            donation_date   TEXT,
            description     TEXT,
            synced_at       TEXT NOT NULL
        );

        -- Persistent security counters shared by every gunicorn worker.
        -- bucket_key is a SHA-256 digest; raw usernames, phone numbers and IPs
        -- are not retained in this operational table.
        CREATE TABLE IF NOT EXISTS rate_limit_events (
            id         TEXT PRIMARY KEY,
            scope      TEXT NOT NULL,
            bucket_key TEXT NOT NULL,
            created_at TEXT NOT NULL
        );

        -- Registry for new files supports per-uploader quotas and safe cleanup
        -- of uploads that were never attached to a receipt.
        CREATE TABLE IF NOT EXISTS uploaded_files (
            filename         TEXT PRIMARY KEY,
            uploader_user_id TEXT,
            volunteer_id     TEXT,
            size_bytes       INTEGER NOT NULL,
            created_at       TEXT NOT NULL,
            claimed_at       TEXT
        );
    ''')

    # ── Performance indexes ───────────────────────────────────────────────────
    _performance_indexes = [
        # volunteer_slots: used in almost every delivery/portal operation
        "CREATE INDEX IF NOT EXISTS idx_vs_cycle_family   ON volunteer_slots(cycle_id, family_id)",
        "CREATE INDEX IF NOT EXISTS idx_vs_cycle_status   ON volunteer_slots(cycle_id, status)",
        "CREATE INDEX IF NOT EXISTS idx_vs_claimed_by     ON volunteer_slots(claimed_by)",
        # food_request_items: N+1 guard — fetched per-order in many list endpoints
        "CREATE INDEX IF NOT EXISTS idx_fri_request_id    ON food_request_items(request_id)",
        # food_requests: core lookup by family or cycle
        "CREATE INDEX IF NOT EXISTS idx_fr_family_id      ON food_requests(family_id)",
        "CREATE INDEX IF NOT EXISTS idx_fr_cycle_id       ON food_requests(cycle_id)",
        # food_request_events: audit log lookup by request
        "CREATE INDEX IF NOT EXISTS idx_fre_request_id    ON food_request_events(request_id)",
        # donations: date-range filters used in all finance views
        "CREATE INDEX IF NOT EXISTS idx_donations_date    ON donations(date)",
        # sessions: looked up by token (PK) + occasionally by user_id
        "CREATE INDEX IF NOT EXISTS idx_sessions_user_id  ON sessions(user_id)",
        # ── audit 3.2 additions (2026-06-10) ──
        # families: phone lookup is the family login/fuzzy-match hot path
        "CREATE INDEX IF NOT EXISTS idx_families_phone    ON families(phone)",
        # volunteer_slots by family alone: list_families runs per-family subqueries
        # (idx_vs_cycle_family leads on cycle_id so it can't serve these)
        "CREATE INDEX IF NOT EXISTS idx_vs_family_id      ON volunteer_slots(family_id)",
        # receipts: joined/filtered by slot, volunteer, and cycle in finance views
        "CREATE INDEX IF NOT EXISTS idx_receipts_slot     ON receipts(slot_id)",
        "CREATE INDEX IF NOT EXISTS idx_receipts_vol      ON receipts(volunteer_id)",
        "CREATE INDEX IF NOT EXISTS idx_receipts_cycle    ON receipts(cycle_id)",
        # sessions: expiry sweeps
        "CREATE INDEX IF NOT EXISTS idx_sessions_expires  ON sessions(expires_at)",
        # donations: per-cycle finance summary
        "CREATE INDEX IF NOT EXISTS idx_donations_cycle   ON donations(cycle_id)",
        # donations: Wix sync dedupes by reference_id (wix order id) on every page —
        # without this it's an O(orders×donations) scan that worsens as history grows (audit P1.7)
        "CREATE INDEX IF NOT EXISTS idx_donations_ref     ON donations(reference_id)",
        # reminder_log: idempotency guards query by (slot_id, sent_to)
        "CREATE INDEX IF NOT EXISTS idx_rl_slot_sent      ON reminder_log(slot_id, sent_to)",
        # persistent request throttling and upload quota/cleanup lookups
        "CREATE INDEX IF NOT EXISTS idx_rate_limit_bucket ON rate_limit_events(scope, bucket_key, created_at)",
        "CREATE INDEX IF NOT EXISTS idx_rate_limit_time   ON rate_limit_events(created_at)",
        "CREATE INDEX IF NOT EXISTS idx_upload_user_time  ON uploaded_files(uploader_user_id, created_at)",
        "CREATE INDEX IF NOT EXISTS idx_upload_vol_time   ON uploaded_files(volunteer_id, created_at)",
        "CREATE INDEX IF NOT EXISTS idx_upload_claimed    ON uploaded_files(claimed_at, created_at)",
        "CREATE INDEX IF NOT EXISTS idx_account_invites_user ON account_invitations(user_id, created_at)",
        "CREATE INDEX IF NOT EXISTS idx_account_invites_expiry ON account_invitations(expires_at)",
        "CREATE INDEX IF NOT EXISTS idx_account_events_user ON account_access_events(user_id, created_at)",
    ]
    for _idx_sql in _performance_indexes:
        try:
            conn.execute(_idx_sql)
        except Exception as _e:
            log.warning(f'Index creation skipped: {_e}')

    # Seed default admin — INSERT OR IGNORE is atomic, safe under concurrent gunicorn workers.
    # Password is read from ADMIN_PASSWORD env var (set in Railway dashboard).
    # In production (Railway) a missing ADMIN_PASSWORD is a fatal misconfiguration —
    # refuse to start rather than silently seeding the well-known 'admin123'.
    admin_pw = os.environ.get('ADMIN_PASSWORD')
    if not admin_pw:
        if os.environ.get('RAILWAY_ENVIRONMENT') or os.environ.get('RAILWAY_PROJECT_ID'):
            raise RuntimeError(
                'ADMIN_PASSWORD env var is not set — refusing to start in production. '
                'Set it in the Railway dashboard (Variables) and redeploy.'
            )
        admin_pw = 'admin123'  # dev/local/test fallback only
    conn.execute(
        "INSERT OR IGNORE INTO users (id, username, password_hash, name, role, created_at) VALUES (?,?,?,?,?,?)",
        (str(uuid.uuid4()), 'admin', generate_password_hash(admin_pw),
         'Administrator', 'admin', now())
    )
    # ADMIN_PASSWORD remains a break-glass recovery credential. Track the last
    # applied environment value separately so an in-app admin reset survives
    # ordinary restarts, while an actual Railway variable change still resets it.
    if os.environ.get('ADMIN_PASSWORD'):
        _sync_admin_password_from_env(conn, admin_pw)
    else:
        log.warning('Admin password is default admin123 — set ADMIN_PASSWORD env var in Railway!')

    # Seed bundle size rules if not present
    if not conn.execute("SELECT id FROM bundle_size_rules LIMIT 1").fetchone():
        rules = [
            (str(uuid.uuid4()), 'S', 'Small',  1, 2),
            (str(uuid.uuid4()), 'M', 'Medium', 3, 5),
            (str(uuid.uuid4()), 'L', 'Large',  6, None),
        ]
        conn.executemany(
            "INSERT INTO bundle_size_rules (id, bundle_size, label, min_household, max_household) VALUES (?,?,?,?,?)",
            rules
        )
        log.info('Bundle size rules seeded.')

    # Seed food catalog if not present
    if not conn.execute("SELECT id FROM food_categories LIMIT 1").fetchone():
        ts = now()
        # Categories
        cats = [
            (str(uuid.uuid4()), 'Grains',  1, ts),
            (str(uuid.uuid4()), 'Protein', 2, ts),
            (str(uuid.uuid4()), 'Produce', 3, ts),
        ]
        conn.executemany(
            "INSERT INTO food_categories (id, name, display_order, is_active, created_at) VALUES (?,?,?,1,?)",
            cats
        )
        grains_id, protein_id, produce_id = cats[0][0], cats[1][0], cats[2][0]

        # Items: (id, category_id, name, unit, display_order)
        items_data = [
            (str(uuid.uuid4()), grains_id,  'Rice',          'lb',     1),
            (str(uuid.uuid4()), grains_id,  'Pasta',         'packet', 2),
            (str(uuid.uuid4()), grains_id,  'Bread',         'loaf',   3),
            (str(uuid.uuid4()), protein_id, 'Eggs',          'dozen',  1),
            (str(uuid.uuid4()), protein_id, 'Canned Beans',  'can',    2),
            (str(uuid.uuid4()), protein_id, 'Whole Chicken', 'each',   3),
            (str(uuid.uuid4()), protein_id, 'Brown Lentils', 'lb',     4),
            (str(uuid.uuid4()), produce_id, 'Potatoes',      'lb',     1),
            (str(uuid.uuid4()), produce_id, 'Oranges',       'bag',    2),
            (str(uuid.uuid4()), produce_id, 'Bananas',       'bunch',  3),
        ]
        conn.executemany(
            "INSERT INTO food_items (id, category_id, name, unit, is_active, display_order, created_at) VALUES (?,?,?,?,1,?,?)",
            [(i[0],i[1],i[2],i[3],i[4],ts) for i in items_data]
        )

        # Bundle quantities per item (S, M, L)
        qty_map = {
            'Rice':          [('S','2 lb'),  ('M','5 lb'),  ('L','10 lb')],
            'Pasta':         [('S','2'),     ('M','3'),     ('L','4')],
            'Bread':         [('S','1'),     ('M','2'),     ('L','3')],
            'Eggs':          [('S','2'),     ('M','3'),     ('L','4')],
            'Canned Beans':  [('S','4'),     ('M','6'),     ('L','8')],
            'Whole Chicken': [('S','1'),     ('M','2'),     ('L','3')],
            'Brown Lentils': [('S','1 lb'),  ('M','2 lb'),  ('L','3 lb')],
            'Potatoes':      [('S','5 lb'),  ('M','8 lb'),  ('L','10 lb')],
            'Oranges':       [('S','1'),     ('M','2'),     ('L','3')],
            'Bananas':       [('S','1'),     ('M','2'),     ('L','2')],
        }
        item_name_to_id = {i[2]: i[0] for i in items_data}
        bq_rows = []
        for name, qtys in qty_map.items():
            iid = item_name_to_id[name]
            for size, qty in qtys:
                bq_rows.append((str(uuid.uuid4()), iid, size, qty))
        conn.executemany(
            "INSERT INTO bundle_quantities (id, food_item_id, bundle_size, quantity) VALUES (?,?,?,?)",
            bq_rows
        )
        log.info('Food catalog seeded with default items and bundle quantities.')

    # ── Idempotent column migrations (alter existing tables safely) ───────────
    for _col, _def in [('wa_phone', 'TEXT'), ('wa_apikey', 'TEXT')]:
        try:
            conn.execute(f'ALTER TABLE volunteers ADD COLUMN {_col} {_def}')
        except sqlite3.OperationalError:
            pass  # Column already exists

    # Add completed_at to volunteer_slots (tracks when a task was marked done)
    try:
        conn.execute('ALTER TABLE volunteer_slots ADD COLUMN completed_at TEXT')
        log.info('Migration: added completed_at to volunteer_slots')
    except sqlite3.OperationalError:
        pass  # Already exists

    # Add family_code — human-readable reference (last 4 digits of phone + bundle size)
    try:
        conn.execute('ALTER TABLE families ADD COLUMN family_code TEXT')
        log.info('Migration: added family_code to families')
    except sqlite3.OperationalError:
        pass  # Already exists

    # Add columns that exist in CREATE TABLE but were missing from live DBs
    for _col, _def in [
        ('bundle_size',         'TEXT'),
        ('updated_at',          'TEXT'),
        ('pending_bundle_size', 'TEXT'),
        ('wa_phone',            'TEXT'),
        ('wa_apikey',           'TEXT'),
    ]:
        try:
            conn.execute(f'ALTER TABLE families ADD COLUMN {_col} {_def}')
            log.info(f'Migration: added families.{_col}')
        except sqlite3.OperationalError:
            pass  # Already exists

    # Ensure all donations columns exist (some only added inside route handlers, not here)
    for _col, _def in [
        ('donor_email',  'TEXT'),
        ('type',         'TEXT'),
        ('reference_id', 'TEXT'),
        ('cycle_id',     'TEXT'),
        ('frequency',    'TEXT'),
        ('source',       'TEXT'),
    ]:
        try:
            conn.execute(f'ALTER TABLE donations ADD COLUMN {_col} {_def}')
            log.info(f'Migration: added donations.{_col}')
        except sqlite3.OperationalError:
            pass

    # ── Phase 4A migrations ───────────────────────────────────────────────────

    # Ensure food_requests has all expected columns (safety net in case table recreation failed)
    for _col, _def in [
        ('confirmation_token',    'TEXT'),
        ('confirmation_expires_at', 'TEXT'),
        ('confirmed_at',          'TEXT'),
        ('confirmation_sent_at',  'TEXT'),
        ('updated_at',            'TEXT'),
        ('family_notes',          'TEXT'),
    ]:
        try:
            conn.execute(f'ALTER TABLE food_requests ADD COLUMN {_col} {_def}')
            log.info(f'Migration: added food_requests.{_col}')
        except sqlite3.OperationalError:
            pass  # Already exists

    # Add slot_id to receipts (links a portal-submitted receipt to a volunteer slot)
    try:
        conn.execute('ALTER TABLE receipts ADD COLUMN slot_id TEXT')
        log.info('Migration: added slot_id to receipts')
    except sqlite3.OperationalError:
        pass

    # ── Receipt vision-parsing columns (Phase A) ──────────────────────────────
    # parsed_* hold what the vision model extracted; the existing store/amount/
    # purchase_date stay the human-confirmed source of truth. amount_mismatch is a
    # review flag when the confirmed amount and parsed_total disagree. parse_status:
    # none|queued|parsing|parsed|failed. parsed_json keeps the raw model output.
    for _col, _def in [
        ('parsed_store',     'TEXT'),
        ('parsed_date',      'TEXT'),
        ('subtotal',         'REAL'),
        ('tax',              'REAL'),
        ('parsed_total',     'REAL'),
        ('parse_status',     "TEXT DEFAULT 'none'"),
        ('parse_confidence', 'REAL'),
        ('parse_model',      'TEXT'),
        ('parsed_at',        'TEXT'),
        ('parsed_json',      'TEXT'),
        ('amount_mismatch',  'INTEGER DEFAULT 0'),
        # reimbursable_amount = confirmed total minus any line items excluded from
        # reimbursement (a volunteer's personal charge on the same receipt). NULL means
        # "no exclusions yet" and callers treat it as equal to amount.
        ('reimbursable_amount', 'REAL'),
    ]:
        try:
            conn.execute(f'ALTER TABLE receipts ADD COLUMN {_col} {_def}')
            log.info(f'Migration: added receipts.{_col}')
        except sqlite3.OperationalError:
            pass

    # Line-item breakdown extracted from each receipt (audit: itemized tracking).
    conn.execute('''
        CREATE TABLE IF NOT EXISTS receipt_items (
            id          TEXT PRIMARY KEY,
            receipt_id  TEXT NOT NULL,
            line_no     INTEGER,
            name        TEXT,
            qty         REAL,
            unit_price  REAL,
            line_total  REAL,
            category    TEXT,
            created_at  TEXT NOT NULL,
            FOREIGN KEY (receipt_id) REFERENCES receipts(id) ON DELETE CASCADE
        )
    ''')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_receipt_items_receipt ON receipt_items(receipt_id)')
    # excluded=1 removes this line's line_total from the receipt's reimbursable amount.
    try:
        conn.execute('ALTER TABLE receipt_items ADD COLUMN excluded INTEGER DEFAULT 0')
        log.info('Migration: added receipt_items.excluded')
    except sqlite3.OperationalError:
        pass

    # Operating expenses — organizational overhead paid by the charity directly
    # (web hosting, supplies, admin fees), separate from volunteer food reimbursements.
    # status: pending (committed, not yet paid) | paid (money out). Both hit the ledger.
    conn.execute('''
        CREATE TABLE IF NOT EXISTS operating_expenses (
            id             TEXT PRIMARY KEY,
            expense_date   TEXT,
            category       TEXT,
            vendor         TEXT,
            description    TEXT,
            amount         REAL NOT NULL,
            payment_method TEXT,
            payment_ref    TEXT,
            status         TEXT NOT NULL DEFAULT 'paid',
            paid_date      TEXT,
            created_by     TEXT,
            created_at     TEXT NOT NULL,
            updated_at     TEXT
        )
    ''')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_opex_status ON operating_expenses(status)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_opex_date   ON operating_expenses(expense_date)')

    # Add email/wa_phone/wa_apikey to users (treasurer notification channels)
    for _col, _def in [('email', 'TEXT'), ('wa_phone', 'TEXT'), ('wa_apikey', 'TEXT')]:
        try:
            conn.execute(f'ALTER TABLE users ADD COLUMN {_col} {_def}')
            log.info(f'Migration: added {_col} to users')
        except sqlite3.OperationalError:
            pass

    # Add new auth columns to users (unified password login system)
    for _col, _def in [
        ('linked_id',            'TEXT'),
        ('linked_type',          'TEXT'),
        ('must_change_password', 'INTEGER NOT NULL DEFAULT 1'),
        ('password_changed_at',  'TEXT'),
        ('last_login_at',        'TEXT'),
    ]:
        try:
            conn.execute(f'ALTER TABLE users ADD COLUMN {_col} {_def}')
            log.info(f'Migration: added users.{_col}')
        except sqlite3.OperationalError:
            pass  # Already exists

    # Existing admin user should NOT be forced to change password on first deploy
    conn.execute(
        "UPDATE users SET must_change_password=0 WHERE username='admin' AND must_change_password=1 AND password_changed_at IS NULL"
    )

    # Migrate users table: add treasurer + family roles + new auth columns
    # (SQLite CHECK constraints require table recreation to modify)
    users_sql = conn.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name='users'"
    ).fetchone()
    if users_sql and ('family' not in users_sql[0] or 'treasurer' not in users_sql[0]):
        log.info('Migration: upgrading users table for family/treasurer roles + new auth columns')
        try:
            conn.execute('PRAGMA foreign_keys=OFF')
            conn.executescript('''
                CREATE TABLE IF NOT EXISTS users_new (
                    id                   TEXT PRIMARY KEY,
                    username             TEXT UNIQUE NOT NULL,
                    password_hash        TEXT NOT NULL,
                    name                 TEXT,
                    role                 TEXT NOT NULL DEFAULT 'viewer'
                                         CHECK(role IN ('admin','volunteer','finance','treasurer','viewer','family')),
                    email                TEXT,
                    wa_phone             TEXT,
                    wa_apikey            TEXT,
                    active               INTEGER NOT NULL DEFAULT 1,
                    linked_id            TEXT,
                    linked_type          TEXT,
                    must_change_password INTEGER NOT NULL DEFAULT 1,
                    password_changed_at  TEXT,
                    last_login_at        TEXT,
                    created_at           TEXT NOT NULL
                );
                INSERT OR IGNORE INTO users_new
                    (id, username, password_hash, name, role, email, wa_phone, wa_apikey, active, created_at)
                SELECT id, username, password_hash, name, role, email, wa_phone, wa_apikey, active, created_at
                FROM users;
                DROP TABLE IF EXISTS users;
                ALTER TABLE users_new RENAME TO users;
            ''')
            conn.execute('PRAGMA foreign_keys=ON')
            log.info('Migration: users table upgraded — family/treasurer roles + auth columns added')
        except Exception as _e:
            log.info(f'Migration: users table already upgraded or in progress — skipping ({_e})')
            conn.execute('PRAGMA foreign_keys=ON')

    # Migrate reimbursements table: add venmo/zelle payment methods + payment_ref column
    reimb_sql = conn.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name='reimbursements'"
    ).fetchone()
    if reimb_sql and 'venmo' not in reimb_sql[0]:
        log.info('Migration: upgrading reimbursements table for new payment methods')
        try:
            conn.executescript('''
                CREATE TABLE IF NOT EXISTS reimbursements_new (
                    id              TEXT PRIMARY KEY,
                    receipt_id      TEXT NOT NULL,
                    volunteer_id    TEXT,
                    amount          REAL,
                    status          TEXT NOT NULL DEFAULT 'pending'
                                    CHECK(status IN ('pending','approved','paid','rejected')),
                    payment_method  TEXT CHECK(payment_method IN ('venmo','zelle','check','cash','bank_transfer','cheque','other')),
                    payment_ref     TEXT,
                    paid_date       TEXT,
                    approved_by     TEXT,
                    notes           TEXT,
                    created_at      TEXT NOT NULL,
                    updated_at      TEXT,
                    FOREIGN KEY (receipt_id) REFERENCES receipts(id)
                );
                INSERT OR IGNORE INTO reimbursements_new
                    (id, receipt_id, volunteer_id, amount, status, payment_method,
                     payment_ref, paid_date, approved_by, notes, created_at, updated_at)
                SELECT id, receipt_id, volunteer_id, amount, status, payment_method,
                       NULL, paid_date, approved_by, notes, created_at, updated_at
                FROM reimbursements;
                DROP TABLE IF EXISTS reimbursements;
                ALTER TABLE reimbursements_new RENAME TO reimbursements;
            ''')
            log.info('Migration: reimbursements table upgraded — venmo/zelle/payment_ref added')
        except Exception as _e:
            # Another worker already ran this migration — safe to skip
            log.info(f'Migration: reimbursements table already upgraded or in progress — skipping ({_e})')

    # ── Phase 6 migrations: phone normalisation + bundle size request ─────────────

    # Add pending_bundle_size to families (coordinator must approve before it takes effect)
    try:
        conn.execute('ALTER TABLE families ADD COLUMN pending_bundle_size TEXT')
        log.info('Migration: added pending_bundle_size to families')
    except sqlite3.OperationalError:
        pass

    # Normalise all existing phone numbers — strip hyphens/spaces/parens to digits only
    rows = conn.execute("SELECT id, phone FROM families WHERE phone IS NOT NULL").fetchall()
    updated = 0
    for row in rows:
        normalized = ''.join(c for c in row[1] if c.isdigit())
        if normalized != row[1]:
            conn.execute("UPDATE families SET phone=? WHERE id=?", (normalized, row[0]))
            updated += 1
    if updated:
        conn.commit()
        log.info(f'Migration: normalised {updated} family phone numbers')

    # Same for volunteers
    vrows = conn.execute("SELECT id, phone FROM volunteers WHERE phone IS NOT NULL").fetchall()
    vupdated = 0
    for row in vrows:
        normalized = ''.join(c for c in row[1] if c.isdigit())
        if normalized != row[1]:
            conn.execute("UPDATE volunteers SET phone=? WHERE id=?", (normalized, row[0]))
            vupdated += 1
    if vupdated:
        conn.commit()
        log.info(f'Migration: normalised {vupdated} volunteer phone numbers')

    # ── Phase 5 migrations: family phone fields + food_request confirmation ───────

    # Add wa_phone / wa_apikey / email to families
    for _col in ['wa_phone', 'wa_apikey', 'email']:
        try:
            conn.execute(f'ALTER TABLE families ADD COLUMN {_col} TEXT')
            log.info(f'Migration: added {_col} to families')
        except sqlite3.OperationalError:
            pass

    # Add confirmation fields to food_requests
    for _col, _def in [('confirmation_token', 'TEXT'),
                       ('confirmation_expires_at', 'TEXT'),
                       ('confirmed_at', 'TEXT'),
                       ('confirmation_sent_at', 'TEXT')]:
        try:
            conn.execute(f'ALTER TABLE food_requests ADD COLUMN {_col} {_def}')
            log.info(f'Migration: added {_col} to food_requests')
        except sqlite3.OperationalError:
            pass

    # Upgrade delivery_cycles status CHECK to include 'upcoming'
    cycles_sql = conn.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name='delivery_cycles'"
    ).fetchone()
    if cycles_sql and 'upcoming' not in cycles_sql[0]:
        log.info('Migration: upgrading delivery_cycles CHECK to include upcoming status')
        try:
            conn.execute('PRAGMA foreign_keys=OFF')
            # Clean up any leftover _new table from a previous failed attempt
            conn.execute('DROP TABLE IF EXISTS delivery_cycles_new')
            conn.execute('''
                CREATE TABLE delivery_cycles_new (
                    id                  TEXT PRIMARY KEY,
                    title               TEXT NOT NULL,
                    delivery_date_start TEXT NOT NULL,
                    delivery_date_end   TEXT NOT NULL,
                    request_open_at     TEXT NOT NULL DEFAULT '',
                    request_close_at    TEXT NOT NULL DEFAULT '',
                    status              TEXT NOT NULL DEFAULT 'upcoming'
                                        CHECK(status IN ('draft','open','closed','upcoming','shopping','delivered')),
                    notes               TEXT,
                    created_by          TEXT,
                    created_at          TEXT NOT NULL,
                    updated_at          TEXT
                )''')
            # Explicit column list so old rows (without updated_at) copy correctly
            conn.execute('''
                INSERT INTO delivery_cycles_new
                    (id, title, delivery_date_start, delivery_date_end,
                     request_open_at, request_close_at, status, notes, created_by, created_at)
                SELECT id, title, delivery_date_start, delivery_date_end,
                       COALESCE(request_open_at,''), COALESCE(request_close_at,''),
                       status, notes, created_by, created_at
                FROM delivery_cycles''')
            conn.execute('DROP TABLE delivery_cycles')
            conn.execute('ALTER TABLE delivery_cycles_new RENAME TO delivery_cycles')
            conn.execute("UPDATE delivery_cycles SET status='upcoming' WHERE status IN ('draft','open','closed')")
            conn.commit()
            conn.execute('PRAGMA foreign_keys=ON')
            log.info('Migration: delivery_cycles upgraded — upcoming status added')
        except Exception as _e:
            conn.execute('PRAGMA foreign_keys=ON')
            log.info(f'Migration: delivery_cycles upgrade failed — {_e}')

    # Upgrade food_requests status CHECK to include confirmation statuses
    freq_sql = conn.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name='food_requests'"
    ).fetchone()
    if freq_sql and 'pending_confirmation' not in freq_sql[0]:
        log.info('Migration: upgrading food_requests for confirmation statuses')
        try:
            conn.execute('PRAGMA foreign_keys=OFF')
            conn.executescript('''
                CREATE TABLE IF NOT EXISTS food_requests_new (
                    id                   TEXT PRIMARY KEY,
                    cycle_id             TEXT NOT NULL,
                    family_id            TEXT NOT NULL,
                    bundle_size          TEXT NOT NULL CHECK(bundle_size IN ('S','M','L')),
                    submitted_at         TEXT NOT NULL,
                    status               TEXT NOT NULL DEFAULT 'pending_confirmation'
                                         CHECK(status IN ('submitted','assigned','delivered','cancelled',
                                                          'pending_confirmation','confirmed','skipped','auto_confirmed')),
                    assigned_volunteer_id TEXT,
                    delivered_at         TEXT,
                    notes                TEXT,
                    confirmation_token   TEXT,
                    confirmation_expires_at TEXT,
                    confirmed_at         TEXT,
                    confirmation_sent_at TEXT,
                    updated_at           TEXT,
                    family_notes         TEXT,
                    UNIQUE(cycle_id, family_id),
                    FOREIGN KEY (cycle_id)  REFERENCES delivery_cycles(id),
                    FOREIGN KEY (family_id) REFERENCES families(id)
                );
                INSERT OR IGNORE INTO food_requests_new
                    (id, cycle_id, family_id, bundle_size, submitted_at, status,
                     assigned_volunteer_id, delivered_at, notes,
                     confirmation_token, confirmation_expires_at,
                     confirmed_at, confirmation_sent_at,
                     updated_at, family_notes)
                SELECT id, cycle_id, family_id, bundle_size, submitted_at, status,
                       assigned_volunteer_id, delivered_at, notes,
                       confirmation_token, confirmation_expires_at,
                       confirmed_at, confirmation_sent_at,
                       NULL, NULL
                FROM food_requests;
                DROP TABLE IF EXISTS food_requests;
                ALTER TABLE food_requests_new RENAME TO food_requests;
            ''')
            conn.execute('PRAGMA foreign_keys=ON')
            log.info('Migration: food_requests upgraded — confirmation statuses added')
        except Exception as _e:
            conn.execute('PRAGMA foreign_keys=ON')
            log.info(f'Migration: food_requests already upgraded — skipping ({_e})')

    conn.commit()

    # Give still-pending legacy confirmation links a short migration grace
    # period. Links for already processed orders remain unusable because the
    # public confirmation route accepts pending_confirmation only.
    conn.execute(
        """UPDATE food_requests
           SET confirmation_expires_at=datetime('now', '+24 hours')
           WHERE confirmation_token IS NOT NULL
             AND confirmation_expires_at IS NULL
             AND status='pending_confirmation'"""
    )

    # Back-fill family_code for existing families that don't have one
    existing = conn.execute(
        "SELECT id, phone, family_size FROM families WHERE family_code IS NULL OR family_code=''"
    ).fetchall()
    for row in existing:
        code = _make_family_code(row[1], row[2], db_conn=conn, exclude_id=row[0])
        conn.execute("UPDATE families SET family_code=? WHERE id=?", (code, row[0]))
    if existing:
        log.info(f'Back-filled family_code for {len(existing)} existing families')

    # ── Phase 3C tables ───────────────────────────────────────────────────────
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS volunteer_slots (
            id           TEXT PRIMARY KEY,
            cycle_id     TEXT NOT NULL,
            family_id    TEXT NOT NULL,
            task_type    TEXT NOT NULL,
            task_date    TEXT,
            claimed_by   TEXT,
            claimed_at   TEXT,
            completed_at TEXT,
            status       TEXT NOT NULL DEFAULT 'claimed'
                         CHECK(status IN ('open','claimed','complete','cancelled')),
            notes        TEXT,
            created_at   TEXT NOT NULL,
            updated_at   TEXT,
            FOREIGN KEY (cycle_id)   REFERENCES delivery_cycles(id),
            FOREIGN KEY (family_id)  REFERENCES families(id),
            FOREIGN KEY (claimed_by) REFERENCES volunteers(id)
        );

        CREATE TABLE IF NOT EXISTS volunteer_task_types (
            slug          TEXT PRIMARY KEY,
            label         TEXT NOT NULL,
            display_order INTEGER NOT NULL DEFAULT 0,
            is_active     INTEGER NOT NULL DEFAULT 1
        );

        CREATE TABLE IF NOT EXISTS portal_sessions (
            token        TEXT PRIMARY KEY,
            volunteer_id TEXT NOT NULL,
            expires_at   TEXT NOT NULL,
            created_at   TEXT NOT NULL,
            FOREIGN KEY (volunteer_id) REFERENCES volunteers(id)
        );

        CREATE TABLE IF NOT EXISTS otp_tokens (
            id         TEXT PRIMARY KEY,
            phone      TEXT NOT NULL,
            code       TEXT NOT NULL,
            type       TEXT NOT NULL DEFAULT 'volunteer',
            expires_at TEXT NOT NULL,
            used       INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS reminder_log (
            id       TEXT PRIMARY KEY,
            slot_id  TEXT NOT NULL,
            sent_to  TEXT NOT NULL,
            sent_at  TEXT NOT NULL,
            UNIQUE(slot_id, sent_to)
        );

        CREATE TABLE IF NOT EXISTS food_request_events (
            id          TEXT PRIMARY KEY,
            request_id  TEXT NOT NULL,
            event_type  TEXT NOT NULL,
            actor       TEXT NOT NULL DEFAULT 'system',
            payload     TEXT NOT NULL DEFAULT '{}',
            created_at  TEXT NOT NULL,
            FOREIGN KEY (request_id) REFERENCES food_requests(id)
        );
    ''')

    # ── Migration #19: backfill synthetic events for existing orders ──────────
    existing_events = conn.execute("SELECT COUNT(*) FROM food_request_events").fetchone()[0]
    if existing_events == 0:
        log.info('Migration #19: backfilling food_request_events for existing orders...')
        import json as _json
        backfill_rows = conn.execute(
            '''SELECT id, status, confirmed_at, updated_at, submitted_at FROM food_requests'''
        ).fetchall()
        _bf_count = 0
        for _r in backfill_rows:
            _etype = None
            _ts    = None
            if _r['status'] in ('confirmed', 'auto_confirmed'):
                _etype = 'confirmed'
                _ts    = _r['confirmed_at'] or _r['submitted_at'] or datetime.utcnow().isoformat()
            elif _r['status'] == 'submitted':
                _etype = 'confirmed'
                _ts    = _r['submitted_at'] or datetime.utcnow().isoformat()
            elif _r['status'] == 'skipped':
                _etype = 'auto_skipped'
                _ts    = _r['updated_at'] or _r['submitted_at'] or datetime.utcnow().isoformat()
            elif _r['status'] == 'cancelled':
                _etype = 'cancelled'
                _ts    = _r['updated_at'] or _r['submitted_at'] or datetime.utcnow().isoformat()
            elif _r['status'] == 'delivered':
                _etype = 'confirmed'
                _ts    = _r['confirmed_at'] or _r['submitted_at'] or datetime.utcnow().isoformat()
            if _etype:
                conn.execute(
                    "INSERT OR IGNORE INTO food_request_events (id, request_id, event_type, actor, payload, created_at) VALUES (?,?,?,?,?,?)",
                    (str(uuid.uuid4()), _r['id'], _etype, 'system',
                     _json.dumps({'note': 'backfilled'}), _ts)
                )
                _bf_count += 1
        conn.commit()
        log.info(f'Migration #19: backfilled {_bf_count} events for existing orders')

    # ── Seed default task types (idempotent) ─────────────────────────────────
    for _slug, _label, _order in [('shopping', 'Shop', 1), ('delivery', 'Delivery', 2), ('stock', 'Stock', 3)]:
        conn.execute(
            "INSERT OR IGNORE INTO volunteer_task_types (slug, label, display_order, is_active) VALUES (?,?,?,1)",
            (_slug, _label, _order)
        )

    # ── Migration: add is_family_slot to volunteer_task_types ─────────────────
    try:
        conn.execute('ALTER TABLE volunteer_task_types ADD COLUMN is_family_slot INTEGER NOT NULL DEFAULT 0')
        log.info('Migration: added volunteer_task_types.is_family_slot')
    except Exception:
        pass
    # Backfill: shopping + delivery are family-specific; stock is not
    conn.execute("UPDATE volunteer_task_types SET is_family_slot=1 WHERE slug IN ('shopping','delivery')")
    conn.execute("UPDATE volunteer_task_types SET is_family_slot=0 WHERE slug='stock'")

    # ── Migration: rebuild volunteer_slots — remove UNIQUE + CHECK constraints ─
    _vs_ddl = conn.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name='volunteer_slots'"
    ).fetchone()
    if _vs_ddl and (
        "UNIQUE(cycle_id, family_id, task_type)" in (_vs_ddl[0] or "") or
        "CHECK(task_type IN" in (_vs_ddl[0] or "")
    ):
        log.info('Migration: rebuilding volunteer_slots — removing UNIQUE + CHECK constraints')
        try:
            conn.execute('PRAGMA foreign_keys=OFF')
            conn.executescript('''
                ALTER TABLE volunteer_slots RENAME TO _vs_backup;
                CREATE TABLE volunteer_slots (
                    id           TEXT PRIMARY KEY,
                    cycle_id     TEXT NOT NULL,
                    family_id    TEXT NOT NULL,
                    task_type    TEXT NOT NULL,
                    task_date    TEXT,
                    claimed_by   TEXT,
                    claimed_at   TEXT,
                    completed_at TEXT,
                    status       TEXT NOT NULL DEFAULT 'claimed'
                                 CHECK(status IN ('open','claimed','complete','cancelled')),
                    notes        TEXT,
                    created_at   TEXT NOT NULL,
                    updated_at   TEXT,
                    FOREIGN KEY (cycle_id)   REFERENCES delivery_cycles(id),
                    FOREIGN KEY (family_id)  REFERENCES families(id),
                    FOREIGN KEY (claimed_by) REFERENCES volunteers(id)
                );
                INSERT INTO volunteer_slots SELECT * FROM _vs_backup;
                DROP TABLE _vs_backup;
            ''')
            conn.execute('PRAGMA foreign_keys=ON')
            log.info('Migration: volunteer_slots rebuilt — multi-volunteer per task now supported')
        except Exception as _e:
            conn.execute('PRAGMA foreign_keys=ON')
            log.info(f'Migration: volunteer_slots rebuild skipped ({_e})')

    # ── Migration: add prev_claimed_by to volunteer_slots ────────────────────────
    try:
        conn.execute('ALTER TABLE volunteer_slots ADD COLUMN prev_claimed_by TEXT')
        log.info('Migration: added volunteer_slots.prev_claimed_by')
    except Exception:
        pass

    # ── Migration: volunteer_slots — expand status CHECK to include 'confirmed' ─
    _vs_ddl2 = conn.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name='volunteer_slots'"
    ).fetchone()
    if _vs_ddl2 and "CHECK(status IN ('open','claimed','complete','cancelled'))" in (_vs_ddl2[0] or ''):
        log.info("Migration: rebuilding volunteer_slots — adding 'confirmed' status")
        try:
            conn.execute('PRAGMA foreign_keys=OFF')
            conn.executescript('''
                ALTER TABLE volunteer_slots RENAME TO _vs_backup2;
                CREATE TABLE volunteer_slots (
                    id              TEXT PRIMARY KEY,
                    cycle_id        TEXT NOT NULL,
                    family_id       TEXT NOT NULL,
                    task_type       TEXT NOT NULL,
                    task_date       TEXT,
                    claimed_by      TEXT,
                    claimed_at      TEXT,
                    completed_at    TEXT,
                    status          TEXT NOT NULL DEFAULT 'open',
                    notes           TEXT,
                    created_at      TEXT NOT NULL,
                    updated_at      TEXT,
                    prev_claimed_by TEXT,
                    FOREIGN KEY (cycle_id)   REFERENCES delivery_cycles(id),
                    FOREIGN KEY (family_id)  REFERENCES families(id),
                    FOREIGN KEY (claimed_by) REFERENCES volunteers(id)
                );
                INSERT INTO volunteer_slots
                    SELECT id, cycle_id, family_id, task_type, task_date,
                           claimed_by, claimed_at, completed_at, status, notes,
                           created_at, updated_at, prev_claimed_by
                    FROM _vs_backup2;
                DROP TABLE _vs_backup2;
            ''')
            conn.execute('PRAGMA foreign_keys=ON')
            log.info("Migration: volunteer_slots rebuilt — 'confirmed' status now supported")
        except Exception as _e:
            conn.execute('PRAGMA foreign_keys=ON')
            log.info(f'Migration: volunteer_slots confirm-rebuild skipped ({_e})')

    # ── order_change_requests table ──────────────────────────────────────────────
    conn.execute('''
        CREATE TABLE IF NOT EXISTS order_change_requests (
            id           TEXT PRIMARY KEY,
            family_id    TEXT NOT NULL,
            cycle_id     TEXT NOT NULL,
            request_id   TEXT,
            status       TEXT NOT NULL DEFAULT 'pending'
                         CHECK(status IN ('pending','approved','rejected','retracted')),
            family_notes TEXT,
            payload      TEXT NOT NULL DEFAULT '{}',
            admin_notes  TEXT,
            reviewed_by  TEXT,
            created_at   TEXT NOT NULL,
            reviewed_at  TEXT,
            FOREIGN KEY (family_id) REFERENCES families(id),
            FOREIGN KEY (cycle_id)  REFERENCES delivery_cycles(id)
        )
    ''')

    # ── Migration: priced bundle selection ───────────────────────────────────────
    # food_items: price per unit + allow_qty flag
    for _col, _def in [('price', 'REAL NOT NULL DEFAULT 0'),
                       ('allow_qty', 'INTEGER NOT NULL DEFAULT 0')]:
        try:
            conn.execute(f'ALTER TABLE food_items ADD COLUMN {_col} {_def}')
            log.info(f'Migration: added food_items.{_col}')
        except Exception:
            pass

    # bundle_size_rules: budget per bundle size
    try:
        conn.execute('ALTER TABLE bundle_size_rules ADD COLUMN budget REAL NOT NULL DEFAULT 0')
        log.info('Migration: added bundle_size_rules.budget')
    except Exception:
        pass

    # food_request_items: quantity per selected item
    try:
        conn.execute('ALTER TABLE food_request_items ADD COLUMN quantity INTEGER NOT NULL DEFAULT 1')
        log.info('Migration: added food_request_items.quantity')
    except Exception:
        pass

    # food_request_items: custom_value for free-text items (e.g. "Other Fruit")
    try:
        conn.execute('ALTER TABLE food_request_items ADD COLUMN custom_value TEXT')
        log.info('Migration: added food_request_items.custom_value')
    except Exception:
        pass

    # food_items: item selection rules (defaults, mutual-exclusion groups, free-text)
    for _col, _def in [
        ('is_default',   'INTEGER NOT NULL DEFAULT 0'),
        ('group_id',     'TEXT'),
        ('group_max',    'INTEGER NOT NULL DEFAULT 1'),
        ('is_free_text', 'INTEGER NOT NULL DEFAULT 0'),
    ]:
        try:
            conn.execute(f'ALTER TABLE food_items ADD COLUMN {_col} {_def}')
            log.info(f'Migration: added food_items.{_col}')
        except Exception:
            pass

    # ── One-time catalog pricing migration (Sam's Club prices) ──────────────────
    # Updates existing seeded items: name, unit, price, allow_qty (only if price=0)
    _catalog_updates = [
        # (old_name,        new_name,            unit,          price,  allow_qty)
        ('Rice',          'Rice',             '4 lb bag',    10.00, 0),
        ('Pasta',         'Pasta',            'box',          8.52, 0),
        ('Eggs',          'Eggs',             '18 ct',        5.87, 0),
        ('Canned Beans',  'Red Kidney Beans', 'lb bag',       5.00, 0),
        ('Whole Chicken', 'Whole Chicken',    'lb',           5.00, 1),
        ('Potatoes',      'Red Potato',       'bag',          4.92, 0),
        ('Bananas',       'Bananas',          'bunch',        2.16, 0),
    ]
    for _old, _new, _unit, _price, _allow_qty in _catalog_updates:
        _row = conn.execute("SELECT id, price FROM food_items WHERE name=?", (_old,)).fetchone()
        if _row and (not _row['price'] or _row['price'] == 0):
            conn.execute(
                "UPDATE food_items SET name=?, unit=?, price=?, allow_qty=? WHERE id=?",
                (_new, _unit, _price, _allow_qty, _row['id'])
            )
            log.info(f'Catalog migration: updated {_old!r} → {_new!r} @ ${_price}')

    # Add missing items (Apples, Red Onion, Italian Dressing, Grapes) to Produce
    _produce_cat = conn.execute("SELECT id FROM food_categories WHERE name='Produce'").fetchone()
    if _produce_cat:
        _produce_id = _produce_cat['id']
        _missing = [
            ('Apples',           'bag',       6.22, 0, 4),
            ('Red Onion',        'each',      5.82, 0, 5),
            ('Italian Dressing', 'bottle',    5.14, 0, 6),
            ('Grapes',           '4 lb box',  7.00, 0, 7),
        ]
        for _name, _unit, _price, _aq, _order in _missing:
            if not conn.execute("SELECT id FROM food_items WHERE name=?", (_name,)).fetchone():
                _nid = str(uuid.uuid4())
                conn.execute(
                    "INSERT INTO food_items (id, category_id, name, unit, is_active, display_order, created_at, price, allow_qty) "
                    "VALUES (?,?,?,?,1,?,?,?,?)",
                    (_nid, _produce_id, _name, _unit, _order, now(), _price, _aq)
                )
                for _sz in ('S', 'M', 'L'):
                    conn.execute(
                        "INSERT OR IGNORE INTO bundle_quantities (id, food_item_id, bundle_size, quantity) VALUES (?,?,?,0)",
                        (str(uuid.uuid4()), _nid, _sz)
                    )
                log.info(f'Catalog migration: added {_name!r} @ ${_price}')

    # Deactivate placeholder items not in approved catalog (only if still unpriced)
    for _inactive in ('Oranges',):
        conn.execute("UPDATE food_items SET is_active=0 WHERE name=? AND (price=0 OR price IS NULL)", (_inactive,))

    # Set S/M/L bundle budgets if still at default 0
    _budgets = {'S': 30.0, 'M': 50.0, 'L': 80.0}
    for _sz, _bud in _budgets.items():
        _br = conn.execute("SELECT budget FROM bundle_size_rules WHERE bundle_size=?", (_sz,)).fetchone()
        if _br and (not _br['budget'] or _br['budget'] == 0):
            conn.execute("UPDATE bundle_size_rules SET budget=? WHERE bundle_size=?", (_bud, _sz))
            log.info(f'Bundle budget set: {_sz}=${_bud}')

    # ── Migration: item selection rules (defaults + mutual-exclusion groups) ──────
    # Only runs once — when all items still have is_default=0 (fresh migration)
    _rules_seeded = conn.execute(
        "SELECT COUNT(*) FROM food_items WHERE is_default=1"
    ).fetchone()[0]
    if not _rules_seeded:
        log.info('Seeding item selection rules…')
        # Per-item: (name, is_default, group_id, is_free_text)
        _item_rules = [
            ('Rice',             1, None,          0),
            ('Pasta',            0, 'bread_pasta',  0),
            ('Eggs',             1, None,          0),
            ('Red Kidney Beans', 0, 'beans',        0),
            ('Whole Chicken',    1, None,          0),
            ('Red Potato',       1, None,          0),
            ('Bananas',          1, None,          0),
            ('Apples',           0, 'fruit',        0),
            ('Red Onion',        1, None,          0),
            ('Grapes',           0, 'fruit',        0),
            ('Brown Lentils',    0, None,          0),
            ('Bread',            1, 'bread_pasta',  0),
            ('Italian Dressing', 0, None,          0),
        ]
        for _nm, _def, _grp, _ft in _item_rules:
            conn.execute(
                "UPDATE food_items SET is_default=?, group_id=?, group_max=1, is_free_text=? WHERE name=?",
                (_def, _grp, _ft, _nm)
            )
            log.info(f'Item rules: {_nm!r} default={_def} group={_grp!r}')

        # Bread: reactivate with price (was deactivated as placeholder)
        _bread = conn.execute("SELECT id, price, is_active FROM food_items WHERE name='Bread'").fetchone()
        if _bread:
            conn.execute(
                "UPDATE food_items SET is_active=1, price=3.50, is_default=1, group_id='bread_pasta', group_max=1 WHERE id=?",
                (_bread['id'],)
            )
            log.info('Catalog: reactivated Bread @ $3.50')
        else:
            # Add Bread if missing — find Staples/Pantry/Grains category
            _bc = conn.execute(
                "SELECT id FROM food_categories WHERE LOWER(name) LIKE '%staple%' OR LOWER(name) LIKE '%pantry%' OR LOWER(name) LIKE '%grain%'"
            ).fetchone() or conn.execute("SELECT id FROM food_categories LIMIT 1").fetchone()
            if _bc:
                _bid = str(uuid.uuid4())
                conn.execute(
                    "INSERT INTO food_items (id,category_id,name,unit,is_active,display_order,created_at,price,allow_qty,is_default,group_id,group_max,is_free_text) "
                    "VALUES (?,?,'Bread','loaf',1,3,?,3.50,0,1,'bread_pasta',1,0)",
                    (_bid, _bc['id'], now())
                )
                for _sz in ('S','M','L'):
                    conn.execute("INSERT OR IGNORE INTO bundle_quantities (id,food_item_id,bundle_size,quantity) VALUES (?,?,?,0)",
                                 (str(uuid.uuid4()), _bid, _sz))
                log.info('Catalog: added Bread @ $3.50')

        # Brown Lentils: reactivate (was deactivated as placeholder)
        _lentils = conn.execute("SELECT id FROM food_items WHERE name='Brown Lentils'").fetchone()
        if _lentils:
            conn.execute(
                "UPDATE food_items SET is_active=1, price=4.00, is_default=0, group_id=NULL, is_free_text=0 WHERE id=?",
                (_lentils['id'],)
            )
            log.info('Catalog: reactivated Brown Lentils @ $4.00')
        else:
            _pc = conn.execute("SELECT id FROM food_categories WHERE name='Produce'").fetchone()
            if _pc:
                _lid = str(uuid.uuid4())
                conn.execute(
                    "INSERT INTO food_items (id,category_id,name,unit,is_active,display_order,created_at,price,allow_qty,is_default,group_id,group_max,is_free_text) "
                    "VALUES (?,?,'Brown Lentils','bag',1,10,?,4.00,0,0,NULL,1,0)",
                    (_lid, _pc['id'], now())
                )
                for _sz in ('S','M','L'):
                    conn.execute("INSERT OR IGNORE INTO bundle_quantities (id,food_item_id,bundle_size,quantity) VALUES (?,?,?,0)",
                                 (str(uuid.uuid4()), _lid, _sz))
                log.info('Catalog: added Brown Lentils @ $4.00')

        # Red Beans Cans: add to beans group (same category as Red Kidney Beans)
        if not conn.execute("SELECT id FROM food_items WHERE name='Red Beans Cans'").fetchone():
            _rk_row = conn.execute("SELECT category_id FROM food_items WHERE name='Red Kidney Beans'").fetchone()
            _rb_cat = _rk_row['category_id'] if _rk_row else None
            if not _rb_cat:
                _rb_cat = conn.execute("SELECT id FROM food_categories LIMIT 1").fetchone()['id']
            _rbid = str(uuid.uuid4())
            conn.execute(
                "INSERT INTO food_items (id,category_id,name,unit,is_active,display_order,created_at,price,allow_qty,is_default,group_id,group_max,is_free_text) "
                "VALUES (?,?,'Red Beans Cans','can',1,12,?,3.50,0,0,'beans',1,0)",
                (_rbid, _rb_cat, now())
            )
            for _sz in ('S','M','L'):
                conn.execute("INSERT OR IGNORE INTO bundle_quantities (id,food_item_id,bundle_size,quantity) VALUES (?,?,?,0)",
                             (str(uuid.uuid4()), _rbid, _sz))
            log.info('Catalog: added Red Beans Cans @ $3.50 (beans group)')

        # Other Fruit: free-text item in fruit group
        if not conn.execute("SELECT id FROM food_items WHERE name='Other Fruit'").fetchone():
            _pcat = conn.execute("SELECT id FROM food_categories WHERE name='Produce'").fetchone()
            if not _pcat:
                _pcat = conn.execute("SELECT id FROM food_categories LIMIT 1").fetchone()
            _ofid = str(uuid.uuid4())
            conn.execute(
                "INSERT INTO food_items (id,category_id,name,unit,is_active,display_order,created_at,price,allow_qty,is_default,group_id,group_max,is_free_text) "
                "VALUES (?,?,'Other Fruit','each',1,9,?,6.00,0,0,'fruit',1,1)",
                (_ofid, _pcat['id'], now())
            )
            for _sz in ('S','M','L'):
                conn.execute("INSERT OR IGNORE INTO bundle_quantities (id,food_item_id,bundle_size,quantity) VALUES (?,?,?,0)",
                             (str(uuid.uuid4()), _ofid, _sz))
            log.info('Catalog: added Other Fruit (free-text, fruit group) @ $6.00')

        conn.commit()
        log.info('Item selection rules seeded.')

    # Migration: remove 'paused' from families.status CHECK constraint
    # First migrate any paused rows to inactive, then recreate the table without 'paused'
    try:
        fam_sql = conn.execute(
            "SELECT sql FROM sqlite_master WHERE type='table' AND name='families'"
        ).fetchone()
        if fam_sql and 'paused' in (fam_sql['sql'] or ''):
            conn.execute("UPDATE families SET status='inactive' WHERE status='paused'")
            conn.executescript('''
                PRAGMA foreign_keys=OFF;
                CREATE TABLE IF NOT EXISTS _families_new (
                    id                  TEXT PRIMARY KEY,
                    name                TEXT NOT NULL,
                    phone               TEXT,
                    address             TEXT,
                    city                TEXT,
                    family_size         INTEGER,
                    children_count      INTEGER,
                    dietary_notes       TEXT,
                    frequency           TEXT,
                    income_range        TEXT,
                    status              TEXT NOT NULL DEFAULT 'pending'
                                        CHECK(status IN ('pending','active','inactive')),
                    notes               TEXT,
                    source              TEXT DEFAULT 'admin',
                    created_at          TEXT NOT NULL,
                    updated_at          TEXT,
                    family_code         TEXT,
                    bundle_size         TEXT,
                    pending_bundle_size TEXT,
                    wa_phone            TEXT,
                    wa_apikey           TEXT,
                    email               TEXT
                );
                INSERT INTO _families_new
                    SELECT id, name, phone, address, city, family_size, children_count,
                           dietary_notes, frequency, income_range, status, notes, source,
                           created_at, updated_at, family_code, bundle_size, pending_bundle_size,
                           wa_phone, wa_apikey, email
                    FROM families;
                DROP TABLE families;
                ALTER TABLE _families_new RENAME TO families;
                PRAGMA foreign_keys=ON;
            ''')
            conn.commit()
            log.info("Migration: removed 'paused' from families.status CHECK constraint")
    except Exception as _e:
        log.warning(f"Migration: families paused-removal skipped ({_e})")

    # Migrate food_request_events: remove FK constraint so audit events persist after order deletion
    fre_sql = conn.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name='food_request_events'"
    ).fetchone()
    if fre_sql and 'FOREIGN KEY' in fre_sql[0]:
        log.info('Migration: removing FK from food_request_events to preserve audit trail')
        try:
            conn.execute('PRAGMA foreign_keys=OFF')
            conn.executescript('''
                CREATE TABLE IF NOT EXISTS food_request_events_new (
                    id          TEXT PRIMARY KEY,
                    request_id  TEXT NOT NULL,
                    event_type  TEXT NOT NULL,
                    actor       TEXT NOT NULL DEFAULT 'system',
                    payload     TEXT NOT NULL DEFAULT '{}',
                    created_at  TEXT NOT NULL
                );
                INSERT OR IGNORE INTO food_request_events_new SELECT * FROM food_request_events;
                DROP TABLE food_request_events;
                ALTER TABLE food_request_events_new RENAME TO food_request_events;
            ''')
            conn.execute('PRAGMA foreign_keys=ON')
            log.info('Migration: food_request_events FK removed — audit log persists after order deletion')
        except Exception as _e:
            conn.execute('PRAGMA foreign_keys=ON')
            log.info(f'Migration: food_request_events already migrated — skipping ({_e})')

    # ── Migration: add cycle_id to receipts (admin historical entry) ────────────
    try:
        conn.execute('ALTER TABLE receipts ADD COLUMN cycle_id TEXT')
        log.info('Migration: added cycle_id to receipts')
    except Exception:
        pass

    # Ensure users CHECK constraint includes ALL roles (treasurer + family)
    # Runs on every boot — idempotent, bails immediately if already correct
    _ensure_treasurer_role(conn)

    # Extra receipt-spend categories (2026-07-12) — so grocery items outside the
    # Grains/Protein/Produce staples get classified instead of falling to "Other".
    # Idempotent: only inserts a category if that name doesn't already exist.
    for _cat in ('Oil & Condiments', 'Household'):
        if not conn.execute("SELECT id FROM food_categories WHERE name=?", (_cat,)).fetchone():
            _mo = conn.execute("SELECT COALESCE(MAX(display_order),0) FROM food_categories").fetchone()[0]
            conn.execute(
                "INSERT INTO food_categories (id, name, display_order, is_active, created_at) VALUES (?,?,?,1,?)",
                (str(uuid.uuid4()), _cat, _mo + 1, now()))
            log.info(f'Seeded food category: {_cat}')

    # Some indexed tables/columns are added by migrations below the first index
    # pass. Retry after all migrations so a brand-new database gets the same
    # indexes as an upgraded one.
    for _idx_sql in _performance_indexes:
        try:
            conn.execute(_idx_sql)
        except Exception as _e:
            log.warning(f'Post-migration index creation skipped: {_e}')

    # ── audit P1.9: one ACTIVE volunteer slot per (cycle, family, task_type) ──
    # Placed at the END of bootstrap so it runs AFTER volunteer_slots is created and
    # rebuilt by the migrations above (the early index-loop runs before the table
    # exists on a fresh DB). _ensure_volunteer_slots was SELECT-then-INSERT with no
    # DB constraint, so two concurrent requests could each create a duplicate open
    # slot for the same family+task → two volunteers claim different rows and both
    # get confirmations. Step 1: cancel redundant UNCLAIMED ('open') duplicates only
    # (never auto-cancel a volunteer's claim). Step 2: add a partial UNIQUE index so
    # the DB rejects the duplicate INSERT and the race can no longer produce two.
    try:
        dup_groups = conn.execute(
            "SELECT cycle_id, family_id, task_type FROM volunteer_slots "
            "WHERE status!='cancelled' "
            "GROUP BY cycle_id, family_id, task_type HAVING COUNT(*)>1"
        ).fetchall()
        for gr in dup_groups:
            rows = conn.execute(
                "SELECT id, status, claimed_by FROM volunteer_slots "
                "WHERE cycle_id=? AND family_id=? AND task_type=? AND status!='cancelled' "
                "ORDER BY CASE status WHEN 'confirmed' THEN 0 WHEN 'complete' THEN 1 "
                "                     WHEN 'claimed' THEN 2 ELSE 3 END, "
                "(claimed_by IS NULL), created_at",
                (gr['cycle_id'], gr['family_id'], gr['task_type'])
            ).fetchall()
            # Keep rows[0] (highest-progress / owned / oldest); cancel the rest ONLY
            # if they are unclaimed 'open' slots. Any surviving claimed duplicate is
            # left for manual review (the unique index is then skipped-and-logged
            # rather than silently dropping someone's claim).
            for extra in rows[1:]:
                if extra['status'] == 'open' and not extra['claimed_by']:
                    conn.execute("UPDATE volunteer_slots SET status='cancelled', updated_at=? WHERE id=?",
                                 (now(), extra['id']))
    except Exception as _e:
        log.warning(f'volunteer_slots duplicate cleanup skipped: {_e}')
    try:
        conn.execute(
            "CREATE UNIQUE INDEX IF NOT EXISTS uq_vs_active_slot "
            "ON volunteer_slots(cycle_id, family_id, task_type) WHERE status!='cancelled'"
        )
    except Exception as _e:
        log.warning(f'Active-slot unique index skipped (duplicate CLAIMED slots need manual review): {_e}')

    conn.commit()
    conn.close()
    final_size_kb = os.path.getsize(abs_db) / 1024
    log.info(f'Database bootstrapped. Size: {final_size_kb:.1f} KB  Path: {abs_db}')

# ── Auth Helpers ──────────────────────────────────────────────────────────────

def get_session(token):
    return get_db().execute(
        '''SELECT s.token, s.expires_at, u.id as user_id, u.username,
                  u.name, u.role, u.active, u.linked_id, u.linked_type,
                  u.must_change_password
           FROM sessions s JOIN users u ON s.user_id = u.id
           WHERE s.token=? AND s.expires_at > ?''',
        (token, now())
    ).fetchone()

def _linked_account_is_active(db, session):
    """Keep application access aligned with the linked family/volunteer record."""
    if session['role'] == 'volunteer':
        row = db.execute(
            "SELECT status FROM volunteers WHERE id=?", (session['linked_id'],)
        ).fetchone()
        return bool(row and row['status'] == 'active')
    if session['role'] == 'family':
        row = db.execute(
            "SELECT status FROM families WHERE id=?", (session['linked_id'],)
        ).fetchone()
        return bool(row and row['status'] == 'active')
    return True

def _revoke_user_sessions(db, user_id, except_token=None):
    """Revoke a user's sessions after credential or account-state changes."""
    if except_token:
        db.execute(
            "DELETE FROM sessions WHERE user_id=? AND token!=?", (user_id, except_token)
        )
    else:
        db.execute("DELETE FROM sessions WHERE user_id=?", (user_id,))

def require_auth(roles=None):
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            auth = request.headers.get('Authorization', '')
            if not auth.startswith('Bearer '):
                return jsonify({'error': 'Unauthorized'}), 401
            token = auth[7:]
            if token.startswith('tmp_'):
                # Temp tokens (must-change-password flow) are only valid for
                # /api/auth/set-password — never for general API access
                return jsonify({'error': 'Password change required. Complete it before using the app.'}), 401
            session = get_session(token)
            if not session:
                return jsonify({'error': 'Session expired or invalid'}), 401
            if not session['active']:
                return jsonify({'error': 'Account inactive'}), 401
            if not _linked_account_is_active(get_db(), session):
                return jsonify({'error': 'Account inactive'}), 401
            if roles and session['role'] not in roles:
                return jsonify({'error': 'Forbidden'}), 403
            # Slide expiry — at most once per hour per session (audit 3.6).
            # Previously this UPDATE+commit ran on EVERY authenticated request,
            # a needless disk write per API call on SQLite.
            new_expiry = (datetime.utcnow() + timedelta(hours=SESSION_HOURS)).isoformat()
            # threshold guards small SESSION_HOURS values (never goes below half-life)
            slide_threshold = (datetime.utcnow() + timedelta(
                hours=max(SESSION_HOURS - 1, SESSION_HOURS * 0.5))).isoformat()
            # Avoid issuing a zero-row UPDATE: Python's sqlite driver still opens
            # a transaction for it, which breaks routes that correctly start an
            # explicit atomic transaction (for example family deletion).
            if session['expires_at'] < slide_threshold:
                get_db().execute(
                    "UPDATE sessions SET expires_at=? WHERE token=?",
                    (new_expiry, token))
                get_db().commit()
            g.user = dict(session)
            return f(*args, **kwargs)
        return wrapper
    return decorator

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXT

def _normalize_upload_url(value):
    """Accept only app-owned upload paths, never arbitrary or cross-origin URLs."""
    if value in (None, ''):
        return None
    if not isinstance(value, str) or not value.startswith('/uploads/'):
        return None
    filename = value[len('/uploads/'):]
    if (not filename or '/' in filename or '\\' in filename
            or secure_filename(filename) != filename or not allowed_file(filename)):
        return None
    return f'/uploads/{filename}'

def _validate_receipt_upload(original_filename, raw):
    """Verify size, signature, decodability and dimensions before persistence."""
    if not original_filename or not allowed_file(original_filename):
        return None, 'Invalid file type. Use JPG, PNG, GIF, PDF, or HEIC.'
    if not raw:
        return None, 'The uploaded file is empty.'
    if len(raw) > MAX_UPLOAD_BYTES:
        return None, f'File is too large. Maximum size is {MAX_UPLOAD_BYTES // (1024 * 1024)} MB.'
    ext = secure_filename(original_filename).rsplit('.', 1)[-1].lower()
    signatures = {
        'jpg': raw.startswith(b'\xff\xd8\xff'),
        'jpeg': raw.startswith(b'\xff\xd8\xff'),
        'png': raw.startswith(b'\x89PNG\r\n\x1a\n'),
        'gif': raw.startswith((b'GIF87a', b'GIF89a')),
        'pdf': raw.startswith(b'%PDF-'),
        'heic': len(raw) >= 12 and raw[4:8] == b'ftyp' and raw[8:12] in {
            b'heic', b'heix', b'hevc', b'hevx', b'mif1', b'msf1'
        },
    }
    if not signatures.get(ext, False):
        return None, 'File contents do not match the selected file type.'
    if ext == 'pdf':
        return ext, None

    try:
        from io import BytesIO
        if ext == 'heic':
            import pillow_heif
            pillow_heif.register_heif_opener()
        from PIL import Image
        with Image.open(BytesIO(raw)) as image:
            width, height = image.size
            if width < 1 or height < 1 or width * height > MAX_IMAGE_PIXELS:
                return None, 'Image dimensions are invalid or too large.'
            image.verify()
    except Exception as exc:
        log.warning(f'Upload image verification failed: {exc}')
        return None, 'The uploaded image is damaged or unsupported.'
    return ext, None

def _cleanup_orphan_uploads(db, older_than_hours=24):
    """Delete only registered files that no receipt references after the grace period."""
    cutoff = (datetime.utcnow() - timedelta(hours=older_than_hours)).isoformat()
    rows = db.execute(
        "SELECT filename FROM uploaded_files WHERE created_at<?",
        (cutoff,)
    ).fetchall()
    removed = 0
    for row in rows:
        filename = row['filename']
        file_url = f'/uploads/{filename}'
        if db.execute("SELECT 1 FROM receipts WHERE file_url=? LIMIT 1", (file_url,)).fetchone():
            db.execute(
                "UPDATE uploaded_files SET claimed_at=? WHERE filename=?", (now(), filename)
            )
            continue
        if secure_filename(filename) == filename:
            try:
                os.remove(os.path.join(UPLOAD_FOLDER, filename))
                removed += 1
            except FileNotFoundError:
                pass
            except OSError as exc:
                log.warning(f'Orphan upload cleanup failed for {filename}: {exc}')
                continue
        db.execute("DELETE FROM uploaded_files WHERE filename=?", (filename,))
    if rows:
        db.commit()
    return removed

def _upload_quota_error(db, size_bytes, uploader_user_id=None, volunteer_id=None):
    cutoff = (datetime.utcnow() - timedelta(hours=24)).isoformat()
    if volunteer_id:
        where, identity = 'volunteer_id=?', volunteer_id
        max_files, max_bytes = UPLOAD_FILES_PER_DAY, UPLOAD_BYTES_PER_DAY
    else:
        where, identity = 'uploader_user_id=?', uploader_user_id
        # Staff may use the bulk uploader, but still receive a finite safety cap.
        max_files, max_bytes = UPLOAD_FILES_PER_DAY * 5, UPLOAD_BYTES_PER_DAY * 4
    usage = db.execute(
        f"SELECT COUNT(*) files, COALESCE(SUM(size_bytes),0) bytes FROM uploaded_files "
        f"WHERE {where} AND created_at>?",
        (identity, cutoff)
    ).fetchone()
    if usage['files'] >= max_files or usage['bytes'] + size_bytes > max_bytes:
        return 'Daily upload quota reached. Please try again later or contact a coordinator.'
    try:
        total_bytes = sum(
            entry.stat().st_size for entry in os.scandir(UPLOAD_FOLDER)
            if entry.is_file(follow_symlinks=False)
        )
    except OSError as exc:
        log.error(f'Unable to inspect upload storage: {exc}')
        return 'Upload storage is temporarily unavailable.'
    if total_bytes + size_bytes > UPLOAD_TOTAL_BYTES:
        return 'Upload storage is full. Please contact a coordinator.'
    return None

def _store_receipt_upload(db, file_storage, uploader_user_id=None, volunteer_id=None):
    raw = file_storage.read()
    ext, validation_error = _validate_receipt_upload(file_storage.filename, raw)
    if validation_error:
        return None, None, validation_error
    _cleanup_orphan_uploads(db)
    quota_error = _upload_quota_error(
        db, len(raw), uploader_user_id=uploader_user_id, volunteer_id=volunteer_id
    )
    if quota_error:
        return None, None, quota_error

    filename = f'{uuid.uuid4()}.{ext}'
    path = os.path.join(UPLOAD_FOLDER, filename)
    try:
        with open(path, 'xb') as out:
            out.write(raw)
        db.execute(
            """INSERT INTO uploaded_files
               (filename,uploader_user_id,volunteer_id,size_bytes,created_at)
               VALUES (?,?,?,?,?)""",
            (filename, uploader_user_id, volunteer_id, len(raw), now())
        )
        db.commit()
    except Exception:
        db.rollback()
        try:
            os.remove(path)
        except OSError:
            pass
        raise
    return filename, raw, None

def _claim_registered_upload(db, file_url, volunteer_id=None):
    """Claim a new upload; volunteer claims must match the uploader identity."""
    if not file_url:
        return True
    filename = file_url.rsplit('/', 1)[-1]
    row = db.execute(
        "SELECT volunteer_id, claimed_at FROM uploaded_files WHERE filename=?", (filename,)
    ).fetchone()
    if not row:
        # Staff may attach pre-registry legacy files; portal callers may not.
        return volunteer_id is None
    if volunteer_id and row['volunteer_id'] != volunteer_id:
        return False
    if row['claimed_at']:
        # A file is a one-receipt capability and cannot be replayed.
        return False
    db.execute("UPDATE uploaded_files SET claimed_at=? WHERE filename=?", (now(), filename))
    return True

# ── Portal Auth (volunteer self-service, phone-based) ─────────────────────────

def get_portal_session(token):
    # Check main sessions table first (new username/password login)
    row = get_db().execute(
        '''SELECT s.token, u.linked_id as volunteer_id, v.name, v.phone, v.role,
                  v.wa_phone, v.wa_apikey
           FROM sessions s
           JOIN users u ON s.user_id = u.id
           JOIN volunteers v ON u.linked_id = v.id
           WHERE s.token=? AND s.expires_at > ? AND u.role='volunteer' AND u.active=1
             AND v.status='active' AND s.token NOT LIKE 'tmp_%' ''',
        (token, now())
    ).fetchone()
    if row:
        return row
    # Fallback: old portal_sessions table (backward compat for existing sessions)
    return get_db().execute(
        '''SELECT ps.token, ps.volunteer_id, v.name, v.phone, v.role,
                  v.wa_phone, v.wa_apikey
           FROM portal_sessions ps JOIN volunteers v ON ps.volunteer_id = v.id
           WHERE ps.token=? AND ps.expires_at > ? AND v.status='active'
             AND ps.token NOT LIKE 'tmp_%' ''',
        (token, now())
    ).fetchone()

def require_portal_auth():
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            auth = request.headers.get('Authorization', '')
            if not auth.startswith('Bearer '):
                return jsonify({'error': 'Unauthorized'}), 401
            token = auth[7:]
            if token.startswith('tmp_'):
                return jsonify({'error': 'Password change required'}), 401
            session = get_portal_session(token)
            if not session:
                return jsonify({'error': 'Session expired — please log in again'}), 401
            g.pv = dict(session)  # portal volunteer
            return f(*args, **kwargs)
        return wrapper
    return decorator

def get_family_session(token):
    """Returns family record for a family-role session token."""
    row = get_db().execute(
        '''SELECT s.token, u.linked_id as family_id, f.name, f.phone, f.status
           FROM sessions s
           JOIN users u ON s.user_id = u.id
           JOIN families f ON u.linked_id = f.id
           WHERE s.token=? AND s.expires_at > ? AND u.role='family' AND u.active=1
             AND f.status='active' AND s.token NOT LIKE 'tmp_%' ''',
        (token, now())
    ).fetchone()
    return row

def require_family_auth():
    """Decorator for family portal API routes. Sets g.fam with family details."""
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            auth = request.headers.get('Authorization', '')
            if not auth.startswith('Bearer '):
                return jsonify({'error': 'Unauthorized'}), 401
            token = auth[7:]
            if token.startswith('tmp_'):
                return jsonify({'error': 'Password change required'}), 401
            session = get_family_session(token)
            if not session:
                return jsonify({'error': 'Session expired — please log in again'}), 401
            g.fam = dict(session)
            return f(*args, **kwargs)
        return wrapper
    return decorator

# ── Schema migration helper ───────────────────────────────────────────────────

VALID_ROLES = {'admin', 'volunteer', 'finance', 'treasurer', 'viewer', 'family'}
PASSWORD_MIN_LEN = 8
PASSWORD_EXPIRY_DAYS = 60

def _validate_password(password):
    """Returns (ok: bool, error: str). Enforces min length, uppercase, digit, special char."""
    import re
    if not password or len(password) < PASSWORD_MIN_LEN:
        return False, f'Password must be at least {PASSWORD_MIN_LEN} characters.'
    if not re.search(r'[A-Z]', password):
        return False, 'Password must contain at least one uppercase letter.'
    if not re.search(r'\d', password):
        return False, 'Password must contain at least one number.'
    if not re.search(r'[!@#$%^&*()_+\-=\[\]{};\':"\\|,.<>\/?]', password):
        return False, 'Password must contain at least one special character (!@#$%^&* etc).'
    return True, ''

def _ensure_treasurer_role(conn):
    """Ensure the users table CHECK constraint includes all required roles.
    Uses safe table-recreation — no PRAGMA writable_schema."""
    REQUIRED_ROLES = ('admin', 'volunteer', 'finance', 'treasurer', 'viewer', 'family')

    row = conn.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name='users'"
    ).fetchone()
    if not row:
        return  # table doesn't exist yet

    old_sql = row[0]
    # Check if ALL required roles are already present — bail early if already correct
    if all(f"'{r}'" in old_sql for r in REQUIRED_ROLES):
        return

    log.info('_ensure_treasurer_role: roles missing from CHECK constraint — recreating users table')
    _recreate_users_table(conn)


def _recreate_users_table(conn):
    """Full table-recreation migration as fallback. Requires an exclusive DB lock."""
    try:
        conn.execute('PRAGMA foreign_keys=OFF')
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS users_new (
                id                   TEXT PRIMARY KEY,
                username             TEXT UNIQUE NOT NULL,
                password_hash        TEXT NOT NULL,
                name                 TEXT,
                role                 TEXT NOT NULL DEFAULT 'viewer'
                                     CHECK(role IN ('admin','volunteer','finance','treasurer','viewer','family')),
                email                TEXT,
                wa_phone             TEXT,
                wa_apikey            TEXT,
                active               INTEGER NOT NULL DEFAULT 1,
                linked_id            TEXT,
                linked_type          TEXT,
                must_change_password INTEGER NOT NULL DEFAULT 1,
                password_changed_at  TEXT,
                last_login_at        TEXT,
                created_at           TEXT NOT NULL
            );
        ''')
        # Copy the INTERSECTION of columns present in both tables (audit 2.4).
        # The previous hardcoded column list silently RESET late-added columns
        # (linked_id, linked_type, must_change_password→1 forcing a password
        # reset for every user, password_changed_at, last_login_at).
        old_cols = {r[1] for r in conn.execute('PRAGMA table_info(users)').fetchall()}
        new_cols = [r[1] for r in conn.execute('PRAGMA table_info(users_new)').fetchall()]
        cols = ', '.join(c for c in new_cols if c in old_cols)  # identifiers from PRAGMA, not user input
        conn.execute(f'INSERT OR IGNORE INTO users_new ({cols}) SELECT {cols} FROM users')
        conn.executescript('''
            DROP TABLE IF EXISTS users;
            ALTER TABLE users_new RENAME TO users;
        ''')
        conn.execute('PRAGMA foreign_keys=ON')
        log.info('_recreate_users_table: migration complete')
    except Exception as _e:
        conn.execute('PRAGMA foreign_keys=ON')
        log.warning(f'_recreate_users_table: failed ({_e})')

def _email_notify(to_email, subject, body):
    """Send a notification email. Wrapper around _email_send with logging.
    Returns True on success, False/None if no email or send fails."""
    if not to_email:
        return False
    result = _email_send(to_email, subject, body)
    if not result:
        log.warning(f'Email notification failed or not configured — to={to_email!r} subject={subject!r}')
    return result


def _email_notify_async(sends):
    """Fire a list of email notifications in a background thread.
    sends: list of (to_email, subject, body) tuples — items with missing email are skipped."""
    import threading as _t
    items = [(e, s, b) for e, s, b in sends if e]
    if not items:
        return
    def _run():
        for email, subject, body in items:
            _email_notify(email, subject, body)
    _t.Thread(target=_run, daemon=True).start()


def _lookup_volunteer_email(conn_or_db, vol_id):
    """Look up a volunteer's email by ID. Works with both Flask db (sqlite3.Row) and direct connections."""
    row = conn_or_db.execute("SELECT email FROM volunteers WHERE id=?", (vol_id,)).fetchone()
    return (row['email'] or '').strip() if row else ''


def _lookup_family_email(conn_or_db, family_id):
    """Look up a family's email by ID."""
    row = conn_or_db.execute("SELECT email FROM families WHERE id=?", (family_id,)).fetchone()
    return (row['email'] or '').strip() if row else ''


def _today_central():
    """Return today's date in US Central time (America/Chicago).
    Uses zoneinfo (Python 3.9+, always available on Railway).
    Falls back to UTC if zoneinfo is unavailable."""
    try:
        from zoneinfo import ZoneInfo
        from datetime import datetime as _dt
        return _dt.now(ZoneInfo('America/Chicago')).date()
    except Exception:
        from datetime import date as _d
        return _d.today()

def _confirmation_expiry_iso(delivery_date_start=None):
    """Create a UTC expiry, capped at the start of delivery day in Chicago."""
    expiry = datetime.utcnow() + timedelta(hours=CONFIRMATION_TOKEN_HOURS)
    if delivery_date_start:
        try:
            from datetime import date as _date, time as _time
            from zoneinfo import ZoneInfo
            local_deadline = datetime.combine(
                _date.fromisoformat(str(delivery_date_start)), _time.min,
                tzinfo=ZoneInfo('America/Chicago')
            )
            utc_deadline = local_deadline.astimezone(ZoneInfo('UTC')).replace(tzinfo=None)
            expiry = min(expiry, utc_deadline)
        except (TypeError, ValueError):
            pass
    return expiry.isoformat()

def _log_order_event(db, request_id, event_type, actor='system', payload=None):
    """Append an event to food_request_events. Never raises — failures are logged only.
    event_type: confirmed | items_edited | cancelled | admin_override | auto_skipped
    actor:      family | admin | scheduler | system
    payload:    dict — e.g. {'removed': ['Whole Chicken'], 'added': ['Brown Lentils']}
    """
    import json as _json
    try:
        db.execute(
            "INSERT INTO food_request_events (id, request_id, event_type, actor, payload, created_at) VALUES (?,?,?,?,?,?)",
            (str(uuid.uuid4()), request_id, event_type, actor,
             _json.dumps(payload or {}), now())
        )
    except Exception as _e:
        log.warning(f'_log_order_event failed ({event_type} on {request_id}): {_e}')

def _notify_coordinators(db, message):
    """Email all active admin users — fires in background thread (non-blocking)."""
    try:
        admins = db.execute(
            "SELECT email FROM users WHERE role='admin' AND active=1 AND email IS NOT NULL AND TRIM(email)!=''"
        ).fetchall()
        import threading as _t
        def _run():
            for a in admins:
                _email_send(a['email'], 'Sihha Ops Alert', message)
        _t.Thread(target=_run, daemon=True).start()
    except Exception as _e:
        log.warning(f'_notify_coordinators failed: {_e}')


def _provider_error_summary(exc):
    """Return a credential-safe provider error label for APIs and logs."""
    status = getattr(exc, 'code', None)
    if status is not None:
        return f'HTTP {status}'
    return type(exc).__name__


def _email_provider_configured(provider=None):
    """Return whether the selected email provider has its required credentials."""
    selected = (provider or EMAIL_PROVIDER).strip().lower()
    if selected == 'sendgrid':
        return bool(SENDGRID_API_KEY)
    if selected == 'twilio':
        return bool(TWILIO_EMAIL_API_KEY_SID and TWILIO_EMAIL_API_KEY_SECRET)
    return False


def _email_send(to_email, subject, text_body, attachment=None, html_body=None):
    """Send email through EMAIL_PROVIDER (sendgrid rollback default or twilio).
    Returns True on success, False on failure (never raises).
    attachment: optional (filename, bytes) tuple — used by the off-site backup job."""
    import base64 as _b64
    import html as _html
    import json as _json
    import urllib.request

    if not _email_provider_configured():
        log.warning('Email not sent — %s email provider is not configured', EMAIL_PROVIDER)
        return False

    if EMAIL_PROVIDER == 'sendgrid':
        msg = {
            'personalizations': [{'to': [{'email': to_email}]}],
            'from': {'email': NOTIFY_FROM_EMAIL, 'name': 'Sihha Ops Hub'},
            'subject': subject,
            'content': [{'type': 'text/plain', 'value': text_body}],
        }
        if html_body:
            msg['content'].append({'type': 'text/html', 'value': html_body})
        if attachment:
            fname, fbytes = attachment
            msg['attachments'] = [{
                'content': _b64.b64encode(fbytes).decode('ascii'),
                'filename': fname,
                'type': 'application/gzip',
                'disposition': 'attachment',
            }]
        endpoint = 'https://api.sendgrid.com/v3/mail/send'
        headers = {
            'Authorization': f'Bearer {SENDGRID_API_KEY}',
            'Content-Type': 'application/json',
        }
    elif EMAIL_PROVIDER == 'twilio':
        msg = {
            'from': {'address': NOTIFY_FROM_EMAIL, 'name': 'Sihha Ops Hub'},
            'to': [{'address': to_email}],
            'content': {
                'subject': subject,
                'text': text_body,
                'html': html_body or _html.escape(text_body).replace('\n', '<br>'),
            },
        }
        if attachment:
            fname, fbytes = attachment
            msg['content']['attachments'] = [{
                'filename': fname,
                'contentType': 'application/gzip',
                'content': _b64.b64encode(fbytes).decode('ascii'),
            }]
        credentials = _b64.b64encode(
            f'{TWILIO_EMAIL_API_KEY_SID}:{TWILIO_EMAIL_API_KEY_SECRET}'.encode('utf-8')
        ).decode('ascii')
        endpoint = TWILIO_EMAIL_ENDPOINT
        headers = {
            'Authorization': f'Basic {credentials}',
            'Content-Type': 'application/json',
        }
    else:
        log.warning('Email not sent — unsupported EMAIL_PROVIDER=%r', EMAIL_PROVIDER)
        return False

    payload = _json.dumps(msg).encode('utf-8')
    if EMAIL_PROVIDER == 'twilio' and len(payload) > TWILIO_EMAIL_MAX_REQUEST_BYTES:
        log.warning('Email not sent — Twilio Email request exceeds the safe request-size limit')
        return False
    req = urllib.request.Request(
        endpoint,
        data=payload,
        headers=headers,
        method='POST',
    )
    try:
        with urllib.request.urlopen(req, timeout=(30 if attachment else 10)) as response:
            sent = 200 <= response.status < 300
        if sent:
            log.info('Email accepted by %s: %s', EMAIL_PROVIDER, subject)
        return sent
    except Exception as e:
        log.warning('Email send failed through %s: %s',
                    EMAIL_PROVIDER, _provider_error_summary(e))
        return False

def _notify_treasurers(db, subject, message):
    """Notify all active treasurer users via email — fired in a background thread
    so the N blocking SendGrid sends never freeze the request path (audit P1.6).
    Used for new reimbursement requests, receipt submissions, etc."""
    treasurers = db.execute(
        "SELECT name, email FROM users WHERE role='treasurer' AND active=1"
    ).fetchall()
    if not treasurers:
        log.info('No active treasurers found to notify.')
        return
    _email_notify_async([(t['email'], subject, message) for t in treasurers if t['email']])

# ── Health ────────────────────────────────────────────────────────────────────

@app.route('/api/health')
def health():
    """Readiness check used by Railway; never creates a missing database."""
    conn = None
    try:
        from urllib.parse import quote as _urlquote
        abs_db = os.path.abspath(DB_PATH)
        if not os.path.isfile(abs_db):
            raise RuntimeError('database file is missing')
        conn = sqlite3.connect(
            f'file:{_urlquote(abs_db)}?mode=rw', uri=True, timeout=2
        )
        required = {'users', 'sessions', 'families', 'volunteers',
                    'delivery_cycles', 'food_requests'}
        found = {
            row[0] for row in conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            ).fetchall()
        }
        if not required.issubset(found):
            raise RuntimeError('required database tables are missing')
        quick_check = conn.execute('PRAGMA quick_check').fetchone()
        if not quick_check or quick_check[0] != 'ok':
            raise RuntimeError('database integrity check failed')
        conn.execute('SELECT 1 FROM users LIMIT 1').fetchone()
        return jsonify({
            'status': 'ok', 'version': '1.2.2', 'time': now(),
            'checks': {'database': 'ok', 'schema': 'ok'},
            'communications': {
                'email_provider': EMAIL_PROVIDER,
                'email_configured': _email_provider_configured(),
                'sendgrid_configured': bool(SENDGRID_API_KEY),
            },
        })
    except Exception as exc:
        log.error(f'Readiness check failed: {exc}')
        return jsonify({
            'status': 'error', 'time': now(),
            'checks': {'database': 'unavailable'}
        }), 503
    finally:
        if conn is not None:
            conn.close()


@app.route('/api/admin/communications/health', methods=['GET'])
@require_auth(roles=['admin'])
def communications_health():
    """Verify the selected email credentials without sending a message."""
    import base64 as _b64
    import urllib.error
    import urllib.request

    result = {
        'email': {
            'provider': EMAIL_PROVIDER,
            'configured': _email_provider_configured(),
            'authenticated': False,
        },
        'sendgrid': {'configured': bool(SENDGRID_API_KEY), 'authenticated': False},
        'twilio_email': {
            'configured': bool(TWILIO_EMAIL_API_KEY_SID and TWILIO_EMAIL_API_KEY_SECRET),
            'authenticated': False,
        },
    }
    if EMAIL_PROVIDER == 'sendgrid' and SENDGRID_API_KEY:
        try:
            req = urllib.request.Request(
                'https://api.sendgrid.com/v3/scopes',
                headers={'Authorization': f'Bearer {SENDGRID_API_KEY}'},
            )
            with urllib.request.urlopen(req, timeout=10) as response:
                result['sendgrid']['authenticated'] = 200 <= response.status < 300
        except Exception as exc:
            result['sendgrid']['error'] = _provider_error_summary(exc)
        result['email']['authenticated'] = result['sendgrid']['authenticated']
        if 'error' in result['sendgrid']:
            result['email']['error'] = result['sendgrid']['error']
    elif EMAIL_PROVIDER == 'twilio' and result['twilio_email']['configured']:
        try:
            credentials = _b64.b64encode(
                f'{TWILIO_EMAIL_API_KEY_SID}:{TWILIO_EMAIL_API_KEY_SECRET}'.encode('utf-8')
            ).decode('ascii')
            req = urllib.request.Request(
                TWILIO_EMAIL_ENDPOINT,
                data=b'{}',
                headers={
                    'Authorization': f'Basic {credentials}',
                    'Content-Type': 'application/json',
                },
                method='POST',
            )
            with urllib.request.urlopen(req, timeout=10) as response:
                result['twilio_email']['authenticated'] = 200 <= response.status < 300
        except urllib.error.HTTPError as exc:
            if exc.code in (400, 422):
                result['twilio_email']['authenticated'] = True
            else:
                result['twilio_email']['error'] = _provider_error_summary(exc)
        except Exception as exc:
            result['twilio_email']['error'] = _provider_error_summary(exc)
        result['email']['authenticated'] = result['twilio_email']['authenticated']
        if 'error' in result['twilio_email']:
            result['email']['error'] = result['twilio_email']['error']
    elif EMAIL_PROVIDER not in ('sendgrid', 'twilio'):
        result['email']['error'] = 'unsupported_provider'

    result['status'] = 'ok' if result['email']['authenticated'] else 'attention_required'
    return jsonify(result), (200 if result['email']['authenticated'] else 503)

# ── Auth Routes ───────────────────────────────────────────────────────────────

# Failed-login throttle shared by every worker through SQLite. The stored key is
# a one-way digest, not a raw username/IP/phone value.
LOGIN_MAX_FAILS       = 5    # per (ip, username)
LOGIN_MAX_FAILS_USER  = 20   # absolute per-username cap across ALL IPs (audit P1.5)
LOGIN_WINDOW_MIN      = 15

def _client_ip():
    # ProxyFix (x_for=1) has already set request.remote_addr to the real client IP
    # from the last trusted proxy hop, so X-Forwarded-For can no longer be spoofed
    # to rotate the throttle key. Read remote_addr directly.
    return request.remote_addr or '?'

def _rate_bucket(value):
    return hashlib.sha256(str(value).encode('utf-8')).hexdigest()

def _rate_limit_count(db, scope, identity, window_seconds):
    cutoff = (datetime.utcnow() - timedelta(seconds=window_seconds)).isoformat()
    return db.execute(
        "SELECT COUNT(*) FROM rate_limit_events WHERE scope=? AND bucket_key=? AND created_at>?",
        (scope, _rate_bucket(identity), cutoff)
    ).fetchone()[0]

def _rate_limit_blocked(db, scope, identity, limit, window_seconds):
    return _rate_limit_count(db, scope, identity, window_seconds) >= limit

def _rate_limit_record(db, scope, identity):
    db.execute(
        "INSERT INTO rate_limit_events (id,scope,bucket_key,created_at) VALUES (?,?,?,?)",
        (str(uuid.uuid4()), scope, _rate_bucket(identity), now())
    )
    # Keep the table bounded. All configured windows are at most 24 hours.
    db.execute(
        "DELETE FROM rate_limit_events WHERE created_at<?",
        ((datetime.utcnow() - timedelta(hours=48)).isoformat(),)
    )

def _rate_limit_reset(db, scope, identity):
    db.execute(
        "DELETE FROM rate_limit_events WHERE scope=? AND bucket_key=?",
        (scope, _rate_bucket(identity))
    )

def _consume_rate_limit(db, scope, identity, limit, window_seconds):
    """Record a public request and return False when the window is exhausted."""
    if _rate_limit_blocked(db, scope, identity, limit, window_seconds):
        return False
    _rate_limit_record(db, scope, identity)
    db.commit()
    return True

def _rate_limit_response(retry_after):
    response = jsonify({'error': 'Too many requests. Please try again later.'})
    response.status_code = 429
    response.headers['Retry-After'] = str(retry_after)
    return response

def _valid_public_email(value):
    """Small syntax check; deliverability is still determined by the mail provider."""
    import re
    if not value:
        return True
    return bool(re.fullmatch(r'[^\s@]{1,64}@[^\s@]{1,190}\.[^\s@]{2,63}', value))

def _account_invitation_digest(token):
    """Return the one-way database representation of an invitation bearer token."""
    return hashlib.sha256(token.encode('utf-8')).hexdigest()

def _account_delivery_email(db, user):
    """Resolve the verified delivery address attached to an account or linked record."""
    direct = ((user['email'] if 'email' in user.keys() else None) or '').strip()
    if direct:
        return direct
    linked_id = user['linked_id'] if 'linked_id' in user.keys() else None
    if not linked_id:
        return ''
    if user['role'] == 'family':
        row = db.execute("SELECT email FROM families WHERE id=?", (linked_id,)).fetchone()
    elif user['role'] == 'volunteer':
        row = db.execute("SELECT email FROM volunteers WHERE id=?", (linked_id,)).fetchone()
    else:
        row = None
    return ((row['email'] if row else None) or '').strip()

def _email_hint(email):
    """Return a non-sensitive destination hint suitable for API responses and UI."""
    local, _, domain = (email or '').partition('@')
    if not domain:
        return ''
    visible = local[:1]
    return f'{visible}{"*" * max(3, len(local) - 1)}@{domain}'

def _record_account_access_event(db, user_id, event_type,
                                 invitation_id=None, actor_user_id=None, detail=None):
    db.execute(
        '''INSERT INTO account_access_events
           (id,user_id,invitation_id,event_type,actor_user_id,detail,created_at)
           VALUES (?,?,?,?,?,?,?)''',
        (str(uuid.uuid4()), user_id, invitation_id, event_type,
         actor_user_id, detail, now())
    )

def _invalidate_account_invitations(db, user_id, actor_user_id=None, detail=None,
                                    except_invitation_id=None):
    """Invalidate every outstanding link for a user and append one audit event."""
    invalidated_at = now()
    sql = '''UPDATE account_invitations SET invalidated_at=?
             WHERE user_id=? AND used_at IS NULL AND invalidated_at IS NULL'''
    params = [invalidated_at, user_id]
    if except_invitation_id:
        sql += ' AND id!=?'
        params.append(except_invitation_id)
    changed = db.execute(sql, tuple(params)).rowcount
    if changed:
        _record_account_access_event(
            db, user_id, 'invitations_invalidated',
            actor_user_id=actor_user_id, detail=detail
        )
    return changed

def _valid_account_invitation(db, token):
    if not token or len(token) > 256:
        return None
    return db.execute(
        '''SELECT ai.*, u.username, u.name, u.role, u.active,
                  u.linked_id, u.linked_type
           FROM account_invitations ai
           JOIN users u ON u.id=ai.user_id
           WHERE ai.token_hash=? AND ai.email_sent_at IS NOT NULL
             AND ai.used_at IS NULL AND ai.invalidated_at IS NULL
             AND ai.expires_at>?''',
        (_account_invitation_digest(token), now())
    ).fetchone()

def _invalid_account_invitation_response():
    return jsonify({
        'error': 'This access link is invalid or has expired. Ask your Sihha coordinator for a new link.'
    }), 400

def _send_account_access_invitation(db, user, actor_user_id=None):
    """Create and email one secure password-creation invitation for an account.

    Callers must enforce their own authorization and any applicable send-rate limit.
    The account's existing password and sessions are deliberately left unchanged
    until the recipient consumes the single-use link.
    """
    if not user['active'] or not _linked_account_is_active(db, user):
        return {
            'error': 'Account must be active before access can be sent.'
        }, 409

    delivery_email = _account_delivery_email(db, user)
    if not delivery_email or not _valid_public_email(delivery_email):
        return {
            'error': 'Add a valid email address before sending access.'
        }, 422

    invitation_id = str(uuid.uuid4())
    raw_token = secrets.token_urlsafe(32)
    created_at = now()
    expires_at = (
        datetime.utcnow() + timedelta(minutes=ACCOUNT_INVITATION_MINUTES)
    ).isoformat()
    db.execute(
        '''INSERT INTO account_invitations
           (id,user_id,token_hash,delivery_email,created_by,created_at,expires_at)
           VALUES (?,?,?,?,?,?,?)''',
        (invitation_id, user['id'], _account_invitation_digest(raw_token),
         delivery_email, actor_user_id, created_at, expires_at)
    )
    _record_account_access_event(
        db, user['id'], 'invitation_created', invitation_id, actor_user_id
    )
    db.commit()

    access_url = f'{ACCOUNT_INVITATION_URL}#token={raw_token}'
    body = (
        f"Hello {user['name'] or user['username']},\n\n"
        f"Sihha has created a secure account-access link for you.\n\n"
        f"Username: {user['username']}\n"
        f"Create your password: {access_url}\n\n"
        f"This link works once and expires in {ACCOUNT_INVITATION_MINUTES} minutes. "
        f"If you were not expecting it, you can ignore this email.\n\n"
        f"— Sihha Food Program"
    )
    import html as _html
    safe_name = _html.escape(user['name'] or user['username'])
    safe_username = _html.escape(user['username'])
    safe_access_url = _html.escape(access_url, quote=True)
    html_body = (
        f'<p>Hello {safe_name},</p>'
        f'<p>Sihha has created a secure account-access link for you.</p>'
        f'<p>Username: <strong>{safe_username}</strong></p>'
        f'<p><a href="{safe_access_url}">Create your Sihha password</a></p>'
        f'<p>This link works once and expires in {ACCOUNT_INVITATION_MINUTES} minutes. '
        f'If you were not expecting it, you can ignore this email.</p>'
        f'<p>— Sihha Food Program</p>'
    )
    email_sent = _email_send(
        delivery_email, 'Create Your Sihha Password', body, html_body=html_body
    )
    if email_sent:
        sent_at = now()
        db.execute(
            "UPDATE account_invitations SET email_sent_at=? WHERE id=?",
            (sent_at, invitation_id)
        )
        _invalidate_account_invitations(
            db, user['id'], actor_user_id=actor_user_id,
            detail='new_invitation_sent', except_invitation_id=invitation_id
        )
        _record_account_access_event(
            db, user['id'], 'invitation_email_sent', invitation_id, actor_user_id
        )
        db.commit()
        log.info('Account invitation email accepted for user_id=%s', user['id'])
        return {
            'username': user['username'],
            'email_sent': True,
            'email_hint': _email_hint(delivery_email),
            'expires_in_minutes': ACCOUNT_INVITATION_MINUTES,
        }, 200

    db.execute(
        "UPDATE account_invitations SET invalidated_at=? WHERE id=?",
        (now(), invitation_id)
    )
    _record_account_access_event(
        db, user['id'], 'invitation_email_failed', invitation_id, actor_user_id
    )
    db.commit()
    return {
        'error': 'The access email could not be sent. Please try again.'
    }, 502

@app.route('/api/auth/access-invitation', methods=['POST'])
def account_invitation_info():
    """Validate an invitation without changing credentials or creating a session."""
    data = request.get_json(silent=True) or {}
    raw_token = data.get('token') if isinstance(data, dict) else None
    token = raw_token.strip() if isinstance(raw_token, str) else ''
    db = get_db()
    if not _consume_rate_limit(
            db, 'account_invitation_info_ip', _client_ip(), 60, 3600):
        return _rate_limit_response(3600)
    invitation = _valid_account_invitation(db, token)
    if (not invitation or not invitation['active']
            or not _linked_account_is_active(db, invitation)):
        return _invalid_account_invitation_response()
    return jsonify({
        'valid': True,
        'username': invitation['username'],
        'name': invitation['name'],
        'expires_at': invitation['expires_at'],
    })

@app.route('/api/auth/access-invitation/activate', methods=['POST'])
def activate_account_invitation():
    """Consume a single-use invitation and let its owner create the account password."""
    data = request.get_json(silent=True) or {}
    if not isinstance(data, dict):
        return _invalid_account_invitation_response()
    raw_token = data.get('token')
    raw_password = data.get('password')
    raw_confirmation = data.get('password_confirmation')
    token = raw_token.strip() if isinstance(raw_token, str) else ''
    password = raw_password if isinstance(raw_password, str) else ''
    password_confirmation = raw_confirmation if isinstance(raw_confirmation, str) else ''
    if password != password_confirmation:
        return jsonify({'error': 'Passwords do not match.'}), 422
    ok, error = _validate_password(password)
    if not ok:
        return jsonify({'error': error}), 422

    db = get_db()
    token_identity = _account_invitation_digest(token) if token else 'missing'
    if (not _consume_rate_limit(
                db, 'account_invitation_activate_ip', _client_ip(), 30, 3600)
            or not _consume_rate_limit(
                db, 'account_invitation_activate_token', token_identity, 10, 3600)):
        return _rate_limit_response(3600)

    invitation = _valid_account_invitation(db, token)
    if (not invitation or not invitation['active']
            or not _linked_account_is_active(db, invitation)):
        return _invalid_account_invitation_response()

    used_at = now()
    consumed = db.execute(
        '''UPDATE account_invitations SET used_at=?
           WHERE id=? AND token_hash=? AND email_sent_at IS NOT NULL
             AND used_at IS NULL AND invalidated_at IS NULL AND expires_at>?''',
        (used_at, invitation['id'], _account_invitation_digest(token), used_at)
    ).rowcount
    if consumed != 1:
        db.rollback()
        return _invalid_account_invitation_response()

    db.execute(
        '''UPDATE users SET password_hash=?, must_change_password=0,
           password_changed_at=? WHERE id=?''',
        (generate_password_hash(password), used_at, invitation['user_id'])
    )
    _revoke_user_sessions(db, invitation['user_id'])
    _invalidate_account_invitations(
        db, invitation['user_id'], detail='password_created',
        except_invitation_id=invitation['id']
    )
    _record_account_access_event(
        db, invitation['user_id'], 'password_created', invitation['id']
    )
    db.commit()

    confirmation_body = (
        f"Hello {invitation['name'] or invitation['username']},\n\n"
        f"Your Sihha account password was created successfully.\n\n"
        f"Sign in at: https://ops.sihha.org/login\n\n"
        f"If you did not make this change, contact your Sihha coordinator immediately "
        f"at info@sihha.org.\n\n"
        f"— Sihha Food Program"
    )
    confirmation_sent = _email_send(
        invitation['delivery_email'], 'Your Sihha Password Was Created', confirmation_body
    )
    _record_account_access_event(
        db, invitation['user_id'],
        'password_confirmation_sent' if confirmation_sent else 'password_confirmation_failed',
        invitation['id']
    )
    db.commit()
    log.info('Account invitation consumed for user_id=%s', invitation['user_id'])
    return jsonify({
        'ok': True,
        'username': invitation['username'],
        'login_url': '/login',
        'confirmation_email_sent': confirmation_sent,
    })

@app.route('/api/auth/login', methods=['POST'])
def login():
    data = request.json or {}
    username = (data.get('username') or '').strip()
    password = data.get('password') or ''
    if not username or not password:
        return jsonify({'error': 'Username and password required'}), 400

    db = get_db()
    normalized_username = username.lower()
    client_ip = _client_ip()
    ip_user_identity = f'{client_ip}\0{normalized_username}'
    if (
        _rate_limit_blocked(db, 'login_ip_user', ip_user_identity,
                            LOGIN_MAX_FAILS, LOGIN_WINDOW_MIN * 60)
        or _rate_limit_blocked(db, 'login_user', normalized_username,
                               LOGIN_MAX_FAILS_USER, LOGIN_WINDOW_MIN * 60)
    ):
        log.warning(f'Login throttled: {username} from {client_ip}')
        return jsonify({'error': 'Too many failed attempts. Try again in 15 minutes.'}), 429

    user = db.execute(
        "SELECT * FROM users WHERE username=? AND active=1", (username,)
    ).fetchone()
    if not user or not check_password_hash(user['password_hash'], password):
        _rate_limit_record(db, 'login_ip_user', ip_user_identity)
        _rate_limit_record(db, 'login_user', normalized_username)
        db.commit()
        return jsonify({'error': 'Invalid credentials'}), 401
    if not _linked_account_is_active(db, user):
        return jsonify({'error': 'Account inactive'}), 403
    _rate_limit_reset(db, 'login_ip_user', ip_user_identity)
    _rate_limit_reset(db, 'login_user', normalized_username)

    # Update last login timestamp
    db.execute("UPDATE users SET last_login_at=? WHERE id=?", (now(), user['id']))

    # Check if password change is required (first login or admin-forced reset)
    must_change = user['must_change_password'] if 'must_change_password' in user.keys() else 0

    # Check 60-day password expiry (skip for admin to avoid lockout on deploy)
    if not must_change and user['role'] != 'admin':
        changed_at = user['password_changed_at'] if 'password_changed_at' in user.keys() else None
        if changed_at:
            try:
                changed_dt = datetime.fromisoformat(changed_at)
                if (datetime.utcnow() - changed_dt).days >= PASSWORD_EXPIRY_DAYS:
                    must_change = 1
            except Exception:
                pass

    # Issue a short-lived temp token for must_change_password flow
    if must_change:
        # 'tmp_' prefix lets require_auth reject temp tokens — they are ONLY
        # valid for /api/auth/set-password (audit: temp token was a full session)
        temp_token = 'tmp_' + secrets.token_urlsafe(32)
        expires_at = (datetime.utcnow() + timedelta(minutes=30)).isoformat()
        db.execute("UPDATE users SET must_change_password=1 WHERE id=?", (user['id'],))
        db.execute(
            "INSERT INTO sessions (token, user_id, expires_at, created_at) VALUES (?,?,?,?)",
            (temp_token, user['id'], expires_at, now())
        )
        db.commit()
        log.info(f'Login (must_change_password): {username} ({user["role"]})')
        return jsonify({
            'must_change_password': True,
            'temp_token': temp_token,
            'user': {'id': user['id'], 'username': user['username'],
                     'name': user['name'], 'role': user['role']}
        })

    token = secrets.token_urlsafe(32)  # CSPRNG session token (was uuid4)
    expires_at = (datetime.utcnow() + timedelta(hours=SESSION_HOURS)).isoformat()
    db.execute(
        "INSERT INTO sessions (token, user_id, expires_at, created_at) VALUES (?,?,?,?)",
        (token, user['id'], expires_at, now())
    )
    db.commit()
    log.info(f'Login: {username} ({user["role"]})')

    # Determine redirect based on role
    redirect_map = {
        'admin': '/', 'treasurer': '/', 'finance': '/', 'viewer': '/',
        'volunteer': '/portal', 'family': '/family'
    }
    redirect = redirect_map.get(user['role'], '/')

    return jsonify({
        'token': token,
        'redirect': redirect,
        'user': {
            'id': user['id'], 'username': user['username'],
            'name': user['name'], 'role': user['role'],
            'linked_id': user['linked_id'] if 'linked_id' in user.keys() else None
        }
    })

@app.route('/api/auth/logout', methods=['POST'])
def logout():
    auth = request.headers.get('Authorization', '')
    if auth.startswith('Bearer '):
        db = get_db()
        db.execute("DELETE FROM sessions WHERE token=?", (auth[7:],))
        db.commit()
    return jsonify({'ok': True})

@app.route('/api/auth/me')
def me():
    auth = request.headers.get('Authorization', '')
    if not auth.startswith('Bearer '):
        return jsonify({'error': 'Unauthorized'}), 401
    token = auth[7:]
    if token.startswith('tmp_'):
        return jsonify({'error': 'Password change required'}), 401
    session = get_session(token)
    if not session or not session['active'] or not _linked_account_is_active(get_db(), session):
        return jsonify({'error': 'Unauthorized'}), 401
    return jsonify({
        'id': session['user_id'], 'username': session['username'],
        'name': session['name'], 'role': session['role'],
        'linked_id': session['linked_id'], 'linked_type': session['linked_type']
    })

@app.route('/api/auth/set-password', methods=['POST'])
def set_password():
    """Set password on first login or after forced reset. Requires temp_token from login response."""
    data = request.json or {}
    temp_token = (data.get('temp_token') or '').strip()
    new_password = data.get('password') or ''
    if not temp_token or not new_password:
        return jsonify({'error': 'temp_token and password required'}), 400
    if not temp_token.startswith('tmp_'):
        return jsonify({'error': 'A valid temporary password-change token is required'}), 401

    ok, err = _validate_password(new_password)
    if not ok:
        return jsonify({'error': err}), 422

    db = get_db()
    session = get_session(temp_token)
    if (not session or not session['active'] or not session['must_change_password']
            or not _linked_account_is_active(db, session)):
        return jsonify({'error': 'Token expired or invalid — please log in again'}), 401

    db.execute(
        '''UPDATE users SET password_hash=?, must_change_password=0,
           password_changed_at=? WHERE id=?''',
        (generate_password_hash(new_password), now(), session['user_id'])
    )
    # A password reset is a security boundary: expire every pre-reset session.
    _revoke_user_sessions(db, session['user_id'])
    _invalidate_account_invitations(
        db, session['user_id'], actor_user_id=session['user_id'],
        detail='legacy_temporary_password_completed'
    )
    _record_account_access_event(
        db, session['user_id'], 'password_created_legacy',
        actor_user_id=session['user_id']
    )

    # Issue a full session
    token = secrets.token_urlsafe(32)  # CSPRNG session token (was uuid4)
    expires_at = (datetime.utcnow() + timedelta(hours=SESSION_HOURS)).isoformat()
    db.execute(
        "INSERT INTO sessions (token, user_id, expires_at, created_at) VALUES (?,?,?,?)",
        (token, session['user_id'], expires_at, now())
    )
    db.commit()

    redirect_map = {
        'admin': '/', 'treasurer': '/', 'finance': '/', 'viewer': '/',
        'volunteer': '/portal', 'family': '/family'
    }
    redirect = redirect_map.get(session['role'], '/')
    log.info(f'Password set for user_id={session["user_id"]}')
    return jsonify({
        'token': token,
        'redirect': redirect,
        'user': {'id': session['user_id'], 'username': session['username'],
                 'name': session['name'], 'role': session['role'],
                 'linked_id': session['linked_id']}
    })

@app.route('/api/auth/change-password', methods=['POST'])
@require_auth()
def change_password():
    """Logged-in user changes their own password. Requires current + new password."""
    data = request.json or {}
    current_password = data.get('current_password') or ''
    new_password = data.get('new_password') or ''
    if not current_password or not new_password:
        return jsonify({'error': 'current_password and new_password required'}), 400

    ok, err = _validate_password(new_password)
    if not ok:
        return jsonify({'error': err}), 422

    db = get_db()
    user = db.execute("SELECT * FROM users WHERE id=?", (g.user['user_id'],)).fetchone()
    if not user or not check_password_hash(user['password_hash'], current_password):
        return jsonify({'error': 'Current password is incorrect'}), 401

    db.execute(
        '''UPDATE users SET password_hash=?, must_change_password=0,
           password_changed_at=? WHERE id=?''',
        (generate_password_hash(new_password), now(), g.user['user_id'])
    )
    current_token = request.headers.get('Authorization', '')[7:]
    _revoke_user_sessions(db, g.user['user_id'], except_token=current_token)
    _invalidate_account_invitations(
        db, g.user['user_id'], actor_user_id=g.user['user_id'],
        detail='password_changed'
    )
    _record_account_access_event(
        db, g.user['user_id'], 'password_changed', actor_user_id=g.user['user_id']
    )
    db.commit()
    return jsonify({'ok': True, 'message': 'Password updated successfully'})

# ── Users (Admin only) ────────────────────────────────────────────────────────

@app.route('/api/users', methods=['GET'])
@require_auth(roles=['admin'])
def list_users():
    db = get_db()
    rows = db.execute(
        '''SELECT id, username, name, role, email, active, linked_id, linked_type,
                  must_change_password, password_changed_at, last_login_at, created_at
           FROM users ORDER BY created_at'''
    ).fetchall()
    result = []
    for row in rows:
        item = dict(row)
        delivery_email = _account_delivery_email(db, row)
        item['has_access_email'] = bool(delivery_email)
        item['access_email_hint'] = _email_hint(delivery_email)
        result.append(item)
    return jsonify(result)

def _generate_temp_password():
    """Generate a cryptographically secure temp password that meets the rules."""
    import secrets, string
    chars = string.ascii_letters + string.digits + '!@#$%'
    while True:
        pw = ''.join(secrets.choice(chars) for _ in range(12))
        ok, _ = _validate_password(pw)
        if ok:
            return pw

def _generate_unclaimed_password_hash():
    """Create an unknown placeholder credential for invitation-only accounts."""
    return generate_password_hash(secrets.token_urlsafe(48))

@app.route('/api/users', methods=['POST'])
@require_auth(roles=['admin'])
def create_user():
    data = request.json or {}
    if not data.get('username'):
        return jsonify({'error': 'Username required'}), 422
    new_role = data.get('role', 'viewer')
    if new_role not in VALID_ROLES:
        return jsonify({'error': f'Invalid role "{new_role}"'}), 400

    portal_account = new_role in ('family', 'volunteer')
    if portal_account and data.get('password'):
        return jsonify({
            'error': 'Family and volunteer passwords are created through a secure access link.'
        }), 422

    # Portal users receive an unknown placeholder credential and create their own
    # password from a secure invitation. Staff accounts retain the existing
    # explicit/temporary-password creation behavior for this focused portal change.
    raw_password = None if portal_account else (data.get('password') or _generate_temp_password())
    if raw_password:
        ok, err = _validate_password(raw_password)
        if not ok:
            return jsonify({'error': err}), 422

    uid = str(uuid.uuid4())
    linked_id = data.get('linked_id')
    linked_type = data.get('linked_type')
    must_change = (
        1 if portal_account or not data.get('password')
        else int(data.get('must_change_password', 1))
    )

    db = get_db()
    try:
        db.execute(
            '''INSERT INTO users (id, username, password_hash, name, role, email,
               linked_id, linked_type, must_change_password, created_at)
               VALUES (?,?,?,?,?,?,?,?,?,?)''',
            (uid, data['username'],
             _generate_unclaimed_password_hash() if portal_account
             else generate_password_hash(raw_password),
             data.get('name'), new_role, data.get('email'),
             linked_id, linked_type, must_change, now())
        )
        db.commit()
    except sqlite3.IntegrityError as e:
        if 'UNIQUE' in str(e):
            return jsonify({'error': 'Username already exists'}), 409
        return jsonify({'error': str(e)}), 400

    result = {
        'id': uid,
        'username': data['username'],
        'must_change_password': bool(must_change),
    }
    if not portal_account:
        result['temp_password'] = raw_password  # Existing staff-account workflow.
        return jsonify(result), 201

    result['access_email_sent'] = False
    user = db.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
    delivery_email = _account_delivery_email(db, user)
    result['access_email_hint'] = _email_hint(delivery_email) if delivery_email else ''
    if user['active'] and _linked_account_is_active(db, user) and delivery_email:
        access_payload, access_status = _send_account_access_invitation(
            db, user, actor_user_id=g.user['user_id']
        )
        result['access_email_sent'] = access_status == 200
        if access_status != 200:
            result['access_email_error'] = access_payload['error']
    return jsonify(result), 201

@app.route('/api/users/<uid>', methods=['PUT'])
@require_auth(roles=['admin'])
def update_user(uid):
    data = request.json or {}
    db = get_db()
    row = db.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
    if not row:
        return jsonify({'error': 'Not found'}), 404
    new_role = data.get('role', row['role'])
    if new_role not in VALID_ROLES:
        return jsonify({'error': f'Invalid role "{new_role}". Must be one of: {", ".join(sorted(VALID_ROLES))}'}), 400
    if data.get('password'):
        return jsonify({
            'error': 'Use the administrator Reset Password action for credential changes.'
        }), 422

    linked_id = data.get('linked_id', row['linked_id'] if 'linked_id' in row.keys() else None)
    linked_type = data.get('linked_type', row['linked_type'] if 'linked_type' in row.keys() else None)
    must_change = int(data.get('must_change_password', row['must_change_password'] if 'must_change_password' in row.keys() else 0))

    try:
        db.execute(
            '''UPDATE users SET name=?, role=?, active=?, email=?,
               linked_id=?, linked_type=?, must_change_password=? WHERE id=?''',
            (data.get('name', row['name']), new_role, data.get('active', row['active']),
             data.get('email', row['email']),
             linked_id, linked_type, must_change, uid)
        )
        if any(k in data for k in (
                'role', 'active', 'linked_id', 'linked_type',
                'must_change_password')):
            _revoke_user_sessions(db, uid)
        if any(k in data for k in (
                'email', 'role', 'active', 'linked_id',
                'linked_type', 'must_change_password')):
            _invalidate_account_invitations(
                db, uid, actor_user_id=g.user['user_id'], detail='account_updated'
            )
        db.commit()
    except sqlite3.IntegrityError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    return jsonify({'ok': True})

@app.route('/api/users/<uid>/admin-reset-password', methods=['POST'])
@require_auth(roles=['admin'])
def admin_reset_password(uid):
    """Set a temporary password for any account without requiring its old one."""
    data = request.json or {}
    password = data.get('password') or ''
    confirmation = data.get('password_confirmation') or ''
    if not password or not confirmation:
        return jsonify({'error': 'Password and confirmation are required'}), 400
    if password != confirmation:
        return jsonify({'error': 'Passwords do not match'}), 422

    ok, err = _validate_password(password)
    if not ok:
        return jsonify({'error': err}), 422

    db = get_db()
    row = db.execute(
        "SELECT id, username FROM users WHERE id=?", (uid,)
    ).fetchone()
    if not row:
        return jsonify({'error': 'Not found'}), 404

    changed_at = now()
    db.execute(
        '''UPDATE users SET password_hash=?, must_change_password=1,
           password_changed_at=? WHERE id=?''',
        (generate_password_hash(password), changed_at, uid)
    )
    _revoke_user_sessions(db, uid)
    _invalidate_account_invitations(
        db, uid, actor_user_id=g.user['user_id'],
        detail='password_reset_by_admin'
    )
    _record_account_access_event(
        db, uid, 'password_reset_by_admin',
        actor_user_id=g.user['user_id'],
        detail='Temporary password set; change required at next login'
    )
    db.commit()
    log.info('Administrator reset password for user_id=%s', uid)
    return jsonify({
        'ok': True,
        'username': row['username'],
        'must_change_password': True,
    })

@app.route('/api/users/<uid>/force-reset', methods=['POST'])
@require_auth(roles=['admin'])
def force_password_reset(uid):
    """Force a user to change their password on next login."""
    db = get_db()
    row = db.execute("SELECT id FROM users WHERE id=?", (uid,)).fetchone()
    if not row:
        return jsonify({'error': 'Not found'}), 404
    db.execute("UPDATE users SET must_change_password=1 WHERE id=?", (uid,))
    _revoke_user_sessions(db, uid)
    _invalidate_account_invitations(
        db, uid, actor_user_id=g.user['user_id'], detail='forced_password_reset'
    )
    db.commit()
    return jsonify({'ok': True})

@app.route('/api/users/<uid>/send-access-link', methods=['POST'])
@app.route('/api/users/<uid>/reset-password', methods=['POST'])
@require_auth(roles=['admin'])
def send_account_access_link(uid):
    """Email a single-use password-creation link without changing the account first.

    The legacy reset-password URL intentionally aliases this safer behavior so
    older admin clients cannot continue emailing generated passwords.
    """
    db = get_db()
    row = db.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
    if not row:
        return jsonify({'error': 'Not found'}), 404
    if not row['active'] or not _linked_account_is_active(db, row):
        return jsonify({'error': 'Account must be active before access can be sent.'}), 409
    delivery_email = _account_delivery_email(db, row)
    if not delivery_email or not _valid_public_email(delivery_email):
        return jsonify({'error': 'Add a valid email address before sending access.'}), 422
    if not _consume_rate_limit(db, 'account_invitation_send_user', uid, 3, 3600):
        return _rate_limit_response(3600)
    payload, status = _send_account_access_invitation(
        db, row, actor_user_id=g.user['user_id']
    )
    return jsonify(payload), status

@app.route('/api/users/bulk-create', methods=['POST'])
@require_auth(roles=['admin'])
def bulk_create_users():
    """Bulk-create user accounts from existing volunteer or family records.
    Body: {type: 'volunteer'|'family'}
    Secure access links are emailed when the linked record has an address."""
    data = request.json or {}
    kind = data.get('type')
    if kind not in ('volunteer', 'family'):
        return jsonify({'error': 'type must be "volunteer" or "family"'}), 400

    db = get_db()
    if kind == 'volunteer':
        records = db.execute(
            "SELECT id, name, email FROM volunteers WHERE status='active'"
        ).fetchall()
        role = 'volunteer'
        linked_type = 'volunteer'
    else:
        records = db.execute(
            "SELECT id, name, email FROM families WHERE status='active'"
        ).fetchall()
        role = 'family'
        linked_type = 'family'

    created = []
    skipped = []
    for rec in records:
        # Generate username from name (firstname.lastname, lowercase, no spaces)
        name_parts = (rec['name'] or 'user').lower().split()
        base_username = '.'.join(name_parts[:2]) if len(name_parts) >= 2 else name_parts[0]
        username = base_username
        # Make unique if taken
        suffix = 1
        while db.execute("SELECT id FROM users WHERE username=?", (username,)).fetchone():
            username = f'{base_username}{suffix}'
            suffix += 1

        # Skip if linked_id already has an account
        existing = db.execute(
            "SELECT id FROM users WHERE linked_id=? AND linked_type=?",
            (rec['id'], linked_type)
        ).fetchone()
        if existing:
            skipped.append({'id': rec['id'], 'name': rec['name'], 'reason': 'account exists'})
            continue

        uid = str(uuid.uuid4())
        delivery_email = (rec['email'] or '').strip() or None
        db.execute(
            '''INSERT INTO users (id, username, password_hash, name, role, email,
               linked_id, linked_type, must_change_password, created_at)
               VALUES (?,?,?,?,?,?,?,?,1,?)''',
            (uid, username, _generate_unclaimed_password_hash(),
             rec['name'], role, delivery_email,
             rec['id'], linked_type, now())
        )
        db.commit()
        created_item = {
            'id': uid,
            'username': username,
            'name': rec['name'],
            'access_email_sent': False,
            'access_email_hint': _email_hint(delivery_email) if delivery_email else '',
        }
        if delivery_email:
            user = db.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
            payload, status = _send_account_access_invitation(
                db, user, actor_user_id=g.user['user_id']
            )
            created_item['access_email_sent'] = status == 200
            if status != 200:
                created_item['access_email_error'] = payload['error']
        created.append(created_item)

    db.commit()
    return jsonify({'created': created, 'skipped': skipped})


@app.route('/api/admin/fix-schema', methods=['POST'])
@require_auth(roles=['admin'])
def fix_schema():
    """Manual schema repair endpoint — call if role migrations fail at startup."""
    db = get_db()
    before = db.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='users'").fetchone()
    _ensure_treasurer_role(db)
    after = db.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='users'").fetchone()
    had_treasurer_before = before and 'treasurer' in before[0]
    has_treasurer_after  = after  and 'treasurer' in after[0]
    return jsonify({
        'had_treasurer_before': had_treasurer_before,
        'has_treasurer_after':  has_treasurer_after,
        'fixed': not had_treasurer_before and has_treasurer_after,
    })

# ── Public Donate Stats (no auth — safe aggregate data only) ─────────────────

@app.route('/api/donate-stats')
def public_donate_stats():
    """Public endpoint: aggregate donation data for the Wix embed."""
    db = get_db()
    this_month = datetime.utcnow().strftime('%Y-%m')

    # Monthly totals — last 8 months for chart
    monthly_rows = db.execute("""
        SELECT substr(date,1,7) AS month,
               COALESCE(SUM(amount),0) AS total,
               COUNT(*) AS count,
               COUNT(DISTINCT donor_name) AS donors
        FROM donations
        WHERE date >= date('now','-8 months') AND amount > 0
        GROUP BY month
        ORDER BY month ASC
    """).fetchall()
    monthly = [dict(r) for r in monthly_rows]

    # Projection: 6-month run rate + trend
    proj_rows = db.execute("""
        SELECT substr(date,1,7) AS month,
               COUNT(DISTINCT donor_name) AS donors,
               COALESCE(SUM(amount),0) AS total
        FROM donations
        WHERE date >= date('now','-6 months') AND amount > 0
        GROUP BY month ORDER BY month ASC
    """).fetchall()
    proj_rows = [dict(r) for r in proj_rows]

    if proj_rows:
        totals         = [r['total']  for r in proj_rows]
        donor_counts   = [r['donors'] for r in proj_rows]
        avg_monthly    = sum(totals) / len(totals)
        total_donors   = sum(donor_counts)
        total_revenue  = sum(totals)
        avg_gift       = round(total_revenue / total_donors, 2) if total_donors else 0
        avg_donors     = round(sum(donor_counts) / len(donor_counts), 1)
        deltas         = [totals[i] - totals[i-1] for i in range(1, len(totals))]
        monthly_trend  = round(sum(deltas) / len(deltas), 2) if deltas else 0
    else:
        avg_monthly = avg_gift = avg_donors = monthly_trend = 0

    # High-level totals
    total_raised  = db.execute("SELECT COALESCE(SUM(amount),0) FROM donations").fetchone()[0]
    month_raised  = db.execute(
        "SELECT COALESCE(SUM(amount),0) FROM donations WHERE date LIKE ?",
        (f'{this_month}%',)
    ).fetchone()[0]
    families_active = db.execute(
        "SELECT COUNT(*) FROM families WHERE status='active'"
    ).fetchone()[0]
    lives_impacted = db.execute(
        "SELECT COALESCE(SUM(family_size), COUNT(*)) FROM families WHERE status='active'"
    ).fetchone()[0]
    volunteers_active = db.execute(
        "SELECT COUNT(*) FROM volunteers WHERE status='active'"
    ).fetchone()[0]

    # Smart goal + surplus carry-forward
    cost_per_family   = 200
    monthly_need      = families_active * cost_per_family
    last_month        = (datetime.utcnow().replace(day=1) - timedelta(days=1)).strftime('%Y-%m')
    last_month_don    = db.execute(
        "SELECT COALESCE(SUM(amount),0) FROM donations WHERE date LIKE ?",
        (f'{last_month}%',)
    ).fetchone()[0]
    last_month_surplus            = max(0, last_month_don - monthly_need)
    this_month_adjusted_target    = max(0, monthly_need - last_month_surplus)

    # Next delivery — earliest open/shopping/upcoming cycle
    next_cycle = db.execute("""
        SELECT title, delivery_date_start, delivery_date_end
        FROM delivery_cycles
        WHERE status IN ('open','shopping','upcoming')
          AND delivery_date_start IS NOT NULL
        ORDER BY delivery_date_start ASC
        LIMIT 1
    """).fetchone()
    if next_cycle:
        next_delivery_date = next_cycle['delivery_date_start']
        next_delivery_end  = next_cycle['delivery_date_end']
        try:
            today      = datetime.utcnow().date()
            nd         = datetime.strptime(next_delivery_date[:10], '%Y-%m-%d').date()
            days_to_delivery = (nd - today).days
            cycle_start = nd - timedelta(days=30)
            cycle_elapsed = max(0, (today - cycle_start).days)
            cycle_pct   = min(100, round(cycle_elapsed / 30 * 100))
        except Exception:
            days_to_delivery = None
            cycle_pct        = None
        next_cycle_title = next_cycle['title']
    else:
        next_delivery_date = None
        next_delivery_end  = None
        days_to_delivery   = None
        cycle_pct          = None
        next_cycle_title   = None

    return jsonify({
        'donations_by_month':           monthly,
        'proj_avg_donors_per_month':    avg_donors,
        'proj_avg_gift':                avg_gift,
        'proj_avg_monthly':             round(avg_monthly, 2),
        'proj_monthly_trend':           monthly_trend,
        'total_raised':                 total_raised,
        'month_raised':                 month_raised,
        'families_active':              families_active,
        'lives_impacted':               lives_impacted,
        'volunteers_active':            volunteers_active,
        'monthly_need':                 monthly_need,
        'last_month_surplus':           last_month_surplus,
        'this_month_adjusted_target':   this_month_adjusted_target,
        'cost_per_family':              cost_per_family,
        'next_delivery_date':           next_delivery_date,
        'next_delivery_end':            next_delivery_end,
        'days_to_delivery':             days_to_delivery,
        'cycle_pct':                    cycle_pct,
        'next_cycle_title':             next_cycle_title,
    })

# ── Dashboard ─────────────────────────────────────────────────────────────────

@app.route('/api/dashboard/stats')
@require_auth(roles=['admin', 'finance', 'treasurer', 'viewer'])
def dashboard_stats():
    db = get_db()
    this_month = datetime.utcnow().strftime('%Y-%m')

    stats = {
        # Families
        'families_total':    db.execute("SELECT COUNT(*) FROM families").fetchone()[0],
        'families_active':   db.execute("SELECT COUNT(*) FROM families WHERE status='active'").fetchone()[0],
        'families_pending':  db.execute("SELECT COUNT(*) FROM families WHERE status='pending'").fetchone()[0],
        'families_no_wa':    db.execute(
            "SELECT COUNT(*) FROM families WHERE status='active' AND (wa_phone IS NULL OR TRIM(wa_phone)='' OR wa_apikey IS NULL OR TRIM(wa_apikey)='')"
        ).fetchone()[0],
        # Volunteers
        'volunteers_total':  db.execute("SELECT COUNT(*) FROM volunteers").fetchone()[0],
        'volunteers_active': db.execute("SELECT COUNT(*) FROM volunteers WHERE status='active'").fetchone()[0],
        'volunteers_pending':db.execute("SELECT COUNT(*) FROM volunteers WHERE status='pending'").fetchone()[0],
        'volunteers_no_wa':  db.execute(
            "SELECT COUNT(*) FROM volunteers WHERE status='active' AND (wa_phone IS NULL OR TRIM(wa_phone)='' OR wa_apikey IS NULL OR TRIM(wa_apikey)='')"
        ).fetchone()[0],
        # Receipts
        'receipts_pending':  db.execute("SELECT COUNT(*) FROM receipts WHERE status='pending'").fetchone()[0],
        # Reimbursements
        'reimb_pending_count': db.execute(
            "SELECT COUNT(*) FROM reimbursements WHERE status='pending'"
        ).fetchone()[0],
        'reimb_pending_amount': db.execute(
            "SELECT COALESCE(SUM(amount),0) FROM reimbursements WHERE status='pending'"
        ).fetchone()[0],
        'reimb_paid_month': db.execute(
            "SELECT COALESCE(SUM(amount),0) FROM reimbursements WHERE status='paid' AND paid_date LIKE ?",
            (f'{this_month}%',)
        ).fetchone()[0],
        'reimb_paid_total': db.execute(
            "SELECT COALESCE(SUM(amount),0) FROM reimbursements WHERE status='paid'"
        ).fetchone()[0],
        # Spend (paid reimbursements = actual money out to volunteers)
        'spend_this_month':  db.execute(
            "SELECT COALESCE(SUM(amount),0) FROM reimbursements WHERE status='paid' AND paid_date LIKE ?",
            (f'{this_month}%',)
        ).fetchone()[0],
        'spend_total':       db.execute(
            "SELECT COALESCE(SUM(amount),0) FROM reimbursements WHERE status='paid'"
        ).fetchone()[0],
        # Donations
        'donations_month':   db.execute(
            "SELECT COALESCE(SUM(amount),0) FROM donations WHERE date LIKE ?",
            (f'{this_month}%',)
        ).fetchone()[0],
        'donations_total':   db.execute(
            "SELECT COALESCE(SUM(amount),0) FROM donations"
        ).fetchone()[0],
        'donations_count':   db.execute("SELECT COUNT(*) FROM donations").fetchone()[0],
        'donations_recurring_month': db.execute(
            "SELECT COALESCE(SUM(amount),0) FROM donations WHERE date LIKE ? AND frequency='monthly'",
            (f'{this_month}%',)
        ).fetchone()[0],
        # Change requests
        'pending_change_requests': db.execute(
            "SELECT COUNT(*) FROM order_change_requests WHERE status='pending'"
        ).fetchone()[0],
    }

    # Monthly donations — last 6 months for trend chart
    monthly_rows = db.execute("""
        SELECT substr(date,1,7) AS month,
               COALESCE(SUM(amount),0) AS total,
               COUNT(*) AS count,
               COALESCE(SUM(CASE WHEN frequency='monthly' THEN amount ELSE 0 END),0) AS recurring,
               COALESCE(SUM(CASE WHEN frequency<>'monthly' OR frequency IS NULL THEN amount ELSE 0 END),0) AS one_time,
               COUNT(DISTINCT CASE WHEN frequency='monthly' THEN donor_name END) AS recurring_donors
        FROM donations
        WHERE date >= date('now','-6 months')
        GROUP BY month
        ORDER BY month ASC
    """).fetchall()
    stats['donations_by_month'] = [dict(r) for r in monthly_rows]

    # Projection stats: last 6 months run rate + trend slope
    proj_rows = db.execute("""
        SELECT substr(date,1,7)           AS month,
               COUNT(DISTINCT donor_name) AS donors,
               COALESCE(SUM(amount),0)    AS total
        FROM donations
        WHERE date >= date('now','-6 months')
          AND amount > 0
        GROUP BY month
        ORDER BY month ASC
    """).fetchall()
    proj_rows = [dict(r) for r in proj_rows]

    if proj_rows:
        totals         = [r['total']  for r in proj_rows]
        donor_counts   = [r['donors'] for r in proj_rows]
        avg_monthly    = sum(totals) / len(totals)
        total_donors   = sum(donor_counts)
        total_revenue  = sum(totals)
        avg_gift_3mo   = round(total_revenue / total_donors, 2) if total_donors else 0
        avg_donors_3mo = round(sum(donor_counts) / len(donor_counts), 1)
        # Monthly trend: average change month-over-month (positive = growing)
        deltas = [totals[i] - totals[i-1] for i in range(1, len(totals))]
        monthly_trend  = round(sum(deltas) / len(deltas), 2) if deltas else 0
    else:
        avg_monthly    = 0
        avg_gift_3mo   = 0
        avg_donors_3mo = 0
        monthly_trend  = 0

    stats['proj_avg_donors_per_month'] = avg_donors_3mo
    stats['proj_avg_gift']             = avg_gift_3mo
    stats['proj_avg_monthly']          = round(avg_monthly, 2)
    stats['proj_monthly_trend']        = monthly_trend  # $ change per month

    # Active cycle stats — prefer open > shopping > upcoming
    active_cycle = (
        db.execute("SELECT id,title,status,delivery_date_start,delivery_date_end FROM delivery_cycles WHERE status='open' ORDER BY delivery_date_start LIMIT 1").fetchone() or
        db.execute("SELECT id,title,status,delivery_date_start,delivery_date_end FROM delivery_cycles WHERE status='shopping' ORDER BY delivery_date_start LIMIT 1").fetchone() or
        db.execute("SELECT id,title,status,delivery_date_start,delivery_date_end FROM delivery_cycles WHERE status='upcoming' ORDER BY delivery_date_start LIMIT 1").fetchone()
    )
    if active_cycle:
        cid = active_cycle['id']
        stats['cycle_id']      = cid
        stats['cycle_title']   = active_cycle['title']
        stats['cycle_status']  = active_cycle['status']
        stats['orders_this_cycle'] = db.execute(
            "SELECT COUNT(*) FROM food_requests WHERE cycle_id=? AND status='confirmed'", (cid,)
        ).fetchone()[0]
        stats['slots_open']    = db.execute(
            "SELECT COUNT(*) FROM volunteer_slots WHERE cycle_id=? AND status='open'", (cid,)
        ).fetchone()[0]
        stats['slots_claimed'] = db.execute(
            "SELECT COUNT(*) FROM volunteer_slots WHERE cycle_id=? AND status IN ('claimed','confirmed')", (cid,)
        ).fetchone()[0]
        stats['slots_complete']= db.execute(
            "SELECT COUNT(*) FROM volunteer_slots WHERE cycle_id=? AND status='complete'", (cid,)
        ).fetchone()[0]
    else:
        stats.update({'cycle_id': None, 'cycle_title': None, 'cycle_status': None,
                      'orders_this_cycle': 0, 'slots_open': 0, 'slots_claimed': 0, 'slots_complete': 0})

    # Upcoming cycles — next 4 (excluding delivered)
    upcoming_rows = db.execute(
        """SELECT dc.id, dc.title, dc.status, dc.delivery_date_start, dc.delivery_date_end,
                  COUNT(DISTINCT CASE WHEN fr.status='confirmed' THEN fr.id END) AS confirmed_orders,
                  COUNT(DISTINCT CASE WHEN vs.status='open' THEN vs.id END) AS slots_open,
                  COUNT(DISTINCT CASE WHEN vs.status IN ('claimed','confirmed') THEN vs.id END) AS slots_filled
           FROM delivery_cycles dc
           LEFT JOIN food_requests fr ON fr.cycle_id = dc.id
           LEFT JOIN volunteer_slots vs ON vs.cycle_id = dc.id
           WHERE dc.status != 'delivered'
           GROUP BY dc.id
           ORDER BY dc.delivery_date_start ASC
           LIMIT 4"""
    ).fetchall()
    stats['upcoming_cycles'] = [dict(r) for r in upcoming_rows]

    # Smart donation goal: active_families × $200/month = monthly need
    # Surplus carry-forward: if last month exceeded the need, reduce this month's target
    last_month = (datetime.utcnow().replace(day=1) - timedelta(days=1)).strftime('%Y-%m')
    last_month_donations = db.execute(
        "SELECT COALESCE(SUM(amount),0) FROM donations WHERE date LIKE ?",
        (f'{last_month}%',)
    ).fetchone()[0]
    cost_per_family                = 200  # $200 per active family per month
    monthly_need                   = stats['families_active'] * cost_per_family
    last_month_surplus             = max(0, last_month_donations - monthly_need)
    this_month_adjusted_target     = max(0, monthly_need - last_month_surplus)
    next_month_target              = max(0, monthly_need - last_month_donations)

    stats['cost_per_family']               = cost_per_family
    stats['monthly_need']                  = monthly_need
    stats['last_month_donations']          = last_month_donations
    stats['last_month_surplus']            = last_month_surplus
    stats['this_month_adjusted_target']    = this_month_adjusted_target
    stats['next_month_target']             = next_month_target
    stats['last_month']                    = last_month

    return jsonify(stats)

# ── Families ──────────────────────────────────────────────────────────────────

@app.route('/api/families', methods=['GET'])
@require_auth(roles=['admin', 'finance', 'treasurer', 'viewer'])
def list_families():
    db = get_db()
    role = g.user['role']
    status = request.args.get('status')
    search = (request.args.get('search') or '').strip()
    q = "SELECT f.* FROM families f WHERE 1=1"
    params = []
    if status == 'needs_wa':
        q += " AND f.status='active' AND (f.wa_phone IS NULL OR TRIM(f.wa_phone)='' OR f.wa_apikey IS NULL OR TRIM(f.wa_apikey)='')"
    elif status:
        q += " AND f.status=?"; params.append(status)
    if search:
        q += " AND (f.name LIKE ? OR f.phone LIKE ? OR f.address LIKE ?)"; params += [f'%{search}%']*3
    q += " ORDER BY f.created_at DESC"
    rows = [dict(r) for r in db.execute(q, params).fetchall()]

    # Last shopper/deliverer per family (audit 3.3 — replaces 4 correlated subqueries
    # per family row). Audit P2: the previous version pulled EVERY claimed slot in
    # history into Python and overwrote to keep the latest; as history grows that's an
    # unbounded transfer. A ROW_NUMBER() window keeps only the most-recent claimed slot
    # per (family, task_type) in SQL, so at most ~2 rows per family reach Python. Result
    # is identical (rn=1 by created_at DESC == the old ASC-overwrite winner).
    last_map = {}
    for s in db.execute(
        '''SELECT family_id, task_type, vol_name, d_date FROM (
               SELECT vs.family_id AS family_id, vs.task_type AS task_type,
                      v.name AS vol_name, dc.delivery_date_start AS d_date,
                      ROW_NUMBER() OVER (PARTITION BY vs.family_id, vs.task_type
                                         ORDER BY vs.created_at DESC) AS rn
               FROM volunteer_slots vs
               JOIN volunteers v ON vs.claimed_by = v.id
               JOIN delivery_cycles dc ON vs.cycle_id = dc.id
               WHERE vs.claimed_by IS NOT NULL AND vs.task_type IN ('shopping','delivery')
           ) WHERE rn = 1'''
    ).fetchall():
        last_map[(s['family_id'], s['task_type'])] = (s['vol_name'], s['d_date'])
    for r in rows:
        dv = last_map.get((r['id'], 'delivery'), (None, None))
        sv = last_map.get((r['id'], 'shopping'), (None, None))
        r['last_delivery_volunteer'], r['last_delivery_date'] = dv
        r['last_shopping_volunteer'], r['last_shopping_date'] = sv

    # Restrict PII fields for viewer role — full data only for admin/finance/treasurer
    if role == 'viewer':
        SAFE_FIELDS = {'id', 'name', 'family_code', 'bundle_size', 'family_size',
                       'city', 'status', 'created_at', 'last_delivery_volunteer',
                       'last_shopping_volunteer', 'last_delivery_date', 'last_shopping_date'}
        rows = [{k: v for k, v in r.items() if k in SAFE_FIELDS} for r in rows]

    return jsonify(rows)

@app.route('/api/families', methods=['POST'])
@require_auth(roles=['admin', 'finance', 'treasurer'])
def create_family():
    data = request.json or {}
    if not data.get('name'):
        return jsonify({'error': 'Name is required'}), 422
    phone = _normalize_phone(data.get('phone'))
    fid = str(uuid.uuid4())
    db = get_db()
    family_code = _make_family_code(phone, data.get('family_size'), db_conn=db)
    family_email = (data.get('email') or '').strip() or None
    family_status = data.get('status', 'pending')
    db.execute(
        '''INSERT INTO families
           (id,name,phone,address,city,family_size,children_count,
            dietary_notes,frequency,income_range,status,notes,source,family_code,email,created_at)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
        (fid, data['name'], phone, data.get('address'), data.get('city'),
         data.get('family_size'), data.get('children_count'), data.get('dietary_notes'),
         data.get('frequency'), data.get('income_range'),
         family_status, data.get('notes'), data.get('source', 'admin'),
         family_code, family_email, now())
    )

    # Auto-create login account for the family
    name_parts = (data['name'] or 'family').lower().split()
    base_username = '.'.join(name_parts[:2]) if len(name_parts) >= 2 else name_parts[0]
    # Make username unique
    username = base_username
    suffix = 1
    while db.execute("SELECT id FROM users WHERE username=?", (username,)).fetchone():
        username = f'{base_username}{suffix}'
        suffix += 1
    uid = str(uuid.uuid4())
    db.execute(
        '''INSERT INTO users (id, username, password_hash, name, role, email,
           active, linked_id, linked_type, must_change_password, created_at)
           VALUES (?,?,?,?,?,?,?,?,?,1,?)''',
        (uid, username, _generate_unclaimed_password_hash(),
         data['name'], 'family', family_email,
         1 if family_status == 'active' else 0, fid, 'family', now())
    )
    db.commit()

    fam = dict(db.execute("SELECT * FROM families WHERE id=?", (fid,)).fetchone())
    fam['login_username'] = username
    log.info(f'Family created: {data["name"]} — account: {username}')

    fam['access_email_sent'] = False
    fam['access_email_hint'] = _email_hint(family_email) if family_email else ''
    if family_status == 'active' and family_email:
        user = db.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
        access_result, access_status = _send_account_access_invitation(
            db, user, actor_user_id=g.user['user_id']
        )
        fam['access_email_sent'] = access_status == 200
        if access_status != 200:
            fam['access_email_error'] = access_result['error']

    return jsonify(fam), 201

@app.route('/api/families/<fid>', methods=['GET'])
@require_auth(roles=['admin', 'finance', 'treasurer', 'viewer'])
def get_family(fid):
    row = get_db().execute("SELECT * FROM families WHERE id=?", (fid,)).fetchone()
    return (jsonify(dict(row)) if row else (jsonify({'error': 'Not found'}), 404))

@app.route('/api/families/<fid>/preview-token', methods=['POST'])
@require_auth(roles=['admin'])
def family_preview_token(fid):
    """Admin-only: mint a short-lived (2h) family session token for portal testing."""
    db = get_db()
    fam = db.execute("SELECT * FROM families WHERE id=?", (fid,)).fetchone()
    if not fam:
        return jsonify({'error': 'Family not found'}), 404
    # Find or create the family's user account
    user = db.execute(
        "SELECT id FROM users WHERE linked_id=? AND role='family'", (fid,)
    ).fetchone()
    import hashlib, secrets
    if not user:
        # Auto-create a family user if none exists
        uid = str(uuid.uuid4())
        tmp_hash = hashlib.sha256(secrets.token_bytes(32)).hexdigest()
        db.execute(
            "INSERT INTO users (id, username, password_hash, name, role, active, linked_id, created_at) "
            "VALUES (?,?,?,?,?,1,?,?)",
            (uid, f'family_{fam["family_code"]}', tmp_hash, fam['name'], 'family', fid, now())
        )
        db.commit()
        user_id = uid
    else:
        user_id = user['id']
    # Create a 2-hour preview session
    token = secrets.token_urlsafe(32)
    expires = (datetime.utcnow() + timedelta(hours=2)).strftime('%Y-%m-%dT%H:%M:%S')
    db.execute(
        "INSERT INTO sessions (token, user_id, expires_at, created_at) VALUES (?,?,?,?)",
        (token, user_id, expires, now())
    )
    db.commit()
    log.info(f'Admin preview token minted for family {fid} ({fam["name"]})')
    return jsonify({'token': token, 'family_name': fam['name'], 'expires_at': expires})

@app.route('/api/volunteers/<vid>/preview-token', methods=['POST'])
@require_auth(roles=['admin'])
def volunteer_preview_token(vid):
    """Admin-only: mint a short-lived (2h) volunteer portal session for testing —
    lets an admin open /portal as this volunteer."""
    db = get_db()
    vol = db.execute("SELECT * FROM volunteers WHERE id=?", (vid,)).fetchone()
    if not vol:
        return jsonify({'error': 'Volunteer not found'}), 404
    import hashlib, secrets, re
    user = db.execute("SELECT id FROM users WHERE linked_id=? AND role='volunteer'", (vid,)).fetchone()
    if not user:
        uid = str(uuid.uuid4())
        base = 'vol_' + re.sub(r'[^a-z0-9]', '', (vol['name'] or 'volunteer').lower())[:16]
        username, n = base, 1
        while db.execute("SELECT id FROM users WHERE username=?", (username,)).fetchone():
            username = f'{base}{n}'; n += 1
        db.execute(
            "INSERT INTO users (id, username, password_hash, name, role, active, linked_id, linked_type, created_at) "
            "VALUES (?,?,?,?,?,1,?,?,?)",
            (uid, username, hashlib.sha256(secrets.token_bytes(32)).hexdigest(),
             vol['name'], 'volunteer', vid, 'volunteer', now()))
        db.commit()
        user_id = uid
    else:
        user_id = user['id']
    token = secrets.token_urlsafe(32)
    expires = (datetime.utcnow() + timedelta(hours=2)).strftime('%Y-%m-%dT%H:%M:%S')
    db.execute("INSERT INTO sessions (token, user_id, expires_at, created_at) VALUES (?,?,?,?)",
               (token, user_id, expires, now()))
    db.commit()
    log.info(f'Admin preview token minted for volunteer {vid} ({vol["name"]})')
    return jsonify({'token': token, 'volunteer_name': vol['name'], 'expires_at': expires})


@app.route('/api/families/<fid>', methods=['PUT'])
@require_auth(roles=['admin', 'finance', 'treasurer'])
def update_family(fid):
    db = get_db()
    row = db.execute("SELECT * FROM families WHERE id=?", (fid,)).fetchone()
    if not row:
        return jsonify({'error': 'Not found'}), 404
    d = request.json or {}
    new_phone = _normalize_phone(d.get('phone', row['phone']))
    new_size  = d.get('family_size', row['family_size'])
    new_code  = _make_family_code(new_phone, new_size, db_conn=db, exclude_id=fid)
    new_email = (d.get('email', row['email']) or '').strip() or None
    prev_status = row['status']
    new_status  = d.get('status', row['status'])
    db.execute(
        '''UPDATE families SET name=?,phone=?,address=?,city=?,family_size=?,children_count=?,
           dietary_notes=?,frequency=?,income_range=?,status=?,bundle_size=?,notes=?,family_code=?,
           wa_phone=?,wa_apikey=?,email=?,updated_at=? WHERE id=?''',
        (d.get('name', row['name']), new_phone,
         d.get('address', row['address']), d.get('city', row['city']),
         new_size, d.get('children_count', row['children_count']),
         d.get('dietary_notes', row['dietary_notes']), d.get('frequency', row['frequency']),
         d.get('income_range', row['income_range']), new_status,
         d.get('bundle_size', row['bundle_size']),
         d.get('notes', row['notes']), new_code,
         d.get('wa_phone', row['wa_phone']), d.get('wa_apikey', row['wa_apikey']),
         new_email, now(), fid)
    )
    linked_user = db.execute(
        "SELECT id FROM users WHERE linked_id=? AND role='family'", (fid,)
    ).fetchone()
    if linked_user:
        is_active = 1 if new_status == 'active' else 0
        db.execute(
            "UPDATE users SET active=?, email=?, name=? WHERE id=?",
            (is_active, new_email, d.get('name', row['name']), linked_user['id'])
        )
        if not is_active:
            _revoke_user_sessions(db, linked_user['id'])
    db.commit()
    access_result = None
    # When a family is approved or reactivated, ensure the portal account exists
    # and deliver the same secure password-creation link used by Send Access.
    if new_status == 'active' and prev_status != 'active':
        try:
            existing_user = db.execute(
                "SELECT * FROM users WHERE linked_id=? AND role='family'", (fid,)
            ).fetchone()
            if not existing_user:
                fam_name      = d.get('name', row['name'])
                name_parts    = (fam_name or 'family').lower().split()
                base_username = '.'.join(name_parts[:2]) if len(name_parts) >= 2 else name_parts[0]
                username      = base_username
                suffix        = 1
                while db.execute("SELECT id FROM users WHERE username=?", (username,)).fetchone():
                    username = f'{base_username}{suffix}'
                    suffix  += 1
                uid     = str(uuid.uuid4())
                db.execute(
                    '''INSERT INTO users (id, username, password_hash, name, role, email,
                       active, linked_id, linked_type, must_change_password, created_at)
                       VALUES (?,?,?,?,?,?,1,?,?,1,?)''',
                    (uid, username, _generate_unclaimed_password_hash(),
                     fam_name, 'family', new_email, fid, 'family', now())
                )
                db.commit()
                log.info(f'update_family: auto-created account "{username}" for newly active family {fid}')
                existing_user = db.execute(
                    "SELECT * FROM users WHERE id=?", (uid,)
                ).fetchone()
            if new_email:
                payload, status = _send_account_access_invitation(
                    db, existing_user, actor_user_id=g.user['user_id']
                )
                access_result = {
                    'access_email_sent': status == 200,
                    'access_email_hint': _email_hint(new_email),
                }
                if status != 200:
                    access_result['access_email_error'] = payload['error']
            else:
                access_result = {'access_email_sent': False, 'access_email_hint': ''}
        except Exception as _e:
            log.warning(f'update_family: secure account onboarding failed for family {fid}: {_e}')
            access_result = {
                'access_email_sent': False,
                'access_email_error': 'Secure access could not be sent. Use Send Access to retry.'
            }
        # Pre-create volunteer slots for the newly active family.
        try:
            slots = _pre_create_slots_for_family(db, fid)
            db.commit()
            log.info(f'update_family: pre-created {slots} volunteer slots for newly active family {fid}')
        except Exception as _e:
            log.warning(f'update_family: slot pre-creation failed for family {fid}: {_e}')
    result = dict(db.execute("SELECT * FROM families WHERE id=?", (fid,)).fetchone())
    if access_result:
        result.update(access_result)
    return jsonify(result)

@app.route('/api/families/<fid>', methods=['DELETE'])
@require_auth(roles=['admin'])
def delete_family(fid):
    db = get_db()
    row = db.execute("SELECT * FROM families WHERE id=?", (fid,)).fetchone()
    if not row:
        return jsonify({'error': 'Not found'}), 404

    # Financial records are audit evidence, not disposable family children. A
    # hard delete is blocked when any receipt is linked directly, through an
    # assignment, or through a volunteer slot. The coordinator can deactivate
    # the family instead without changing receipt/reimbursement history.
    financial_records = db.execute(
        """SELECT COUNT(DISTINCT r.id)
           FROM receipts r
           LEFT JOIN assignments a ON r.assignment_id=a.id
           LEFT JOIN volunteer_slots vs ON r.slot_id=vs.id
           WHERE r.family_id=? OR a.family_id=? OR vs.family_id=?""",
        (fid, fid, fid)
    ).fetchone()[0]
    if financial_records:
        return jsonify({
            'error': 'This family has financial records and cannot be permanently deleted. '
                     'Set the family to inactive instead.',
            'financial_records': financial_records,
        }), 409

    try:
        db.execute('BEGIN IMMEDIATE')
        request_ids = [r['id'] for r in db.execute(
            "SELECT id FROM food_requests WHERE family_id=?", (fid,)
        ).fetchall()]
        slot_ids = [r['id'] for r in db.execute(
            "SELECT id FROM volunteer_slots WHERE family_id=?", (fid,)
        ).fetchall()]
        user_ids = [r['id'] for r in db.execute(
            "SELECT id FROM users WHERE role='family' AND linked_id=?", (fid,)
        ).fetchall()]

        for rid in request_ids:
            db.execute("DELETE FROM food_request_events WHERE request_id=?", (rid,))
            db.execute("DELETE FROM food_request_items WHERE request_id=?", (rid,))
        for sid in slot_ids:
            db.execute("DELETE FROM reminder_log WHERE slot_id=?", (sid,))
        for uid in user_ids:
            db.execute("DELETE FROM sessions WHERE user_id=?", (uid,))

        db.execute("DELETE FROM order_change_requests WHERE family_id=?", (fid,))
        db.execute("DELETE FROM food_requests          WHERE family_id=?", (fid,))
        db.execute("DELETE FROM volunteer_slots        WHERE family_id=?", (fid,))
        db.execute("DELETE FROM cycle_assignments      WHERE family_id=?", (fid,))
        db.execute("DELETE FROM assignments            WHERE family_id=?", (fid,))
        db.execute("DELETE FROM users WHERE role='family' AND linked_id=?", (fid,))
        db.execute("DELETE FROM families WHERE id=?", (fid,))
        db.commit()
    except sqlite3.IntegrityError as exc:
        db.rollback()
        log.warning(f'delete_family blocked by related records for {fid}: {exc}')
        return jsonify({
            'error': 'This family still has related records and cannot be permanently deleted. '
                     'Set the family to inactive instead.'
        }), 409
    except Exception:
        db.rollback()
        raise
    log.info(f'delete_family: family {fid} ({row["name"]}) permanently deleted by admin')
    return jsonify({'ok': True})

# ── Bundle size change request (family self-serve, coordinator must approve) ──

@app.route('/api/families/<fid>/request-bundle-change', methods=['POST'])
@require_family_auth()
def request_bundle_change(fid):
    """Family portal: request a bundle size change. Stored as pending until coordinator approves."""
    if str(fid) != str(g.fam['family_id']):
        return jsonify({'error': 'Forbidden'}), 403
    data = request.json or {}
    new_size = (data.get('bundle_size') or '').upper().strip()
    if new_size not in ('S', 'M', 'L'):
        return jsonify({'error': 'Bundle size must be S, M, or L'}), 422
    db = get_db()
    family = db.execute("SELECT * FROM families WHERE id=?", (fid,)).fetchone()
    if not family:
        return jsonify({'error': 'Family not found'}), 404
    if family['bundle_size'] == new_size:
        return jsonify({'error': 'That is already your current bundle size'}), 409
    if family['pending_bundle_size'] == new_size:
        return jsonify({'ok': True, 'message': 'Your request is already pending coordinator approval.'}), 200
    db.execute(
        "UPDATE families SET pending_bundle_size=?, updated_at=? WHERE id=?",
        (new_size, now(), fid)
    )
    db.commit()
    # Notify coordinator via email
    sizes = {'S': 'Small', 'M': 'Medium', 'L': 'Large'}
    _notify_coordinators(db,
        f"Bundle size change request:\n"
        f"Family: {family['name']} ({family['family_code']})\n"
        f"Current: {sizes.get(family['bundle_size'] or 'M', family['bundle_size'])}\n"
        f"Requested: {sizes.get(new_size, new_size)}\n"
        f"Please log in to approve or deny."
    )
    log.info(f'Bundle change request: family {fid} → {new_size}')
    return jsonify({'ok': True, 'message': 'Your request has been sent. The coordinator will review it and let you know.'}), 200


@app.route('/api/families/<fid>/approve-bundle-change', methods=['POST'])
@require_auth(roles=['admin'])
def approve_bundle_change(fid):
    """Admin: approve or deny a pending bundle size change request."""
    data = request.json or {}
    action = data.get('action')  # 'approve' or 'deny'
    if action not in ('approve', 'deny'):
        return jsonify({'error': 'action must be approve or deny'}), 422
    db = get_db()
    family = db.execute("SELECT * FROM families WHERE id=?", (fid,)).fetchone()
    if not family:
        return jsonify({'error': 'Not found'}), 404
    if not family['pending_bundle_size']:
        return jsonify({'error': 'No pending bundle size request for this family'}), 409
    if action == 'approve':
        db.execute(
            "UPDATE families SET bundle_size=?, pending_bundle_size=NULL, updated_at=? WHERE id=?",
            (family['pending_bundle_size'], now(), fid)
        )
        msg = f"Approved. Bundle size changed to {family['pending_bundle_size']}."
    else:
        db.execute(
            "UPDATE families SET pending_bundle_size=NULL, updated_at=? WHERE id=?",
            (now(), fid)
        )
        msg = "Bundle size change request denied. Current size kept."
    db.commit()
    log.info(f'Bundle change {action}: family {fid}')
    return jsonify({'ok': True, 'message': msg})


# ── Volunteers ────────────────────────────────────────────────────────────────

@app.route('/api/volunteers', methods=['GET'])
@require_auth(roles=['admin', 'finance', 'treasurer', 'viewer'])
def list_volunteers():
    db = get_db()
    status = request.args.get('status')
    search = (request.args.get('search') or '').strip()
    q = "SELECT * FROM volunteers WHERE 1=1"
    params = []
    if status == 'needs_wa':
        q += " AND status='active' AND (wa_phone IS NULL OR TRIM(wa_phone)='' OR wa_apikey IS NULL OR TRIM(wa_apikey)='')"
    elif status:
        q += " AND status=?"; params.append(status)
    if search:
        q += " AND (name LIKE ? OR phone LIKE ? OR email LIKE ?)"; params += [f'%{search}%']*3
    q += " ORDER BY created_at DESC"
    return jsonify([dict(r) for r in db.execute(q, params).fetchall()])

@app.route('/api/volunteers', methods=['POST'])
@require_auth(roles=['admin'])
def create_volunteer():
    data = request.json or {}
    if not data.get('name'):
        return jsonify({'error': 'Name is required'}), 422
    volunteer_role = _normalize_volunteer_role(data.get('role', 'shopper'))
    if volunteer_role not in VALID_VOLUNTEER_ROLES:
        return jsonify({'error': 'Invalid volunteer role'}), 422
    vid = str(uuid.uuid4())
    db = get_db()
    volunteer_email = (data.get('email') or '').strip() or None
    volunteer_status = data.get('status', 'pending')
    db.execute(
        '''INSERT INTO volunteers
           (id,name,phone,email,role,availability,service_area,
            wa_phone,wa_apikey,status,notes,source,created_at)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)''',
        (vid, data['name'], data.get('phone'), volunteer_email,
         volunteer_role, data.get('availability'), data.get('service_area'),
         data.get('wa_phone'), data.get('wa_apikey'),
         volunteer_status, data.get('notes'), data.get('source', 'admin'), now())
    )

    name_parts = (data['name'] or 'volunteer').lower().split()
    base_username = '.'.join(name_parts[:2]) if len(name_parts) >= 2 else name_parts[0]
    username = base_username
    suffix = 1
    while db.execute("SELECT id FROM users WHERE username=?", (username,)).fetchone():
        username = f'{base_username}{suffix}'
        suffix += 1
    uid = str(uuid.uuid4())
    db.execute(
        '''INSERT INTO users (id, username, password_hash, name, role, email,
           active, linked_id, linked_type, must_change_password, created_at)
           VALUES (?,?,?,?,?,?,?,?,?,1,?)''',
        (uid, username, _generate_unclaimed_password_hash(), data['name'],
         'volunteer', volunteer_email,
         1 if volunteer_status == 'active' else 0, vid, 'volunteer', now())
    )
    db.commit()

    result = dict(db.execute("SELECT * FROM volunteers WHERE id=?", (vid,)).fetchone())
    result['login_username'] = username
    result['access_email_sent'] = False
    result['access_email_hint'] = _email_hint(volunteer_email) if volunteer_email else ''
    if volunteer_status == 'active' and volunteer_email:
        user = db.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
        access_payload, access_status = _send_account_access_invitation(
            db, user, actor_user_id=g.user['user_id']
        )
        result['access_email_sent'] = access_status == 200
        if access_status != 200:
            result['access_email_error'] = access_payload['error']
    log.info(f'Volunteer created: {data["name"]} — account: {username}')
    return jsonify(result), 201

@app.route('/api/volunteers/<vid>', methods=['GET'])
@require_auth(roles=['admin', 'finance', 'treasurer', 'viewer'])
def get_volunteer(vid):
    row = get_db().execute("SELECT * FROM volunteers WHERE id=?", (vid,)).fetchone()
    return (jsonify(dict(row)) if row else (jsonify({'error': 'Not found'}), 404))

@app.route('/api/volunteers/<vid>', methods=['PUT'])
@require_auth(roles=['admin'])
def update_volunteer(vid):
    db = get_db()
    row = db.execute("SELECT * FROM volunteers WHERE id=?", (vid,)).fetchone()
    if not row:
        return jsonify({'error': 'Not found'}), 404
    d = request.json or {}
    prev_status = row['status']
    new_status  = d.get('status', row['status'])
    volunteer_role = _normalize_volunteer_role(d.get('role', row['role']))
    if volunteer_role not in VALID_VOLUNTEER_ROLES:
        return jsonify({'error': 'Invalid volunteer role'}), 422
    new_email = (d.get('email', row['email']) or '').strip() or None
    db.execute(
        '''UPDATE volunteers SET name=?,phone=?,email=?,role=?,availability=?,
           service_area=?,wa_phone=?,wa_apikey=?,status=?,notes=?,updated_at=? WHERE id=?''',
        (d.get('name', row['name']), d.get('phone', row['phone']),
         new_email, volunteer_role,
         d.get('availability', row['availability']), d.get('service_area', row['service_area']),
         d.get('wa_phone', row['wa_phone']), d.get('wa_apikey', row['wa_apikey']),
         new_status, d.get('notes', row['notes']), now(), vid)
    )
    linked_user = db.execute(
        "SELECT id FROM users WHERE linked_id=? AND role='volunteer'", (vid,)
    ).fetchone()
    if linked_user:
        is_active = 1 if new_status == 'active' else 0
        db.execute(
            "UPDATE users SET active=?, email=?, name=? WHERE id=?",
            (is_active, new_email, d.get('name', row['name']), linked_user['id'])
        )
        if not is_active:
            _revoke_user_sessions(db, linked_user['id'])
    db.commit()

    access_result = None
    # Approval and reactivation use the same secure invitation as family accounts.
    if new_status == 'active' and prev_status != 'active':
        vol_name  = d.get('name', row['name'])
        try:
            existing_user = db.execute(
                "SELECT * FROM users WHERE linked_id=? AND role='volunteer'", (vid,)
            ).fetchone()
            if not existing_user:
                name_parts    = (vol_name or 'volunteer').lower().split()
                base_username = '.'.join(name_parts[:2]) if len(name_parts) >= 2 else name_parts[0]
                username      = base_username
                suffix        = 1
                while db.execute("SELECT id FROM users WHERE username=?", (username,)).fetchone():
                    username = f'{base_username}{suffix}'
                    suffix  += 1
                uid     = str(uuid.uuid4())
                db.execute(
                    '''INSERT INTO users (id, username, password_hash, name, role, email,
                       active, linked_id, linked_type, must_change_password, created_at)
                       VALUES (?,?,?,?,?,?,1,?,?,1,?)''',
                    (uid, username, _generate_unclaimed_password_hash(),
                     vol_name, 'volunteer', new_email, vid, 'volunteer', now())
                )
                db.commit()
                log.info(f'update_volunteer: auto-created account "{username}" for newly active volunteer {vid}')
                existing_user = db.execute(
                    "SELECT * FROM users WHERE id=?", (uid,)
                ).fetchone()
            if new_email:
                payload, status = _send_account_access_invitation(
                    db, existing_user, actor_user_id=g.user['user_id']
                )
                access_result = {
                    'access_email_sent': status == 200,
                    'access_email_hint': _email_hint(new_email),
                }
                if status != 200:
                    access_result['access_email_error'] = payload['error']
            else:
                access_result = {'access_email_sent': False, 'access_email_hint': ''}
        except Exception as _e:
            log.warning(f'update_volunteer: secure account onboarding failed for volunteer {vid}: {_e}')
            access_result = {
                'access_email_sent': False,
                'access_email_error': 'Secure access could not be sent. Use Send Access to retry.'
            }

    result = dict(db.execute("SELECT * FROM volunteers WHERE id=?", (vid,)).fetchone())
    if access_result:
        result.update(access_result)
    return jsonify(result)

# ── Assignments routes removed 2026-06-11 (audit 3.4) ────────────────────────
# Legacy /api/assignments CRUD deleted — zero frontend callers; superseded by
# volunteer_slots (Phase 3C). The `assignments` table is retained (historical
# data); drop it in a future migration once confirmed empty/unneeded.

# ── Receipts ──────────────────────────────────────────────────────────────────

# ── Receipt vision parsing (Phase A) ──────────────────────────────────────────

def _active_food_categories_with_items():
    """Live catalog categories, each with its example item names — used to teach the
    model what belongs in each category. Returns [] outside a request / if tables are
    missing. Shape: [{'name': 'Grains', 'items': ['Rice','Pasta','Bread']}, ...]."""
    try:
        db = get_db()
        cats = db.execute(
            "SELECT id, name FROM food_categories WHERE is_active=1 ORDER BY display_order, name"
        ).fetchall()
        out = []
        for c in cats:
            if not (c['name'] or '').strip():
                continue
            items = db.execute(
                "SELECT name FROM food_items WHERE category_id=? AND is_active=1 ORDER BY display_order, name",
                (c['id'],)
            ).fetchall()
            out.append({'name': c['name'], 'items': [i['name'] for i in items if (i['name'] or '').strip()]})
        return out
    except Exception:
        return []


def _build_receipt_prompt(categories):
    """Extraction prompt. `categories` is [{'name','items':[...]}]; when present, teach
    the model each category with its example products and ask it to tag each line item."""
    category_names = [c['name'] for c in categories] if categories else []
    cat_field = ', "category":string|null' if category_names else ''
    cat_rule = ''
    if category_names:
        catalog = '; '.join(
            c['name'] + (' (e.g. ' + ', '.join(c['items'][:8]) + ')' if c['items'] else '')
            for c in categories
        )
        cat_rule = (" For each line item also set \"category\" to the SINGLE best match "
                    "from EXACTLY this list of food-program categories: [" + ", ".join(category_names) + "]. "
                    "Here is what each category contains — match by the KIND of product, even if the "
                    "receipt uses an abbreviated brand name: " + catalog + ". "
                    "Use \"Other\" ONLY for products that clearly don't belong to any of these "
                    "(e.g. household/cleaning supplies, toiletries, bags).")
    return (
        "You are extracting structured data from a photo of a store receipt. "
        "Return ONLY a single minified JSON object, no prose, no markdown fences. "
        "Schema: {\"store\":string|null, \"purchase_date\":\"YYYY-MM-DD\"|null, "
        "\"subtotal\":number|null, \"tax\":number|null, \"total\":number|null, "
        "\"currency\":string|null, \"confidence\":number (0..1), "
        "\"line_items\":[{\"name\":string, \"qty\":number|null, "
        "\"unit_price\":number|null, \"line_total\":number|null" + cat_field + "}]}. "
        "Use the receipt's grand total for \"total\". Omit loyalty/discount summary lines "
        "that are not products. For purchase_date, use ONLY a date actually printed on the "
        "receipt and read the year exactly as shown — NEVER guess a date or year; if no date "
        "is clearly visible, set purchase_date to null." + cat_rule +
        " The image or PDF may be rotated, sideways, or UPSIDE DOWN, or a low-quality "
        "scan — mentally rotate it and read it regardless of orientation. "
        "Only if it is genuinely unreadable, set confidence to 0 and null fields."
    )


def _prepare_receipt_image(image_bytes, ext):
    """Return (base64_str, media_type, block_type) ready for the API, or None if the
    file can't be used. block_type is 'document' for PDFs (Claude reads PDFs natively)
    or 'image' for photos. Converts HEIC→JPEG and downscales large phone photos."""
    import base64 as _b64
    ext = (ext or '').lower().lstrip('.')
    if ext == 'pdf':
        # Claude reads PDFs directly via a base64 document block — no conversion.
        return _b64.b64encode(image_bytes).decode('ascii'), 'application/pdf', 'document'
    try:
        from PIL import Image
        try:
            import pillow_heif  # noqa: registers HEIF/HEIC opener
            pillow_heif.register_heif_opener()
        except Exception:
            if ext in ('heic', 'heif'):
                return None  # can't decode HEIC without the plugin
        import io as _io
        from PIL import ImageOps
        img = Image.open(_io.BytesIO(image_bytes))
        # Respect the camera's EXIF orientation so sideways/upside-down phone photos
        # are uprighted before we send them (the model reads upright far better).
        try:
            img = ImageOps.exif_transpose(img)
        except Exception:
            pass
        if img.mode not in ('RGB', 'L'):
            img = img.convert('RGB')
        # Downscale so the longest side is <= 1600px (plenty for OCR, bounds cost)
        max_side = 1600
        if max(img.size) > max_side:
            ratio = max_side / float(max(img.size))
            img = img.resize((int(img.size[0] * ratio), int(img.size[1] * ratio)))
        buf = _io.BytesIO()
        img.save(buf, format='JPEG', quality=80)
        return _b64.b64encode(buf.getvalue()).decode('ascii'), 'image/jpeg', 'image'
    except Exception as e:
        # PIL missing or decode failed — try sending common web formats as-is.
        if ext in ('jpg', 'jpeg'):
            return _b64.b64encode(image_bytes).decode('ascii'), 'image/jpeg', 'image'
        if ext == 'png':
            return _b64.b64encode(image_bytes).decode('ascii'), 'image/png', 'image'
        log.warning(f'_prepare_receipt_image: could not prepare .{ext} image ({e})')
        return None


def _extract_json(text):
    """Pull the first balanced {...} JSON object out of a model reply and parse it."""
    import json as _json
    if not text:
        return None
    start = text.find('{')
    if start < 0:
        return None
    depth = 0
    for i in range(start, len(text)):
        if text[i] == '{':
            depth += 1
        elif text[i] == '}':
            depth -= 1
            if depth == 0:
                try:
                    return _json.loads(text[start:i + 1])
                except Exception:
                    return None
    return None


def _parse_receipt_image(image_bytes, filename):
    """Back-compat wrapper — returns just the parsed dict (or None)."""
    parsed, _err = _parse_receipt_image_ex(image_bytes, filename)
    return parsed


def _parse_receipt_image_ex(image_bytes, filename):
    """Call the Anthropic vision API to extract receipt data. Returns (parsed_dict, None)
    on success or (None, error_message) on failure. NEVER raises — the app works without
    it. No-op unless RECEIPT_PARSING_ACTIVE (key present + not disabled)."""
    if not RECEIPT_PARSING_ACTIVE:
        return None, 'auto-read is off (no ANTHROPIC_API_KEY)'
    ext = filename.rsplit('.', 1)[-1] if '.' in filename else ''
    prepared = _prepare_receipt_image(image_bytes, ext)
    if not prepared:
        return None, f'could not decode a .{ext or "?"} file (HEIC needs the plugin)'
    b64, media_type, block_type = prepared
    categories = _active_food_categories_with_items()
    category_names = [c['name'] for c in categories]
    import urllib.request as _req, urllib.error as _uerr, json as _json
    body = {
        'model': RECEIPT_PARSE_MODEL,
        'max_tokens': 1500,
        'messages': [{
            'role': 'user',
            'content': [
                {'type': block_type, 'source': {'type': 'base64', 'media_type': media_type, 'data': b64}},
                {'type': 'text', 'text': _build_receipt_prompt(categories)},
            ],
        }],
    }
    req = _req.Request(
        'https://api.anthropic.com/v1/messages',
        data=_json.dumps(body).encode('utf-8'),
        headers={
            'x-api-key': ANTHROPIC_API_KEY,
            'anthropic-version': '2023-06-01',
            'content-type': 'application/json',
        },
        method='POST',
    )
    try:
        with _req.urlopen(req, timeout=45) as resp:
            data = _json.loads(resp.read())
        text = ''.join(b.get('text', '') for b in data.get('content', []) if b.get('type') == 'text')
        parsed = _extract_json(text)
        if not parsed:
            log.warning('_parse_receipt_image: model returned no parseable JSON')
            return None, 'the model reply had no readable data'
        norm = _normalize_parsed_receipt(parsed)
        # Snap each item's category to a canonical catalog name (or 'Other'); leave
        # None if the model gave nothing. Keeps analytics from fragmenting on variants.
        if category_names:
            allowed = {c.lower(): c for c in category_names}
            for it in norm['line_items']:
                c = (it.get('category') or '').strip().lower()
                it['category'] = None if not c else allowed.get(c, 'Other')
        return norm, None
    except _uerr.HTTPError as e:
        detail = e.read().decode('utf-8', 'replace')[:200]
        log.warning(f'_parse_receipt_image HTTP {e.code}: {detail}')
        return None, f'API {e.code}: {detail}'
    except Exception as e:
        log.warning(f'_parse_receipt_image failed: {e}')
        return None, str(e)[:200]


def _to_float(v):
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return None


_STORE_STOPWORDS = {
    'grocer', 'grocers', 'grocery', 'groceries', 'market', 'markets', 'supermarket',
    'store', 'stores', 'foods', 'food', 'wholesale', 'warehouse', 'supercenter',
    'super', 'center', 'centre', 'inc', 'llc', 'co', 'company', 'the', 'and', 'of',
}

def _store_tokens(name):
    """Significant lowercase tokens of a store name (apostrophes dropped, punctuation
    split, generic grocery filler words removed). 'Sam's Club' -> ['sams','club']."""
    import re
    if not name:
        return []
    s = name.lower().replace("'", "").replace("’", "")
    s = re.sub(r'[^a-z0-9 ]', ' ', s)
    toks = [t for t in s.split() if t and t not in _STORE_STOPWORDS]
    if not toks:  # name was entirely filler/punctuation — fall back to raw alnum
        toks = [re.sub(r'[^a-z0-9]', '', name.lower())] if name.strip() else []
    return toks


def _store_key(name):
    """Exact normalized key (used for first-pass grouping + tests)."""
    return ' '.join(_store_tokens(name))


def _stores_similar(a, b):
    """Fuzzy match two (token_set, compact_str) stores. Merges location suffixes,
    spacing/hyphen/typo variants — but keeps genuinely different names apart."""
    ta, sa = a
    tb, sb = b
    if ta and tb:
        shared = ta & tb
        if shared:
            jac = len(shared) / len(ta | tb)
            if jac >= 0.5:                       # strong token overlap / subset
                return True
    import difflib
    return difflib.SequenceMatcher(None, sa, sb).ratio() >= 0.88   # spacing/typos


def _group_by_store(rows, limit=50):
    """Aggregate (store, amount) rows and fuzzily cluster near-duplicate store names.
    Each cluster is labelled with its most-common original spelling. Grouping-only —
    the stored receipt data is never modified."""
    # Pass 1: collapse exact-normalized duplicates and tally raw spellings.
    exact = {}
    for r in rows:
        raw = (r['store'] or '').strip() or '(unknown)'
        toks = _store_tokens(raw)
        key = ' '.join(toks) or raw.lower()
        e = exact.setdefault(key, {'total': 0.0, 'count': 0, 'labels': {},
                                   'tokens': set(toks), 'compact': ''.join(toks) or key})
        e['total'] += (r['amount'] or 0)
        e['count'] += 1
        e['labels'][raw] = e['labels'].get(raw, 0) + 1

    # Pass 2: seed clusters from the largest groups first (so the biggest becomes the
    # canonical), match each remaining group against existing SEEDS (non-transitive).
    clusters = []
    for g in sorted(exact.values(), key=lambda x: -x['total']):
        for cl in clusters:
            if _stores_similar((g['tokens'], g['compact']), (cl['tokens'], cl['compact'])):
                cl['total'] += g['total']; cl['count'] += g['count']
                for lbl, c in g['labels'].items():
                    cl['labels'][lbl] = cl['labels'].get(lbl, 0) + c
                break
        else:
            clusters.append(dict(g))

    out = [{'store': max(cl['labels'].items(), key=lambda kv: kv[1])[0],
            'total': round(cl['total'], 2), 'count': cl['count'],
            'variants': [v for v in cl['labels'].keys() if v != '(unknown)']}
           for cl in clusters]
    out.sort(key=lambda x: -x['total'])
    return out[:limit]


def _cycle_for_date(db, date_str, max_days=10):
    """Best delivery cycle for a receipt purchased on date_str — the cycle whose
    delivery date is nearest the purchase date (volunteers shop right around delivery),
    but only if within max_days. Returns a cycle id or None."""
    if not date_str:
        return None
    from datetime import date as _date
    try:
        d = _date.fromisoformat(str(date_str)[:10])
    except Exception:
        return None
    best, best_diff = None, None
    for r in db.execute("SELECT id, delivery_date_start FROM delivery_cycles WHERE delivery_date_start IS NOT NULL").fetchall():
        try:
            dd = _date.fromisoformat(str(r['delivery_date_start'])[:10])
        except Exception:
            continue
        diff = abs((dd - d).days)
        if best_diff is None or diff < best_diff:
            best, best_diff = r['id'], diff
    return best if (best and best_diff is not None and best_diff <= max_days) else None


def _sane_receipt_date(s):
    """Return a plausible YYYY-MM-DD or None. Guards against the vision model guessing
    a wrong year (e.g. defaulting to 2023) — a receipt date more than ~13 months old or
    in the future is treated as a misread and dropped (falls back to the upload date)."""
    if not s:
        return None
    from datetime import timedelta as _td
    try:
        d = datetime.fromisoformat(str(s)[:10]).date()
    except Exception:
        return None
    today = datetime.utcnow().date()
    if d > today + _td(days=2) or d < today - _td(days=400):
        return None
    return d.isoformat()


def _normalize_parsed_receipt(p):
    """Coerce the model's JSON into a clean, typed dict we can trust downstream."""
    items = []
    for it in (p.get('line_items') or [])[:200]:
        if not isinstance(it, dict):
            continue
        name = (it.get('name') or '').strip()
        if not name:
            continue
        items.append({
            'name':       name[:200],
            'qty':        _to_float(it.get('qty')),
            'unit_price': _to_float(it.get('unit_price')),
            'line_total': _to_float(it.get('line_total')),
            'category':   (it.get('category') or '').strip()[:60] or None,
        })
    conf = _to_float(p.get('confidence'))
    if conf is not None:
        conf = max(0.0, min(1.0, conf))
    return {
        'store':         (p.get('store') or '').strip()[:200] or None,
        'purchase_date': _sane_receipt_date((p.get('purchase_date') or '').strip()[:10]),
        'subtotal':      _to_float(p.get('subtotal')),
        'tax':           _to_float(p.get('tax')),
        'total':         _to_float(p.get('total')),
        'currency':      (p.get('currency') or '').strip()[:8] or None,
        'confidence':    conf,
        'line_items':    items,
    }


def _persist_receipt_parse(db, receipt_id, parsed, confirmed_amount=None):
    """Write parsed_* fields + receipt_items for a receipt and flag amount mismatches.
    Safe to call repeatedly (clears prior items first). Returns the mismatch flag."""
    import json as _json
    if parsed is None:
        db.execute("UPDATE receipts SET parse_status='failed', parsed_at=? WHERE id=?",
                   (now(), receipt_id))
        return 0
    mismatch = 0
    total = parsed.get('total')
    if confirmed_amount is not None and total is not None:
        try:
            mismatch = 1 if abs(float(confirmed_amount) - total) > 0.02 else 0
        except (TypeError, ValueError):
            mismatch = 0
    db.execute(
        "UPDATE receipts SET parsed_store=?, parsed_date=?, subtotal=?, tax=?, parsed_total=?, "
        "parse_status='parsed', parse_confidence=?, parse_model=?, parsed_at=?, parsed_json=?, "
        "amount_mismatch=? WHERE id=?",
        (parsed.get('store'), parsed.get('purchase_date'), parsed.get('subtotal'),
         parsed.get('tax'), parsed.get('total'), parsed.get('confidence'), RECEIPT_PARSE_MODEL,
         now(), _json.dumps(parsed), mismatch, receipt_id)
    )
    db.execute("DELETE FROM receipt_items WHERE receipt_id=?", (receipt_id,))
    for i, it in enumerate(parsed.get('line_items') or []):
        db.execute(
            "INSERT INTO receipt_items (id, receipt_id, line_no, name, qty, unit_price, line_total, category, created_at) "
            "VALUES (?,?,?,?,?,?,?,?,?)",
            (str(uuid.uuid4()), receipt_id, i + 1, it['name'], it['qty'],
             it['unit_price'], it['line_total'], it.get('category'), now())
        )
    return mismatch


@app.route('/api/receipts', methods=['GET'])
@require_auth(roles=['admin', 'finance', 'treasurer'])  # audit: was open to any
def list_receipts():                                    # authenticated role

    db = get_db()
    status = request.args.get('status')
    q = '''SELECT r.*, f.name as family_name, v.name as volunteer_name,
              dc.title as cycle_title,
              (SELECT COUNT(*) FROM receipt_items ri WHERE ri.receipt_id = r.id) as item_count,
              COALESCE(r.cycle_id, vs.cycle_id) as resolved_cycle_id
           FROM receipts r
           LEFT JOIN families f ON r.family_id = f.id
           LEFT JOIN volunteers v ON r.volunteer_id = v.id
           LEFT JOIN volunteer_slots vs ON r.slot_id = vs.id
           LEFT JOIN delivery_cycles dc ON COALESCE(r.cycle_id, vs.cycle_id) = dc.id
           WHERE 1=1'''
    params = []
    if status:
        q += " AND r.status=?"; params.append(status)
    q += " ORDER BY r.created_at DESC"
    return jsonify([dict(r) for r in db.execute(q, params).fetchall()])

@app.route('/api/receipts/parse-diagnostics', methods=['GET'])
@require_auth(roles=['admin'])
def receipt_parse_diagnostics():
    """Tell the admin exactly why receipt auto-read is or isn't working. Does a live
    text-only call to the Anthropic API so key/model problems surface with the real
    error (e.g. 401 bad key, 404 model-not-found)."""
    info = {
        'enable_flag':  ENABLE_RECEIPT_PARSING,
        'has_api_key':  bool(ANTHROPIC_API_KEY),
        'active':       RECEIPT_PARSING_ACTIVE,
        'model':        RECEIPT_PARSE_MODEL,
        'heic_support': False,
    }
    try:
        from PIL import Image  # noqa
        import pillow_heif      # noqa
        info['heic_support'] = True
    except Exception:
        info['heic_support'] = False
    if not RECEIPT_PARSING_ACTIVE:
        if not ANTHROPIC_API_KEY:
            reason = 'Auto-read is OFF: no ANTHROPIC_API_KEY set in Railway. Add it and redeploy.'
        else:
            reason = 'Auto-read is OFF: ENABLE_RECEIPT_PARSING is set to a disabling value ' \
                     '(0/false/no/off). Remove it (or set 1) and redeploy.'
        info['test'] = {'ok': False, 'reason': reason}
        return jsonify(info)
    import urllib.request as _req, urllib.error as _uerr, json as _json
    body = {'model': RECEIPT_PARSE_MODEL, 'max_tokens': 16,
            'messages': [{'role': 'user', 'content': 'Reply with just: OK'}]}
    req = _req.Request('https://api.anthropic.com/v1/messages',
        data=_json.dumps(body).encode(), method='POST',
        headers={'x-api-key': ANTHROPIC_API_KEY, 'anthropic-version': '2023-06-01',
                 'content-type': 'application/json'})
    try:
        with _req.urlopen(req, timeout=30) as resp:
            data = _json.loads(resp.read())
        txt = ''.join(b.get('text', '') for b in data.get('content', []) if b.get('type') == 'text')
        info['test'] = {'ok': True, 'model_replied': txt[:50]}
    except _uerr.HTTPError as e:
        info['test'] = {'ok': False, 'status': e.code,
                        'error': e.read().decode('utf-8', 'replace')[:300]}
    except Exception as e:
        info['test'] = {'ok': False, 'error': str(e)[:300]}

    # Image/vision test — this is what actually matters for receipts. Text can pass
    # while image input fails (e.g. a model without vision), so test a real image.
    try:
        from PIL import Image as _Img
        import io as _io, base64 as _b64
        buf = _io.BytesIO(); _Img.new('RGB', (64, 64), (255, 255, 255)).save(buf, format='JPEG')
        ib64 = _b64.b64encode(buf.getvalue()).decode('ascii')
        ibody = {'model': RECEIPT_PARSE_MODEL, 'max_tokens': 16, 'messages': [{'role': 'user', 'content': [
            {'type': 'image', 'source': {'type': 'base64', 'media_type': 'image/jpeg', 'data': ib64}},
            {'type': 'text', 'text': 'Reply with just: OK'}]}]}
        ireq = _req.Request('https://api.anthropic.com/v1/messages',
            data=_json.dumps(ibody).encode(), method='POST',
            headers={'x-api-key': ANTHROPIC_API_KEY, 'anthropic-version': '2023-06-01',
                     'content-type': 'application/json'})
        with _req.urlopen(ireq, timeout=30) as r:
            _json.loads(r.read())
        info['image_test'] = {'ok': True}
    except _uerr.HTTPError as e:
        info['image_test'] = {'ok': False, 'status': e.code,
                              'error': e.read().decode('utf-8', 'replace')[:300]}
    except Exception as e:
        info['image_test'] = {'ok': False, 'error': str(e)[:300]}
    return jsonify(info)


@app.route('/api/receipts', methods=['POST'])
@require_auth(roles=['admin', 'finance', 'treasurer'])
def create_receipt():
    data = request.json or {}
    file_url = _normalize_upload_url(data.get('file_url'))
    if data.get('file_url') and not file_url:
        return jsonify({'error': 'Invalid receipt file URL'}), 422
    if data.get('amount') is not None:
        try:
            if float(data['amount']) < 0:
                return jsonify({'error': 'Amount cannot be negative'}), 422
        except (TypeError, ValueError):
            return jsonify({'error': 'Invalid amount'}), 422
    rid = str(uuid.uuid4())
    db = get_db()
    if not _claim_registered_upload(db, file_url):
        return jsonify({'error': 'Receipt file is already attached or unavailable'}), 422
    # Auto-match a delivery cycle from the purchase date when the caller didn't set one
    # (dashboard uploads) — keeps them out of the "Unassigned" bucket.
    cycle_id = data.get('cycle_id')
    if not cycle_id and not data.get('slot_id'):
        # Match on the purchase date; fall back to today's upload date when the date is
        # missing or doesn't land near a delivery (a good proxy — receipts are uploaded
        # around when the volunteer shopped).
        cycle_id = _cycle_for_date(db, data.get('purchase_date')) or _cycle_for_date(db, now())
    db.execute(
        '''INSERT INTO receipts
           (id,assignment_id,volunteer_id,family_id,store,purchase_date,amount,file_url,slot_id,cycle_id,status,notes,created_at)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)''',
        (rid, data.get('assignment_id'), data.get('volunteer_id'), data.get('family_id'),
         data.get('store'), data.get('purchase_date'), data.get('amount'),
         file_url, data.get('slot_id'), cycle_id, 'pending', data.get('notes'), now())
    )
    if data.get('parsed'):
        try:
            _persist_receipt_parse(db, rid, _normalize_parsed_receipt(data['parsed']), data.get('amount'))
        except Exception as _e:
            log.warning(f'create_receipt: parse persist failed for {rid}: {_e}')
    db.commit()
    # Notify treasurers of new receipt submission (skip for admin bulk quick-upload,
    # which passes notify=false so N uploads don't send N emails).
    if data.get('notify', True):
        try:
            vol = db.execute("SELECT name FROM volunteers WHERE id=?", (data.get('volunteer_id'),)).fetchone()
            vol_name = vol['name'] if vol else 'A volunteer'
            amount = data.get('amount') or 0
            store  = data.get('store') or 'unknown store'
            subject = f'New Reimbursement Request — ${amount:.2f} from {vol_name}'
            msg = (f'New receipt submitted on Sihha Ops Hub.\n'
                   f'Volunteer: {vol_name}\n'
                   f'Store: {store}\n'
                   f'Amount: ${amount:.2f}\n'
                   f'Date: {data.get("purchase_date","")}\n\n'
                   f'Log in to review and pay: https://sihha-ops-hub-production.up.railway.app')
            _notify_treasurers(db, subject, msg)
        except Exception as e:
            log.warning(f'Treasurer notification failed: {e}')
    return jsonify({'id': rid}), 201

def _recompute_reimbursable(db, rid):
    """Reimbursable = confirmed receipt total minus the line_total of any excluded
    items (clamped at 0). Persist it on the receipt and keep an unpaid (pending)
    reimbursement's amount in sync. Paid reimbursements are never touched."""
    row = db.execute("SELECT amount FROM receipts WHERE id=?", (rid,)).fetchone()
    if not row:
        return None
    amt = float(row['amount'] or 0)
    excl = db.execute(
        "SELECT COALESCE(SUM(line_total),0) s FROM receipt_items "
        "WHERE receipt_id=? AND excluded=1", (rid,)
    ).fetchone()['s'] or 0
    reimb = round(max(0.0, amt - float(excl)), 2)
    db.execute("UPDATE receipts SET reimbursable_amount=? WHERE id=?", (reimb, rid))
    db.execute("UPDATE reimbursements SET amount=?, updated_at=? "
               "WHERE receipt_id=? AND status='pending'", (reimb, now(), rid))
    return reimb


@app.route('/api/receipts/<rid>', methods=['GET'])
@require_auth(roles=['admin', 'finance', 'treasurer'])
def get_receipt(rid):
    """Full receipt detail incl. line items — powers the editable detail view."""
    db = get_db()
    r = db.execute(
        '''SELECT r.*, v.name as volunteer_name, f.name as family_name,
                  dc.title as cycle_title, COALESCE(r.cycle_id, vs.cycle_id) as resolved_cycle_id
           FROM receipts r
           LEFT JOIN volunteers v ON r.volunteer_id = v.id
           LEFT JOIN families f ON r.family_id = f.id
           LEFT JOIN volunteer_slots vs ON r.slot_id = vs.id
           LEFT JOIN delivery_cycles dc ON COALESCE(r.cycle_id, vs.cycle_id) = dc.id
           WHERE r.id=?''', (rid,)
    ).fetchone()
    if not r:
        return jsonify({'error': 'Not found'}), 404
    items = db.execute(
        "SELECT id, line_no, name, qty, unit_price, line_total, category, "
        "COALESCE(excluded,0) as excluded "
        "FROM receipt_items WHERE receipt_id=? ORDER BY line_no", (rid,)
    ).fetchall()
    out = dict(r)
    out['items'] = [dict(i) for i in items]
    # Effective reimbursable (falls back to amount when no exclusions recorded yet).
    out['reimbursable_amount'] = (r['reimbursable_amount']
                                  if r['reimbursable_amount'] is not None else r['amount'])
    out['excluded_total'] = round(sum((i['line_total'] or 0) for i in items if i['excluded']), 2)
    # Is the linked reimbursement already paid? (locks excluded-item editing)
    rb = db.execute("SELECT status FROM reimbursements WHERE receipt_id=? "
                    "ORDER BY (status='paid') DESC LIMIT 1", (rid,)).fetchone()
    out['reimb_status'] = rb['status'] if rb else None
    out['reimb_paid'] = bool(rb and rb['status'] == 'paid')
    return jsonify(out)


@app.route('/api/receipts/<rid>', methods=['PUT'])
@require_auth(roles=['admin', 'finance', 'treasurer'])
def update_receipt(rid):
    db = get_db()
    row = db.execute("SELECT * FROM receipts WHERE id=?", (rid,)).fetchone()
    if not row:
        return jsonify({'error': 'Not found'}), 404
    d = request.json or {}

    # Editable header fields — assign a volunteer, fix store/date/amount/cycle, etc.
    # Only overwrite a field when the caller actually sent it (partial update).
    def pick(key):
        return d[key] if key in d else row[key]
    new_amount = pick('amount')
    if new_amount is not None:
        try:
            if float(new_amount) < 0:
                return jsonify({'error': 'Amount cannot be negative'}), 422
        except (TypeError, ValueError):
            return jsonify({'error': 'Invalid amount'}), 422

    # Recompute the parsed-vs-confirmed mismatch flag if we have a parsed total.
    mismatch = row['amount_mismatch'] if 'amount_mismatch' in row.keys() else 0
    parsed_total = row['parsed_total'] if 'parsed_total' in row.keys() else None
    if parsed_total is not None and new_amount is not None:
        try:
            mismatch = 1 if abs(float(new_amount) - float(parsed_total)) > 0.02 else 0
        except (TypeError, ValueError):
            pass

    db.execute(
        "UPDATE receipts SET volunteer_id=?, family_id=?, cycle_id=?, store=?, "
        "purchase_date=?, amount=?, status=?, notes=?, amount_mismatch=?, updated_at=? WHERE id=?",
        (pick('volunteer_id'), pick('family_id'), pick('cycle_id'), pick('store'),
         pick('purchase_date'), new_amount, pick('status'), pick('notes'),
         mismatch, now(), rid)
    )

    # Keep a pending reimbursement's volunteer in sync with the receipt.
    if 'volunteer_id' in d:
        db.execute(
            "UPDATE reimbursements SET volunteer_id=?, updated_at=? "
            "WHERE receipt_id=? AND status='pending'",
            (pick('volunteer_id'), now(), rid)
        )
    # If the total changed, recompute reimbursable (respects excluded items) and
    # push it onto any pending reimbursement.
    if 'amount' in d:
        _recompute_reimbursable(db, rid)

    # Approval is the single point money is committed: create the payable (unpaid)
    # exactly once when the receipt first becomes 'approved'. Amount owed is the
    # reimbursable total, not the raw receipt total.
    if d.get('status') == 'approved' and row['status'] != 'approved':
        existing = db.execute("SELECT id FROM reimbursements WHERE receipt_id=?", (rid,)).fetchone()
        if not existing:
            reimb_amt = _recompute_reimbursable(db, rid)   # no pending reimb yet, just sets the receipt figure
            db.execute(
                '''INSERT INTO reimbursements
                   (id,receipt_id,volunteer_id,amount,status,approved_by,created_at)
                   VALUES (?,?,?,?,?,?,?)''',
                (str(uuid.uuid4()), rid, pick('volunteer_id'),
                 reimb_amt if reimb_amt is not None else new_amount,
                 'pending', g.user['user_id'], now())
            )
    # Un-approving (back to pending) or rejecting removes the UNPAID payable.
    if d.get('status') in ('pending', 'rejected') and row['status'] == 'approved':
        db.execute("DELETE FROM reimbursements WHERE receipt_id=? AND status!='paid'", (rid,))
    db.commit()
    return jsonify(dict(db.execute("SELECT * FROM receipts WHERE id=?", (rid,)).fetchone()))


@app.route('/api/receipts/<rid>/reparse', methods=['POST'])
@require_auth(roles=['admin', 'finance', 'treasurer'])
def reparse_receipt(rid):
    """Re-run vision parsing on an existing receipt's stored photo. Fills empty
    store/date/amount from the parse. Used for receipts uploaded before the API key
    was working, or to retry a failed read."""
    if not RECEIPT_PARSING_ACTIVE:
        return jsonify({'error': 'Auto-read is off — no valid ANTHROPIC_API_KEY.'}), 400
    db = get_db()
    row = db.execute("SELECT * FROM receipts WHERE id=?", (rid,)).fetchone()
    if not row:
        return jsonify({'error': 'Not found'}), 404
    if not row['file_url']:
        return jsonify({'error': 'This receipt has no photo to read.'}), 400
    fname = row['file_url'].rsplit('/', 1)[-1]
    path = os.path.join(UPLOAD_FOLDER, fname)
    if not os.path.exists(path):
        return jsonify({'error': 'Photo file not found on the server.'}), 404
    with open(path, 'rb') as f:
        raw = f.read()
    parsed, perr = _parse_receipt_image_ex(raw, fname)
    if not parsed:
        db.execute("UPDATE receipts SET parse_status='failed', parsed_at=? WHERE id=?", (now(), rid))
        db.commit()
        return jsonify({'error': 'Could not read this photo — ' + (perr or 'unknown reason')}), 422
    _persist_receipt_parse(db, rid, parsed, row['amount'])
    # Fill any empty header fields from the parse (don't clobber values already set).
    sets, vals = [], []
    if not row['store'] and parsed.get('store'):
        sets.append('store=?'); vals.append(parsed['store'])
    if not row['purchase_date'] and parsed.get('purchase_date'):
        sets.append('purchase_date=?'); vals.append(parsed['purchase_date'])
    if row['amount'] is None and parsed.get('total') is not None:
        sets.append('amount=?'); vals.append(parsed['total'])
    if sets:
        sets.append('updated_at=?'); vals.append(now()); vals.append(rid)
        db.execute(f"UPDATE receipts SET {', '.join(sets)} WHERE id=?", vals)
    db.commit()
    return jsonify({'ok': True, 'parsed': parsed})


@app.route('/api/receipts/delete-all', methods=['POST'])
@require_auth(roles=['admin'])
def delete_all_receipts():
    """Wipe every receipt + its line items + linked reimbursements. Admin-only and
    requires an explicit confirm phrase so it can't fire by accident."""
    if (request.json or {}).get('confirm') != 'DELETE ALL':
        return jsonify({'error': 'Confirmation required (confirm="DELETE ALL").'}), 400
    db = get_db()
    n = db.execute("SELECT COUNT(*) FROM receipts").fetchone()[0]
    db.execute("DELETE FROM receipt_items")
    db.execute("DELETE FROM reimbursements")   # every reimbursement is receipt-linked
    db.execute("DELETE FROM receipts")
    db.commit()
    log.info(f'delete_all_receipts: wiped {n} receipts by admin {g.user["username"]}')
    return jsonify({'ok': True, 'deleted': n})


@app.route('/api/receipts/bulk-approve', methods=['POST'])
@require_auth(roles=['admin', 'finance', 'treasurer'])
def bulk_approve_receipts():
    """Approve several pending receipts at once — creates a payable for each. Skips
    any id that isn't currently pending. Returns how many were approved."""
    ids = (request.json or {}).get('ids') or []
    if not isinstance(ids, list) or not ids:
        return jsonify({'error': 'No receipts selected'}), 400
    db = get_db()
    approved = 0
    for rid in ids:
        row = db.execute("SELECT * FROM receipts WHERE id=?", (rid,)).fetchone()
        if not row or row['status'] != 'pending':
            continue
        db.execute("UPDATE receipts SET status='approved', updated_at=? WHERE id=?", (now(), rid))
        if not db.execute("SELECT id FROM reimbursements WHERE receipt_id=?", (rid,)).fetchone():
            db.execute(
                '''INSERT INTO reimbursements
                   (id,receipt_id,volunteer_id,amount,status,approved_by,created_at)
                   VALUES (?,?,?,?,?,?,?)''',
                (str(uuid.uuid4()), rid, row['volunteer_id'], row['amount'],
                 'pending', g.user['user_id'], now())
            )
        approved += 1
    db.commit()
    return jsonify({'ok': True, 'approved': approved})


@app.route('/api/receipts/bulk-assign', methods=['POST'])
@require_auth(roles=['admin', 'finance', 'treasurer'])
def bulk_assign_receipts():
    """Assign (or clear) a volunteer on several receipts at once. Also updates any
    UNPAID payable so the reimbursement follows the assignment."""
    d = request.json or {}
    ids = d.get('ids') or []
    vid = d.get('volunteer_id') or None
    if not isinstance(ids, list) or not ids:
        return jsonify({'error': 'No receipts selected'}), 400
    if vid and not get_db().execute("SELECT id FROM volunteers WHERE id=?", (vid,)).fetchone():
        return jsonify({'error': 'Volunteer not found'}), 404
    db = get_db()
    n = 0
    for rid in ids:
        if db.execute("SELECT id FROM receipts WHERE id=?", (rid,)).fetchone():
            db.execute("UPDATE receipts SET volunteer_id=?, updated_at=? WHERE id=?", (vid, now(), rid))
            db.execute("UPDATE reimbursements SET volunteer_id=?, updated_at=? WHERE receipt_id=? AND status!='paid'",
                       (vid, now(), rid))
            n += 1
    db.commit()
    return jsonify({'ok': True, 'assigned': n})


@app.route('/api/receipt-items/<item_id>', methods=['PUT'])
@require_auth(roles=['admin', 'finance', 'treasurer'])
def update_receipt_item(item_id):
    """Edit a line item: override its spending category and/or exclude it from
    reimbursement (a volunteer's personal charge on a shared receipt). Excluding
    subtracts the line's total from what's owed; blocked once the receipt is paid."""
    db = get_db()
    item = db.execute("SELECT id, receipt_id FROM receipt_items WHERE id=?", (item_id,)).fetchone()
    if not item:
        return jsonify({'error': 'Not found'}), 404
    d = request.json or {}
    rid = item['receipt_id']

    resp = {'ok': True}
    if 'category' in d:
        cat = (d.get('category') or '').strip()[:60] or None
        db.execute("UPDATE receipt_items SET category=? WHERE id=?", (cat, item_id))
        resp['category'] = cat

    if 'excluded' in d:
        # Can't change what's owed after it's been paid out.
        paid = db.execute("SELECT 1 FROM reimbursements WHERE receipt_id=? AND status='paid'",
                          (rid,)).fetchone()
        if paid:
            return jsonify({'error': 'This receipt has already been paid; undo the payment before changing excluded items.'}), 409
        db.execute("UPDATE receipt_items SET excluded=? WHERE id=?",
                   (1 if d['excluded'] else 0, item_id))
        resp['reimbursable_amount'] = _recompute_reimbursable(db, rid)

    db.commit()
    return jsonify(resp)


@app.route('/api/receipts/rename-store', methods=['POST'])
@require_auth(roles=['admin', 'finance', 'treasurer'])
def rename_store():
    """Rename a store across every receipt that uses one of its spelling variants —
    standardizes the underlying data (list, detail, exports all update). `from` is the
    exact list of raw store strings to replace (the cluster's variants)."""
    d = request.json or {}
    variants = d.get('from') or []
    to = (d.get('to') or '').strip()
    if not isinstance(variants, list) or not variants or not to:
        return jsonify({'error': 'Provide the store variants and a new name.'}), 400
    db = get_db()
    ph = ','.join('?' * len(variants))
    cur = db.execute(f"UPDATE receipts SET store=?, updated_at=? WHERE store IN ({ph})",
                     [to, now(), *variants])
    db.commit()
    return jsonify({'ok': True, 'updated': cur.rowcount})


@app.route('/api/receipts/auto-match-cycle', methods=['POST'])
@require_auth(roles=['admin', 'finance', 'treasurer'])
def auto_match_cycle():
    """Auto-assign a delivery cycle from each receipt's purchase date. With `ids`, only
    those; otherwise every currently-unassigned, non-rejected receipt with a date."""
    ids = (request.json or {}).get('ids')
    db = get_db()
    if ids:
        rows = db.execute(f"SELECT id, purchase_date, created_at FROM receipts WHERE id IN ({','.join('?'*len(ids))})", ids).fetchall()
    else:
        rows = db.execute("SELECT id, purchase_date, created_at FROM receipts "
                          "WHERE cycle_id IS NULL AND slot_id IS NULL AND status!='rejected'").fetchall()
    matched = 0
    for r in rows:
        # Purchase date first; fall back to the upload date when it's missing or a
        # misread (e.g. a stray 2023) that lands near no delivery.
        cid = _cycle_for_date(db, r['purchase_date']) or _cycle_for_date(db, r['created_at'])
        if cid:
            db.execute("UPDATE receipts SET cycle_id=?, updated_at=? WHERE id=?", (cid, now(), r['id']))
            matched += 1
    db.commit()
    return jsonify({'ok': True, 'matched': matched, 'checked': len(rows)})


@app.route('/api/receipts/bulk-assign-cycle', methods=['POST'])
@require_auth(roles=['admin', 'finance', 'treasurer'])
def bulk_assign_cycle():
    """Attach several receipts to a delivery cycle at once (or clear the cycle)."""
    d = request.json or {}
    ids = d.get('ids') or []
    cid = d.get('cycle_id') or None
    if not isinstance(ids, list) or not ids:
        return jsonify({'error': 'No receipts selected'}), 400
    if cid and not get_db().execute("SELECT id FROM delivery_cycles WHERE id=?", (cid,)).fetchone():
        return jsonify({'error': 'Cycle not found'}), 404
    db = get_db()
    n = 0
    for rid in ids:
        if db.execute("SELECT id FROM receipts WHERE id=?", (rid,)).fetchone():
            db.execute("UPDATE receipts SET cycle_id=?, updated_at=? WHERE id=?", (cid, now(), rid))
            n += 1
    db.commit()
    return jsonify({'ok': True, 'assigned': n})


@app.route('/api/receipts/<rid>', methods=['DELETE'])
@require_auth(roles=['admin', 'finance', 'treasurer'])
def delete_receipt(rid):
    """Delete a single receipt and its line items + linked reimbursement."""
    db = get_db()
    if not db.execute("SELECT id FROM receipts WHERE id=?", (rid,)).fetchone():
        return jsonify({'error': 'Not found'}), 404
    db.execute("DELETE FROM receipt_items WHERE receipt_id=?", (rid,))
    db.execute("DELETE FROM reimbursements WHERE receipt_id=?", (rid,))
    db.execute("DELETE FROM receipts WHERE id=?", (rid,))
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/receipts/upload', methods=['POST'])
@require_auth(roles=['admin', 'finance', 'treasurer'])
def upload_receipt():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    db = get_db()
    filename, raw, upload_error = _store_receipt_upload(
        db, request.files['file'], uploader_user_id=g.user['user_id']
    )
    if upload_error:
        status = 429 if 'quota' in upload_error.lower() else 422
        return jsonify({'error': upload_error}), status
    parsed, perr = _parse_receipt_image_ex(raw, filename)  # (None, reason) unless active
    return jsonify({'file_url': f'/uploads/{filename}', 'parsed': parsed,
                    'parse_error': (perr if not parsed else None)}), 201

# ── Finance Summary ───────────────────────────────────────────────────────────

OPEX_CATEGORIES = ['Web hosting / software', 'Supplies', 'Admin / fees', 'Other']


@app.route('/api/expenses', methods=['GET'])
@require_auth(roles=['admin', 'finance', 'treasurer'])
def list_expenses():
    """Operating expenses (org overhead), newest first. Optional ?status= and ?category=."""
    db = get_db()
    where, params = [], []
    st = (request.args.get('status') or '').strip()
    if st in ('paid', 'pending'):
        where.append('status=?'); params.append(st)
    cat = (request.args.get('category') or '').strip()
    if cat:
        where.append('category=?'); params.append(cat)
    sql = 'SELECT * FROM operating_expenses'
    if where:
        sql += ' WHERE ' + ' AND '.join(where)
    sql += " ORDER BY COALESCE(NULLIF(expense_date,''), substr(created_at,1,10)) DESC, created_at DESC"
    return jsonify([dict(r) for r in db.execute(sql, params).fetchall()])


@app.route('/api/expenses/summary', methods=['GET'])
@require_auth(roles=['admin', 'finance', 'treasurer'])
def expenses_summary():
    db = get_db()
    paid    = db.execute("SELECT COALESCE(SUM(amount),0) s, COUNT(*) c FROM operating_expenses WHERE status='paid'").fetchone()
    pending = db.execute("SELECT COALESCE(SUM(amount),0) s, COUNT(*) c FROM operating_expenses WHERE status='pending'").fetchone()
    by_cat = [dict(r) for r in db.execute('''
        SELECT COALESCE(NULLIF(TRIM(category),''),'Other') as category,
               ROUND(COALESCE(SUM(amount),0),2) as total, COUNT(*) as count
        FROM operating_expenses GROUP BY category ORDER BY total DESC''').fetchall()]
    return jsonify({
        'total_paid':    round(paid['s'], 2),    'paid_count':    paid['c'],
        'total_pending': round(pending['s'], 2), 'pending_count': pending['c'],
        'by_category':   by_cat, 'categories': OPEX_CATEGORIES,
    })


def _expense_payload(d, existing=None):
    """Validate + assemble an operating-expense row from request JSON."""
    def val(k, default=None):
        return d[k] if k in d else (existing[k] if existing else default)
    amount = val('amount')
    try:
        amount = float(amount)
        if amount < 0:
            return None, 'Amount cannot be negative'
    except (TypeError, ValueError):
        return None, 'Amount is required and must be a number'
    status = (val('status') or 'paid').strip()
    if status not in ('paid', 'pending'):
        return None, 'Status must be paid or pending'
    row = {
        'expense_date':   (val('expense_date') or None),
        'category':       (val('category') or 'Other'),
        'vendor':         (val('vendor') or None),
        'description':    (val('description') or None),
        'amount':         round(amount, 2),
        'payment_method': (val('payment_method') or None),
        'payment_ref':    (val('payment_ref') or None),
        'status':         status,
    }
    # paid_date: set when paid (keep provided or default to today), clear when pending
    row['paid_date'] = (val('paid_date') or (val('expense_date') or now()[:10])) if status == 'paid' else None
    return row, None


@app.route('/api/expenses', methods=['POST'])
@require_auth(roles=['admin', 'finance', 'treasurer'])
def create_expense():
    row, err = _expense_payload(request.json or {})
    if err:
        return jsonify({'error': err}), 422
    db = get_db()
    eid = str(uuid.uuid4())
    db.execute('''INSERT INTO operating_expenses
        (id, expense_date, category, vendor, description, amount, payment_method,
         payment_ref, status, paid_date, created_by, created_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)''',
        (eid, row['expense_date'], row['category'], row['vendor'], row['description'],
         row['amount'], row['payment_method'], row['payment_ref'], row['status'],
         row['paid_date'], g.user['user_id'], now()))
    db.commit()
    return jsonify(dict(db.execute("SELECT * FROM operating_expenses WHERE id=?", (eid,)).fetchone())), 201


@app.route('/api/expenses/<eid>', methods=['PUT'])
@require_auth(roles=['admin', 'finance', 'treasurer'])
def update_expense(eid):
    db = get_db()
    existing = db.execute("SELECT * FROM operating_expenses WHERE id=?", (eid,)).fetchone()
    if not existing:
        return jsonify({'error': 'Not found'}), 404
    row, err = _expense_payload(request.json or {}, existing=existing)
    if err:
        return jsonify({'error': err}), 422
    db.execute('''UPDATE operating_expenses SET expense_date=?, category=?, vendor=?,
        description=?, amount=?, payment_method=?, payment_ref=?, status=?, paid_date=?,
        updated_at=? WHERE id=?''',
        (row['expense_date'], row['category'], row['vendor'], row['description'],
         row['amount'], row['payment_method'], row['payment_ref'], row['status'],
         row['paid_date'], now(), eid))
    db.commit()
    return jsonify(dict(db.execute("SELECT * FROM operating_expenses WHERE id=?", (eid,)).fetchone()))


@app.route('/api/expenses/<eid>', methods=['DELETE'])
@require_auth(roles=['admin', 'finance', 'treasurer'])
def delete_expense(eid):
    db = get_db()
    if not db.execute("SELECT id FROM operating_expenses WHERE id=?", (eid,)).fetchone():
        return jsonify({'error': 'Not found'}), 404
    db.execute("DELETE FROM operating_expenses WHERE id=?", (eid,))
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/finance/summary', methods=['GET'])
@require_auth(roles=['admin', 'finance', 'treasurer'])
def finance_summary():
    """Ledger view of the charity's money.
      income               = donations received
      pending_review       = receipts submitted but not yet approved (potential spend)
      committed            = approved expenses (paid + still-owed) — every approved receipt
      paid_out             = reimbursements actually disbursed to volunteers
      outstanding_payable  = approved but unpaid = what we owe volunteers right now
      cash_balance         = income − paid_out   (cash actually left)
      available            = income − committed  (left after honoring approved payables)
    committed always reconciles to paid_out + outstanding_payable."""
    db = get_db()

    def _sum(sql, *p):
        return db.execute(sql, p).fetchone()[0]

    income          = _sum("SELECT COALESCE(SUM(amount),0) FROM donations")
    pending_review  = _sum("SELECT COALESCE(SUM(amount),0) FROM receipts WHERE status='pending'")
    paid_out        = _sum("SELECT COALESCE(SUM(amount),0) FROM reimbursements WHERE status='paid'")
    outstanding     = _sum("SELECT COALESCE(SUM(amount),0) FROM reimbursements WHERE status='pending'")
    committed       = paid_out + outstanding
    approved_count  = _sum("SELECT COUNT(*) FROM receipts WHERE status='approved'")
    pending_count   = _sum("SELECT COUNT(*) FROM receipts WHERE status='pending'")
    owed_count      = _sum("SELECT COUNT(*) FROM reimbursements WHERE status='pending'")
    mismatch_count  = _sum("SELECT COUNT(*) FROM receipts WHERE status!='rejected' AND amount_mismatch=1")

    # Operating expenses (org overhead) — paid ones spend cash now, pending ones are
    # committed but not yet disbursed. Both reduce what's available.
    opex_paid          = _sum("SELECT COALESCE(SUM(amount),0) FROM operating_expenses WHERE status='paid'")
    opex_pending       = _sum("SELECT COALESCE(SUM(amount),0) FROM operating_expenses WHERE status='pending'")
    opex_pending_count = _sum("SELECT COUNT(*) FROM operating_expenses WHERE status='pending'")

    # Per-cycle spend (approved receipts, resolved to a cycle via cycle_id or slot).
    cycles = db.execute('''
        SELECT
            dc.id, dc.title, dc.delivery_date_start, dc.status as cycle_status,
            COUNT(DISTINCT CASE WHEN r.status='approved' THEN r.id END) as approved_count,
            COUNT(DISTINCT CASE WHEN r.status='pending'  THEN r.id END) as pending_count,
            COALESCE(SUM(CASE WHEN r.status='pending' THEN r.amount ELSE 0 END),0) as pending_total,
            COALESCE(SUM(CASE WHEN reimb.status IN ('pending','paid') THEN reimb.amount ELSE 0 END),0) as committed_total,
            COALESCE(SUM(CASE WHEN reimb.status='paid'    THEN reimb.amount ELSE 0 END),0) as paid_total,
            COALESCE(SUM(CASE WHEN reimb.status='pending' THEN reimb.amount ELSE 0 END),0) as outstanding_total
        FROM delivery_cycles dc
        LEFT JOIN receipts r ON (
            (r.cycle_id = dc.id AND r.status != 'rejected')
            OR (r.cycle_id IS NULL AND r.slot_id IS NOT NULL AND r.status != 'rejected'
                AND EXISTS(SELECT 1 FROM volunteer_slots vs
                           WHERE vs.id = r.slot_id AND vs.cycle_id = dc.id))
        )
        LEFT JOIN reimbursements reimb ON reimb.receipt_id = r.id
        GROUP BY dc.id
        ORDER BY dc.delivery_date_start
    ''').fetchall()

    # Receipts not tied to any cycle (e.g. admin dashboard uploads) — surface them as
    # an "Unassigned" row so per-cycle totals reconcile with the overall committed.
    unassigned = db.execute('''
        SELECT
            COUNT(DISTINCT CASE WHEN r.status='approved' THEN r.id END) as approved_count,
            COUNT(DISTINCT CASE WHEN r.status='pending'  THEN r.id END) as pending_count,
            COALESCE(SUM(CASE WHEN r.status='pending' THEN r.amount ELSE 0 END),0) as pending_total,
            COALESCE(SUM(CASE WHEN reimb.status IN ('pending','paid') THEN reimb.amount ELSE 0 END),0) as committed_total,
            COALESCE(SUM(CASE WHEN reimb.status='paid'    THEN reimb.amount ELSE 0 END),0) as paid_total,
            COALESCE(SUM(CASE WHEN reimb.status='pending' THEN reimb.amount ELSE 0 END),0) as outstanding_total
        FROM receipts r
        LEFT JOIN reimbursements reimb ON reimb.receipt_id = r.id
        WHERE r.status != 'rejected'
          AND COALESCE(r.cycle_id, (SELECT vs.cycle_id FROM volunteer_slots vs WHERE vs.id = r.slot_id)) IS NULL
    ''').fetchone()

    cyc_list = [dict(c) for c in cycles]
    if unassigned and (unassigned['approved_count'] or unassigned['pending_count']):
        cyc_list.append({
            'id': None, 'title': 'Unassigned (no cycle)', 'delivery_date_start': None,
            'cycle_status': None,
            'approved_count':   unassigned['approved_count'],
            'pending_count':    unassigned['pending_count'],
            'pending_total':    round(unassigned['pending_total'], 2),
            'committed_total':  round(unassigned['committed_total'], 2),
            'paid_total':       round(unassigned['paid_total'], 2),
            'outstanding_total': round(unassigned['outstanding_total'], 2),
        })

    return jsonify({
        'totals': {
            'income':              round(income, 2),
            'pending_review':      round(pending_review, 2),
            'committed':           round(committed, 2),
            'paid_out':            round(paid_out, 2),
            'outstanding_payable': round(outstanding, 2),
            'opex_paid':           round(opex_paid, 2),
            'opex_pending':        round(opex_pending, 2),
            # Cash truly left after every disbursement (reimbursements + operating costs).
            'cash_balance':        round(income - paid_out - opex_paid, 2),
            # Left after honoring all commitments (approved payables + operating costs).
            'available':           round(income - committed - opex_paid - opex_pending, 2),
            'pending_count':       pending_count,
            'approved_count':      approved_count,
            'owed_count':          owed_count,
            'mismatch_count':      mismatch_count,
            'opex_pending_count':  opex_pending_count,
        },
        'cycles': cyc_list
    })


@app.route('/api/receipts/analytics', methods=['GET'])
@require_auth(roles=['admin', 'finance', 'treasurer'])
def receipts_analytics():
    """Spend breakdowns from approved receipts + their parsed line items:
    by store, by volunteer (with amount still owed), top items, and monthly trend."""
    db = get_db()
    by_store = _group_by_store(
        db.execute("SELECT store, amount FROM receipts WHERE status='approved'").fetchall(), limit=50)
    by_volunteer = [dict(r) for r in db.execute('''
        SELECT COALESCE(v.name,'(unassigned)') as volunteer_name,
               ROUND(COALESCE(SUM(r.amount),0),2) as total, COUNT(*) as count,
               ROUND(COALESCE(SUM(CASE WHEN rb.status='pending' THEN rb.amount ELSE 0 END),0),2) as owed
        FROM receipts r
        LEFT JOIN volunteers v ON r.volunteer_id = v.id
        LEFT JOIN reimbursements rb ON rb.receipt_id = r.id
        WHERE r.status='approved'
        GROUP BY r.volunteer_id ORDER BY total DESC LIMIT 50
    ''').fetchall()]
    top_items = [dict(r) for r in db.execute('''
        SELECT ri.name, COUNT(*) as count, ROUND(COALESCE(SUM(ri.line_total),0),2) as total
        FROM receipt_items ri JOIN receipts r ON ri.receipt_id = r.id
        WHERE r.status='approved' AND ri.name IS NOT NULL AND TRIM(ri.name) != ''
        GROUP BY LOWER(TRIM(ri.name)) ORDER BY total DESC LIMIT 25
    ''').fetchall()]
    by_month = [dict(r) for r in db.execute('''
        SELECT substr(COALESCE(NULLIF(purchase_date,''), created_at), 1, 7) as month,
               ROUND(COALESCE(SUM(amount),0),2) as total, COUNT(*) as count
        FROM receipts WHERE status='approved'
        GROUP BY month ORDER BY month
    ''').fetchall()]
    # Spend by food-catalog category (from the model-tagged line items).
    by_category = [dict(r) for r in db.execute('''
        SELECT COALESCE(NULLIF(TRIM(ri.category),''),'(uncategorized)') as category,
               ROUND(COALESCE(SUM(ri.line_total),0),2) as total, COUNT(*) as count
        FROM receipt_items ri JOIN receipts r ON ri.receipt_id = r.id
        WHERE r.status='approved'
        GROUP BY LOWER(TRIM(COALESCE(ri.category,''))) ORDER BY total DESC
    ''').fetchall()]
    return jsonify({'by_store': by_store, 'by_volunteer': by_volunteer,
                    'top_items': top_items, 'by_month': by_month, 'by_category': by_category})


def _spend_report_data(db, since, until):
    """Assemble the spend-report dashboard data (approved receipts, optional date
    window on purchase_date→created_at fallback). Used by both the JSON + Excel routes."""
    date_expr = "COALESCE(NULLIF(r.purchase_date,''), substr(r.created_at,1,10))"
    where, params = ["r.status='approved'"], []
    if since:
        where.append(f"{date_expr} >= ?"); params.append(since)
    if until:
        where.append(f"{date_expr} <= ?"); params.append(until)
    w = " AND ".join(where)

    head = db.execute(f"SELECT COALESCE(SUM(r.amount),0) t, COUNT(*) c FROM receipts r WHERE {w}", params).fetchone()
    total_spend, receipt_count = round(head['t'], 2), head['c']
    avg_receipt = round(total_spend / receipt_count, 2) if receipt_count else 0

    cov = db.execute(
        f"""SELECT COALESCE(SUM(CASE WHEN ri.category IS NOT NULL AND TRIM(ri.category)!=''
                       AND ri.category!='Other' THEN ri.line_total ELSE 0 END),0) cat,
                   COALESCE(SUM(ri.line_total),0) allv
            FROM receipt_items ri JOIN receipts r ON ri.receipt_id=r.id WHERE {w}""", params).fetchone()
    itemized_spend = round(cov['allv'], 2)
    categorized_pct = round(100 * cov['cat'] / cov['allv'], 1) if cov['allv'] else 0

    cat_rows = db.execute(
        f"""SELECT COALESCE(NULLIF(TRIM(ri.category),''),'(uncategorized)') category,
                   ROUND(SUM(ri.line_total),2) total, COUNT(*) count
            FROM receipt_items ri JOIN receipts r ON ri.receipt_id=r.id WHERE {w}
            GROUP BY LOWER(TRIM(COALESCE(ri.category,''))) ORDER BY total DESC""", params).fetchall()
    item_rows = db.execute(
        f"""SELECT COALESCE(NULLIF(TRIM(ri.category),''),'(uncategorized)') category,
                   ri.name, ri.qty, ROUND(ri.line_total,2) line_total, r.store,
                   {date_expr} as date
            FROM receipt_items ri JOIN receipts r ON ri.receipt_id=r.id WHERE {w}
            ORDER BY ri.line_total DESC""", params).fetchall()
    items_by_cat = {}
    for it in item_rows:
        items_by_cat.setdefault(it['category'], []).append(
            {'name': it['name'], 'qty': it['qty'], 'line_total': it['line_total'],
             'store': it['store'], 'date': it['date']})
    denom = sum(c['total'] for c in cat_rows) or 1
    categories = [{
        'category': c['category'], 'total': c['total'], 'count': c['count'],
        'pct': round(100 * c['total'] / denom, 1),
        'items': items_by_cat.get(c['category'], [])[:100],
    } for c in cat_rows]

    by_month = [dict(r) for r in db.execute(
        f"""SELECT substr({date_expr},1,7) as month, ROUND(SUM(r.amount),2) total, COUNT(*) count
            FROM receipts r WHERE {w} GROUP BY month ORDER BY month""", params).fetchall()]
    by_store = _group_by_store(
        db.execute(f"SELECT r.store as store, r.amount as amount FROM receipts r WHERE {w}", params).fetchall(),
        limit=12)
    top_items = [dict(r) for r in db.execute(
        f"""SELECT ri.name, COUNT(*) count, ROUND(SUM(ri.line_total),2) total
            FROM receipt_items ri JOIN receipts r ON ri.receipt_id=r.id WHERE {w}
            AND ri.name IS NOT NULL AND TRIM(ri.name)!=''
            GROUP BY LOWER(TRIM(ri.name)) ORDER BY total DESC LIMIT 15""", params).fetchall()]

    return {
        'metrics': {
            'total_spend': total_spend, 'receipt_count': receipt_count, 'avg_receipt': avg_receipt,
            'itemized_spend': itemized_spend, 'categorized_pct': categorized_pct,
            'category_count': len([c for c in categories if c['category'] != '(uncategorized)']),
            'top_category': categories[0]['category'] if categories else None,
            'top_store': by_store[0]['store'] if by_store else None,
        },
        'categories': categories, 'by_month': by_month, 'by_store': by_store, 'top_items': top_items,
    }


@app.route('/api/finance/spend-report', methods=['GET'])
@require_auth(roles=['admin', 'finance', 'treasurer'])
def spend_report():
    since = (request.args.get('since') or '').strip()
    until = (request.args.get('until') or '').strip()
    return jsonify(_spend_report_data(get_db(), since, until))


@app.route('/api/finance/spend-report.xlsx', methods=['GET'])
@require_auth(roles=['admin', 'finance', 'treasurer'])
def spend_report_xlsx():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from flask import send_file
    import io
    since = (request.args.get('since') or '').strip()
    until = (request.args.get('until') or '').strip()
    d = _spend_report_data(get_db(), since, until)

    wb = openpyxl.Workbook()
    hf = Font(bold=True, color='FFFFFF'); hfill = PatternFill('solid', fgColor='1A3A2A')
    def _hdr(ws, cols):
        for i, h in enumerate(cols, 1):
            c = ws.cell(row=1, column=i, value=h); c.font = hf; c.fill = hfill
            c.alignment = Alignment(horizontal='center')

    ws1 = wb.active; ws1.title = 'By Category'
    _hdr(ws1, ['Category', 'Total ($)', 'Items', '% of itemized'])
    for r, c in enumerate(d['categories'], 2):
        ws1.cell(row=r, column=1, value=c['category']); ws1.cell(row=r, column=2, value=c['total'])
        ws1.cell(row=r, column=3, value=c['count']); ws1.cell(row=r, column=4, value=c['pct'])
    for i, wdt in enumerate([26, 14, 10, 14], 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = wdt

    ws2 = wb.create_sheet('Line Items')
    _hdr(ws2, ['Category', 'Item', 'Qty', 'Amount ($)', 'Store', 'Date'])
    r = 2
    for cat in d['categories']:
        for it in cat['items']:
            ws2.cell(row=r, column=1, value=cat['category']); ws2.cell(row=r, column=2, value=it['name'])
            ws2.cell(row=r, column=3, value=it['qty']); ws2.cell(row=r, column=4, value=it['line_total'])
            ws2.cell(row=r, column=5, value=it['store']); ws2.cell(row=r, column=6, value=it['date']); r += 1
    for i, wdt in enumerate([22, 30, 8, 12, 20, 12], 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = wdt

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"sihaa_spend_report_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── Reimbursements ────────────────────────────────────────────────────────────

@app.route('/api/reimbursements', methods=['GET'])
@require_auth(roles=['admin', 'finance', 'treasurer'])
def list_reimbursements():
    db = get_db()
    status = request.args.get('status')
    try:
        q = '''SELECT rb.*, v.name as volunteer_name,
                      r.store, r.purchase_date, r.file_url as receipt_photo, r.amount as receipt_amount
               FROM reimbursements rb
               LEFT JOIN volunteers v ON rb.volunteer_id = v.id
               LEFT JOIN receipts r ON rb.receipt_id = r.id
               WHERE 1=1'''
        params = []
        if status:
            q += " AND rb.status=?"; params.append(status)
        q += " ORDER BY rb.created_at DESC"
        return jsonify([dict(r) for r in db.execute(q, params).fetchall()])
    except Exception as e:
        log.error(f'list_reimbursements error: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/reimbursements/by-volunteer', methods=['GET'])
@require_auth(roles=['admin', 'finance', 'treasurer'])
def reimbursements_by_volunteer():
    """Reimbursements grouped by volunteer → delivery cycle, for the accordion view.
    ?filter=all|owed|paid. Each volunteer + cycle carries owed/paid subtotals and the
    list of still-owed reimbursement ids (for 'pay all owed')."""
    db = get_db()
    filt = (request.args.get('filter') or 'all').strip()
    where = "rb.status IN ('pending','paid')"
    if filt == 'owed': where = "rb.status='pending'"
    elif filt == 'paid': where = "rb.status='paid'"
    cyc = "COALESCE(r.cycle_id,(SELECT vs.cycle_id FROM volunteer_slots vs WHERE vs.id=r.slot_id))"
    rows = db.execute(f'''
        SELECT rb.id reimb_id, rb.status reimb_status, rb.amount, rb.payment_method, rb.payment_ref, rb.paid_date,
               r.id receipt_id, COALESCE(NULLIF(r.purchase_date,''),substr(r.created_at,1,10)) date, r.store, r.file_url,
               rb.volunteer_id, v.name volunteer_name,
               f.family_code, f.name family_name,
               {cyc} cycle_id, dc.title cycle_title, dc.delivery_date_start cycle_date
        FROM reimbursements rb
        JOIN receipts r ON rb.receipt_id = r.id
        LEFT JOIN volunteers v ON rb.volunteer_id = v.id
        LEFT JOIN families f ON r.family_id = f.id
        LEFT JOIN delivery_cycles dc ON dc.id = {cyc}
        WHERE {where}
        ORDER BY v.name
    ''').fetchall()

    vols = {}
    for r in rows:
        amt = r['amount'] or 0
        vk = r['volunteer_id'] or '__none__'
        v = vols.setdefault(vk, {'volunteer_id': r['volunteer_id'],
                                 'volunteer_name': r['volunteer_name'] or '(unassigned)',
                                 'owed': 0.0, 'paid': 0.0, 'count': 0, 'owed_ids': [], 'cycles': {}})
        v['count'] += 1
        if r['reimb_status'] == 'paid':
            v['paid'] += amt
        else:
            v['owed'] += amt; v['owed_ids'].append(r['reimb_id'])
        ck = r['cycle_id'] or '__nocycle__'
        c = v['cycles'].setdefault(ck, {'cycle_id': r['cycle_id'],
                                        'cycle_title': r['cycle_title'] or 'No cycle',
                                        'cycle_date': r['cycle_date'], 'owed': 0.0, 'paid': 0.0,
                                        'owed_ids': [], 'receipts': []})
        if r['reimb_status'] == 'paid':
            c['paid'] += amt
        else:
            c['owed'] += amt; c['owed_ids'].append(r['reimb_id'])
        fam = r['family_code'] or ''
        if r['family_name']:
            fam = (fam + ' · ' if fam else '') + r['family_name']
        c['receipts'].append({'reimb_id': r['reimb_id'], 'receipt_id': r['receipt_id'],
                              'date': r['date'], 'family': fam or None, 'store': r['store'],
                              'file_url': r['file_url'], 'amount': round(amt, 2),
                              'status': r['reimb_status'], 'payment_method': r['payment_method'],
                              'payment_ref': r['payment_ref'], 'paid_date': r['paid_date']})

    out = []
    for v in vols.values():
        cycles = sorted(v['cycles'].values(), key=lambda c: (c['cycle_date'] or ''), reverse=True)
        for c in cycles:
            c['owed'] = round(c['owed'], 2); c['paid'] = round(c['paid'], 2)
        out.append({'volunteer_id': v['volunteer_id'], 'volunteer_name': v['volunteer_name'],
                    'owed': round(v['owed'], 2), 'paid': round(v['paid'], 2), 'count': v['count'],
                    'owed_ids': v['owed_ids'], 'cycles': cycles})
    out.sort(key=lambda x: (-x['owed'], -x['paid'], x['volunteer_name'].lower()))
    totals = {'owed': round(sum(v['owed'] for v in out), 2), 'paid': round(sum(v['paid'] for v in out), 2),
              'receipts': len(rows), 'vol_with_balance': sum(1 for v in out if v['owed'] > 0)}
    return jsonify({'volunteers': out, 'totals': totals})


@app.route('/api/reimbursements/bulk-pay', methods=['POST'])
@require_auth(roles=['admin', 'finance', 'treasurer'])
def bulk_pay_reimbursements():
    """Mark several pending reimbursements paid in one action (a cycle or a volunteer's
    whole run). Skips any that aren't currently pending."""
    d = request.json or {}
    ids = d.get('ids') or []
    method = (d.get('payment_method') or '').strip()
    ref = (d.get('payment_ref') or '').strip() or None
    valid = ('venmo', 'zelle', 'check', 'cash', 'bank_transfer', 'cheque', 'other')
    if not isinstance(ids, list) or not ids:
        return jsonify({'error': 'No reimbursements selected'}), 400
    if method not in valid:
        return jsonify({'error': f'Invalid payment method. Use one of: {", ".join(valid)}'}), 422
    db = get_db()
    paid = 0
    for rid in ids:
        row = db.execute("SELECT status FROM reimbursements WHERE id=?", (rid,)).fetchone()
        if row and row['status'] == 'pending':
            db.execute("UPDATE reimbursements SET status='paid', payment_method=?, payment_ref=?, "
                       "paid_date=?, updated_at=? WHERE id=?", (method, ref, now()[:10], now(), rid))
            paid += 1
    db.commit()
    return jsonify({'ok': True, 'paid': paid})


@app.route('/api/reimbursements/<rid>', methods=['PUT'])
@require_auth(roles=['admin', 'finance', 'treasurer'])
def update_reimbursement(rid):
    db = get_db()
    row = db.execute("SELECT * FROM reimbursements WHERE id=?", (rid,)).fetchone()
    if not row:
        return jsonify({'error': 'Not found'}), 404
    d = request.json or {}
    # Validate up front — previously an invalid method hit the DB CHECK constraint
    # and surfaced as a generic 500
    _VALID_PAYMENT_METHODS = ('venmo', 'zelle', 'check', 'cash', 'bank_transfer', 'cheque', 'other')
    if d.get('payment_method') is not None and d['payment_method'] not in _VALID_PAYMENT_METHODS:
        return jsonify({'error': f'Invalid payment method. Use one of: {", ".join(_VALID_PAYMENT_METHODS)}'}), 422
    new_status = d.get('status', row['status'])
    # Reverting a paid reimbursement back to unpaid (e.g. it was marked paid by mistake)
    # clears the payment details so it shows as owed again and stops counting as paid_out.
    reverting = (new_status != 'paid' and row['status'] == 'paid')
    if reverting:
        payment_method = d.get('payment_method', None)
        payment_ref    = d.get('payment_ref', None)
        paid_date      = d.get('paid_date', None)
    else:
        payment_method = d.get('payment_method', row['payment_method'])
        payment_ref    = d.get('payment_ref', row['payment_ref'])
        paid_date      = d.get('paid_date', row['paid_date']) or (now()[:10] if new_status == 'paid' else row['paid_date'])
    db.execute(
        '''UPDATE reimbursements SET status=?,payment_method=?,payment_ref=?,paid_date=?,
           approved_by=?,notes=?,updated_at=? WHERE id=?''',
        (new_status, payment_method, payment_ref, paid_date,
         d.get('approved_by', row['approved_by']) or g.user['user_id'],
         d.get('notes', row['notes']), now(), rid)
    )
    # (A payable only exists after the receipt is approved, so paying it no longer
    # needs to touch the receipt's status — it's already 'approved'.)
    db.commit()
    # Notify volunteer via SMS when payment is sent
    if new_status == 'paid' and row['status'] != 'paid':
        try:
            vol = db.execute(
                "SELECT name, email FROM volunteers WHERE id=?", (row['volunteer_id'],)
            ).fetchone()  # was SELECT phone — vol['email'] raised IndexError, silently killing this notification
            if vol and vol['email']:
                method = d.get('payment_method', row['payment_method']) or 'bank transfer'
                ref    = d.get('payment_ref', row['payment_ref'])
                amount = row['amount'] or 0
                ref_line = f'\nReference: {ref}' if ref else ''
                body = (f'Assalamu Alaikum {vol["name"]},\n\n'
                        f'Your reimbursement has been sent!\n\n'
                        f'Amount: ${amount:.2f}\n'
                        f'Method: {method.title()}{ref_line}\n\n'
                        f'JazakAllah Khair for your service!\n\n— Sihha Food Program')
                _email_notify(vol['email'], 'Sihha Reimbursement Sent', body)
        except Exception as e:
            log.warning(f'Volunteer payment email notification failed: {e}')
    return jsonify(dict(db.execute("SELECT * FROM reimbursements WHERE id=?", (rid,)).fetchone()))

# ── Donations ─────────────────────────────────────────────────────────────────

DONATIONS_MAX_ROWS = 10000  # safety ceiling on donations reads (audit P2: bound the
                            # fastest-growing table). ~decades of headroom for this
                            # charity, so today's results are unchanged; prevents an
                            # unbounded SELECT * from OOM-ing the worker years from now.

def _donations_where(args):
    """Build an optional date-window WHERE clause + effective LIMIT from query args
    (?since=YYYY-MM-DD, ?until=YYYY-MM-DD, ?limit=N). Shared by the JSON list and the
    Excel export so both are bounded identically."""
    where, params = [], []
    since = (args.get('since') or '').strip()
    until = (args.get('until') or '').strip()
    if since:
        where.append("date >= ?"); params.append(since)
    if until:
        where.append("date <= ?"); params.append(until)
    wsql = (" WHERE " + " AND ".join(where)) if where else ""
    try:
        req_limit = int(args.get('limit') or 0)
    except ValueError:
        req_limit = 0
    eff_limit = min(req_limit, DONATIONS_MAX_ROWS) if req_limit > 0 else DONATIONS_MAX_ROWS
    return wsql, params, eff_limit

@app.route('/api/donations', methods=['GET'])
@require_auth(roles=['admin', 'finance', 'treasurer'])
def list_donations():
    wsql, params, limit = _donations_where(request.args)
    rows = get_db().execute(
        f"SELECT * FROM donations{wsql} ORDER BY date DESC, created_at DESC LIMIT ?",
        params + [limit]
    ).fetchall()
    if len(rows) >= DONATIONS_MAX_ROWS:
        log.warning(f'list_donations hit the {DONATIONS_MAX_ROWS}-row ceiling — '
                    f'client-side totals may under-count; add date filters or paginate.')
    return jsonify([dict(r) for r in rows])

@app.route('/api/donations/export', methods=['GET'])
@require_auth(roles=['admin', 'finance', 'treasurer'])
def export_donations():
    """Download all donations as an Excel file."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from flask import send_file
    import io

    db   = get_db()
    wsql, params, limit = _donations_where(request.args)
    rows = db.execute(
        "SELECT date, donor_name, donor_email, amount, frequency, type, source, notes, reference_id, created_at "
        f"FROM donations{wsql} ORDER BY date DESC, created_at DESC LIMIT ?",
        params + [limit]
    ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Donations'

    # Header styling
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1A3A2A')
    headers = ['Date', 'Donor', 'Email', 'Amount ($)', 'Frequency', 'Type', 'Source', 'Campaign / Notes', 'Reference ID', 'Imported At']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Column widths
    col_widths = [12, 18, 24, 12, 12, 10, 10, 28, 36, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Data rows
    alt_fill = PatternFill('solid', fgColor='F0FAF5')
    for row_idx, r in enumerate(rows, 2):
        vals = [
            r['date'] or '',
            r['donor_name'] or 'Anonymous',
            r['donor_email'] or '',
            round(r['amount'] or 0, 2),
            r['frequency'] or 'one-time',
            r['type'] or 'online',
            r['source'] or 'manual',
            r['notes'] or '',
            r['reference_id'] or '',
            (r['created_at'] or '')[:10],
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Freeze header row
    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"sihaa_donations_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)

@app.route('/api/donations', methods=['POST'])
@require_auth(roles=['admin', 'finance', 'treasurer'])
def create_donation():
    data = request.json or {}
    did = str(uuid.uuid4())
    db = get_db()
    db.execute(
        '''INSERT INTO donations (id,donor_name,donor_email,amount,type,date,source,reference_id,notes,created_at)
           VALUES (?,?,?,?,?,?,?,?,?,?)''',
        (did, data.get('donor_name'), data.get('donor_email'), data.get('amount'),
         data.get('type'), data.get('date'), data.get('source'),
         data.get('reference_id'), data.get('notes'), now())
    )
    db.commit()
    return jsonify({'id': did}), 201

@app.route('/api/donations/<did>', methods=['PUT'])
@require_auth(roles=['admin', 'finance', 'treasurer'])
def update_donation(did):
    data = request.json or {}
    db = get_db()
    if not db.execute("SELECT id FROM donations WHERE id=?", (did,)).fetchone():
        return jsonify({'error': 'Not found'}), 404
    db.execute("""
        UPDATE donations
           SET donor_name=?, donor_email=?, amount=?, type=?, frequency=?,
               date=?, source=?, notes=?
         WHERE id=?
    """, (
        data.get('donor_name'), data.get('donor_email'),
        data.get('amount'),     data.get('type'),
        data.get('frequency'),  data.get('date'),
        data.get('source'),     data.get('notes'),
        did
    ))
    db.commit()
    return jsonify({'ok': True})

@app.route('/api/donations/<did>', methods=['DELETE'])
@require_auth(roles=['admin', 'treasurer'])
def delete_donation(did):
    db = get_db()
    if not db.execute("SELECT id FROM donations WHERE id=?", (did,)).fetchone():
        return jsonify({'error': 'Not found'}), 404
    db.execute("DELETE FROM donations WHERE id=?", (did,))
    db.commit()
    return jsonify({'ok': True})

class WixSyncError(Exception):
    """Raised by the Wix sync core for config/API errors. Carries an HTTP status and
    the number imported before the failure so callers can report partial progress."""
    def __init__(self, message, status=502, imported=0):
        super().__init__(message)
        self.message  = message
        self.status   = status
        self.imported = imported


def _sync_wix_donations_core(db):
    """Pull all PAID donation orders from the Wix eCommerce API and upsert into the
    donations table (deduped by Wix order id). Returns {'imported': int, 'skipped': int}.
    Raises WixSyncError on missing config or Wix API errors. Safe for both request
    context (pass get_db()) and scheduler jobs (pass make_conn())."""
    try:
        import urllib.request as _req
        import urllib.error as _ureq
        import json as _json

        api_key = os.environ.get('WIX_API_KEY', '').strip()
        site_id = os.environ.get('WIX_SITE_ID', '038c9d97-1ce8-4495-982b-37591dce50ee').strip()

        if not api_key:
            raise WixSyncError('WIX_API_KEY not configured in environment variables', status=400)

        # Anonymize any previously synced full names (one-time cleanup)
        # Detects unabbreviated names: no period, source='wix', more than one word
        existing_full = db.execute(
            "SELECT id, donor_name, donor_email FROM donations WHERE source='wix' AND donor_name NOT LIKE '%.%'"
        ).fetchall()
        for row in existing_full:
            parts = (row['donor_name'] or '').split()
            if len(parts) >= 2:
                f_abbr = parts[0][:3].capitalize()
                l_abbr = parts[-1][:1].upper()
                new_name = f"{f_abbr}. {l_abbr}."
            elif len(parts) == 1:
                new_name = parts[0][:3].capitalize() + '.'
            else:
                new_name = 'Anonymous'
            # Mask email domain if it looks like a full address
            raw_e = row['donor_email'] or ''
            if '@' in raw_e and not raw_e.startswith('***'):
                new_email = '***@' + raw_e.split('@', 1)[1]
            else:
                new_email = raw_e
            db.execute(
                "UPDATE donations SET donor_name=?, donor_email=? WHERE id=?",
                (new_name, new_email, row['id'])
            )
        if existing_full:
            db.commit()
            log.info(f'Anonymized {len(existing_full)} existing Wix donor records')

        url      = 'https://www.wixapis.com/ecom/v1/orders/search'
        headers  = {'Content-Type': 'application/json',
                    'Authorization': api_key,
                    'wix-site-id': site_id}
        cursor   = None
        imported = 0
        skipped  = 0

        while True:
            page_imported = 0
            body = {'search': {'cursorPaging': {'limit': 100}},
                    'sort': [{'fieldName': 'createdDate', 'order': 'DESC'}]}
            if cursor:
                body['search']['cursorPaging']['cursor'] = cursor

            req_obj = _req.Request(url, data=_json.dumps(body).encode(),
                                   headers=headers, method='POST')
            try:
                with _req.urlopen(req_obj, timeout=20) as resp:
                    result = _json.loads(resp.read())
            except _ureq.HTTPError as e:
                body_text = e.read().decode('utf-8', errors='replace')
                log.error(f'Wix HTTPError {e.code}: {body_text}')
                raise WixSyncError(f'Wix API {e.code}: {body_text[:200]}', status=502, imported=imported)
            except Exception as e:
                log.error(f'Wix request error: {e}')
                raise WixSyncError(f'Wix request error: {str(e)}', status=502, imported=imported)

            orders = result.get('orders', [])
            for order in orders:
                if order.get('paymentStatus') != 'PAID':
                    continue
                line_items = order.get('lineItems', [])
                is_donation = any(
                    li.get('itemType', {}).get('custom') == 'DONATION'
                    for li in line_items
                )
                if not is_donation:
                    continue

                wix_order_id = order['id']
                existing = db.execute(
                    "SELECT id FROM donations WHERE reference_id=?", (wix_order_id,)
                ).fetchone()
                if existing:
                    skipped += 1
                    continue

                li        = line_items[0]
                opts      = li.get('catalogReference', {}).get('options', {})
                amount    = float(opts.get('amount') or li.get('price', {}).get('amount') or 0)
                freq_raw  = opts.get('frequency', 'ONE_TIME')
                frequency = 'monthly' if freq_raw == 'MONTH' else 'one-time'
                product   = li.get('productName', {}).get('original', 'Food Donation')

                buyer    = order.get('buyerInfo', {})
                billing  = order.get('billingInfo', {}).get('contactDetails', {})
                fname    = billing.get('firstName', '').strip()
                lname    = billing.get('lastName', '').strip()
                # Abbreviate for privacy: first 2 letters of first name + first letter of last name
                # e.g. "Ahmer Kamal" → "Ah. K."
                f_abbr   = fname[:3].capitalize() if fname else ''
                l_abbr   = lname[:1].upper() if lname else ''
                if f_abbr or l_abbr:
                    donor = f"{f_abbr}. {l_abbr}.".strip() if (f_abbr and l_abbr) else f"{f_abbr or l_abbr}."
                else:
                    donor = 'Anonymous'
                # Store only email domain for privacy (e.g. "***@gmail.com")
                raw_email = buyer.get('email', '')
                if '@' in raw_email:
                    email = '***@' + raw_email.split('@', 1)[1]
                else:
                    email = ''
                date_str = (order.get('purchasedDate') or order.get('createdDate', ''))[:10]

                did = str(uuid.uuid4())
                db.execute(
                    '''INSERT INTO donations
                       (id,donor_name,donor_email,amount,type,frequency,date,source,reference_id,notes,created_at)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
                    (did, donor, email, amount, 'online', frequency, date_str,
                     'wix', wix_order_id, product, now())
                )
                imported += 1
                page_imported += 1

            db.commit()

            # Early-exit (audit P1.7): orders come newest-first (createdDate DESC),
            # and imports form a contiguous newest-first prefix — so once a full page
            # of 100 orders yields ZERO new imports, everything older is already
            # imported and there is nothing left to find. The full 100-order page acts
            # as a look-behind buffer against an order whose payment flips to PAID late.
            # Without this, the hourly job re-walked all history every run.
            if orders and page_imported == 0:
                log.info(f'Wix sync early-exit: page of {len(orders)} orders yielded 0 new imports')
                break

            # Next page
            meta   = result.get('metadata', {})
            cursor = meta.get('cursors', {}).get('next')
            if not cursor or not orders:
                break

        return {'imported': imported, 'skipped': skipped}

    except WixSyncError:
        raise
    except Exception as e:
        log.error(f'sync_wix_donations_core unhandled error: {e}', exc_info=True)
        raise WixSyncError(f'Server error: {str(e)}', status=500, imported=0)


@app.route('/api/donations/sync-wix', methods=['POST'])
@require_auth(roles=['admin', 'treasurer'])
def sync_wix_donations():
    """Manual trigger: sync Wix donations now. The hourly scheduler calls the same core."""
    try:
        result = _sync_wix_donations_core(get_db())
        return jsonify({'ok': True, 'imported': result['imported'],
                        'skipped_duplicates': result['skipped']})
    except WixSyncError as e:
        return jsonify({'error': e.message, 'imported': e.imported}), e.status


def _sync_wix_donations_job():
    """Hourly scheduler job — sync Wix donations using a standalone DB connection.
    No-op when WIX_API_KEY is unset so non-configured deploys stay quiet. Idempotent:
    the core dedupes by Wix order id, so running in each gunicorn worker is safe."""
    if not os.environ.get('WIX_API_KEY', '').strip():
        return
    conn = make_conn()
    try:
        result = _sync_wix_donations_core(conn)
        if result['imported']:
            log.info(f"Wix hourly sync: imported {result['imported']} new donation(s), "
                     f"skipped {result['skipped']} duplicate(s)")
    except WixSyncError as e:
        log.error(f'Wix hourly sync failed ({e.status}): {e.message}')
    except Exception as e:
        log.error(f'Wix hourly sync unhandled error: {e}', exc_info=True)
    finally:
        conn.close()

# ── Food Categories ───────────────────────────────────────────────────────────

@app.route('/api/food-categories', methods=['GET'])
@require_auth(roles=['admin', 'finance', 'treasurer', 'viewer'])
def list_food_categories():
    rows = get_db().execute(
        "SELECT * FROM food_categories ORDER BY display_order, name"
    ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/food-categories', methods=['POST'])
@require_auth(roles=['admin'])
def create_food_category():
    data = request.json or {}
    if not data.get('name'):
        return jsonify({'error': 'Name is required'}), 422
    cid = str(uuid.uuid4())
    max_order = get_db().execute("SELECT COALESCE(MAX(display_order),0) FROM food_categories").fetchone()[0]
    get_db().execute(
        "INSERT INTO food_categories (id, name, display_order, is_active, created_at) VALUES (?,?,?,?,?)",
        (cid, data['name'].strip(), data.get('display_order', max_order + 1),
         data.get('is_active', 1), now())
    )
    get_db().commit()
    return jsonify(dict(get_db().execute("SELECT * FROM food_categories WHERE id=?", (cid,)).fetchone())), 201

@app.route('/api/food-categories/<cid>', methods=['PUT'])
@require_auth(roles=['admin'])
def update_food_category(cid):
    db = get_db()
    row = db.execute("SELECT * FROM food_categories WHERE id=?", (cid,)).fetchone()
    if not row:
        return jsonify({'error': 'Not found'}), 404
    d = request.json or {}
    db.execute(
        "UPDATE food_categories SET name=?, display_order=?, is_active=? WHERE id=?",
        (d.get('name', row['name']), d.get('display_order', row['display_order']),
         d.get('is_active', row['is_active']), cid)
    )
    db.commit()
    return jsonify(dict(db.execute("SELECT * FROM food_categories WHERE id=?", (cid,)).fetchone()))

@app.route('/api/food-categories/<cid>', methods=['DELETE'])
@require_auth(roles=['admin'])
def delete_food_category(cid):
    db = get_db()
    if db.execute("SELECT id FROM food_items WHERE category_id=? AND is_active=1", (cid,)).fetchone():
        return jsonify({'error': 'Cannot delete category with active items'}), 409
    db.execute("UPDATE food_categories SET is_active=0 WHERE id=?", (cid,))
    db.commit()
    return jsonify({'ok': True})

# ── Food Items ────────────────────────────────────────────────────────────────

@app.route('/api/food-items', methods=['GET'])
@require_auth(roles=['admin', 'finance', 'treasurer', 'viewer'])
def list_food_items():
    db = get_db()
    active_only = request.args.get('active') == '1'
    cat_id = request.args.get('category_id')
    q = '''SELECT fi.*, fc.name as category_name
           FROM food_items fi
           JOIN food_categories fc ON fi.category_id = fc.id
           WHERE 1=1'''
    params = []
    if active_only:
        q += " AND fi.is_active=1 AND fc.is_active=1"
    if cat_id:
        q += " AND fi.category_id=?"; params.append(cat_id)
    q += " ORDER BY fc.display_order, fi.display_order, fi.name"
    return jsonify([dict(r) for r in db.execute(q, params).fetchall()])

@app.route('/api/food-items', methods=['POST'])
@require_auth(roles=['admin'])
def create_food_item():
    data = request.json or {}
    if not data.get('name') or not data.get('category_id'):
        return jsonify({'error': 'Name and category_id are required'}), 422
    iid = str(uuid.uuid4())
    db = get_db()
    if not db.execute("SELECT id FROM food_categories WHERE id=?", (data['category_id'],)).fetchone():
        return jsonify({'error': 'Category not found'}), 404
    max_order = db.execute(
        "SELECT COALESCE(MAX(display_order),0) FROM food_items WHERE category_id=?",
        (data['category_id'],)
    ).fetchone()[0]
    db.execute(
        "INSERT INTO food_items (id, category_id, name, unit, is_active, display_order, created_at, price, allow_qty) VALUES (?,?,?,?,?,?,?,?,?)",
        (iid, data['category_id'], data['name'].strip(),
         data.get('unit', 'each'), data.get('is_active', 1),
         data.get('display_order', max_order + 1), now(),
         float(data.get('price', 0) or 0),
         1 if data.get('allow_qty') else 0)
    )
    db.commit()

    # Seed empty bundle quantities for S/M/L
    for size in ('S', 'M', 'L'):
        db.execute(
            "INSERT OR IGNORE INTO bundle_quantities (id, food_item_id, bundle_size, quantity) VALUES (?,?,?,?)",
            (str(uuid.uuid4()), iid, size, '0')
        )
    db.commit()

    item = dict(db.execute(
        '''SELECT fi.*, fc.name as category_name FROM food_items fi
           JOIN food_categories fc ON fi.category_id = fc.id WHERE fi.id=?''', (iid,)
    ).fetchone())
    return jsonify(item), 201

@app.route('/api/food-items/<iid>', methods=['PUT'])
@require_auth(roles=['admin'])
def update_food_item(iid):
    db = get_db()
    row = db.execute("SELECT * FROM food_items WHERE id=?", (iid,)).fetchone()
    if not row:
        return jsonify({'error': 'Not found'}), 404
    d = request.json or {}
    rk = row.keys()
    try:
        price_val     = float(d['price'] or 0)       if 'price'     in d else float(row['price']     if 'price'     in rk else 0)
        allow_qty_val = (1 if d['allow_qty'] else 0)  if 'allow_qty' in d else int(row['allow_qty']   if 'allow_qty' in rk else 0)
        is_default_v  = (1 if d['is_default'] else 0) if 'is_default' in d else int(row['is_default'] if 'is_default' in rk else 0)
        group_id_v    = d.get('group_id',     row['group_id']    if 'group_id'    in rk else None)
        group_max_v   = int(d.get('group_max', row['group_max']  if 'group_max'   in rk else 1) or 1)
        is_free_text_v= (1 if d['is_free_text'] else 0) if 'is_free_text' in d else int(row['is_free_text'] if 'is_free_text' in rk else 0)
        db.execute(
            "UPDATE food_items SET name=?, unit=?, is_active=?, display_order=?, category_id=?, "
            "price=?, allow_qty=?, is_default=?, group_id=?, group_max=?, is_free_text=? WHERE id=?",
            (d.get('name', row['name']), d.get('unit', row['unit']),
             d.get('is_active', row['is_active']), d.get('display_order', row['display_order']),
             d.get('category_id', row['category_id']),
             price_val, allow_qty_val, is_default_v, group_id_v or None, group_max_v, is_free_text_v, iid)
        )
        db.commit()
    except Exception as e:
        log.exception(f'update_food_item error: {e}')
        return jsonify({'error': 'Save failed — please try again.'}), 500
    return jsonify(dict(db.execute("SELECT * FROM food_items WHERE id=?", (iid,)).fetchone()))

# ── Bundle Quantities ─────────────────────────────────────────────────────────

@app.route('/api/bundle-quantities', methods=['GET'])
@require_auth(roles=['admin', 'finance', 'treasurer', 'viewer'])
def get_bundle_quantities():
    db = get_db()
    item_id = request.args.get('item_id')
    if item_id:
        rows = db.execute(
            "SELECT * FROM bundle_quantities WHERE food_item_id=? ORDER BY bundle_size",
            (item_id,)
        ).fetchall()
    else:
        rows = db.execute(
            '''SELECT bq.*, fi.name as item_name, fi.unit, fc.name as category_name
               FROM bundle_quantities bq
               JOIN food_items fi ON bq.food_item_id = fi.id
               JOIN food_categories fc ON fi.category_id = fc.id
               ORDER BY fc.display_order, fi.display_order, bq.bundle_size'''
        ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/bundle-quantities', methods=['PUT'])
@require_auth(roles=['admin'])
def update_bundle_quantities():
    """Bulk update: expects list of {food_item_id, bundle_size, quantity}"""
    items = request.json or []
    if not isinstance(items, list):
        return jsonify({'error': 'Expected array'}), 422
    db = get_db()
    for item in items:
        if not all(k in item for k in ('food_item_id', 'bundle_size', 'quantity')):
            continue
        db.execute(
            '''INSERT INTO bundle_quantities (id, food_item_id, bundle_size, quantity)
               VALUES (?,?,?,?)
               ON CONFLICT(food_item_id, bundle_size) DO UPDATE SET quantity=excluded.quantity''',
            (str(uuid.uuid4()), item['food_item_id'], item['bundle_size'], item['quantity'])
        )
    db.commit()
    return jsonify({'ok': True})

# ── Bundle Size Rules ─────────────────────────────────────────────────────────

@app.route('/api/bundle-size-rules', methods=['GET'])
@require_auth(roles=['admin', 'finance', 'treasurer', 'viewer'])
def get_bundle_size_rules():
    rows = get_db().execute(
        "SELECT * FROM bundle_size_rules ORDER BY min_household"
    ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/bundle-size-rules', methods=['PUT'])
@require_auth(roles=['admin'])
def update_bundle_size_rules():
    """Bulk update: expects list of {bundle_size, min_household, max_household, label}"""
    items = request.json or []
    db = get_db()
    for item in items:
        db.execute(
            '''UPDATE bundle_size_rules
               SET min_household=?, max_household=?, label=?, budget=?
               WHERE bundle_size=?''',
            (item.get('min_household'), item.get('max_household'),
             item.get('label'),
             float(item['budget']) if 'budget' in item else 0,
             item.get('bundle_size'))
        )
    db.commit()
    return jsonify([dict(r) for r in db.execute(
        "SELECT * FROM bundle_size_rules ORDER BY min_household"
    ).fetchall()])

# ── Delivery Cycles ───────────────────────────────────────────────────────────

def _ensure_volunteer_slots(db, cycle_id, family_id):
    """Ensure open slots exist for a family in a cycle for all is_family_slot task types.
    Idempotent — SELECT-first to avoid duplicates. Creates open slots only if absent.
    Returns the number of new slots created (0 if they already existed).
    Called as a safety net on order confirmation, cycle creation, and family activation.
    """
    cycle = db.execute("SELECT * FROM delivery_cycles WHERE id=?", (cycle_id,)).fetchone()
    if not cycle:
        return 0
    delivery_date = cycle['delivery_date_start']
    task_types = db.execute(
        "SELECT slug FROM volunteer_task_types WHERE is_active=1 AND is_family_slot=1 ORDER BY display_order"
    ).fetchall()
    created = 0
    for tt_row in task_types:
        task_type = tt_row['slug']
        task_date = delivery_date if task_type == 'delivery' else None
        try:
            existing = db.execute(
                "SELECT id FROM volunteer_slots WHERE cycle_id=? AND family_id=? AND task_type=? AND status!='cancelled'",
                (cycle_id, family_id, task_type)
            ).fetchone()
            if not existing:
                db.execute(
                    "INSERT INTO volunteer_slots (id,cycle_id,family_id,task_type,task_date,status,created_at) VALUES (?,?,?,?,?,?,?)",
                    (str(uuid.uuid4()), cycle_id, family_id, task_type, task_date, 'open', now())
                )
                created += 1
        except sqlite3.IntegrityError:
            # uq_vs_active_slot (audit P1.9): a concurrent request already created the
            # active slot in the SELECT→INSERT window. Not an error — just not ours.
            log.info(f'_ensure_volunteer_slots: {task_type} slot for family {family_id} already created concurrently')
        except Exception as e:
            log.warning(f'_ensure_volunteer_slots: could not create {task_type} slot for family {family_id}: {e}')
    return created


def _pre_create_slots_for_cycle(db, cycle_id):
    """Pre-create open slots for ALL active families in a cycle.
    Called when a new cycle is created or seeded. Returns total slots created.
    """
    families = db.execute(
        "SELECT id FROM families WHERE status='active'"
    ).fetchall()
    total = 0
    for fam in families:
        total += _ensure_volunteer_slots(db, cycle_id, fam['id'])
    return total


def _pre_create_slots_for_family(db, family_id):
    """Pre-create open slots for a newly activated family across all future cycles.
    Called when a family's status is set to 'active'. Returns total slots created.
    """
    future_cycles = db.execute(
        "SELECT id FROM delivery_cycles WHERE status NOT IN ('delivered') AND delivery_date_start >= date('now')"
    ).fetchall()
    total = 0
    for cyc in future_cycles:
        total += _ensure_volunteer_slots(db, cyc['id'], family_id)
    return total


def _enroll_families_in_cycle(db, cycle_id, delivery_date_start):
    """Auto-create food_requests for all active families in a new cycle."""
    families = db.execute(
        "SELECT id, family_size FROM families WHERE status='active'"
    ).fetchall()
    items = db.execute(
        "SELECT id FROM food_items WHERE is_active=1"
    ).fetchall()
    enrolled = 0
    for fam in families:
        # Compute bundle size from household size
        bundle = db.execute(
            "SELECT bundle_size FROM bundle_size_rules WHERE min_household<=? AND (max_household IS NULL OR max_household>=?) ORDER BY min_household DESC LIMIT 1",
            (fam['family_size'] or 1, fam['family_size'] or 1)
        ).fetchone()
        bsize = bundle['bundle_size'] if bundle else 'M'
        token = secrets.token_urlsafe(32)
        token_expires = _confirmation_expiry_iso(delivery_date_start)
        rid = str(uuid.uuid4())
        try:
            db.execute(
                '''INSERT INTO food_requests
                   (id, cycle_id, family_id, bundle_size, submitted_at, status,
                    confirmation_token, confirmation_expires_at)
                   VALUES (?,?,?,?,?,?,?,?)''',
                (rid, cycle_id, fam['id'], bsize, now(), 'pending_confirmation',
                 token, token_expires)
            )
            # Pre-populate all food items as selected
            for item in items:
                db.execute(
                    'INSERT OR IGNORE INTO food_request_items (id, request_id, food_item_id, selected) VALUES (?,?,?,1)',
                    (str(uuid.uuid4()), rid, item['id'])
                )
            enrolled += 1
        except Exception:
            pass  # Already enrolled (UNIQUE constraint)
    db.commit()
    log.info(f'Auto-enrolled {enrolled} families in cycle {cycle_id}')
    return enrolled

@app.route('/api/delivery-cycles', methods=['GET'])
@require_auth(roles=['admin', 'finance', 'treasurer', 'viewer'])
def list_delivery_cycles():
    db = get_db()
    status = request.args.get('status')
    q = "SELECT * FROM delivery_cycles WHERE 1=1"
    params = []
    if status:
        q += " AND status=?"; params.append(status)
    q += " ORDER BY delivery_date_start ASC"
    return jsonify([dict(r) for r in db.execute(q, params).fetchall()])

def _fix_delivery_cycles_schema(db):
    """
    Ensure delivery_cycles.status CHECK includes 'upcoming'.
    Works regardless of which columns exist on the live DB.
    Idempotent — no-ops if schema is already correct.
    """
    schema_row = db.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name='delivery_cycles'"
    ).fetchone()
    if not schema_row or 'upcoming' in schema_row[0]:
        return  # nothing to do

    log.info('_fix_delivery_cycles_schema: rebuilding table to add upcoming status')
    db.execute('PRAGMA foreign_keys=OFF')

    # Discover which columns actually exist in the live table
    col_info  = db.execute('PRAGMA table_info(delivery_cycles)').fetchall()
    live_cols = {row[1] for row in col_info}   # row[1] = column name

    # Build the SELECT list, supplying '' for missing TEXT NOT NULL columns
    wanted = [
        ('id',                  'id'),
        ('title',               'title'),
        ('delivery_date_start', 'delivery_date_start'),
        ('delivery_date_end',   'delivery_date_end'),
        ('request_open_at',     "COALESCE(request_open_at,'')"),
        ('request_close_at',    "COALESCE(request_close_at,'')"),
        ('status',              'status'),
        ('notes',               'notes'),
        ('created_by',          'created_by'),
        ('created_at',          'created_at'),
    ]
    dst_cols = []
    src_exprs = []
    for col_name, expr in wanted:
        raw = col_name if 'COALESCE' not in expr else col_name
        if raw in live_cols:
            dst_cols.append(col_name)
            src_exprs.append(expr)
        elif col_name in ('request_open_at', 'request_close_at'):
            # Column doesn't exist at all — supply empty string
            dst_cols.append(col_name)
            src_exprs.append("''")

    db.execute('DROP TABLE IF EXISTS delivery_cycles_new')
    db.execute('''
        CREATE TABLE delivery_cycles_new (
            id                  TEXT PRIMARY KEY,
            title               TEXT NOT NULL,
            delivery_date_start TEXT NOT NULL,
            delivery_date_end   TEXT NOT NULL,
            request_open_at     TEXT NOT NULL DEFAULT '',
            request_close_at    TEXT NOT NULL DEFAULT '',
            status              TEXT NOT NULL DEFAULT 'upcoming'
                                CHECK(status IN
                                  ('draft','open','closed','upcoming','shopping','delivered')),
            notes               TEXT,
            created_by          TEXT,
            created_at          TEXT NOT NULL,
            updated_at          TEXT
        )''')

    db.execute(
        f"INSERT INTO delivery_cycles_new ({', '.join(dst_cols)}) "
        f"SELECT {', '.join(src_exprs)} FROM delivery_cycles"
    )
    db.execute('DROP TABLE delivery_cycles')
    db.execute('ALTER TABLE delivery_cycles_new RENAME TO delivery_cycles')
    db.commit()
    db.execute('PRAGMA foreign_keys=ON')
    log.info('_fix_delivery_cycles_schema: done')


@app.route('/api/delivery-cycles', methods=['POST'])
@require_auth(roles=['admin'])
def create_delivery_cycle():
    data = request.json or {}
    if not all(data.get(k) for k in ('title', 'delivery_date_start', 'delivery_date_end')):
        return jsonify({'error': 'title, delivery_date_start and delivery_date_end are required'}), 422
    cid = str(uuid.uuid4())
    db  = get_db()

    # Ensure schema supports 'upcoming' before inserting
    try:
        _fix_delivery_cycles_schema(db)
    except Exception as _e:
        log.error(f'create_delivery_cycle: schema fix failed: {_e}')
        return jsonify({'error': f'Schema migration failed: {_e}'}), 500

    req_open  = data.get('request_open_at')  or ''
    req_close = data.get('request_close_at') or ''
    db.execute(
        '''INSERT INTO delivery_cycles
           (id, title, delivery_date_start, delivery_date_end,
            request_open_at, request_close_at, status, notes, created_by, created_at)
           VALUES (?,?,?,?,?,?,?,?,?,?)''',
        (cid, data['title'], data['delivery_date_start'], data['delivery_date_end'],
         req_open, req_close,
         data.get('status') or 'upcoming', data.get('notes'),
         g.user['user_id'], now())
    )
    db.commit()
    # Pre-create volunteer slots for all active families in this new cycle
    try:
        slots_created = _pre_create_slots_for_cycle(db, cid)
        db.commit()
        log.info(f'create_delivery_cycle: pre-created {slots_created} volunteer slots for cycle {cid}')
    except Exception as _e:
        log.warning(f'create_delivery_cycle: slot pre-creation failed: {_e}')
    result = dict(db.execute("SELECT * FROM delivery_cycles WHERE id=?", (cid,)).fetchone())
    return jsonify(result), 201

@app.route('/api/delivery-cycles/<cid>', methods=['PUT'])
@require_auth(roles=['admin'])
def update_delivery_cycle(cid):
    db = get_db()
    row = db.execute("SELECT * FROM delivery_cycles WHERE id=?", (cid,)).fetchone()
    if not row:
        return jsonify({'error': 'Not found'}), 404
    d = request.json or {}
    db.execute(
        '''UPDATE delivery_cycles SET title=?, delivery_date_start=?, delivery_date_end=?,
           request_open_at=?, request_close_at=?, status=?, notes=?, updated_at=? WHERE id=?''',
        (d.get('title', row['title']),
         d.get('delivery_date_start', row['delivery_date_start']),
         d.get('delivery_date_end', row['delivery_date_end']),
         d.get('request_open_at', row['request_open_at']),
         d.get('request_close_at', row['request_close_at']),
         d.get('status', row['status']),
         d.get('notes', row['notes']), now(), cid)
    )
    db.commit()
    return jsonify(dict(db.execute("SELECT * FROM delivery_cycles WHERE id=?", (cid,)).fetchone()))

@app.route('/api/orders', methods=['GET'])
@require_auth(roles=['admin', 'finance', 'treasurer', 'viewer'])
def get_orders():
    """Orders module: returns orders for a cycle enriched with volunteer slot info.
    Supports status=no_order to return active families without an order."""
    db        = get_db()
    cycle_id  = request.args.get('cycle_id')
    status_f  = request.args.get('status', 'all')
    search    = (request.args.get('search') or '').strip().lower()

    if not cycle_id:
        row = db.execute(
            "SELECT id FROM delivery_cycles WHERE status IN ('open','shopping') "
            "ORDER BY delivery_date_start LIMIT 1"
        ).fetchone()
        if not row:
            row = db.execute(
                "SELECT id FROM delivery_cycles ORDER BY delivery_date_start DESC LIMIT 1"
            ).fetchone()
        cycle_id = row['id'] if row else None
    if not cycle_id:
        return jsonify([])

    if status_f == 'no_order':
        ordered_ids = {r['family_id'] for r in db.execute(
            "SELECT family_id FROM food_requests WHERE cycle_id=?", (cycle_id,)
        ).fetchall()}
        families = db.execute(
            "SELECT id, name, family_code, bundle_size FROM families "
            "WHERE status='active' ORDER BY name"
        ).fetchall()
        result = []
        for f in families:
            if f['id'] in ordered_ids:
                continue
            if search and search not in f['name'].lower() and search not in (f['family_code'] or '').lower():
                continue
            result.append({**dict(f), 'status': 'no_order', 'items': [],
                           'shopper': None, 'deliverer': None, 'request_id': None})
        return jsonify(result)

    orders = db.execute(
        '''SELECT fr.id, fr.status, fr.bundle_size, fr.family_id, fr.cycle_id,
                  fr.family_notes,
                  f.name as family_name, f.family_code
           FROM food_requests fr
           JOIN families f ON fr.family_id = f.id
           WHERE fr.cycle_id=? ORDER BY f.name''', (cycle_id,)
    ).fetchall()

    slots = db.execute(
        '''SELECT vs.family_id, vs.task_type, v.name as vol_name
           FROM volunteer_slots vs
           LEFT JOIN volunteers v ON vs.claimed_by = v.id
           WHERE vs.cycle_id=? AND vs.status IN ('claimed','confirmed','complete')''',
        (cycle_id,)
    ).fetchall()
    slot_map = {}
    for s in slots:
        fid = s['family_id']
        if fid not in slot_map:
            slot_map[fid] = {}
        slot_map[fid][s['task_type']] = s['vol_name']

    # Batch item names for ALL orders in one query (audit 3.3 — was one query
    # per order row)
    items_map = {}
    if orders:
        ph = ','.join('?' * len(orders))
        for r in db.execute(
            f'''SELECT fri.request_id, fi.name FROM food_request_items fri
                JOIN food_items fi ON fri.food_item_id = fi.id
                WHERE fri.request_id IN ({ph}) AND fri.selected=1
                ORDER BY fi.display_order''',
            [o['id'] for o in orders]
        ).fetchall():
            items_map.setdefault(r['request_id'], []).append(r['name'])

    result = []
    for order in orders:
        o = dict(order)
        if status_f != 'all' and o['status'] != status_f:
            continue
        if search and search not in o['family_name'].lower() and search not in (o['family_code'] or '').lower():
            continue
        o['items'] = items_map.get(o['id'], [])
        fam_slots     = slot_map.get(o['family_id'], {})
        o['shopper']  = fam_slots.get('shopping')
        o['deliverer']= fam_slots.get('delivery')
        result.append(o)
    return jsonify(result)


@app.route('/api/delivery-cycles/<cid>/orders', methods=['GET'])
@require_auth(roles=['admin', 'finance', 'treasurer', 'viewer'])
def get_cycle_orders(cid):
    db = get_db()
    orders = db.execute(
        '''SELECT fr.*, f.name as family_name, f.phone as family_phone,
                  f.address as family_address, f.city as family_city, f.family_code
           FROM food_requests fr
           JOIN families f ON fr.family_id = f.id
           WHERE fr.cycle_id=?
           ORDER BY fr.submitted_at''', (cid,)
    ).fetchall()
    result = []
    for order in orders:
        o = dict(order)
        items = db.execute(
            f'''SELECT fri.id, fri.request_id, fri.food_item_id, fri.selected,
                      {_EFFECTIVE_ORDER_QTY_SQL} as quantity,
                      fri.custom_value, fi.name, fi.unit, fc.name as category
               FROM food_request_items fri
               JOIN food_requests fr ON fr.id=fri.request_id
               JOIN food_items fi ON fri.food_item_id = fi.id
               JOIN food_categories fc ON fi.category_id = fc.id
               LEFT JOIN bundle_quantities bq
                 ON bq.food_item_id=fi.id AND bq.bundle_size=fr.bundle_size
               WHERE fri.request_id=? AND fri.selected=1
               ORDER BY fc.display_order, fi.display_order''',
            (o['id'],)
        ).fetchall()
        o['selected_items'] = [dict(i) for i in items]
        result.append(o)
    return jsonify(result)

@app.route('/api/food-requests/<rid>', methods=['PUT'])
@require_auth(roles=['admin'])
def update_food_request_status(rid):
    """Admin manually overrides a family's confirmation status (confirmed or skipped)."""
    db  = get_db()
    row = db.execute("SELECT * FROM food_requests WHERE id=?", (rid,)).fetchone()
    if not row:
        return jsonify({'error': 'Not found'}), 404
    d      = request.json or {}
    status = d.get('status')
    if status not in ('confirmed', 'skipped', 'pending_confirmation'):
        return jsonify({'error': 'status must be confirmed, skipped, or pending_confirmation'}), 422
    ts = now()
    try:
        db.execute(
            "UPDATE food_requests SET status=?, confirmed_at=?, updated_at=? WHERE id=?",
            (status, ts if status == 'confirmed' else None, ts, rid)
        )
    except sqlite3.OperationalError:
        # Fallback: confirmed_at/updated_at columns may not exist yet on old DB
        db.execute("UPDATE food_requests SET status=? WHERE id=?", (status, rid))

    # Auto-create volunteer slots when coordinator confirms a family
    if status == 'confirmed':
        _ensure_volunteer_slots(db, row['cycle_id'], row['family_id'])

    _log_order_event(db, rid, 'admin_override', actor='admin',
                     payload={'new_status': status})
    db.commit()
    return jsonify(dict(db.execute("SELECT * FROM food_requests WHERE id=?", (rid,)).fetchone()))


@app.route('/api/food-requests/<rid>/items', methods=['PUT'])
@require_auth(roles=['admin'])
def admin_edit_order_items(rid):
    """Admin replaces the item list for any food request.
    Body: [{ food_item_id, quantity }]  — only selected items (qty >= 1).
    Deselects everything first, then marks supplied items as selected with qty.
    """
    db  = get_db()
    row = db.execute("SELECT * FROM food_requests WHERE id=?", (rid,)).fetchone()
    if not row:
        return jsonify({'error': 'Not found'}), 404
    items = request.json
    if not isinstance(items, list):
        return jsonify({'error': 'Expected array of {food_item_id, quantity}'}), 422

    # Deselect all existing items for this request
    db.execute("UPDATE food_request_items SET selected=0, quantity=1 WHERE request_id=?", (rid,))

    for item in items:
        iid = item.get('food_item_id')
        qty = max(1, int(item.get('quantity') or 1))
        if not iid:
            continue
        # Upsert: update if exists, insert if not
        existing = db.execute(
            "SELECT id FROM food_request_items WHERE request_id=? AND food_item_id=?",
            (rid, iid)
        ).fetchone()
        if existing:
            db.execute(
                "UPDATE food_request_items SET selected=1, quantity=? WHERE request_id=? AND food_item_id=?",
                (qty, rid, iid)
            )
        else:
            db.execute(
                "INSERT INTO food_request_items (id, request_id, food_item_id, selected, quantity) VALUES (?,?,?,1,?)",
                (str(uuid.uuid4()), rid, iid, qty)
            )

    _log_order_event(db, rid, 'admin_edit_items', actor='admin',
                     payload={'item_count': len(items)})
    db.commit()

    # Return updated item list
    updated = db.execute(
        '''SELECT fri.*, fi.name, fi.unit, fi.price, fc.name as category
           FROM food_request_items fri
           JOIN food_items fi ON fri.food_item_id = fi.id
           JOIN food_categories fc ON fi.category_id = fc.id
           WHERE fri.request_id=? AND fri.selected=1
           ORDER BY fc.display_order, fi.display_order''',
        (rid,)
    ).fetchall()
    return jsonify([dict(i) for i in updated])


@app.route('/api/families/<fid>/manual-confirm', methods=['POST'])
@require_auth(roles=['admin'])
def manual_confirm_family(fid):
    """Coordinator manually adds a family to the current open cycle as confirmed."""
    db = get_db()
    family = db.execute("SELECT * FROM families WHERE id=?", (fid,)).fetchone()
    if not family:
        return jsonify({'error': 'Family not found'}), 404

    # Find open or shopping cycle (coordinator may be confirming mid-cycle)
    cycle = db.execute(
        "SELECT * FROM delivery_cycles WHERE status IN ('open','shopping') ORDER BY delivery_date_start LIMIT 1"
    ).fetchone()
    if not cycle:
        return jsonify({'error': 'No active delivery cycle (open or shopping). Create or open a cycle first.'}), 409

    # Don't duplicate
    existing = db.execute(
        "SELECT id FROM food_requests WHERE cycle_id=? AND family_id=?",
        (cycle['id'], fid)
    ).fetchone()
    if existing:
        return jsonify({'error': 'Family already has an order for this cycle.', 'request_id': existing['id']}), 409

    # Determine bundle size
    bundle_size = family['bundle_size'] or None
    if not bundle_size:
        size = db.execute(
            "SELECT bundle_size FROM bundle_size_rules WHERE min_household<=? AND (max_household IS NULL OR max_household>=?) ORDER BY min_household DESC LIMIT 1",
            (family['family_size'] or 1, family['family_size'] or 1)
        ).fetchone()
        bundle_size = size['bundle_size'] if size else 'M'

    ts  = now()
    rid = str(uuid.uuid4())
    try:
        db.execute(
            '''INSERT INTO food_requests
               (id, cycle_id, family_id, bundle_size, submitted_at, status, confirmed_at, updated_at)
               VALUES (?,?,?,?,?,?,?,?)''',
            (rid, cycle['id'], fid, bundle_size, ts, 'confirmed', ts, ts)
        )
    except Exception:
        # Fallback if extra columns don't exist
        db.execute(
            '''INSERT INTO food_requests
               (id, cycle_id, family_id, bundle_size, submitted_at, status)
               VALUES (?,?,?,?,?,?)''',
            (rid, cycle['id'], fid, bundle_size, ts, 'confirmed')
        )

    # Record all items as selected by default
    all_items = db.execute("SELECT id FROM food_items WHERE is_active=1").fetchall()
    for item in all_items:
        db.execute(
            "INSERT OR IGNORE INTO food_request_items (id, request_id, food_item_id, selected) VALUES (?,?,?,1)",
            (str(uuid.uuid4()), rid, item['id'])
        )

    # Auto-create volunteer slots immediately (creates open slots if absent)
    slots_created = _ensure_volunteer_slots(db, cycle['id'], fid)

    # Flip any already-claimed slots to confirmed — volunteer signed up before admin added the family
    claimed_slots = db.execute(
        '''SELECT vs.*, v.name as vol_name, v.phone as vol_phone
           FROM volunteer_slots vs
           JOIN volunteers v ON vs.claimed_by = v.id
           WHERE vs.cycle_id=? AND vs.family_id=? AND vs.status='claimed' ''',
        (cycle['id'], fid)
    ).fetchall()
    for slot in claimed_slots:
        db.execute(
            "UPDATE volunteer_slots SET status='confirmed', updated_at=? WHERE id=?",
            (now(), slot['id'])
        )
    slots_confirmed = len(claimed_slots)

    _log_order_event(db, rid, 'admin_override', actor='admin',
                     payload={'new_status': 'confirmed', 'note': 'manual confirm by coordinator'})
    db.commit()

    # Email volunteers whose slots just became confirmed
    email_sends = []
    for slot in claimed_slots:
        vol_email = _lookup_volunteer_email(db, slot['claimed_by']) if slot.get('claimed_by') else ''
        if vol_email:
            address_line = f"\nAddress: {family['address']}, {family['city']}" if slot['task_type'] == 'delivery' and family.get('address') else ''
            body = (f"Assalamu Alaikum,\n\n"
                    f"Your {slot['task_type']} slot for family {family['family_code'] or fid[:8]} "
                    f"({cycle['title']}) is now confirmed — coordinator added them to this delivery.{address_line}\n\n"
                    f"JazakAllah Khair!\n\n— Sihha Food Program")
            email_sends.append((vol_email, f'Sihha Slot Confirmed — {cycle["title"]}', body))
    _email_notify_async(email_sends)

    log.info(f'Manual confirm: family {fid} added to cycle {cycle["id"]} — {slots_created} slots created, {slots_confirmed} slots confirmed')
    return jsonify({'ok': True, 'request_id': rid, 'cycle_title': cycle['title'], 'bundle_size': bundle_size,
                    'slots_created': slots_created, 'slots_confirmed': slots_confirmed}), 201

@app.route('/api/delivery-cycles/<cid>/shopping-list', methods=['GET'])
@require_auth(roles=['admin', 'finance', 'treasurer', 'viewer'])
def get_cycle_shopping_list(cid):
    db = get_db()
    # Get all selected items across all orders for this cycle, with bundle quantities
    rows = db.execute(
        f'''SELECT fi.id as item_id, fi.name as item_name, fi.unit,
                  fc.name as category, fc.display_order as cat_order, fi.display_order as item_order,
                  SUM({_EFFECTIVE_ORDER_QTY_SQL}) as total_qty,
                  COUNT(DISTINCT fr.id) as order_count
           FROM food_requests fr
           JOIN food_request_items fri ON fri.request_id = fr.id AND fri.selected = 1
           JOIN food_items fi ON fri.food_item_id = fi.id
           JOIN food_categories fc ON fi.category_id = fc.id
           LEFT JOIN bundle_quantities bq
             ON bq.food_item_id=fi.id AND bq.bundle_size=fr.bundle_size
           WHERE fr.cycle_id=? AND fr.status = 'confirmed'
           GROUP BY fi.id
           ORDER BY fc.display_order, fi.display_order''',
        (cid,)
    ).fetchall()

    shopping_list = []
    for r in rows:
        shopping_list.append({
            'item_name':   r['item_name'],
            'category':    r['category'],
            'unit':        r['unit'],
            'total_qty':   r['total_qty'],
            'order_count': r['order_count'],
        })

    cycle = db.execute("SELECT * FROM delivery_cycles WHERE id=?", (cid,)).fetchone()
    total_orders = db.execute(
        "SELECT COUNT(*) FROM food_requests WHERE cycle_id=? AND status = 'confirmed'", (cid,)
    ).fetchone()[0]

    return jsonify({
        'cycle': dict(cycle) if cycle else {},
        'total_orders': total_orders,
        'shopping_list': shopping_list,
        'generated_at': now()  # UTC — volunteers can see if list was generated before recent edits
    })

# ── Volunteer Activity Report ─────────────────────────────────────────────────

@app.route('/api/reports/volunteer-activity', methods=['GET'])
@require_auth(roles=['admin', 'finance', 'treasurer', 'viewer'])
def report_volunteer_activity():
    """Per-volunteer lifetime stats: tasks completed, shopping/delivery breakdown,
    cycles participated in, unique families served, last active date."""
    db = get_db()
    rows = db.execute(
        '''SELECT v.id, v.name, v.phone, v.status,
                  COUNT(DISTINCT CASE WHEN vs.status='complete' THEN vs.id END)          AS total_tasks,
                  COUNT(DISTINCT CASE WHEN vs.status='complete' AND vs.task_type='shopping'  THEN vs.id END) AS shopping_count,
                  COUNT(DISTINCT CASE WHEN vs.status='complete' AND vs.task_type='delivery'  THEN vs.id END) AS delivery_count,
                  COUNT(DISTINCT CASE WHEN vs.status='complete' THEN vs.cycle_id END)    AS cycles_count,
                  COUNT(DISTINCT CASE WHEN vs.status='complete' THEN vs.family_id END)   AS families_served,
                  MAX(vs.completed_at)                                                    AS last_active
           FROM volunteers v
           LEFT JOIN volunteer_slots vs ON vs.claimed_by = v.id
           GROUP BY v.id
           ORDER BY total_tasks DESC, v.name ASC'''
    ).fetchall()
    return jsonify([dict(r) for r in rows])


# ── Print Reports (HTML → browser PDF) ────────────────────────────────────────

PRINT_CSS = """
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
         font-size: 12px; color: #111; background: #fff; padding: 24px; }
  .header { background: #1a3a2a; color: #fff; padding: 14px 20px;
            border-radius: 6px; margin-bottom: 20px;
            display: flex; justify-content: space-between; align-items: center; }
  .header h1 { font-size: 18px; font-weight: 700; }
  .header .sub { font-size: 12px; opacity: 0.8; margin-top: 3px; }
  .header .right { text-align: right; font-size: 11px; opacity: 0.8; }
  .meta { display: flex; gap: 16px; margin-bottom: 18px; flex-wrap: wrap; }
  .meta-box { background: #f5f5f0; border-radius: 6px; padding: 10px 16px;
              border-left: 3px solid #1a3a2a; }
  .meta-box .label { font-size: 10px; text-transform: uppercase;
                     letter-spacing: 0.5px; color: #888; margin-bottom: 3px; }
  .meta-box .value { font-size: 20px; font-weight: 700; color: #1a3a2a; }
  .meta-box .hint  { font-size: 11px; color: #666; }
  h2 { font-size: 13px; font-weight: 700; color: #1a3a2a; text-transform: uppercase;
       letter-spacing: 0.5px; margin: 20px 0 8px; border-bottom: 2px solid #1a3a2a;
       padding-bottom: 4px; }
  table { width: 100%; border-collapse: collapse; margin-bottom: 16px; }
  th { background: #1a3a2a; color: #fff; font-weight: 700; font-size: 11px;
       padding: 7px 10px; text-align: left; }
  td { padding: 6px 10px; border-bottom: 1px solid #eee; vertical-align: top; }
  tr:nth-child(even) td { background: #fafaf8; }
  .total { font-weight: 700; color: #1a3a2a; }
  .badge { display: inline-block; padding: 2px 7px; border-radius: 10px;
           font-size: 10px; font-weight: 700; text-transform: uppercase; }
  .badge-pending  { background: #fff3cd; color: #856404; }
  .badge-delivered{ background: #d1f5e0; color: #0a5c2e; }
  .badge-complete { background: #d1f5e0; color: #0a5c2e; }
  .badge-claimed  { background: #cce5ff; color: #004085; }
  .badge-open     { background: #f8d7da; color: #721c24; }
  .footer { margin-top: 24px; font-size: 10px; color: #aaa;
            border-top: 1px solid #eee; padding-top: 10px;
            display: flex; justify-content: space-between; }
  @media print {
    body { padding: 0; }
    .no-print { display: none !important; }
    @page { margin: 15mm 12mm; }
  }
</style>
"""

def _print_page(title, subtitle, body_html, cycle_title=''):
    generated = datetime.utcnow().strftime('%B %d, %Y at %H:%M UTC')
    return f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="UTF-8">
<title>{title} — {cycle_title}</title>
{PRINT_CSS}
</head><body>
<div class="no-print" style="margin-bottom:16px;">
  <button onclick="window.print()" style="background:#1a3a2a;color:#fff;border:none;
    padding:8px 18px;border-radius:6px;cursor:pointer;font-size:13px;font-weight:600;">
    Print / Save as PDF
  </button>
  <button onclick="window.close()" style="background:#eee;border:none;
    padding:8px 14px;border-radius:6px;cursor:pointer;font-size:13px;margin-left:8px;">
    Close
  </button>
</div>
<div class="header">
  <div><div class="sub">Sihha Food Charity — Operations Hub</div>
    <h1>{cycle_title}</h1>
    <div class="sub">{subtitle}</div></div>
  <div class="right"><strong>{title}</strong><br>Generated {generated}</div>
</div>
{body_html}
<div class="footer">
  <span>Sihha Food Charity — Operations Hub</span>
  <span>Generated {generated}</span>
</div>
<script>
  // Auto-trigger print dialog after a short delay so page is fully rendered
  // Remove this line if you prefer to click the button manually
  // setTimeout(() => window.print(), 600);
</script>
</body></html>"""


@app.route('/api/reports/shopping-list/<cid>', methods=['GET'])
@require_auth(roles=['admin', 'finance', 'treasurer', 'viewer'])
def report_shopping_list(cid):
    """Printable shopping list — returns HTML page, browser prints to PDF."""
    from collections import defaultdict
    db = get_db()
    cycle = db.execute("SELECT * FROM delivery_cycles WHERE id=?", (cid,)).fetchone()
    if not cycle:
        return jsonify({'error': 'Cycle not found'}), 404

    bundle_counts = {r['bundle_size']: r['cnt'] for r in db.execute(
        "SELECT bundle_size, COUNT(*) as cnt FROM food_requests WHERE cycle_id=? AND status='confirmed' GROUP BY bundle_size",
        (cid,)
    ).fetchall()}
    total_orders = sum(bundle_counts.values())

    rows = db.execute(
        f'''SELECT fi.name as item_name, fi.unit,
                  fc.name as category, fc.display_order as cat_order, fi.display_order as item_order,
                  fr.bundle_size,
                  MIN({_EFFECTIVE_ORDER_QTY_SQL}) as min_qty,
                  MAX({_EFFECTIVE_ORDER_QTY_SQL}) as max_qty,
                  SUM({_EFFECTIVE_ORDER_QTY_SQL}) as total_qty,
                  COUNT(DISTINCT fr.id) as order_count
           FROM food_requests fr
           JOIN food_request_items fri ON fri.request_id = fr.id AND fri.selected = 1
           JOIN food_items fi ON fri.food_item_id = fi.id
           JOIN food_categories fc ON fi.category_id = fc.id
           LEFT JOIN bundle_quantities bq ON bq.food_item_id = fi.id AND bq.bundle_size = fr.bundle_size
           WHERE fr.cycle_id=? AND fr.status = 'confirmed'
           GROUP BY fi.id, fr.bundle_size
           ORDER BY fc.display_order, fi.display_order, fr.bundle_size''',
        (cid,)
    ).fetchall()

    items = defaultdict(lambda: {'category':'','unit':'','cat_order':0,'item_order':0,'sizes':{}})
    for r in rows:
        k = r['item_name']
        items[k].update({'category': r['category'], 'unit': r['unit'],
                         'cat_order': r['cat_order'], 'item_order': r['item_order']})
        count = r['order_count'] or 0
        qty = (str(r['min_qty']) if r['min_qty'] == r['max_qty']
               else f"{r['min_qty']}–{r['max_qty']}")
        items[k]['sizes'][r['bundle_size']] = {
            'qty': qty, 'count': count, 'total': r['total_qty'] or 0
        }

    by_cat = defaultdict(list)
    for name, info in sorted(items.items(), key=lambda x: (x[1]['cat_order'], x[1]['item_order'])):
        by_cat[info['category']].append((name, info['unit'], info['sizes'],
                                         sum(v['total'] for v in info['sizes'].values())))

    # Bundle summary
    body = f"""
    <div class="meta">
      <div class="meta-box"><div class="label">Small (S) · 1–2 people</div>
        <div class="value">{bundle_counts.get('S',0)}</div><div class="hint">families</div></div>
      <div class="meta-box"><div class="label">Medium (M) · 3–5 people</div>
        <div class="value">{bundle_counts.get('M',0)}</div><div class="hint">families</div></div>
      <div class="meta-box"><div class="label">Large (L) · 6+ people</div>
        <div class="value">{bundle_counts.get('L',0)}</div><div class="hint">families</div></div>
      <div class="meta-box"><div class="label">Total Orders</div>
        <div class="value">{total_orders}</div><div class="hint">families</div></div>
    </div>"""

    if not rows:
        body += '<p style="color:#888;padding:20px 0;">No orders submitted for this cycle yet.</p>'
    else:
        def fmt(sizes, sz):
            if sz not in sizes: return '—'
            d = sizes[sz]
            return f"{d['qty']} × {d['count']}" if d['count'] else '—'
        def fmt_total(sizes, sz):
            if sz not in sizes: return '—'
            d = sizes[sz]
            return str(d['total']) if d['total'] else '—'

        for cat_name, cat_items in by_cat.items():
            rows_html = ''
            for item_name, unit, sizes, row_total in cat_items:
                s_qty = fmt_total(sizes, 'S')
                m_qty = fmt_total(sizes, 'M')
                l_qty = fmt_total(sizes, 'L')
                rows_html += f"""<tr>
                  <td>{item_name}</td><td>{unit}</td>
                  <td style="text-align:center">{fmt(sizes,'S')}</td>
                  <td style="text-align:center">{fmt(sizes,'M')}</td>
                  <td style="text-align:center">{fmt(sizes,'L')}</td>
                  <td class="total" style="text-align:center">{row_total if row_total else '—'}</td>
                </tr>"""
            body += f"""
            <h2>{cat_name}</h2>
            <table>
              <thead><tr>
                <th>Item</th><th>Unit</th>
                <th style="text-align:center">S (qty×families)</th>
                <th style="text-align:center">M (qty×families)</th>
                <th style="text-align:center">L (qty×families)</th>
                <th style="text-align:center">TOTAL</th>
              </tr></thead>
              <tbody>{rows_html}</tbody>
            </table>"""

    subtitle = f"Delivery: {cycle['delivery_date_start']} – {cycle['delivery_date_end']}  ·  {total_orders} orders"
    html = _print_page('Shopping List', subtitle, body, cycle['title'])
    return html, 200, {'Content-Type': 'text/html; charset=utf-8'}


@app.route('/api/reports/reimbursements', methods=['GET'])
@require_auth(roles=['admin', 'finance', 'treasurer'])
def report_reimbursements():
    """Printable reimbursement report grouped by volunteer — date, family, store,
    cycle, amount spent, and pay status. Filters: ?since, &until (purchase date),
    &status (all|pending|approved|paid), &cycle_id. Browser prints to PDF."""
    from collections import defaultdict
    db = get_db()
    since    = (request.args.get('since') or '').strip()
    until    = (request.args.get('until') or '').strip()
    status_f = (request.args.get('status') or 'all').strip()
    cycle_id = (request.args.get('cycle_id') or '').strip()
    date_expr = "COALESCE(NULLIF(r.purchase_date,''), substr(r.created_at,1,10))"

    where, params = ["r.status != 'rejected'"], []
    if since:    where.append(f"{date_expr} >= ?"); params.append(since)
    if until:    where.append(f"{date_expr} <= ?"); params.append(until)
    if cycle_id:
        where.append("COALESCE(r.cycle_id,(SELECT vs2.cycle_id FROM volunteer_slots vs2 WHERE vs2.id=r.slot_id))=?")
        params.append(cycle_id)
    if status_f == 'pending':
        where.append("r.status='pending'")
    elif status_f == 'approved':
        where.append("r.status='approved' AND (rb.status IS NULL OR rb.status='pending')")
    elif status_f == 'paid':
        where.append("rb.status='paid'")
    w = " AND ".join(where)

    rows = db.execute(f'''
        SELECT r.id, {date_expr} as date, r.store, r.amount, r.status as rstatus,
               v.name as volunteer_name,
               f.family_code as family_code, f.name as family_name,
               rb.status as reimb_status, rb.payment_method, rb.payment_ref, rb.paid_date,
               COALESCE(dc.title, dc2.title) as cycle_title
        FROM receipts r
        LEFT JOIN volunteers v  ON r.volunteer_id = v.id
        LEFT JOIN families   f  ON r.family_id = f.id
        LEFT JOIN reimbursements rb ON rb.receipt_id = r.id
        LEFT JOIN delivery_cycles dc  ON r.cycle_id = dc.id
        LEFT JOIN volunteer_slots vs  ON r.slot_id = vs.id
        LEFT JOIN delivery_cycles dc2 ON vs.cycle_id = dc2.id
        WHERE {w}
        ORDER BY (v.name IS NULL), v.name, date
    ''', params).fetchall()

    def _money(x):
        return f"${(x or 0):,.2f}"

    def _status(r):
        if r['rstatus'] == 'pending':
            return ('Pending review', '#C4772E')
        if r['reimb_status'] == 'paid':
            m = (r['payment_method'] or '').title()
            ref = f" ({r['payment_ref']})" if r['payment_ref'] else ''
            dt = f" · {r['paid_date']}" if r['paid_date'] else ''
            return (f"Paid · {m}{ref}{dt}", '#4E8A5E')
        return ('Approved — owed', '#B0863C')

    groups = defaultdict(list)
    for r in rows:
        groups[r['volunteer_name'] or '(unassigned)'].append(r)

    tot_spent = sum((r['amount'] or 0) for r in rows)
    tot_owed  = sum((r['amount'] or 0) for r in rows if r['rstatus'] == 'approved' and (r['reimb_status'] or 'pending') == 'pending')
    tot_paid  = sum((r['amount'] or 0) for r in rows if r['reimb_status'] == 'paid')

    body = f"""
    <div class="meta">
      <div class="meta-box"><div class="label">Volunteers</div><div class="value">{len(groups)}</div><div class="hint">{len(rows)} receipts</div></div>
      <div class="meta-box"><div class="label">Total spent</div><div class="value">{_money(tot_spent)}</div></div>
      <div class="meta-box"><div class="label">Owed (unpaid)</div><div class="value">{_money(tot_owed)}</div></div>
      <div class="meta-box"><div class="label">Paid out</div><div class="value">{_money(tot_paid)}</div></div>
    </div>"""

    if not rows:
        body += '<p style="color:#888;padding:20px 0;">No receipts match this filter.</p>'
    else:
        for vol in sorted(groups.keys(), key=lambda k: (k == '(unassigned)', k.lower())):
            recs = groups[vol]
            v_spent = sum((r['amount'] or 0) for r in recs)
            v_owed  = sum((r['amount'] or 0) for r in recs if r['rstatus'] == 'approved' and (r['reimb_status'] or 'pending') == 'pending')
            v_paid  = sum((r['amount'] or 0) for r in recs if r['reimb_status'] == 'paid')
            rows_html = ''
            for r in recs:
                label, color = _status(r)
                fam = r['family_code'] or ''
                if r['family_name']:
                    fam = (fam + ' · ' if fam else '') + r['family_name']
                rows_html += f"""<tr>
                  <td style="white-space:nowrap">{r['date'] or '—'}</td>
                  <td>{fam or '—'}</td>
                  <td>{r['store'] or '—'}</td>
                  <td>{r['cycle_title'] or '—'}</td>
                  <td style="text-align:right">{_money(r['amount'])}</td>
                  <td style="color:{color};font-weight:600">{label}</td>
                </tr>"""
            body += f"""
            <h2>{vol} <span style="font-weight:400;color:#888;font-size:13px">· spent {_money(v_spent)} · owed {_money(v_owed)} · paid {_money(v_paid)}</span></h2>
            <table>
              <thead><tr><th>Date</th><th>Family</th><th>Store</th><th>Cycle</th>
                <th style="text-align:right">Amount</th><th>Status</th></tr></thead>
              <tbody>{rows_html}
                <tr style="background:#f5f2ea;font-weight:600">
                  <td colspan="4" style="text-align:right">Subtotal</td>
                  <td style="text-align:right">{_money(v_spent)}</td><td></td>
                </tr>
              </tbody>
            </table>"""

    rng = []
    if since: rng.append(f"from {since}")
    if until: rng.append(f"to {until}")
    sub = 'Reimbursements by volunteer'
    if status_f != 'all': sub += f' · {status_f}'
    if rng: sub += ' · ' + ' '.join(rng)
    html = _print_page('Reimbursement Report', sub, body, 'Reimbursement Report')
    return html, 200, {'Content-Type': 'text/html; charset=utf-8'}


@app.route('/api/reports/cycle-summary/<cid>', methods=['GET'])
@require_auth(roles=['admin', 'finance', 'treasurer', 'viewer'])
def report_cycle_summary(cid):
    """Printable cycle summary — returns HTML page, browser prints to PDF."""
    db = get_db()
    cycle = db.execute("SELECT * FROM delivery_cycles WHERE id=?", (cid,)).fetchone()
    if not cycle:
        return jsonify({'error': 'Cycle not found'}), 404

    orders = db.execute(
        '''SELECT fr.bundle_size, fr.status, fr.delivered_at,
                  f.name as family_name, f.family_size
           FROM food_requests fr
           JOIN families f ON fr.family_id = f.id
           WHERE fr.cycle_id=? AND fr.status != 'cancelled'
           ORDER BY fr.bundle_size, f.name''', (cid,)
    ).fetchall()

    slots = db.execute(
        '''SELECT vs.task_type, vs.status, vs.task_date,
                  f.name as family_name,
                  v.name as volunteer_name, v.phone as volunteer_phone
           FROM volunteer_slots vs
           LEFT JOIN families f ON vs.family_id = f.id
           LEFT JOIN volunteers v ON vs.claimed_by = v.id
           WHERE vs.cycle_id=?
           ORDER BY vs.task_type, vs.task_date, f.name''', (cid,)
    ).fetchall()

    bundle_counts = {}
    for o in orders:
        bundle_counts[o['bundle_size']] = bundle_counts.get(o['bundle_size'], 0) + 1

    body = f"""
    <div class="meta">
      <div class="meta-box"><div class="label">Total Families</div>
        <div class="value">{len(orders)}</div></div>
      <div class="meta-box"><div class="label">Small Bundles</div>
        <div class="value">{bundle_counts.get('S',0)}</div></div>
      <div class="meta-box"><div class="label">Medium Bundles</div>
        <div class="value">{bundle_counts.get('M',0)}</div></div>
      <div class="meta-box"><div class="label">Large Bundles</div>
        <div class="value">{bundle_counts.get('L',0)}</div></div>
      <div class="meta-box"><div class="label">Volunteer Slots</div>
        <div class="value">{len(slots)}</div></div>
    </div>"""

    # Orders table
    order_rows = ''
    for o in orders:
        delivered = o['delivered_at'][:10] if o['delivered_at'] else ('Yes' if o['status']=='delivered' else '—')
        status_class = o['status'].lower()
        order_rows += f"""<tr>
          <td>{o['family_name']}</td>
          <td style="text-align:center">{o['family_size'] or '—'}</td>
          <td style="text-align:center">{o['bundle_size'] or '—'}</td>
          <td><span class="badge badge-{status_class}">{o['status']}</span></td>
          <td>{delivered}</td>
        </tr>"""
    body += f"""
    <h2>Family Orders ({len(orders)})</h2>
    <table>
      <thead><tr><th>Family</th><th style="text-align:center">HH Size</th>
        <th style="text-align:center">Bundle</th><th>Status</th><th>Delivered</th></tr></thead>
      <tbody>{order_rows if order_rows else '<tr><td colspan="5" style="color:#888">No orders</td></tr>'}</tbody>
    </table>"""

    # Slots table
    if slots:
        slot_rows = ''
        for s in slots:
            status_class = s['status'].lower()
            slot_rows += f"""<tr>
              <td>{s['task_type'].capitalize()}</td>
              <td>{s['task_date'] or '—'}</td>
              <td>{s['family_name'] or '—'}</td>
              <td>{s['volunteer_name'] or '<em style="color:#aaa">Unclaimed</em>'}</td>
              <td>{s['volunteer_phone'] or '—'}</td>
              <td><span class="badge badge-{status_class}">{s['status']}</span></td>
            </tr>"""
        body += f"""
        <h2>Volunteer Assignments ({len(slots)})</h2>
        <table>
          <thead><tr><th>Task</th><th>Date</th><th>Family</th>
            <th>Volunteer</th><th>Phone</th><th>Status</th></tr></thead>
          <tbody>{slot_rows}</tbody>
        </table>"""

    subtitle = f"Delivery: {cycle['delivery_date_start']} – {cycle['delivery_date_end']}  ·  Status: {cycle['status'].upper()}"
    html = _print_page('Cycle Summary', subtitle, body, cycle['title'])
    return html, 200, {'Content-Type': 'text/html; charset=utf-8'}

# ── Cycle Assignments routes removed 2026-06-11 (audit 3.4) ──────────────────
# Legacy /api/cycle-assignments CRUD deleted — zero frontend callers; superseded
# by volunteer_slots. Table retained for historical data.

@app.route('/login')
def login_page():
    return send_from_directory('public', 'login.html')

@app.route('/activate')
def activate_page():
    return send_from_directory('public', 'activate.html')

@app.route('/family')
def family_page():
    return send_from_directory('public', 'family.html')

@app.route('/my-order')
def my_order_redirect():
    """Legacy redirect — keep so old bookmarks/SMS links still work."""
    from flask import redirect
    return redirect('/family', code=301)

@app.route('/volunteer-signup')
def volunteer_signup_page():
    return send_from_directory('public', 'volunteer-signup.html')

# ── Public Intake (no auth) ───────────────────────────────────────────────────

@app.route('/api/intake', methods=['POST'])
def public_intake():
    data = request.get_json(silent=True) or {}
    if not isinstance(data, dict):
        return jsonify({'error': 'Invalid request'}), 400
    db = get_db()
    client_ip = _client_ip()
    if not _consume_rate_limit(db, 'public_intake_ip', client_ip, 5, 3600):
        return _rate_limit_response(3600)
    # Honeypot: real users never see or fill this field. Return the same success
    # shape so basic bots cannot tune around the trap.
    if data.get('_website'):
        return jsonify({'ok': True, 'message': 'Thank you. We will be in touch within 48 hours.'}), 201

    name = data.get('name') if isinstance(data.get('name'), str) else ''
    name = name.strip()
    raw_phone = data.get('phone') if isinstance(data.get('phone'), str) else ''
    if not name or not raw_phone:
        return jsonify({'error': 'Name and phone are required'}), 422
    if len(name) > 120:
        return jsonify({'error': 'Name is too long'}), 422
    phone = _normalize_phone(raw_phone)
    if not 7 <= len(phone) <= 15:
        return jsonify({'error': 'A valid phone number is required'}), 422
    if not _consume_rate_limit(db, 'public_intake_phone', phone, 2, 86400):
        return _rate_limit_response(86400)
    email = data.get('email') if isinstance(data.get('email'), str) else ''
    email = email.strip().lower()
    if len(email) > 254 or not _valid_public_email(email):
        return jsonify({'error': 'A valid email address is required'}), 422
    try:
        family_size = int(data.get('family_size')) if data.get('family_size') not in (None, '') else None
        children_count = int(data.get('children_count')) if data.get('children_count') not in (None, '') else None
    except (TypeError, ValueError):
        return jsonify({'error': 'Household counts must be whole numbers'}), 422
    if family_size is not None and not 1 <= family_size <= 30:
        return jsonify({'error': 'Household size must be between 1 and 30'}), 422
    if children_count is not None and not 0 <= children_count <= 30:
        return jsonify({'error': 'Children count must be between 0 and 30'}), 422
    if family_size is not None and children_count is not None and children_count > family_size:
        return jsonify({'error': 'Children count cannot exceed household size'}), 422
    bounded_fields = {
        'address': 300, 'city': 100, 'dietary_notes': 1000,
        'frequency': 80, 'income_range': 80,
    }
    clean = {}
    for field, max_len in bounded_fields.items():
        value = data.get(field)
        if value is not None and not isinstance(value, str):
            return jsonify({'error': f'{field} must be text'}), 422
        clean[field] = (value or '').strip()
        if len(clean[field]) > max_len:
            return jsonify({'error': f'{field} is too long'}), 422

    # Duplicate guard — block a second record for the same phone number
    existing = db.execute(
        "SELECT id, status FROM families WHERE phone=?", (phone,)
    ).fetchone()
    if existing:
        # Identical response for new and existing applicants prevents public
        # membership/status enumeration.
        return jsonify({'ok': True, 'message': 'Thank you. We will be in touch within 48 hours.'}), 201
    fid = str(uuid.uuid4())
    family_code = _make_family_code(phone, family_size, db_conn=db)
    db.execute(
        '''INSERT INTO families
           (id,name,phone,email,address,city,family_size,children_count,
            dietary_notes,frequency,income_range,status,source,family_code,created_at)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
        (fid, name, phone, email or None, clean['address'], clean['city'],
         family_size, children_count, clean['dietary_notes'],
         clean['frequency'], clean['income_range'],
         'pending', 'intake_form', family_code, now())
    )
    db.commit()
    log.info(f'New intake: {name} ({phone})')
    try:
        _notify_coordinators(db,
            f"New family intake submitted:\n"
            f"Name: {name}\n"
            f"Phone: {phone}\n"
            f"City: {clean['city'] or '—'}\n"
            f"Family size: {family_size or '—'}\n"
            f"Please log in to review and approve."
        )
    except Exception as _e:
        log.warning(f'Intake notify failed: {_e}')
    # Send confirmation email to family if email provided — async so a slow
    # SendGrid response can't freeze this UNAUTHENTICATED public handler (audit P1.6).
    fam_email = email
    if fam_email:
        _email_notify_async([(fam_email, 'We received your Sihha application',
            f"Assalamu Alaikum {name},\n\n"
            f"Thank you for applying to the Sihha Food Program.\n\n"
            f"We have received your application and will review it within 48 hours. "
            f"Once approved, you will receive a separate email with your login credentials.\n\n"
            f"If you have any questions, please contact your coordinator.\n\n"
            f"— Sihha Food Program"
        )])
    return jsonify({'ok': True, 'message': 'Thank you. We will be in touch within 48 hours.'}), 201

@app.route('/api/volunteer-signup', methods=['POST'])
def public_volunteer_signup():
    data = request.get_json(silent=True) or {}
    if not isinstance(data, dict):
        return jsonify({'error': 'Invalid request'}), 400
    db = get_db()
    client_ip = _client_ip()
    if not _consume_rate_limit(db, 'public_volunteer_ip', client_ip, 10, 3600):
        return _rate_limit_response(3600)
    if data.get('_website'):
        return jsonify({'ok': True, 'message': 'Thank you for signing up. We will be in touch soon.'}), 201

    name = data.get('name') if isinstance(data.get('name'), str) else ''
    name = name.strip()
    raw_phone = data.get('phone') if isinstance(data.get('phone'), str) else ''
    if not name or not raw_phone:
        return jsonify({'error': 'Name and phone are required'}), 422
    if len(name) > 120:
        return jsonify({'error': 'Name is too long'}), 422
    if not isinstance(data.get('role'), str) or not data.get('role'):
        return jsonify({'error': 'Please select a role'}), 422
    volunteer_role = _normalize_volunteer_role(data.get('role'))
    if volunteer_role not in VALID_VOLUNTEER_ROLES:
        return jsonify({'error': 'Please select a valid role'}), 422
    phone = _normalize_phone(raw_phone)
    if not 7 <= len(phone) <= 15:
        return jsonify({'error': 'A valid phone number is required'}), 422
    if not _consume_rate_limit(db, 'public_volunteer_phone', phone, 2, 86400):
        return _rate_limit_response(86400)
    email = data.get('email') if isinstance(data.get('email'), str) else ''
    email = email.strip().lower()
    if len(email) > 254 or not _valid_public_email(email):
        return jsonify({'error': 'A valid email address is required'}), 422
    availability = data.get('availability') if isinstance(data.get('availability'), str) else ''
    notes = data.get('notes') if isinstance(data.get('notes'), str) else ''
    availability = availability.strip()
    notes = notes.strip()
    if len(availability) > 300 or len(notes) > 1000:
        return jsonify({'error': 'Availability or notes are too long'}), 422
    existing = db.execute("SELECT id, status FROM volunteers WHERE phone=?", (phone,)).fetchone()
    if existing:
        return jsonify({'ok': True, 'message': 'Thank you for signing up. We will be in touch soon.'}), 201
    vid = str(uuid.uuid4())
    db.execute(
        '''INSERT INTO volunteers
           (id,name,phone,email,role,availability,notes,status,source,created_at)
           VALUES (?,?,?,?,?,?,?,?,?,?)''',
        (vid, name, phone, email or None,
         volunteer_role, availability or None,
         notes or None, 'pending', 'signup_form', now())
    )
    db.commit()
    log.info(f'New volunteer signup: {name} ({phone})')
    try:
        role_label = {'shopper':'Shopper','delivery':'Delivery','both':'Shopper + Delivery','general':'General'}.get(volunteer_role, volunteer_role)
        _notify_coordinators(db,
            f"New volunteer signed up:\n"
            f"Name: {name}\n"
            f"Phone: {phone}\n"
            f"Role: {role_label}\n"
            f"Please log in to review and activate."
        )
    except Exception as _e:
        log.warning(f'Volunteer signup notify failed: {_e}')
    # Send confirmation email to volunteer if email provided — async so a slow
    # SendGrid response can't freeze this UNAUTHENTICATED public handler (audit P1.6).
    vol_email = email
    if vol_email:
        _email_notify_async([(vol_email, 'Thank you for signing up to volunteer with Sihha',
            f"Assalamu Alaikum {name},\n\n"
            f"Thank you for signing up to volunteer with the Sihha Food Program!\n\n"
            f"We have received your application and will review it shortly. "
            f"Once approved, you will receive a separate email with your login credentials "
            f"for the volunteer portal.\n\n"
            f"JazakAllah Khair for your generosity!\n\n"
            f"— Sihha Food Program"
        )])
    return jsonify({'ok': True, 'message': 'Thank you for signing up. We will be in touch soon.'}), 201

# ── Static Pages ──────────────────────────────────────────────────────────────

@app.route('/')
def admin_index():
    resp = send_from_directory('public', 'index.html')
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    resp.headers['Pragma'] = 'no-cache'
    return resp

@app.route('/donate-stats')
def donate_stats_page():
    return send_from_directory('public', 'donate-stats.html')

@app.route('/widget')
def widget_page():
    return send_from_directory('public', 'widget.html')

@app.route('/widget/progress')
def widget_progress_page():
    return send_from_directory('public', 'widget-progress.html')

@app.route('/widget/stats')
def widget_stats_page():
    return send_from_directory('public', 'widget-stats.html')

@app.route('/widget/trend')
def widget_trend_page():
    return send_from_directory('public', 'widget-trend.html')

@app.route('/intake')
def intake_page():
    return send_from_directory('public', 'intake.html')

@app.route('/volunteer')
def volunteer_page():
    from flask import redirect
    return redirect('/portal', code=301)

@app.route('/order')
def order_page():
    from flask import redirect
    return redirect('/intake', code=301)

@app.route('/confirm/<token>')
def confirm_page(token):
    return send_from_directory('public', 'confirm.html')

def _active_confirmation_request(db, token):
    """Resolve a single-use, unexpired legacy confirmation capability."""
    if not isinstance(token, str) or not 20 <= len(token) <= 200:
        return None, ('Invalid or expired link', 404)
    req = db.execute(
        '''SELECT fr.*, f.name as family_name, f.family_size, f.dietary_notes,
                  f.status as family_status, dc.title as cycle_title,
                  dc.delivery_date_start, dc.delivery_date_end,
                  dc.request_close_at, dc.status as cycle_status
           FROM food_requests fr
           JOIN families f ON fr.family_id=f.id
           JOIN delivery_cycles dc ON fr.cycle_id=dc.id
           WHERE fr.confirmation_token=?''',
        (token,)
    ).fetchone()
    if not req:
        return None, ('Invalid or expired link', 404)
    if req['status'] != 'pending_confirmation' or req['family_status'] != 'active':
        return None, ('This confirmation link has already been used or is no longer active.', 410)
    try:
        expires_at = datetime.fromisoformat(req['confirmation_expires_at'])
    except (TypeError, ValueError):
        return None, ('This confirmation link has expired.', 410)
    if expires_at <= datetime.utcnow():
        return None, ('This confirmation link has expired.', 410)
    if req['cycle_status'] not in ('upcoming', 'open'):
        return None, ('Confirmation is closed for this delivery.', 410)
    try:
        from datetime import date as _date
        if (_date.fromisoformat(req['delivery_date_start']) - _today_central()).days < 1:
            return None, ('Confirmation is closed for this delivery.', 410)
    except (TypeError, ValueError):
        log.error(f'Confirmation request {req["id"]} has an invalid delivery date')
        return None, ('Confirmation is unavailable. Please contact a coordinator.', 410)
    if req['request_close_at']:
        try:
            from zoneinfo import ZoneInfo
            close_at = datetime.fromisoformat(str(req['request_close_at']).replace('Z', '+00:00'))
            now_for_close = (datetime.now(close_at.tzinfo) if close_at.tzinfo
                             else datetime.now(ZoneInfo('America/Chicago')).replace(tzinfo=None))
            if now_for_close > close_at:
                return None, ('Confirmation is closed for this delivery.', 410)
        except (TypeError, ValueError):
            log.error(f'Confirmation request {req["id"]} has an invalid request_close_at')
            return None, ('Confirmation is unavailable. Please contact a coordinator.', 410)
    return req, None

@app.route('/api/family/confirm/<token>', methods=['GET'])
def get_family_confirmation(token):
    """Public — family views their pre-populated bundle via confirmation token."""
    db = get_db()
    if not _consume_rate_limit(db, 'confirmation_ip', _client_ip(), 60, 3600):
        return _rate_limit_response(3600)
    req, confirmation_error = _active_confirmation_request(db, token)
    if confirmation_error:
        return jsonify({'error': confirmation_error[0]}), confirmation_error[1]

    # Get all active food items with bundle quantities and current selection
    items = db.execute(
        '''SELECT fi.id, fi.name, fi.unit, fc.name as category, fc.display_order as cat_order,
                  fi.display_order as item_order,
                  bq.quantity,
                  COALESCE(fri.selected, 1) as selected
           FROM food_items fi
           JOIN food_categories fc ON fi.category_id = fc.id
           LEFT JOIN bundle_quantities bq ON bq.food_item_id = fi.id AND bq.bundle_size = ?
           LEFT JOIN food_request_items fri ON fri.food_item_id = fi.id AND fri.request_id = ?
           WHERE fi.is_active = 1
           ORDER BY fc.display_order, fi.display_order''',
        (req['bundle_size'], req['id'])
    ).fetchall()

    return jsonify({
        'family_name': req['family_name'],
        'cycle_title': req['cycle_title'],
        'delivery_date_start': req['delivery_date_start'],
        'delivery_date_end': req['delivery_date_end'],
        'bundle_size': req['bundle_size'],
        'notes': req['notes'],
        'dietary_notes': req['dietary_notes'],
        'items': [dict(i) for i in items],
    })

@app.route('/api/family/confirm/<token>', methods=['POST'])
def submit_family_confirmation(token):
    """Public — family confirms, modifies, or skips their bundle."""
    db  = get_db()
    if not _consume_rate_limit(db, 'confirmation_ip', _client_ip(), 60, 3600):
        return _rate_limit_response(3600)
    data   = request.get_json(silent=True) or {}
    if not isinstance(data, dict):
        return jsonify({'error': 'Invalid request'}), 400
    action = data.get('action', 'confirm')  # confirm | skip
    if action not in ('confirm', 'skip'):
        return jsonify({'error': 'action must be confirm or skip'}), 422
    notes = data.get('notes')
    if notes is not None and not isinstance(notes, str):
        return jsonify({'error': 'notes must be text'}), 422
    notes = (notes or '').strip()
    if len(notes) > 1000:
        return jsonify({'error': 'notes must be 1000 characters or fewer'}), 422

    # Serialize token consumption across workers. The second concurrent submit
    # waits, then observes the token cleared by the first transaction.
    db.execute('BEGIN IMMEDIATE')
    req, confirmation_error = _active_confirmation_request(db, token)
    if confirmation_error:
        db.rollback()
        return jsonify({'error': confirmation_error[0]}), confirmation_error[1]

    if action == 'skip':
        db.execute(
            "UPDATE food_requests SET status='skipped', confirmed_at=?, notes=?, updated_at=?, "
            "confirmation_token=NULL, confirmation_expires_at=NULL WHERE id=?",
            (now(), notes, now(), req['id'])
        )
        _log_order_event(db, req['id'], 'auto_skipped', actor='family')
        db.commit()
        return jsonify({'ok': True, 'action': 'skipped'})

    # Capture previous selections for diff (items_edited vs first-time confirmed)
    was_confirmed = req['status'] == 'confirmed'
    prev_items = {}
    if was_confirmed:
        for _pi in db.execute(
            "SELECT fi.name, fri.selected FROM food_request_items fri JOIN food_items fi ON fri.food_item_id=fi.id WHERE fri.request_id=?",
            (req['id'],)
        ).fetchall():
            prev_items[_pi['name']] = _pi['selected']

    # Save item selections
    selection, selection_error = _validate_order_selection(
        db, data.get('selected_items'), {}, {}, req['bundle_size'],
        enforce_budget=False
    )
    if selection_error:
        db.rollback()
        return jsonify({'error': selection_error}), 422
    selected_ids = selection['selected_ids']
    all_items = db.execute("SELECT id, name FROM food_items WHERE is_active=1").fetchall()
    for item in all_items:
        is_selected = 1 if item['id'] in selected_ids else 0
        db.execute(
            '''INSERT INTO food_request_items (id, request_id, food_item_id, selected)
               VALUES (?,?,?,?)
               ON CONFLICT(request_id, food_item_id) DO UPDATE SET selected=?''',
            (str(uuid.uuid4()), req['id'], item['id'], is_selected, is_selected)
        )

    db.execute(
        "UPDATE food_requests SET status='confirmed', confirmed_at=?, notes=?, updated_at=?, "
        "confirmation_token=NULL, confirmation_expires_at=NULL WHERE id=?",
        (now(), notes, now(), req['id'])
    )

    # Log event — items_edited if re-confirming, confirmed if first time
    if was_confirmed and prev_items:
        _added   = [it['name'] for it in all_items if it['id'] in selected_ids and prev_items.get(it['name'], 0) == 0]
        _removed = [n for n, sel in prev_items.items() if sel == 1 and n not in {it['name'] for it in all_items if it['id'] in selected_ids}]
        _log_order_event(db, req['id'], 'items_edited', actor='family',
                         payload={'added': _added, 'removed': _removed})
    else:
        _log_order_event(db, req['id'], 'confirmed', actor='family')

    # Ensure slots exist (safety net)
    _ensure_volunteer_slots(db, req['cycle_id'], req['family_id'])

    # Confirm any claimed volunteer slots and notify those volunteers
    cycle_row  = db.execute("SELECT * FROM delivery_cycles WHERE id=?", (req['cycle_id'],)).fetchone()
    family_row = db.execute("SELECT * FROM families WHERE id=?", (req['family_id'],)).fetchone()
    claimed_slots_conf = db.execute(
        '''SELECT vs.id, vs.task_type, vs.claimed_by, v.name as vol_name, v.phone as vol_phone,
                  f.address, f.city, f.name as family_name
           FROM volunteer_slots vs
           JOIN volunteers v ON vs.claimed_by = v.id
           JOIN families f ON vs.family_id = f.id
           WHERE vs.cycle_id=? AND vs.family_id=? AND vs.status IN ('claimed','confirmed') ''',
        (req['cycle_id'], req['family_id'])
    ).fetchall()
    bundle_sz = family_row['bundle_size'] if family_row else 'M'
    item_lines_conf = []
    if claimed_slots_conf:
        item_rows_conf = db.execute(
            f'''SELECT fi.name, fi.unit, {_EFFECTIVE_ORDER_QTY_SQL} as quantity,
                       fc.name as category
               FROM food_request_items fri
               JOIN food_requests fr ON fr.id=fri.request_id
               JOIN food_items fi ON fri.food_item_id = fi.id
               JOIN food_categories fc ON fi.category_id = fc.id
               LEFT JOIN bundle_quantities bq
                 ON bq.food_item_id = fi.id AND bq.bundle_size = fr.bundle_size
               WHERE fri.request_id = ? AND fri.selected = 1
               ORDER BY fc.display_order, fi.name''',
            (req['id'],)
        ).fetchall()
        for ir in item_rows_conf:
            qty_str = f"{ir['quantity']} {ir['unit']}" if ir['quantity'] else ir['unit'] or ''
            item_lines_conf.append(f"  - {ir['name']}{(' - ' + qty_str) if qty_str else ''}")
    sms_sends_conf = []
    email_sends_conf = []
    for slot in claimed_slots_conf:
        db.execute("UPDATE volunteer_slots SET status='confirmed', updated_at=? WHERE id=?", (now(), slot['id']))
        vol_email = _lookup_volunteer_email(db, slot['claimed_by']) if slot.get('claimed_by') else ''
        if vol_email:
            if slot['task_type'] == 'shopping':
                items_text = '\n'.join(item_lines_conf) if item_lines_conf else '  (no items selected)'
                body = (f"Assalamu Alaikum,\n\nOrder Confirmed — Shopping Task\n\n"
                        f"Family: {slot['family_name']}\n"
                        f"Delivery: {cycle_row['delivery_date_start'] if cycle_row else 'TBD'}\n\n"
                        f"Shopping list:\n{items_text}\n\nJazakAllah Khair!\n\n— Sihha Food Program")
            else:
                body = (f"Assalamu Alaikum,\n\nOrder Confirmed — Delivery Task\n\n"
                        f"Family: {slot['family_name']}\n"
                        f"Delivery: {cycle_row['delivery_date_start'] if cycle_row else 'TBD'}\n"
                        f"Address: {slot['address'] or 'TBD'}, {slot['city'] or ''}\n\n"
                        f"JazakAllah Khair!\n\n— Sihha Food Program")
            email_sends_conf.append((vol_email, f'Sihha Order Confirmed — {slot["family_name"]}', body))

    db.commit()
    _email_notify_async(email_sends_conf)
    return jsonify({'ok': True, 'action': 'confirmed'})

# ── PWA assets ────────────────────────────────────────────────────────────────

@app.route('/sw.js')
def service_worker():
    resp = send_from_directory('public', 'sw.js', mimetype='application/javascript')
    resp.headers['Service-Worker-Allowed'] = '/'
    resp.headers['Cache-Control'] = 'no-cache'
    return resp

@app.route('/manifest-family.json')
def manifest_family():
    return send_from_directory('public', 'manifest-family.json', mimetype='application/manifest+json')

@app.route('/manifest-volunteer.json')
def manifest_volunteer():
    return send_from_directory('public', 'manifest-volunteer.json', mimetype='application/manifest+json')

@app.route('/icons/<path:filename>')
def pwa_icons(filename):
    return send_from_directory('public/icons', filename)

@app.route('/css/<path:filename>')
def shared_css(filename):
    return send_from_directory('public/css', filename)

@app.route('/js/<path:filename>')
def shared_js(filename):
    return send_from_directory('public/js', filename)

# ── Public Food Order (no auth) ───────────────────────────────────────────────

@app.route('/api/public/bundle-items', methods=['GET'])
def public_bundle_items():
    """Public — return active food items with bundle quantities for a given size.
    Used by the family My Order portal so families can select/deselect items."""
    size = (request.args.get('size') or 'M').upper()
    if size not in ('S', 'M', 'L'):
        size = 'M'
    db = get_db()
    rows = db.execute(
        '''SELECT fi.id, fi.name, fi.unit,
                  fc.name as category, fc.display_order as cat_order,
                  fi.display_order as item_order,
                  COALESCE(bq.quantity, '0') as quantity
           FROM food_items fi
           JOIN food_categories fc ON fi.category_id = fc.id
           LEFT JOIN bundle_quantities bq
                  ON bq.food_item_id = fi.id AND bq.bundle_size = ?
           WHERE fi.is_active = 1 AND fc.is_active = 1
           ORDER BY fc.display_order, fi.display_order''',
        (size,)
    ).fetchall()
    # Group by category
    cats = {}
    for r in rows:
        cat = r['category']
        if cat not in cats:
            cats[cat] = []
        cats[cat].append({
            'id': r['id'],
            'name': r['name'],
            'unit': r['unit'],
            'quantity': r['quantity']
        })
    return jsonify([{'category': k, 'items': v} for k, v in cats.items()])

@app.errorhandler(Exception)
def handle_unhandled_exception(e):
    """Return JSON for unhandled Python exceptions; pass HTTP exceptions through normally."""
    if isinstance(e, HTTPException):
        return e  # 404, 405, etc. keep their proper status codes
    log.exception(f'Unhandled exception: {e}')
    return jsonify({'error': 'Internal server error'}), 500

@app.errorhandler(413)
def handle_request_too_large(_error):
    return jsonify({
        'error': f'Request is too large. Maximum size is '
                 f'{app.config["MAX_CONTENT_LENGTH"] // (1024 * 1024)} MB.'
    }), 413

def _cycle_order_window_error(cycle):
    """Return a user-facing error when a cycle cannot safely accept an order."""
    if not cycle or cycle['status'] != 'open':
        return 'This delivery is not currently accepting orders.'
    try:
        from datetime import date as _date
        delivery_date = _date.fromisoformat(str(cycle['delivery_date_start']))
    except (TypeError, ValueError):
        log.error(f'Cycle {cycle["id"] if cycle else "?"} has an invalid delivery_date_start')
        return 'This delivery is misconfigured. Please contact a coordinator.'
    if (delivery_date - _today_central()).days < 1:
        return 'Orders are closed for this delivery.'

    # request_open/close are stored as Central-time wall-clock values. Empty
    # legacy values are allowed, but malformed non-empty values fail closed.
    try:
        from zoneinfo import ZoneInfo
        now_local = datetime.now(ZoneInfo('America/Chicago')).replace(tzinfo=None)
    except Exception:
        now_local = datetime.utcnow()
    for field, is_open in (('request_open_at', True), ('request_close_at', False)):
        value = cycle[field] if field in cycle.keys() else None
        if not value:
            continue
        try:
            boundary = datetime.fromisoformat(str(value).replace('Z', '+00:00'))
        except (TypeError, ValueError):
            log.error(f'Cycle {cycle["id"]} has invalid {field}={value!r}')
            return 'This delivery is misconfigured. Please contact a coordinator.'
        compare_now = datetime.now(boundary.tzinfo) if boundary.tzinfo else now_local
        if is_open and compare_now < boundary:
            return 'Ordering has not opened for this delivery yet.'
        if not is_open and compare_now > boundary:
            return 'Orders are closed for this delivery.'
    return None

_EFFECTIVE_ORDER_QTY_SQL = '''
    CAST(CASE
      WHEN COALESCE(fi.allow_qty, 0) = 1
        THEN COALESCE(NULLIF(fri.quantity, 0), 1)
      ELSE COALESCE(NULLIF(bq.quantity, 0), NULLIF(fri.quantity, 0), 1)
    END AS INTEGER)
'''

def _validate_order_selection(db, raw_selected, raw_quantities, raw_custom_values,
                              bundle_size, enforce_budget=True):
    """Validate and normalize family-controlled item, quantity and free-text data."""
    if not isinstance(raw_selected, list) or len(raw_selected) > 100:
        return None, 'selected_items must be a list of valid item IDs.'
    if not all(isinstance(item_id, str) and 1 <= len(item_id) <= 100
               for item_id in raw_selected):
        return None, 'selected_items contains an invalid item ID.'
    if raw_quantities is None:
        raw_quantities = {}
    if raw_custom_values is None:
        raw_custom_values = {}
    if not isinstance(raw_quantities, dict) or len(raw_quantities) > 100:
        return None, 'item_quantities must be an object.'
    if not isinstance(raw_custom_values, dict) or len(raw_custom_values) > 100:
        return None, 'item_custom_values must be an object.'

    selected_ids = set(raw_selected)
    rows = db.execute(
        '''SELECT fi.id, COALESCE(fi.price,0) price,
                  COALESCE(fi.allow_qty,0) allow_qty,
                  COALESCE(fi.is_free_text,0) is_free_text,
                  COALESCE(NULLIF(CAST(bq.quantity AS INTEGER),0),0) bundle_qty
           FROM food_items fi
           LEFT JOIN bundle_quantities bq
             ON bq.food_item_id=fi.id AND bq.bundle_size=?
           WHERE fi.is_active=1''',
        (bundle_size,)
    ).fetchall()
    active = {row['id']: row for row in rows}
    unknown = selected_ids - set(active)
    if unknown:
        return None, 'One or more selected items are unavailable.'

    quantities = {}
    custom_values = {}
    total_cost = 0.0
    for item_id in selected_ids:
        item = active[item_id]
        raw_qty = raw_quantities.get(item_id, 1)
        try:
            if isinstance(raw_qty, bool):
                raise ValueError
            qty = int(raw_qty)
        except (TypeError, ValueError):
            return None, 'Item quantities must be whole numbers.'
        # Fixed bundle lines cannot be manipulated by the browser. Persist the
        # configured bundle quantity so every downstream shopping view sees the
        # same amount; adjustable lines keep the family's bounded selection.
        if not item['allow_qty']:
            qty = int(item['bundle_qty'] or 0) or 1
        if not 1 <= qty <= 20:
            return None, 'Item quantities must be between 1 and 20.'
        quantities[item_id] = qty
        cost_qty = qty if item['allow_qty'] else 1
        total_cost += float(item['price'] or 0) * cost_qty

        raw_custom = raw_custom_values.get(item_id, '')
        if raw_custom is not None and not isinstance(raw_custom, str):
            return None, 'Custom item values must be text.'
        custom = (raw_custom or '').strip()
        if len(custom) > 120:
            return None, 'Custom item values must be 120 characters or fewer.'
        if custom and not item['is_free_text']:
            return None, 'A custom value was supplied for an item that does not accept one.'
        custom_values[item_id] = custom or None

    budget_row = db.execute(
        "SELECT COALESCE(budget,0) budget FROM bundle_size_rules WHERE bundle_size=?",
        (bundle_size,)
    ).fetchone()
    budget = float(budget_row['budget']) if budget_row else 0.0
    if enforce_budget and budget > 0 and total_cost > budget:
        return None, 'Your selection exceeds your bundle limit. Please remove some items.'
    return {
        'selected_ids': selected_ids,
        'quantities': quantities,
        'custom_values': custom_values,
    }, None

@app.route('/api/food-order/check', methods=['GET'])
def check_food_order_eligibility():
    """Return family info + all delivery cycles within next 12 months with per-cycle order state.
    Accepts either:
      - Authorization: Bearer <token>  (new session-based auth for family users)
      - ?phone=<phone>                 (legacy phone lookup — kept for backward compat)
    """
    import json as _json
    from datetime import timedelta, date as _date

    db = get_db()
    family = None

    # --- Session-based auth (new) ---
    auth = request.headers.get('Authorization', '')
    if auth.startswith('Bearer '):
        fam_session = get_family_session(auth[7:])
        if not fam_session:
            return jsonify({'error': 'Session expired — please log in again'}), 401
        fam_row = db.execute(
            "SELECT id, name, family_size, family_code, bundle_size, pending_bundle_size, phone, status "
            "FROM families WHERE id=?", (fam_session['family_id'],)
        ).fetchone()
        if fam_row:
            if fam_row['status'] != 'active':
                return jsonify({'error': 'account_pending',
                                'message': 'Your account is pending approval. Please contact SIHAA.'}), 403
            family = dict(fam_row)

    if not family:
        return jsonify({'error': 'Authentication required — please log in.'}), 401

    # Resolve bundle size for this family
    bundle_size = family['bundle_size'] or None
    if not bundle_size:
        sz = db.execute(
            "SELECT bundle_size FROM bundle_size_rules WHERE min_household<=? AND (max_household IS NULL OR max_household>=?) ORDER BY min_household DESC LIMIT 1",
            (family['family_size'] or 1, family['family_size'] or 1)
        ).fetchone()
        bundle_size = sz['bundle_size'] if sz else 'M'

    today   = _today_central()
    cutoff  = today + timedelta(days=365)  # 12-month visibility window

    # All non-delivered cycles within next 12 months
    upcoming_rows = db.execute(
        """SELECT * FROM delivery_cycles
           WHERE delivery_date_start >= ? AND delivery_date_start <= ?
             AND status NOT IN ('delivered')
           ORDER BY delivery_date_start""",
        (today.isoformat(), cutoff.isoformat())
    ).fetchall()

    # Also include any cycle outside that window where the family has a non-terminal active order
    extra_rows = db.execute(
        """SELECT dc.* FROM food_requests fr
           JOIN delivery_cycles dc ON fr.cycle_id = dc.id
           WHERE fr.family_id=? AND fr.status NOT IN ('skipped','cancelled','delivered')
             AND dc.status NOT IN ('delivered')
             AND (dc.delivery_date_start < ? OR dc.delivery_date_start > ?)
           ORDER BY dc.delivery_date_start""",
        (family['id'], today.isoformat(), cutoff.isoformat())
    ).fetchall()

    seen = {r['id'] for r in upcoming_rows}
    all_cycles = list(upcoming_rows)
    for r in extra_rows:
        if r['id'] not in seen:
            all_cycles.append(r)
            seen.add(r['id'])
    all_cycles.sort(key=lambda r: r['delivery_date_start'])

    def _build_items_for_selection(bsize):
        """Return bundle item list grouped by category for the order placement form.
        NOTE: price is included for silent client-side budget math — never display it to families.
        allow_qty controls whether +/- qty stepper is shown (vs simple checkbox).
        is_default: pre-checked when family opens the form.
        group_id/group_max: mutual-exclusion group (e.g. 'beans', 'fruit', 'bread_pasta').
        is_free_text: show a text input alongside the checkbox (for 'Other Fruit' etc).
        """
        rows = db.execute(
            '''SELECT fi.id, fi.name, fi.unit,
                      COALESCE(fi.price, 0) as price,
                      COALESCE(fi.allow_qty, 0) as allow_qty,
                      COALESCE(fi.is_default, 0) as is_default,
                      fi.group_id,
                      COALESCE(fi.group_max, 1) as group_max,
                      COALESCE(fi.is_free_text, 0) as is_free_text,
                      fc.name as category, fc.display_order as cat_order,
                      COALESCE(bq.quantity,'') as quantity
               FROM food_items fi
               JOIN food_categories fc ON fi.category_id=fc.id
               LEFT JOIN bundle_quantities bq ON bq.food_item_id=fi.id AND bq.bundle_size=?
               WHERE fi.is_active=1 AND fc.is_active=1
               ORDER BY fc.display_order, fi.display_order''',
            (bsize,)
        ).fetchall()
        cats = {}
        for r in rows:
            # Use bundle_quantities as the per-size default.
            # If bq.quantity > 0, the item is pre-selected for this bundle size.
            # Fall back to global is_default only when no bundle quantity is configured.
            bq_qty = int(r['quantity']) if str(r['quantity']).isdigit() else 0
            has_bq = bq_qty > 0
            is_default = has_bq or (not r['quantity'] and r['is_default'])
            default_qty = bq_qty if has_bq else (1 if r['is_default'] else 0)
            cats.setdefault(r['category'], []).append({
                'id':          r['id'],
                'name':        r['name'],
                'unit':        r['unit'],
                'quantity':    r['quantity'],
                'default_qty': default_qty,      # pre-filled qty from bundle_quantities
                'price':       r['price'],       # budget math only — never shown to family
                'allow_qty':   r['allow_qty'],   # 1 = +/- stepper; 0 = checkbox
                'is_default':  1 if is_default else 0,
                'group_id':    r['group_id'],    # mutual-exclusion group key
                'group_max':   r['group_max'],   # max items selectable from this group
                'is_free_text':r['is_free_text'],# show text input when checked
            })
        return [{'category': k, 'items': v} for k, v in cats.items()]

    def _build_order_obj(existing, cycle):
        """Build the full order object for a cycle where an order exists."""
        bsize = existing['bundle_size'] or bundle_size

        # Backfill items if the order has no item rows
        item_count = db.execute(
            "SELECT COUNT(*) FROM food_request_items WHERE request_id=?", (existing['id'],)
        ).fetchone()[0]
        if item_count == 0:
            all_items = db.execute("SELECT id FROM food_items WHERE is_active=1").fetchall()
            for it in all_items:
                try:
                    db.execute(
                        "INSERT OR IGNORE INTO food_request_items (id,request_id,food_item_id,selected) VALUES (?,?,?,1)",
                        (str(uuid.uuid4()), existing['id'], it['id'])
                    )
                except Exception:
                    pass
            db.commit()

        # Selected items grouped by category
        sel_rows = db.execute(
            f'''SELECT fi.name, fi.unit, fc.name as category,
                      {_EFFECTIVE_ORDER_QTY_SQL} as quantity
               FROM food_request_items fri
               JOIN food_items fi ON fri.food_item_id=fi.id
               JOIN food_categories fc ON fi.category_id=fc.id
               LEFT JOIN bundle_quantities bq ON bq.food_item_id=fi.id AND bq.bundle_size=?
               WHERE fri.request_id=? AND fri.selected=1
               ORDER BY fc.display_order, fi.display_order''',
            (bsize, existing['id'])
        ).fetchall()
        sel_cats = {}
        for r in sel_rows:
            sel_cats.setdefault(r['category'], []).append(
                {'name': r['name'], 'unit': r['unit'], 'quantity': r['quantity']}
            )

        # Full bundle list (for change-request checklist)
        full_rows = db.execute(
            '''SELECT fi.id, fi.name, fi.unit, fc.name as category,
                      COALESCE(bq.quantity,'') as quantity
               FROM food_items fi
               JOIN food_categories fc ON fi.category_id=fc.id
               LEFT JOIN bundle_quantities bq ON bq.food_item_id=fi.id AND bq.bundle_size=?
               ORDER BY fc.display_order, fi.display_order''',
            (bsize,)
        ).fetchall()
        full_cats = {}
        for r in full_rows:
            full_cats.setdefault(r['category'], []).append(
                {'id': r['id'], 'name': r['name'], 'unit': r['unit'], 'quantity': r['quantity']}
            )

        # Cancel / change-request eligibility
        try:
            ddt = _date.fromisoformat(cycle['delivery_date_start'])
            days_until  = (ddt - today).days
            _terminal   = existing['status'] in ('skipped', 'delivered', 'cancelled')
            can_cancel  = days_until >= 1 and not _terminal
            can_request_change = (
                not _terminal
                and 1 <= days_until <= 30
                and cycle['status'] not in ('shopping', 'delivered')
            )
        except Exception:
            can_cancel = can_request_change = False

        # Pending change request
        pending_cr = db.execute(
            "SELECT * FROM order_change_requests WHERE request_id=? AND status='pending' ORDER BY created_at DESC LIMIT 1",
            (existing['id'],)
        ).fetchone()
        pending_change_request = None
        if pending_cr:
            try:
                cr_payload = _json.loads(pending_cr['payload'])
            except Exception:
                cr_payload = {}
            pending_change_request = {
                'id':           pending_cr['id'],
                'family_notes': pending_cr['family_notes'],
                'payload':      cr_payload,
                'created_at':   pending_cr['created_at'],
            }
            can_request_change = False

        # Event timeline
        ev_rows = db.execute(
            "SELECT event_type, actor, payload, created_at FROM food_request_events WHERE request_id=? ORDER BY created_at ASC",
            (existing['id'],)
        ).fetchall()
        events_list = []
        for ev in ev_rows:
            try:
                pl = _json.loads(ev['payload'])
            except Exception:
                pl = {}
            events_list.append({
                'event_type': ev['event_type'],
                'actor':      ev['actor'],
                'payload':    pl,
                'created_at': ev['created_at'],
            })

        fn = None
        try:
            fn = existing['family_notes']
        except Exception:
            pass

        # Volunteer assignment — show who's signed up (name only, no contact info)
        vol_slots = db.execute(
            '''SELECT vs.task_type, vs.status as slot_status, v.name as vol_name
               FROM volunteer_slots vs
               JOIN volunteers v ON vs.claimed_by = v.id
               WHERE vs.cycle_id=? AND vs.family_id=? AND vs.status IN ('claimed','confirmed')
               ORDER BY vs.task_type''',
            (cycle['id'], family['id'])
        ).fetchall()
        volunteers = {}
        for vs in vol_slots:
            volunteers[vs['task_type']] = {
                'name': vs['vol_name'],
                'confirmed': vs['slot_status'] == 'confirmed',
            }

        return {
            'id':                    existing['id'],
            'status':                existing['status'],
            'bundle_size':           bsize,
            'family_notes':          fn,
            'selected_categories':   [{'category': k, 'items': v} for k, v in sel_cats.items()],
            'bundle_categories':     [{'category': k, 'items': v} for k, v in full_cats.items()],
            'can_cancel':            can_cancel,
            'can_request_change':    can_request_change,
            'pending_change_request': pending_change_request,
            'events':                events_list,
            'volunteers':            volunteers,  # {shopping: {name, confirmed}, delivery: {name, confirmed}}
        }

    cycles_data = []
    for cycle in all_cycles:
        existing = db.execute(
            "SELECT id, bundle_size, status, family_notes FROM food_requests WHERE cycle_id=? AND family_id=?",
            (cycle['id'], family['id'])
        ).fetchone()

        # Fetch budget for this family's bundle size (never expose to family UI — for internal use)
        _budget_row = db.execute(
            "SELECT COALESCE(budget, 0) as budget FROM bundle_size_rules WHERE bundle_size=?",
            (bundle_size,)
        ).fetchone()
        _bundle_budget = float(_budget_row['budget']) if _budget_row else 0.0

        cycle_obj = {
            'id':                   cycle['id'],
            'title':                cycle['title'],
            'status':               cycle['status'],
            'delivery_date_start':  cycle['delivery_date_start'],
            'delivery_date_end':    cycle['delivery_date_end'],
            'request_close_at':     cycle['request_close_at'],
            'order':                None,
            'can_place_order':      False,
            'items_for_selection':  [],
            'bundle_budget':        _bundle_budget,  # used for silent client-side budget check
        }

        if existing:
            cycle_obj['order'] = _build_order_obj(existing, cycle)
        elif cycle['status'] == 'open':
            cycle_obj['can_place_order']     = True
            cycle_obj['items_for_selection'] = _build_items_for_selection(bundle_size)

        cycles_data.append(cycle_obj)

    # History: all past orders (delivered cycles OR terminal order statuses)
    history_rows = db.execute(
        """SELECT fr.id, fr.status, fr.bundle_size, fr.submitted_at,
                  dc.title as cycle_title, dc.delivery_date_start, dc.delivery_date_end
           FROM food_requests fr
           JOIN delivery_cycles dc ON fr.cycle_id=dc.id
           WHERE fr.family_id=? AND (dc.status='delivered' OR fr.status IN ('delivered','cancelled','skipped'))
           ORDER BY dc.delivery_date_start DESC LIMIT 30""",
        (family['id'],)
    ).fetchall()

    return jsonify({
        'registered':         True,
        'family_name':        family['name'],
        'family_id':          family['id'],
        'bundle_size':        bundle_size,
        'pending_bundle_size': family['pending_bundle_size'],
        'cycles':             cycles_data,
        'history':            [dict(r) for r in history_rows],
    })

@app.route('/api/food-order', methods=['POST'])
@require_family_auth()
def submit_food_order():
    """Place a food order for a family. Accepts optional notes field."""
    data = request.get_json(silent=True) or {}
    if not isinstance(data, dict):
        return jsonify({'error': 'Invalid request'}), 400
    # selected_items can be [] (family deselects all) — check key presence, not truthiness
    if not data.get('family_id') or not data.get('cycle_id') or 'selected_items' not in data:
        return jsonify({'error': 'family_id, cycle_id, and selected_items required'}), 422
    if (not isinstance(data['family_id'], str) or len(data['family_id']) > 100
            or not isinstance(data['cycle_id'], str) or len(data['cycle_id']) > 100):
        return jsonify({'error': 'family_id and cycle_id must be valid IDs'}), 422
    if str(data['family_id']) != str(g.fam['family_id']):
        return jsonify({'error': 'Forbidden'}), 403

    db = get_db()

    # Validate the cycle status, delivery date and configured order window.
    cycle = db.execute(
        "SELECT * FROM delivery_cycles WHERE id=?", (data['cycle_id'],)
    ).fetchone()
    cycle_error = _cycle_order_window_error(cycle)
    if cycle_error:
        return jsonify({'error': cycle_error}), 409

    # Validate family
    family = db.execute("SELECT * FROM families WHERE id=?", (data['family_id'],)).fetchone()
    if not family:
        return jsonify({'error': 'Family not found.'}), 404
    if family['status'] != 'active':
        return jsonify({'error': 'This family account is not active.'}), 403

    # Enforce one order per family per cycle
    if db.execute("SELECT id FROM food_requests WHERE cycle_id=? AND family_id=?",
                  (data['cycle_id'], data['family_id'])).fetchone():
        return jsonify({'error': 'You have already placed an order for this delivery.'}), 409

    # Determine bundle size — family override takes priority over size rules
    bundle_size = family['bundle_size'] or None
    if not bundle_size:
        size = db.execute(
            "SELECT bundle_size FROM bundle_size_rules WHERE min_household<=? AND (max_household IS NULL OR max_household>=?) ORDER BY min_household DESC LIMIT 1",
            (family['family_size'] or 1, family['family_size'] or 1)
        ).fetchone()
        bundle_size = size['bundle_size'] if size else 'M'

    notes_value = data.get('notes')
    if notes_value is not None and not isinstance(notes_value, str):
        return jsonify({'error': 'notes must be text'}), 422
    family_notes = (notes_value or '').strip()
    if len(family_notes) > 1000:
        return jsonify({'error': 'notes must be 1000 characters or fewer'}), 422

    selection, selection_error = _validate_order_selection(
        db, data.get('selected_items'), data.get('item_quantities'),
        data.get('item_custom_values'), bundle_size
    )
    if selection_error:
        return jsonify({'error': selection_error}), 422
    selected_ids = selection['selected_ids']
    item_quantities = selection['quantities']
    item_custom_vals = selection['custom_values']

    ts  = now()
    rid = str(uuid.uuid4())

    # Insert food request. bootstrap_db guarantees family_notes exists; a UNIQUE
    # race is translated into the same controlled duplicate response.
    try:
        db.execute(
            '''INSERT INTO food_requests
               (id, cycle_id, family_id, bundle_size, submitted_at, status, confirmed_at, family_notes)
               VALUES (?,?,?,?,?,?,?,?)''',
            (rid, data['cycle_id'], data['family_id'], bundle_size, ts, 'confirmed', ts, family_notes or None)
        )
    except sqlite3.IntegrityError:
        db.rollback()
        return jsonify({'error': 'You have already placed an order for this delivery.'}), 409

    # Save item selections with quantities and custom values
    all_items = db.execute("SELECT id FROM food_items WHERE is_active=1").fetchall()
    for item in all_items:
        is_selected  = 1 if item['id'] in selected_ids else 0
        qty          = item_quantities.get(item['id'], 1) if is_selected else 1
        custom_val   = item_custom_vals.get(item['id']) if is_selected else None
        db.execute(
            "INSERT INTO food_request_items (id, request_id, food_item_id, selected, quantity, custom_value) VALUES (?,?,?,?,?,?)",
            (str(uuid.uuid4()), rid, item['id'], is_selected, qty, custom_val)
        )

    # Ensure slots exist (safety net — should already be pre-created)
    slots_created = _ensure_volunteer_slots(db, data['cycle_id'], data['family_id'])

    # Confirm any claimed volunteer slots and notify those volunteers
    claimed_slots = db.execute(
        '''SELECT vs.id, vs.task_type, v.name as vol_name, v.phone as vol_phone,
                  f.address, f.city, f.name as family_name, f.bundle_size as fam_bundle
           FROM volunteer_slots vs
           JOIN volunteers v ON vs.claimed_by = v.id
           JOIN families f ON vs.family_id = f.id
           WHERE vs.cycle_id=? AND vs.family_id=? AND vs.status IN ('claimed','confirmed') ''',
        (data['cycle_id'], data['family_id'])
    ).fetchall()

    # Build item list for shoppers (fetch selected items + quantities)
    item_lines = []
    if claimed_slots:
        item_rows = db.execute(
            f'''SELECT fi.name, fi.unit, {_EFFECTIVE_ORDER_QTY_SQL} as ord_qty,
                       fc.name as category
               FROM food_request_items fri
               JOIN food_requests fr ON fr.id=fri.request_id
               JOIN food_items fi ON fri.food_item_id = fi.id
               JOIN food_categories fc ON fi.category_id = fc.id
               LEFT JOIN bundle_quantities bq
                 ON bq.food_item_id=fi.id AND bq.bundle_size=fr.bundle_size
               WHERE fri.request_id = ? AND fri.selected = 1
               ORDER BY fc.display_order, fi.name''',
            (rid,)
        ).fetchall()
        for ir in item_rows:
            qty_label = f"×{ir['ord_qty']}" if ir['ord_qty'] and ir['ord_qty'] > 1 else ''
            unit_str = ir['unit'] or ''
            item_lines.append(f"  • {ir['name']}{(' ' + qty_label) if qty_label else ''}{(' (' + unit_str + ')') if unit_str else ''}")

    # Update slot statuses to confirmed (DB only — no WA yet)
    for slot in claimed_slots:
        db.execute(
            "UPDATE volunteer_slots SET status='confirmed', updated_at=? WHERE id=?",
            (now(), slot['id'])
        )

    # Audit log (before commit so it's in the same transaction)
    try:
        _log_order_event(db, rid, 'confirmed', 'family', {
            'source': 'portal',
            'items_count': len(selected_ids),
            'notes': family_notes or None,
        })
    except Exception:
        pass

    db.commit()  # ← commit everything first, then notify in background

    # Notify volunteers via email + coordinators — fire-and-forget
    cycle_start = cycle['delivery_date_start']
    items_text  = '\n'.join(item_lines) if item_lines else '  (no items selected)'
    email_sends = []
    for slot in [dict(s) for s in claimed_slots]:
        vol_email = _lookup_volunteer_email(db, slot.get('claimed_by') or '') if slot.get('claimed_by') else ''
        if vol_email:
            if slot['task_type'] == 'shopping':
                body = (f"Assalamu Alaikum,\n\nOrder Confirmed — Shopping Task\n\n"
                        f"Family: {slot['family_name']}\n"
                        f"Delivery: {cycle_start}\n\n"
                        f"Shopping list:\n{items_text}\n\nJazakAllah Khair!\n\n— Sihha Food Program")
            else:
                body = (f"Assalamu Alaikum,\n\nOrder Confirmed — Delivery Task\n\n"
                        f"Family: {slot['family_name']}\n"
                        f"Delivery: {cycle_start}\n"
                        f"Address: {slot.get('address') or 'TBD'}, {slot.get('city') or ''}\n\n"
                        f"JazakAllah Khair!\n\n— Sihha Food Program")
            email_sends.append((vol_email, f'Sihha Order Confirmed — {slot["family_name"]}', body))
    _email_notify_async(email_sends)

    coord_msg = (
        f"New order placed via portal:\n"
        f"Family: {family['name']}\n"
        f"Cycle: {cycle['title']}\n"
        f"Items selected: {len(selected_ids)}"
        + (f"\nNotes: {family_notes}" if family_notes else '')
    )
    try:
        _notify_coordinators(db, coord_msg)
    except Exception:
        pass

    log.info(f'Food order placed: family {data["family_id"]} cycle {data["cycle_id"]} — {slots_created} new slots, {len(claimed_slots)} slots confirmed')
    return jsonify({
        'ok': True,
        'request_id': rid,
        'message': 'Your order has been placed.',
        'delivery_start': cycle['delivery_date_start'],
        'delivery_end':   cycle['delivery_date_end'],
    }), 201

@app.route('/api/food-order/cancel', methods=['POST'])
@require_family_auth()
def cancel_food_order():
    """Family cancels their confirmed order — allowed up to 24 hours before delivery (Central time)."""
    family_id = request_id = None
    try:
        data      = request.get_json(silent=True) or {}
        if not isinstance(data, dict):
            return jsonify({'error': 'Invalid request'}), 400
        family_id = data.get('family_id')
        request_id = data.get('request_id')
        if not family_id or not request_id:
            return jsonify({'error': 'family_id and request_id required'}), 422
        if (not isinstance(family_id, str) or len(family_id) > 100
                or not isinstance(request_id, str) or len(request_id) > 100):
            return jsonify({'error': 'family_id and request_id must be valid IDs'}), 422
        if str(family_id) != str(g.fam['family_id']):
            return jsonify({'error': 'Forbidden'}), 403

        db  = get_db()
        req = db.execute(
            '''SELECT fr.*, dc.delivery_date_start, dc.title as cycle_title,
                      f.name as family_name, f.family_code
               FROM food_requests fr
               JOIN delivery_cycles dc ON fr.cycle_id=dc.id
               JOIN families f ON fr.family_id=f.id
               WHERE fr.id=? AND fr.family_id=?''',
            (request_id, family_id)
        ).fetchone()
        if not req:
            return jsonify({'error': 'Order not found'}), 404
        if req['status'] in ('skipped', 'delivered', 'cancelled'):
            return jsonify({'error': 'Order cannot be cancelled in its current state'}), 409

        # Enforce 24-hour cutoff using Central time
        try:
            from datetime import date as _date
            delivery_dt = _date.fromisoformat(req['delivery_date_start'])
            days_until  = (delivery_dt - _today_central()).days
        except (TypeError, ValueError):
            log.error(f'cancel_food_order: invalid delivery date on cycle {req["cycle_id"]}')
            return jsonify({
                'error': 'This delivery is misconfigured. Please contact a coordinator before cancelling.'
            }), 409
        if days_until < 1:
            return jsonify({'error': 'Orders can only be cancelled at least 1 day before delivery'}), 409

        # Find claimed/confirmed volunteers BEFORE releasing slots (for notification)
        claimed_volunteers = db.execute(
            '''SELECT v.id, v.name, v.email, vs.task_type
               FROM volunteer_slots vs
               JOIN volunteers v ON vs.claimed_by = v.id
               WHERE vs.cycle_id=? AND vs.family_id=? AND vs.status IN ('claimed','confirmed') ''',
            (req['cycle_id'], family_id)
        ).fetchall()

        # Release claimed + confirmed volunteer slots back to open
        try:
            db.execute(
                "UPDATE volunteer_slots SET prev_claimed_by=claimed_by, claimed_by=NULL, claimed_at=NULL, status='open', updated_at=? WHERE cycle_id=? AND family_id=? AND status IN ('claimed','confirmed')",
                (now(), req['cycle_id'], family_id)
            )
        except Exception:
            db.execute(
                "UPDATE volunteer_slots SET claimed_by=NULL, claimed_at=NULL, status='open' WHERE cycle_id=? AND family_id=? AND status IN ('claimed','confirmed')",
                (req['cycle_id'], family_id)
            )

        # Log event + hard-delete in ONE transaction (audit 3.7) — previously a
        # crash between the two commits left slots released and the event logged
        # while the order still existed. Events survive the delete (FK removed).
        _log_order_event(db, request_id, 'cancelled', actor='family',
                         payload={'days_until_delivery': days_until})
        db.execute("DELETE FROM food_request_items    WHERE request_id=?", (request_id,))
        db.execute("DELETE FROM order_change_requests WHERE request_id=?", (request_id,))
        db.execute("DELETE FROM food_requests         WHERE id=?",         (request_id,))
        db.commit()

        # Fire notifications in background
        try:
            coord_msg = (
                f"Order cancelled by family:\n"
                f"Family: {req['family_name']} ({req['family_code']})\n"
                f"Cycle: {req['cycle_title']}\n"
                f"Days until delivery: {days_until}\n"
                f"Volunteer slots released back to open."
            )
            _notify_coordinators(db, coord_msg)
        except Exception:
            pass

        vol_emails = []
        for vol in claimed_volunteers:
            vol_email = _lookup_volunteer_email(db, vol['id']) if vol.get('id') else ''
            if vol_email:
                body = (f"Assalamu Alaikum {vol['name']},\n\n"
                        f"{req['family_name']} has cancelled their food order for {req['cycle_title']}.\n"
                        f"Your {vol['task_type']} slot has been released — no action needed.\n\n"
                        f"JazakAllah Khair!\n\n— Sihha Food Program")
                vol_emails.append((vol_email, f'Sihha Slot Released — {req["cycle_title"]}', body))
        _email_notify_async(vol_emails)

        log.info(f'Family {family_id} cancelled order {request_id} — {days_until} days before delivery')
        return jsonify({'ok': True, 'message': 'Your order has been cancelled. You can place a new order if needed.'})

    except Exception as _e:
        log.exception(f'cancel_food_order ERROR — family={family_id!r} request={request_id!r}')
        return jsonify({'error': 'Server error — please try again. If it persists, contact a coordinator.'}), 500


# ── Family Change Requests ────────────────────────────────────────────────────

@app.route('/api/family-request', methods=['POST'])
@require_family_auth()
def submit_family_change_request():
    """Family submits a change request for their current order.
    One pending request per order at a time. Cycle must be open/upcoming, not shopping, within 30 days."""
    import json as _json
    try:
        data       = request.get_json(silent=True) or {}
        if not isinstance(data, dict):
            return jsonify({'error': 'Invalid request'}), 400
        family_id  = data.get('family_id')
        request_id = data.get('request_id')
        raw_notes = data.get('family_notes')
        if raw_notes is not None and not isinstance(raw_notes, str):
            return jsonify({'error': 'family_notes must be text'}), 422
        family_notes = (raw_notes or '').strip()
        if len(family_notes) > 1000:
            return jsonify({'error': 'family_notes must be 1000 characters or fewer'}), 422
        selected_item_ids = data.get('selected_item_ids')

        if not family_id or not request_id:
            return jsonify({'error': 'family_id and request_id required'}), 422
        if (not isinstance(family_id, str) or len(family_id) > 100
                or not isinstance(request_id, str) or len(request_id) > 100):
            return jsonify({'error': 'family_id and request_id must be valid IDs'}), 422
        if str(family_id) != str(g.fam['family_id']):
            return jsonify({'error': 'Forbidden'}), 403

        db = get_db()

        # Verify the order belongs to this family
        req = db.execute(
            '''SELECT fr.*, dc.delivery_date_start, dc.status as cycle_status, dc.title as cycle_title,
                      f.name as family_name, f.wa_phone, f.wa_apikey
               FROM food_requests fr
               JOIN delivery_cycles dc ON fr.cycle_id = dc.id
               JOIN families f ON fr.family_id = f.id
               WHERE fr.id=? AND fr.family_id=?''',
            (request_id, family_id)
        ).fetchone()
        if not req:
            return jsonify({'error': 'Order not found'}), 404
        if req['status'] in ('cancelled', 'skipped', 'delivered'):
            return jsonify({'error': 'Cannot request changes for this order'}), 409
        if req['cycle_status'] in ('shopping', 'delivered'):
            return jsonify({'error': 'Changes are not allowed once shopping has started'}), 409

        # 30-day window check
        try:
            from datetime import date as _d
            days_until = (_d.fromisoformat(req['delivery_date_start']) - _today_central()).days
        except (TypeError, ValueError):
            log.error(f'submit_family_change_request: invalid delivery date on cycle {req["cycle_id"]}')
            return jsonify({
                'error': 'This delivery is misconfigured. Please contact a coordinator.'
            }), 409
        if days_until > 30:
            return jsonify({'error': 'Change requests can only be submitted within 30 days of delivery'}), 409
        if days_until < 1:
            return jsonify({'error': 'Delivery is too soon to request changes'}), 409

        # No duplicate pending request for this order
        existing = db.execute(
            "SELECT id FROM order_change_requests WHERE request_id=? AND status='pending'",
            (request_id,)
        ).fetchone()
        if existing:
            return jsonify({'error': 'You already have a pending change request for this order'}), 409

        selection, selection_error = _validate_order_selection(
            db, selected_item_ids, {}, {}, req['bundle_size']
        )
        if selection_error:
            return jsonify({'error': selection_error}), 422
        selected_item_ids = sorted(selection['selected_ids'])

        # Build payload — item selections
        payload = _json.dumps({'selected_item_ids': selected_item_ids})

        # Resolve item names for the event log (so history shows human-readable items)
        requested_item_names = []
        if selected_item_ids:
            try:
                placeholders = ','.join('?' * len(selected_item_ids))
                requested_item_names = [
                    r['name'] for r in db.execute(
                        f"SELECT name FROM food_items WHERE id IN ({placeholders})",
                        list(selected_item_ids)
                    ).fetchall()
                ]
            except Exception:
                pass

        cr_id = str(uuid.uuid4())
        db.execute(
            '''INSERT INTO order_change_requests
               (id, family_id, cycle_id, request_id, status, family_notes, payload, created_at)
               VALUES (?,?,?,?,?,?,?,?)''',
            (cr_id, family_id, req['cycle_id'], request_id, 'pending', family_notes, payload, now())
        )
        _log_order_event(db, request_id, 'change_requested', actor='family',
                         payload={
                             'change_request_id': cr_id,
                             'notes':             family_notes,
                             'requested_items':   requested_item_names,
                         })
        db.commit()

        # Notify coordinators
        _notify_coordinators(db,
            f"Change request from family:\n"
            f"Family: {req['family_name']}\n"
            f"Cycle: {req['cycle_title']}\n"
            f"Notes: {family_notes or '(no notes)'}\n"
            f"Items selected: {len(selected_item_ids)}\n"
            f"Review in admin → Requests"
        )

        log.info(f'Change request {cr_id} submitted by family {family_id} for order {request_id}')
        return jsonify({'ok': True, 'change_request_id': cr_id})

    except Exception as _e:
        log.exception(f'submit_family_change_request ERROR')
        return jsonify({'error': 'Server error — please try again. If it persists, contact a coordinator.'}), 500


@app.route('/api/family-request/<cr_id>/retract', methods=['POST'])
@require_family_auth()
def retract_family_change_request(cr_id):
    """Family retracts their pending change request."""
    try:
        data      = request.json or {}
        family_id = data.get('family_id')
        if not family_id:
            return jsonify({'error': 'family_id required'}), 422
        if str(family_id) != str(g.fam['family_id']):
            return jsonify({'error': 'Forbidden'}), 403

        db = get_db()
        cr = db.execute(
            "SELECT * FROM order_change_requests WHERE id=? AND family_id=? AND status='pending'",
            (cr_id, family_id)
        ).fetchone()
        if not cr:
            return jsonify({'error': 'Request not found or already reviewed'}), 404

        db.execute(
            "UPDATE order_change_requests SET status='retracted', reviewed_at=? WHERE id=?",
            (now(), cr_id)
        )
        _log_order_event(db, cr['request_id'], 'change_retracted', actor='family',
                         payload={'change_request_id': cr_id})
        db.commit()
        return jsonify({'ok': True})

    except Exception as _e:
        log.exception('retract_family_change_request ERROR')
        return jsonify({'error': 'Server error — please try again. If it persists, contact a coordinator.'}), 500


# ── Admin Change Request Routes ───────────────────────────────────────────────

@app.route('/api/admin/change-requests')
@require_auth(roles=['admin', 'finance', 'treasurer', 'viewer'])
def list_change_requests():
    """Admin: list change requests. Default: pending only. ?status=all for all."""
    db = get_db()
    status_filter = request.args.get('status', 'pending')
    if status_filter == 'all':
        rows = db.execute(
            '''SELECT ocr.*, f.name as family_name, f.family_code,
                      dc.title as cycle_title, dc.delivery_date_start,
                      u.name as reviewed_by_name
               FROM order_change_requests ocr
               JOIN families f ON ocr.family_id = f.id
               JOIN delivery_cycles dc ON ocr.cycle_id = dc.id
               LEFT JOIN users u ON ocr.reviewed_by = u.id
               ORDER BY ocr.created_at DESC LIMIT 100''',
        ).fetchall()
    else:
        rows = db.execute(
            '''SELECT ocr.*, f.name as family_name, f.family_code,
                      dc.title as cycle_title, dc.delivery_date_start,
                      u.name as reviewed_by_name
               FROM order_change_requests ocr
               JOIN families f ON ocr.family_id = f.id
               JOIN delivery_cycles dc ON ocr.cycle_id = dc.id
               LEFT JOIN users u ON ocr.reviewed_by = u.id
               WHERE ocr.status=?
               ORDER BY ocr.created_at DESC''',
            (status_filter,)
        ).fetchall()

    import json as _json
    result = []
    for r in rows:
        row = dict(r)
        try:
            row['payload'] = _json.loads(row['payload'])
        except Exception:
            row['payload'] = {}
        # Attach current item names for the selected IDs
        selected_ids = row['payload'].get('selected_item_ids', [])
        if selected_ids:
            placeholders = ','.join('?' * len(selected_ids))
            item_rows = db.execute(
                f"SELECT fi.id, fi.name, fi.unit, fc.name as category FROM food_items fi JOIN food_categories fc ON fi.category_id=fc.id WHERE fi.id IN ({placeholders})",
                selected_ids
            ).fetchall()
            row['selected_items'] = [dict(i) for i in item_rows]
        else:
            row['selected_items'] = []
        result.append(row)
    return jsonify(result)


@app.route('/api/admin/change-requests/<cr_id>/approve', methods=['POST'])
@require_auth(roles=['admin'])
def approve_change_request(cr_id):
    """Admin approves a change request — automatically applies item changes to the order."""
    import json as _json
    try:
        data = request.json or {}
        admin_notes = (data.get('admin_notes') or '').strip()
        db = get_db()

        cr = db.execute(
            '''SELECT ocr.*, f.name as family_name, f.phone as family_phone,
                      dc.title as cycle_title, dc.status as cycle_status
               FROM order_change_requests ocr
               JOIN families f ON ocr.family_id = f.id
               JOIN delivery_cycles dc ON ocr.cycle_id = dc.id
               WHERE ocr.id=? AND ocr.status='pending' ''',
            (cr_id,)
        ).fetchone()
        if not cr:
            return jsonify({'error': 'Request not found or already reviewed'}), 404

        # Parse requested items
        try:
            payload = _json.loads(cr['payload'])
        except Exception:
            payload = {}
        selected_ids = set(payload.get('selected_item_ids', []))

        # Capture item names BEFORE applying changes (for the event log)
        def _item_names_for_request(rid, selected_only=True):
            cond = "AND fri.selected=1" if selected_only else ""
            rows = db.execute(
                f'''SELECT fi.name FROM food_request_items fri
                    JOIN food_items fi ON fri.food_item_id=fi.id
                    WHERE fri.request_id=? {cond}
                    ORDER BY fi.name''',
                (rid,)
            ).fetchall()
            return [r['name'] for r in rows]

        items_before = _item_names_for_request(cr['request_id']) if cr['request_id'] else []

        # Apply item changes to the order
        if cr['request_id']:
            all_items = db.execute(
                "SELECT food_item_id FROM food_request_items WHERE request_id=?",
                (cr['request_id'],)
            ).fetchall()
            for item in all_items:
                new_selected = 1 if item['food_item_id'] in selected_ids else 0
                db.execute(
                    "UPDATE food_request_items SET selected=? WHERE request_id=? AND food_item_id=?",
                    (new_selected, cr['request_id'], item['food_item_id'])
                )

        # Capture item names AFTER applying changes
        items_after = _item_names_for_request(cr['request_id']) if cr['request_id'] else []
        added   = [n for n in items_after  if n not in set(items_before)]
        removed = [n for n in items_before if n not in set(items_after)]

        # Mark request approved
        db.execute(
            '''UPDATE order_change_requests
               SET status='approved', admin_notes=?, reviewed_by=?, reviewed_at=?
               WHERE id=?''',
            (admin_notes, g.user['user_id'], now(), cr_id)
        )

        _log_order_event(db, cr['request_id'], 'change_approved', actor='admin',
                         payload={
                             'change_request_id': cr_id,
                             'admin_notes':       admin_notes,
                             'items_before':      items_before,
                             'items_after':       items_after,
                             'added':             added,
                             'removed':           removed,
                         })
        db.commit()

        # Email family
        fam_email = _lookup_family_email(db, cr['family_id']) if cr.get('family_id') else ''
        if fam_email:
            body = f"Assalamu Alaikum {cr['family_name']},\n\nYour change request for {cr['cycle_title']} has been approved."
            if admin_notes:
                body += f"\nCoordinator note: {admin_notes}"
            body += "\n\n— Sihha Food Program"
            _email_notify(fam_email, f'Sihha Change Request Approved — {cr["cycle_title"]}', body)

        log.info(f'Change request {cr_id} approved by {g.user["username"]}')
        return jsonify({'ok': True})

    except Exception as _e:
        log.exception(f'approve_change_request ERROR cr_id={cr_id}')
        return jsonify({'error': 'Server error — please try again. If it persists, contact a coordinator.'}), 500


@app.route('/api/admin/change-requests/<cr_id>/reject', methods=['POST'])
@require_auth(roles=['admin'])
def reject_change_request(cr_id):
    """Admin rejects a change request — order stays unchanged."""
    try:
        data = request.json or {}
        admin_notes = (data.get('admin_notes') or '').strip()
        db = get_db()

        cr = db.execute(
            '''SELECT ocr.*, f.name as family_name, f.phone as family_phone,
                      dc.title as cycle_title
               FROM order_change_requests ocr
               JOIN families f ON ocr.family_id = f.id
               JOIN delivery_cycles dc ON ocr.cycle_id = dc.id
               WHERE ocr.id=? AND ocr.status='pending' ''',
            (cr_id,)
        ).fetchone()
        if not cr:
            return jsonify({'error': 'Request not found or already reviewed'}), 404

        db.execute(
            '''UPDATE order_change_requests
               SET status='rejected', admin_notes=?, reviewed_by=?, reviewed_at=?
               WHERE id=?''',
            (admin_notes, g.user['user_id'], now(), cr_id)
        )
        _log_order_event(db, cr['request_id'], 'change_rejected', actor='admin',
                         payload={'change_request_id': cr_id, 'admin_notes': admin_notes})
        db.commit()

        # Email family
        fam_email = _lookup_family_email(db, cr['family_id']) if cr.get('family_id') else ''
        if fam_email:
            body = f"Assalamu Alaikum {cr['family_name']},\n\nYour change request for {cr['cycle_title']} was not approved."
            if admin_notes:
                body += f"\nCoordinator note: {admin_notes}"
            body += "\n\n— Sihha Food Program"
            _email_notify(fam_email, f'Sihha Change Request Update — {cr["cycle_title"]}', body)

        log.info(f'Change request {cr_id} rejected by {g.user["username"]}')
        return jsonify({'ok': True})

    except Exception as _e:
        log.exception(f'reject_change_request ERROR cr_id={cr_id}')
        return jsonify({'error': 'Server error — please try again. If it persists, contact a coordinator.'}), 500


@app.route('/api/families/<fid>/reset-order', methods=['POST'])
@require_auth(roles=['admin'])
def reset_family_order(fid):
    """Admin resets a family's order for the most recent active cycle.
    - Cancelled / skipped orders: DELETED entirely so the family can place a fresh order.
    - Confirmed orders: reset to pending_confirmation (items cleared, no WA sent).
    """
    try:
        db = get_db()
        family = db.execute("SELECT * FROM families WHERE id=?", (fid,)).fetchone()
        if not family:
            return jsonify({'error': 'Family not found'}), 404

        # Find most recent non-delivered order
        req = db.execute(
            '''SELECT fr.* FROM food_requests fr
               JOIN delivery_cycles dc ON fr.cycle_id = dc.id
               WHERE fr.family_id=? AND fr.status != 'delivered'
               ORDER BY dc.delivery_date_start DESC LIMIT 1''',
            (fid,)
        ).fetchone()
        if not req:
            return jsonify({'error': 'No order found to reset'}), 404

        ts = now()

        # All statuses: hard-delete the order so family can place a fresh one via the portal
        # Release any claimed/confirmed volunteer slots back to open first
        db.execute(
            "UPDATE volunteer_slots SET prev_claimed_by=claimed_by, claimed_by=NULL, claimed_at=NULL, status='open', updated_at=? WHERE cycle_id=? AND family_id=? AND status IN ('claimed','confirmed')",
            (ts, req['cycle_id'], fid)
        )
        # NOTE: food_request_events are intentionally kept — they form the permanent audit trail
        db.execute("DELETE FROM food_request_items    WHERE request_id=?", (req['id'],))
        db.execute("DELETE FROM order_change_requests WHERE request_id=?", (req['id'],))
        db.execute("DELETE FROM food_requests         WHERE id=?",         (req['id'],))
        db.execute("UPDATE families SET pending_bundle_size=NULL WHERE id=? AND pending_bundle_size IS NOT NULL", (fid,))
        db.commit()
        log.info(f'Order {req["id"]} DELETED (status={req["status"]}) by admin {g.user["username"]} for family {fid}')
        return jsonify({'ok': True, 'message': 'Order cleared. Family can now place a fresh order.'})

    except Exception as _e:
        log.exception(f'reset_family_order ERROR fid={fid}')
        return jsonify({'error': 'Server error — please try again. If it persists, contact a coordinator.'}), 500


@app.route('/api/families/<fid>/cancel-order', methods=['POST'])
@require_auth(roles=['admin'])
def admin_cancel_family_order(fid):
    """Admin cancels a family's order for a specific cycle (or most recent non-delivered).
    Logs the cancellation with actor='admin' so the family portal can distinguish it from
    a family-initiated cancellation. The order row is then deleted so the family can
    re-order for that delivery slot.
    Optional JSON body: { cycle_id: <id>, reason: <str> }
    """
    try:
        db     = get_db()
        data   = request.json or {}
        reason = (data.get('reason') or '').strip()

        family = db.execute("SELECT * FROM families WHERE id=?", (fid,)).fetchone()
        if not family:
            return jsonify({'error': 'Family not found'}), 404

        # Find the target order
        if data.get('cycle_id'):
            req = db.execute(
                "SELECT fr.*, dc.title as cycle_title FROM food_requests fr "
                "JOIN delivery_cycles dc ON fr.cycle_id=dc.id "
                "WHERE fr.family_id=? AND fr.cycle_id=? AND fr.status != 'delivered' LIMIT 1",
                (fid, data['cycle_id'])
            ).fetchone()
        else:
            req = db.execute(
                '''SELECT fr.*, dc.title as cycle_title FROM food_requests fr
                   JOIN delivery_cycles dc ON fr.cycle_id = dc.id
                   WHERE fr.family_id=? AND fr.status != 'delivered'
                   ORDER BY dc.delivery_date_start DESC LIMIT 1''',
                (fid,)
            ).fetchone()

        if not req:
            return jsonify({'error': 'No active order found for this family'}), 404

        # Release any claimed/confirmed volunteer slots back to open
        try:
            db.execute(
                "UPDATE volunteer_slots SET prev_claimed_by=claimed_by, claimed_by=NULL, "
                "claimed_at=NULL, status='open', updated_at=? "
                "WHERE cycle_id=? AND family_id=? AND status IN ('claimed','confirmed')",
                (now(), req['cycle_id'], fid)
            )
        except Exception:
            db.execute(
                "UPDATE volunteer_slots SET claimed_by=NULL, claimed_at=NULL, status='open' "
                "WHERE cycle_id=? AND family_id=? AND status IN ('claimed','confirmed')",
                (req['cycle_id'], fid)
            )

        # Log event + hard-delete in ONE transaction (audit 3.7); events survive
        # the delete (FK removed) and form the permanent audit trail
        _log_order_event(db, req['id'], 'cancelled', actor='admin',
                         payload={'cancelled_by': g.user['username'],
                                  'reason': reason or None,
                                  'prev_status': req['status']})
        db.execute("DELETE FROM food_request_items    WHERE request_id=?", (req['id'],))
        db.execute("DELETE FROM order_change_requests WHERE request_id=?", (req['id'],))
        db.execute("DELETE FROM food_requests         WHERE id=?",         (req['id'],))
        db.commit()

        log.info(f'Order {req["id"]} admin-cancelled by {g.user["username"]} for family {fid}')

        try:
            _notify_coordinators(db,
                f"Order admin-cancelled:\n"
                f"Family: {family['name']}\n"
                f"Cycle: {req['cycle_title']}\n"
                f"Cancelled by: {g.user['username']}"
                + (f"\nReason: {reason}" if reason else "")
            )
        except Exception:
            pass

        return jsonify({'ok': True, 'message': 'Order cancelled. Family can now place a fresh order.'})

    except Exception as _e:
        log.exception(f'admin_cancel_family_order ERROR fid={fid}')
        return jsonify({'error': 'Server error — please try again. If it persists, contact a coordinator.'}), 500


@app.route('/api/food-order/items', methods=['PUT'])
@require_family_auth()
def edit_food_order_items():
    """Family edits their item selections — allowed up to 48 hours before delivery (Central time).
    Cycle must still be open or upcoming (not shopping/delivered).
    Cancel is final — cancelled orders cannot be edited."""
    import json as _json
    data       = request.get_json(silent=True) or {}
    if not isinstance(data, dict):
        return jsonify({'error': 'Invalid request'}), 400
    request_id = data.get('request_id')
    raw_selected = data.get('selected_item_ids')

    if not request_id:
        return jsonify({'error': 'request_id required'}), 422
    if not isinstance(request_id, str) or len(request_id) > 100:
        return jsonify({'error': 'request_id must be a valid ID'}), 422

    db = get_db()
    family_id = g.fam['family_id']

    # Load request and its cycle — scoped to session's family
    req = db.execute(
        '''SELECT fr.*, dc.delivery_date_start, dc.status as cycle_status, dc.title as cycle_title,
                  f.name as family_name, f.family_code
           FROM food_requests fr
           JOIN delivery_cycles dc ON fr.cycle_id=dc.id
           JOIN families f ON fr.family_id=f.id
           WHERE fr.id=? AND fr.family_id=?''',
        (request_id, family_id)
    ).fetchone()
    if not req:
        return jsonify({'error': 'Order not found'}), 404
    if req['status'] in ('cancelled', 'skipped', 'delivered'):
        return jsonify({'error': 'This order can no longer be edited'}), 409
    if req['cycle_status'] in ('shopping', 'delivered'):
        return jsonify({'error': 'Editing is no longer available — the shopping cycle has started'}), 409

    # Enforce 48-hour edit window using Central time
    try:
        from datetime import date as _date
        delivery_dt = _date.fromisoformat(req['delivery_date_start'])
        days_until  = (delivery_dt - _today_central()).days
    except (TypeError, ValueError):
        log.error(f'edit_food_order_items: invalid delivery date on cycle {req["cycle_id"]}')
        return jsonify({
            'error': 'This delivery is misconfigured. Please contact a coordinator.'
        }), 409
    if days_until < 2:
        return jsonify({'error': 'Item editing closes 48 hours before delivery'}), 409

    # Capture previous selections for diff (item names, not IDs)
    prev_rows = db.execute(
        "SELECT fi.id, fi.name, fri.selected, COALESCE(fri.quantity,1) quantity, "
        "fri.custom_value FROM food_request_items fri JOIN food_items fi ON fri.food_item_id=fi.id "
        "WHERE fri.request_id=?",
        (request_id,)
    ).fetchall()
    prev_by_id = {r['id']: (r['name'], r['selected']) for r in prev_rows}

    selection, selection_error = _validate_order_selection(
        db, raw_selected,
        {r['id']: r['quantity'] for r in prev_rows},
        {r['id']: r['custom_value'] for r in prev_rows},
        req['bundle_size']
    )
    if selection_error:
        return jsonify({'error': selection_error}), 422
    selected_ids = selection['selected_ids']

    # Get all active items to upsert
    all_items = db.execute("SELECT id, name FROM food_items WHERE is_active=1").fetchall()
    for item in all_items:
        is_sel = 1 if item['id'] in selected_ids else 0
        db.execute(
            '''INSERT INTO food_request_items
               (id, request_id, food_item_id, selected, quantity, custom_value)
               VALUES (?,?,?,?,?,?)
               ON CONFLICT(request_id, food_item_id) DO UPDATE SET
                   selected=excluded.selected,
                   quantity=excluded.quantity,
                   custom_value=excluded.custom_value''',
            (str(uuid.uuid4()), request_id, item['id'], is_sel,
             selection['quantities'].get(item['id'], 1),
             selection['custom_values'].get(item['id']))
        )

    try:
        db.execute("UPDATE food_requests SET updated_at=? WHERE id=?", (now(), request_id))
    except Exception:
        pass

    # Compute diff using names
    added   = [it['name'] for it in all_items if it['id'] in selected_ids and prev_by_id.get(it['id'], ('', 0))[1] == 0]
    removed = [name for iid, (name, sel) in prev_by_id.items() if sel == 1 and iid not in selected_ids]

    _log_order_event(db, request_id, 'items_edited', actor='family',
                     payload={'added': added, 'removed': removed, 'days_until_delivery': days_until})
    db.commit()

    # Notify coordinators if items changed
    if added or removed:
        added_str   = ', '.join(added)   if added   else 'none'
        removed_str = ', '.join(removed) if removed else 'none'
        _notify_coordinators(db,
            f"Order items updated by family:\n"
            f"Family: {req['family_name']} ({req['family_code']})\n"
            f"Cycle: {req['cycle_title']}\n"
            f"Added: {added_str}\n"
            f"Removed: {removed_str}"
        )
        # Notify claimed shopping volunteers via email — fire-and-forget
        claimed_vols = db.execute(
            '''SELECT v.id, v.name, v.email, vs.task_type
               FROM volunteer_slots vs JOIN volunteers v ON vs.claimed_by=v.id
               WHERE vs.cycle_id=? AND vs.family_id=? AND vs.status IN ('claimed','confirmed') AND vs.task_type='shopping' ''',
            (req['cycle_id'], family_id)
        ).fetchall()
        vol_email_sends = [
            (v['email'],
             f'Sihha Shopping List Update — {req["family_name"]}',
             f"Assalamu Alaikum {v['name']},\n\n"
             f"Shopping list update: {req['family_name']} edited their order for {req['cycle_title']}.\n"
             f"Added: {added_str}\nRemoved: {removed_str}\n"
             f"Please check the updated shopping list.\n\n— Sihha Food Program")
            for v in claimed_vols if v['email']
        ]
        _email_notify_async(vol_email_sends)

    log.info(f'Family {family_id} edited items for order {request_id}: +{added} -{removed}')
    return jsonify({'ok': True, 'added': added, 'removed': removed,
                    'message': 'Your order has been updated.'})


@app.route('/uploads/<path:filename>')
def serve_upload(filename):
    """Serve receipt photos to finance staff or the volunteer who owns the receipt."""
    if (secure_filename(filename) != filename or not allowed_file(filename)):
        return jsonify({'error': 'Not found'}), 404
    auth = request.headers.get('Authorization', '')
    token = auth[7:] if auth.startswith('Bearer ') else None
    if not token or token.startswith('tmp_'):
        return jsonify({'error': 'Unauthorized'}), 401
    db = get_db()
    file_url = f'/uploads/{filename}'
    session = get_session(token)
    if session and session['active'] and _linked_account_is_active(db, session):
        if session['role'] in ('admin', 'finance', 'treasurer'):
            return send_from_directory(UPLOAD_FOLDER, filename)
        if session['role'] == 'volunteer':
            owned = db.execute(
                "SELECT 1 FROM receipts WHERE file_url=? AND volunteer_id=? LIMIT 1",
                (file_url, session['linked_id'])
            ).fetchone()
            if owned:
                return send_from_directory(UPLOAD_FOLDER, filename)
            return jsonify({'error': 'Forbidden'}), 403
        return jsonify({'error': 'Forbidden'}), 403

    portal_session = get_portal_session(token)
    if portal_session:
        owned = db.execute(
            "SELECT 1 FROM receipts WHERE file_url=? AND volunteer_id=? LIMIT 1",
            (file_url, portal_session['volunteer_id'])
        ).fetchone()
        if owned:
            return send_from_directory(UPLOAD_FOLDER, filename)
        return jsonify({'error': 'Forbidden'}), 403

    if not session:
        return jsonify({'error': 'Unauthorized'}), 401
    return jsonify({'error': 'Unauthorized'}), 401

# ── Public Volunteer Portal ───────────────────────────────────────────────────

@app.route('/portal')
def portal_page():
    return send_from_directory('public', 'portal.html')

# ── OTP Authentication removed — all auth via username/password login ──────────

@app.route('/api/otp/request', methods=['POST'])
def otp_request():
    return jsonify({'error': 'OTP login removed. Please use username/password at /login.'}), 410

@app.route('/api/otp/verify', methods=['POST'])
def otp_verify():
    return jsonify({'error': 'OTP login removed. Please use username/password at /login.'}), 410


@app.route('/api/portal/login', methods=['POST'])
def portal_login():
    """Legacy phone-only portal login — removed. Use /api/login with username + password."""
    return jsonify({'error': 'This login method is no longer supported. Please use the main login page.'}), 410

@app.route('/api/portal/cycles')
@require_portal_auth()
def portal_list_cycles():
    """Return cycles that are currently eligible for volunteer sign-up."""
    cutoff = (datetime.utcnow() + timedelta(days=365)).strftime('%Y-%m-%d')
    today  = datetime.utcnow().strftime('%Y-%m-%d')
    rows = get_db().execute(
        """SELECT * FROM delivery_cycles
           WHERE status IN ('upcoming','open','shopping')
             AND delivery_date_start >= ?
             AND delivery_date_start <= ?
           ORDER BY delivery_date_start ASC""",
        (today, cutoff)
    ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/portal/slots/<cycle_id>')
@require_portal_auth()
def portal_get_slots(cycle_id):
    db = get_db()
    cycle = db.execute("SELECT * FROM delivery_cycles WHERE id=?", (cycle_id,)).fetchone()
    if not cycle:
        return jsonify({'error': 'Cycle not found'}), 404
    vol_id = g.pv['volunteer_id']
    slots = db.execute(
        '''SELECT vs.*, f.name as family_name, f.family_size, f.family_code,
                  1 as hide_address,
                  CASE WHEN vs.claimed_by=? THEN 1 ELSE 0 END as is_mine,
                  v.name as claimed_by_name
           FROM volunteer_slots vs
           JOIN families f ON vs.family_id = f.id
           LEFT JOIN volunteers v ON vs.claimed_by = v.id
           WHERE vs.cycle_id=? AND vs.status != 'cancelled'
           ORDER BY vs.task_type, f.name''',
        (vol_id, cycle_id)
    ).fetchall()
    result = []
    for s in slots:
        row = dict(s)
        # Delivery volunteers see address only for their own claimed slots
        confirmed_order = db.execute(
            '''SELECT 1 FROM food_requests
               WHERE cycle_id=? AND family_id=?
                 AND status IN ('confirmed','auto_confirmed','submitted','delivered')
               LIMIT 1''',
            (cycle_id, s['family_id'])
        ).fetchone()
        if (row['task_type'] == 'delivery' and row['claimed_by'] == vol_id
                and row['status'] in ('claimed', 'confirmed')
                and confirmed_order):
            family = db.execute("SELECT address, city FROM families WHERE id=?", (s['family_id'],)).fetchone()
            row['family_address'] = f"{family['address']}, {family['city']}" if family else ''
        result.append(row)
    return jsonify({'cycle': dict(cycle), 'slots': result, 'volunteer_id': vol_id})

# /api/portal/claim removed — superseded by /api/portal/signup (Sprint 2)


@app.route('/api/portal/my-tasks')
@require_portal_auth()
def portal_my_tasks():
    vol_id = g.pv['volunteer_id']
    db = get_db()

    # Active + completed assignments (including confirmed slots)
    rows = db.execute(
        '''SELECT vs.*, f.name as family_name, f.address, f.city, f.family_size, f.family_code,
                  dc.title as cycle_title, dc.delivery_date_start, dc.delivery_date_end
           FROM volunteer_slots vs
           JOIN families f ON vs.family_id = f.id
           JOIN delivery_cycles dc ON vs.cycle_id = dc.id
           WHERE vs.claimed_by=? AND vs.status IN ('claimed','confirmed','complete')
           ORDER BY dc.delivery_date_start ASC, vs.task_type''',
        (vol_id,)
    ).fetchall()
    result = []
    for r in rows:
        row = dict(r)
        row['was_released'] = False
        if row['task_type'] == 'shopping':
            row['address'] = None
            row['city'] = None
            # For confirmed shopping slots, include item list so shopper knows what to buy
            if row['status'] == 'confirmed':
                items = db.execute(
                    f'''SELECT fi.name, {_EFFECTIVE_ORDER_QTY_SQL} as qty
                       FROM food_requests fr
                       JOIN food_request_items fri ON fri.request_id = fr.id
                       JOIN food_items fi ON fi.id = fri.food_item_id
                       LEFT JOIN bundle_quantities bq
                         ON bq.food_item_id=fi.id AND bq.bundle_size=fr.bundle_size
                       WHERE fr.cycle_id=? AND fr.family_id=? AND fri.selected=1
                       ORDER BY fi.name''',
                    (row['cycle_id'], row['family_id'])
                ).fetchall()
                row['shopping_items'] = [{'name': it['name'], 'qty': it['qty']} for it in items]
            else:
                row['shopping_items'] = None
        elif row['task_type'] == 'delivery':
            confirmed_order = db.execute(
                '''SELECT 1 FROM food_requests
                   WHERE cycle_id=? AND family_id=?
                     AND status IN ('confirmed','auto_confirmed','submitted','delivered')
                   LIMIT 1''',
                (row['cycle_id'], row['family_id'])
            ).fetchone()
            if row['status'] not in ('claimed', 'confirmed') or not confirmed_order:
                row['address'] = None
                row['city'] = None
        result.append(row)

    # Recently released slots — where this volunteer was the last holder
    # Only show slots from cycles in the last 60 days to avoid stale history
    released = db.execute(
        '''SELECT vs.*, f.name as family_name, f.family_code,
                  dc.title as cycle_title, dc.delivery_date_start, dc.delivery_date_end
           FROM volunteer_slots vs
           JOIN families f ON vs.family_id = f.id
           JOIN delivery_cycles dc ON vs.cycle_id = dc.id
           WHERE vs.prev_claimed_by=? AND vs.claimed_by != ?
             AND vs.status IN ('open','claimed')
             AND dc.delivery_date_start >= date('now', '-60 days')
           ORDER BY dc.delivery_date_start DESC, vs.task_type''',
        (vol_id, vol_id)
    ).fetchall()
    for r in released:
        row = dict(r)
        row['was_released'] = True
        row['address'] = None  # never show address for released assignments
        row['city'] = None
        result.append(row)

    return jsonify(result)

@app.route('/api/portal/complete/<slot_id>', methods=['POST'])
@require_portal_auth()
def portal_complete_slot(slot_id):
    db = get_db()
    vol_id = g.pv['volunteer_id']
    slot = db.execute(
        '''SELECT vs.*, dc.status as cycle_status
           FROM volunteer_slots vs
           JOIN delivery_cycles dc ON dc.id=vs.cycle_id
           WHERE vs.id=? AND vs.claimed_by=?''',
        (slot_id, vol_id)
    ).fetchone()
    if not slot:
        return jsonify({'error': 'Slot not found or not yours'}), 404
    if slot['status'] != 'confirmed':
        return jsonify({'error': 'Only confirmed tasks can be completed'}), 409
    if slot['cycle_status'] == 'delivered':
        return jsonify({'error': 'This delivery cycle is already closed'}), 409
    if slot['task_type'] == 'delivery':
        order = db.execute(
            '''SELECT id FROM food_requests
               WHERE cycle_id=? AND family_id=?
                 AND status IN ('confirmed','auto_confirmed','submitted')''',
            (slot['cycle_id'], slot['family_id'])
        ).fetchone()
        if not order:
            return jsonify({'error': 'The family order is not confirmed'}), 409
    ts = now()
    db.execute(
        "UPDATE volunteer_slots SET status='complete', completed_at=?, updated_at=? WHERE id=?",
        (ts, ts, slot_id)
    )
    # When a delivery slot is marked complete, auto-set delivered_at on the food_request
    if slot['task_type'] == 'delivery':
        db.execute(
            "UPDATE food_requests SET delivered_at=?, status='delivered' WHERE cycle_id=? AND family_id=? AND delivered_at IS NULL",
            (ts, slot['cycle_id'], slot['family_id'])
        )
    db.commit()
    return jsonify({'ok': True})

# ── Portal: Receipt Submission ────────────────────────────────────────────────

@app.route('/api/portal/receipts/upload', methods=['POST'])
@require_portal_auth()
def portal_upload_receipt_file():
    """Upload a receipt photo from the volunteer portal. If vision parsing is active,
    also returns a `parsed` preview (store/date/total/line-items) so the submit form can
    pre-fill — the volunteer still confirms and a treasurer/admin still approves."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    db = get_db()
    filename, raw, upload_error = _store_receipt_upload(
        db, request.files['file'], volunteer_id=g.pv['volunteer_id']
    )
    if upload_error:
        status = 429 if 'quota' in upload_error.lower() else 422
        return jsonify({'error': upload_error}), status
    parsed, perr = _parse_receipt_image_ex(raw, filename)  # (None, reason) unless active
    return jsonify({'file_url': f'/uploads/{filename}', 'parsed': parsed,
                    'parse_error': (perr if not parsed else None)}), 201

@app.route('/api/portal/receipts', methods=['POST'])
@require_portal_auth()
def portal_submit_receipt():
    """Volunteer submits a receipt for a shopping task via the portal.
    Auto-creates a receipt record + pending reimbursement + notifies treasurer."""
    data   = request.json or {}
    vol_id = g.pv['volunteer_id']
    slot_id = data.get('slot_id')
    db = get_db()

    # Validate the slot belongs to this volunteer and is a shopping task
    if slot_id:
        slot = db.execute(
            "SELECT * FROM volunteer_slots WHERE id=? AND claimed_by=? AND task_type='shopping'",
            (slot_id, vol_id)
        ).fetchone()
        if not slot:
            return jsonify({'error': 'Shopping slot not found or not yours'}), 404

    try:
        amount = float(data.get('amount') or 0)
    except (TypeError, ValueError):
        return jsonify({'error': 'Invalid amount'}), 422
    if amount < 0:
        return jsonify({'error': 'Amount cannot be negative'}), 422
    store  = (data.get('store') or '').strip()
    pdate  = data.get('purchase_date') or now()[:10]
    furl   = _normalize_upload_url(data.get('file_url'))
    if data.get('file_url') and not furl:
        return jsonify({'error': 'Invalid receipt file URL'}), 422
    if furl and not _claim_registered_upload(db, furl, volunteer_id=vol_id):
        return jsonify({'error': 'Receipt file is already attached or does not belong to you'}), 422
    fid    = slot['family_id'] if slot_id and slot else None

    # Check for existing receipt for this slot — update instead of reject
    existing = None
    if slot_id:
        existing = db.execute(
            "SELECT id FROM receipts WHERE slot_id=? AND volunteer_id=?", (slot_id, vol_id)
        ).fetchone()

    if existing:
        # Update the existing receipt — reset status to pending for re-review
        rid = existing['id']
        update_fields = [
            ('amount', amount), ('store', store), ('purchase_date', pdate),
            ('status', 'pending'), ('updated_at', now())
        ]
        if furl:  # only overwrite photo if a new one was uploaded
            update_fields.append(('file_url', furl))
        set_clause = ', '.join(f'{col}=?' for col, _ in update_fields)
        vals = [v for _, v in update_fields] + [rid]
        db.execute(f'UPDATE receipts SET {set_clause} WHERE id=?', vals)

        # Receipt goes back to pending review — remove any UNPAID payable (a paid one
        # is left untouched). The payable is recreated if/when it's approved again.
        db.execute("DELETE FROM reimbursements WHERE receipt_id=? AND status!='paid'", (rid,))
        reimb_id = None
        if data.get('parsed'):
            try:
                _persist_receipt_parse(db, rid, _normalize_parsed_receipt(data['parsed']), amount)
            except Exception as _e:
                log.warning(f'portal_submit_receipt(update): parse persist failed for {rid}: {_e}')
        db.commit()

        try:
            vol_name = g.pv['name']
            subject  = f'Receipt Updated — ${amount:.2f} from {vol_name}'
            msg = (f'Volunteer updated their receipt via the Portal.\n'
                   f'Volunteer: {vol_name}\n'
                   f'Store: {store or "not specified"}\n'
                   f'Amount: ${amount:.2f}\n'
                   f'Date: {pdate}\n\n'
                   f'Log in to review: https://sihha-ops-hub-production.up.railway.app')
            _notify_treasurers(db, subject, msg)
        except Exception as e:
            log.warning(f'Treasurer notification failed: {e}')

        return jsonify({'receipt_id': rid, 'reimbursement_id': reimb_id, 'updated': True}), 200

    # New receipt
    rid    = str(uuid.uuid4())
    db.execute(
        '''INSERT INTO receipts
           (id, volunteer_id, family_id, store, purchase_date, amount, file_url, slot_id, status, notes, created_at)
           VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
        (rid, vol_id, fid, store, pdate, amount, furl, slot_id, 'pending', data.get('notes'), now())
    )

    # No reimbursement (payable) yet — one is created only when a treasurer/admin
    # APPROVES the receipt. A pending receipt is a claim under review, not money owed.
    reimb_id = None

    # Persist vision-parse data (if the upload returned a preview) — stores line items
    # + parsed_total and flags a mismatch vs the confirmed amount for treasurer review.
    if data.get('parsed'):
        try:
            _persist_receipt_parse(db, rid, _normalize_parsed_receipt(data['parsed']), amount)
        except Exception as _e:
            log.warning(f'portal_submit_receipt: parse persist failed for {rid}: {_e}')

    # Auto-complete the slot — submitting receipt IS the completion signal.
    # Covers 'confirmed' too: since the 2026-06-09 auto-confirm redesign, slots go
    # straight to 'confirmed', so the old 'claimed'-only guard silently no-op'd.
    if slot_id:
        db.execute(
            "UPDATE volunteer_slots SET status='complete', completed_at=?, updated_at=? "
            "WHERE id=? AND status IN ('claimed','confirmed')",
            (now(), now(), slot_id)
        )

    db.commit()

    # Notify treasurers
    try:
        vol_name = g.pv['name']
        subject  = f'New Reimbursement Request — ${amount:.2f} from {vol_name}'
        msg = (f'New receipt submitted via Volunteer Portal.\n'
               f'Volunteer: {vol_name}\n'
               f'Store: {store or "not specified"}\n'
               f'Amount: ${amount:.2f}\n'
               f'Date: {pdate}\n\n'
               f'Log in to review: https://ops.sihha.org')
        _notify_treasurers(db, subject, msg)
    except Exception as e:
        log.warning(f'Treasurer notification failed: {e}')

    return jsonify({'receipt_id': rid, 'reimbursement_id': reimb_id}), 201

@app.route('/api/portal/receipts', methods=['GET'])
@require_portal_auth()
def portal_list_receipts():
    """Volunteer sees their own receipt submissions + reimbursement status."""
    vol_id = g.pv['volunteer_id']
    rows = get_db().execute(
        '''SELECT r.id, r.store, r.purchase_date, r.amount, r.file_url, r.slot_id,
                  r.status as receipt_status, r.created_at,
                  rb.id as reimbursement_id, rb.status as reimbursement_status,
                  rb.payment_method, rb.payment_ref, rb.paid_date
           FROM receipts r
           LEFT JOIN reimbursements rb ON rb.receipt_id = r.id
           WHERE r.volunteer_id=?
           ORDER BY r.created_at DESC''',
        (vol_id,)
    ).fetchall()
    return jsonify([dict(r) for r in rows])

# ── History Endpoints ─────────────────────────────────────────────────────────

@app.route('/api/families/<fid>/history')
@require_auth(roles=['admin', 'finance', 'treasurer', 'viewer'])
def family_history(fid):
    """Per-cycle order history for a family. Includes items and volunteer info."""
    db = get_db()
    family = db.execute("SELECT id, name, family_size FROM families WHERE id=?", (fid,)).fetchone()
    if not family:
        return jsonify({'error': 'Not found'}), 404

    orders = db.execute(
        '''SELECT fr.id, fr.cycle_id, fr.bundle_size, fr.submitted_at,
                  fr.status, fr.delivered_at,
                  dc.title as cycle_title,
                  dc.delivery_date_start, dc.delivery_date_end
           FROM food_requests fr
           JOIN delivery_cycles dc ON fr.cycle_id = dc.id
           WHERE fr.family_id=?
           ORDER BY fr.submitted_at DESC''',
        (fid,)
    ).fetchall()

    result = []
    for order in orders:
        o = dict(order)

        # Selected items
        items = db.execute(
            '''SELECT fi.name, fi.unit, fc.name as category
               FROM food_request_items fri
               JOIN food_items fi ON fri.food_item_id = fi.id
               JOIN food_categories fc ON fi.category_id = fc.id
               WHERE fri.request_id=? AND fri.selected=1
               ORDER BY fc.display_order, fi.display_order''',
            (o['id'],)
        ).fetchall()
        o['selected_items'] = [dict(i) for i in items]

        # Volunteer slots for this order
        slots = db.execute(
            '''SELECT vs.id, vs.task_type, vs.status, vs.completed_at,
                      v.name as volunteer_name, v.id as volunteer_id
               FROM volunteer_slots vs
               LEFT JOIN volunteers v ON vs.claimed_by = v.id
               WHERE vs.cycle_id=? AND vs.family_id=?''',
            (o['cycle_id'], fid)
        ).fetchall()
        o['slots'] = [dict(s) for s in slots]

        # Order event log
        import json as _json
        ev_rows = db.execute(
            "SELECT event_type, actor, payload, created_at FROM food_request_events WHERE request_id=? ORDER BY created_at ASC",
            (o['id'],)
        ).fetchall()
        o['events'] = []
        for ev in ev_rows:
            try:
                payload = _json.loads(ev['payload'])
            except Exception:
                payload = {}
            o['events'].append({
                'event_type': ev['event_type'],
                'actor': ev['actor'],
                'payload': payload,
                'created_at': ev['created_at']
            })

        result.append(o)

    # Current active cycle (open or shopping) — used by admin UI to show "Add to Cycle" button
    active_cycle = db.execute(
        "SELECT id, title, status, delivery_date_start FROM delivery_cycles WHERE status IN ('open','shopping') ORDER BY delivery_date_start LIMIT 1"
    ).fetchone()
    active_cycle_data = dict(active_cycle) if active_cycle else None

    return jsonify({'family': dict(family), 'orders': result, 'active_cycle': active_cycle_data})


@app.route('/api/volunteers/<vid>/history')
@require_auth(roles=['admin', 'finance', 'treasurer', 'viewer'])
def volunteer_history(vid):
    """Task history for a volunteer: lifetime stats + per-task log."""
    db = get_db()
    vol = db.execute("SELECT id, name, role FROM volunteers WHERE id=?", (vid,)).fetchone()
    if not vol:
        return jsonify({'error': 'Not found'}), 404

    tasks = db.execute(
        '''SELECT vs.id, vs.task_type, vs.status, vs.task_date,
                  vs.claimed_at, vs.completed_at,
                  dc.title as cycle_title, dc.delivery_date_start,
                  f.id as family_id, f.family_size
           FROM volunteer_slots vs
           JOIN delivery_cycles dc ON vs.cycle_id = dc.id
           JOIN families f ON vs.family_id = f.id
           WHERE vs.claimed_by=?
           ORDER BY vs.claimed_at DESC''',
        (vid,)
    ).fetchall()

    task_list = [dict(t) for t in tasks]

    # Lifetime stats
    total      = len(task_list)
    completed  = sum(1 for t in task_list if t['status'] == 'complete')
    shopping   = sum(1 for t in task_list if t['task_type'] == 'shopping')
    delivery   = sum(1 for t in task_list if t['task_type'] == 'delivery')
    cycles_served = len({t['cycle_title'] for t in task_list})
    families_served = len({t['family_id'] for t in task_list if t['status'] == 'complete'})

    return jsonify({
        'volunteer': dict(vol),
        'stats': {
            'total_tasks': total,
            'completed':   completed,
            'shopping':    shopping,
            'delivery':    delivery,
            'cycles_served':   cycles_served,
            'families_served': families_served,
        },
        'tasks': task_list
    })


@app.route('/api/portal/history')
@require_portal_auth()
def portal_history():
    """Volunteer's own history — privacy-safe: family_id only, no names or addresses."""
    db = get_db()
    vol_id = g.pv['volunteer_id']

    tasks = db.execute(
        '''SELECT vs.id, vs.task_type, vs.status, vs.task_date,
                  vs.claimed_at, vs.completed_at,
                  dc.title as cycle_title, dc.delivery_date_start,
                  f.id as family_id, f.family_size
           FROM volunteer_slots vs
           JOIN delivery_cycles dc ON vs.cycle_id = dc.id
           JOIN families f ON vs.family_id = f.id
           WHERE vs.claimed_by=? AND vs.status='complete'
           ORDER BY vs.completed_at DESC''',
        (vol_id,)
    ).fetchall()

    task_list = [dict(t) for t in tasks]

    completed  = len(task_list)
    shopping   = sum(1 for t in task_list if t['task_type'] == 'shopping')
    delivery   = sum(1 for t in task_list if t['task_type'] == 'delivery')
    cycles_served   = len({t['cycle_title'] for t in task_list})
    families_served = len({t['family_id'] for t in task_list})

    return jsonify({
        'stats': {
            'completed':       completed,
            'shopping':        shopping,
            'delivery':        delivery,
            'cycles_served':   cycles_served,
            'families_served': families_served,
        },
        'tasks': task_list
    })


# ── Admin: Generate Slots for a Cycle ─────────────────────────────────────────

@app.route('/api/delivery-cycles/<cid>/generate-slots', methods=['POST'])
@require_auth(roles=['admin'])
def generate_cycle_slots(cid):
    db = get_db()
    cycle = db.execute("SELECT * FROM delivery_cycles WHERE id=?", (cid,)).fetchone()
    if not cycle:
        return jsonify({'error': 'Cycle not found'}), 404
    requests = db.execute(
        "SELECT * FROM food_requests WHERE cycle_id=? AND status IN ('confirmed','auto_confirmed','submitted')", (cid,)
    ).fetchall()
    created = 0
    for req in requests:
        # Delegate to _ensure_volunteer_slots — it SELECT-firsts to prevent duplicates
        created += _ensure_volunteer_slots(db, cid, req['family_id'])
    db.commit()
    total_slots = db.execute(
        "SELECT COUNT(*) FROM volunteer_slots WHERE cycle_id=? AND status!='cancelled'", (cid,)
    ).fetchone()[0]
    return jsonify({'ok': True, 'slots_created': created, 'slots_total': total_slots, 'total_requests': len(requests)})

@app.route('/api/volunteer-slots')
@require_auth(roles=['admin', 'finance', 'treasurer', 'viewer'])
def list_volunteer_slots():
    db = get_db()
    cycle_id = request.args.get('cycle_id')
    q = '''SELECT vs.*, f.name as family_name, f.family_size, f.family_code,
                  v.name as claimed_by_name
           FROM volunteer_slots vs
           JOIN families f ON vs.family_id = f.id
           LEFT JOIN volunteers v ON vs.claimed_by = v.id
           WHERE 1=1'''
    params = []
    if cycle_id:
        q += " AND vs.cycle_id=?"; params.append(cycle_id)
    q += " ORDER BY vs.task_type, f.name"
    return jsonify([dict(r) for r in db.execute(q, params).fetchall()])

@app.route('/api/volunteer-slots/<sid>', methods=['PUT'])
@require_auth(roles=['admin'])
def update_volunteer_slot(sid):
    db = get_db()
    slot = db.execute("SELECT * FROM volunteer_slots WHERE id=?", (sid,)).fetchone()
    if not slot:
        return jsonify({'error': 'Not found'}), 404
    d = request.json or {}

    # Support assigning / unassigning a volunteer
    old_claimed_by = slot['claimed_by']
    claimed_by = d.get('claimed_by', old_claimed_by)
    # If a volunteer is being assigned, auto-set status to claimed
    if 'claimed_by' in d:
        default_status = 'claimed' if d['claimed_by'] else 'open'
    else:
        default_status = slot['status']

    # Track previous holder when slot is released or reassigned
    prev_claimed_by = slot.get('prev_claimed_by')
    if 'claimed_by' in d and old_claimed_by and old_claimed_by != d.get('claimed_by'):
        prev_claimed_by = old_claimed_by

    db.execute(
        """UPDATE volunteer_slots
           SET status=?, notes=?, task_date=?, claimed_by=?, prev_claimed_by=?, updated_at=?
           WHERE id=?""",
        (d.get('status', default_status), d.get('notes', slot['notes']),
         d.get('task_date', slot['task_date']), claimed_by, prev_claimed_by, now(), sid)
    )
    db.commit()

    # Email the displaced volunteer
    if 'claimed_by' in d and old_claimed_by and old_claimed_by != d.get('claimed_by'):
        try:
            old_vol = db.execute(
                "SELECT name, email FROM volunteers WHERE id=?", (old_claimed_by,)
            ).fetchone()
            if old_vol and old_vol['email']:
                fam = db.execute(
                    "SELECT f.name, dc.title FROM families f JOIN delivery_cycles dc ON dc.id=? WHERE f.id=?",
                    (slot['cycle_id'], slot['family_id'])
                ).fetchone()
                fam_name    = fam['name']  if fam else 'a family'
                cycle_title = fam['title'] if fam else ''
                action = 'reassigned to another volunteer' if d.get('claimed_by') else 'released back to open'
                body = (f"Sihha Update: Your {slot['task_type']} assignment for {fam_name} ({cycle_title}) "
                        f"has been {action} by a coordinator. No action needed.")
                _email_send(old_vol['email'], 'Sihha Slot Update', body)
        except Exception as _e:
            log.warning(f'update_volunteer_slot: email notify failed: {_e}')

    row = db.execute(
        """SELECT vs.*, v.name as volunteer_name
           FROM volunteer_slots vs
           LEFT JOIN volunteers v ON vs.claimed_by = v.id
           WHERE vs.id=?""", (sid,)
    ).fetchone()
    return jsonify(dict(row))

# ── Task Types ────────────────────────────────────────────────────────────────

@app.route('/api/task-types')
def list_task_types():
    db = get_db()
    rows = db.execute(
        "SELECT slug, label, display_order, is_active, is_family_slot FROM volunteer_task_types ORDER BY display_order, label"
    ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/task-types', methods=['POST'])
@require_auth(roles=['admin'])
def create_task_type():
    d = request.json or {}
    label = (d.get('label') or '').strip()
    if not label:
        return jsonify({'error': 'label required'}), 422
    slug = label.lower().replace(' ', '_').replace('-', '_')
    db = get_db()
    if db.execute("SELECT slug FROM volunteer_task_types WHERE slug=?", (slug,)).fetchone():
        return jsonify({'error': 'Task type already exists'}), 409
    order = db.execute("SELECT COALESCE(MAX(display_order),0)+1 FROM volunteer_task_types").fetchone()[0]
    is_family_slot = 1 if d.get('is_family_slot') else 0
    db.execute(
        "INSERT INTO volunteer_task_types (slug, label, display_order, is_active, is_family_slot) VALUES (?,?,?,1,?)",
        (slug, label, order, is_family_slot)
    )
    db.commit()
    return jsonify({'slug': slug, 'label': label, 'display_order': order, 'is_active': 1, 'is_family_slot': is_family_slot}), 201

@app.route('/api/task-types/<slug>', methods=['PUT'])
@require_auth(roles=['admin'])
def update_task_type(slug):
    db = get_db()
    tt = db.execute("SELECT * FROM volunteer_task_types WHERE slug=?", (slug,)).fetchone()
    if not tt:
        return jsonify({'error': 'Not found'}), 404
    d = request.json or {}
    db.execute(
        "UPDATE volunteer_task_types SET label=?, display_order=?, is_active=?, is_family_slot=? WHERE slug=?",
        (d.get('label', tt['label']), d.get('display_order', tt['display_order']),
         int(d.get('is_active', tt['is_active'])),
         int(d.get('is_family_slot', tt['is_family_slot'] if 'is_family_slot' in tt.keys() else 0)),
         slug)
    )
    db.commit()
    return jsonify(dict(db.execute("SELECT * FROM volunteer_task_types WHERE slug=?", (slug,)).fetchone()))

# ── Volunteer Slot Admin ───────────────────────────────────────────────────────

@app.route('/api/volunteer-slots', methods=['POST'])
@require_auth(roles=['admin'])
def create_volunteer_slot():
    db = get_db()
    d = request.json or {}
    cycle_id   = d.get('cycle_id')
    family_id  = d.get('family_id')
    task_type  = d.get('task_type')
    claimed_by = d.get('claimed_by')
    if not all([cycle_id, family_id, task_type, claimed_by]):
        return jsonify({'error': 'cycle_id, family_id, task_type, claimed_by required'}), 422
    # Prevent duplicate assignment
    if db.execute(
        "SELECT id FROM volunteer_slots WHERE cycle_id=? AND family_id=? AND task_type=? AND claimed_by=?",
        (cycle_id, family_id, task_type, claimed_by)
    ).fetchone():
        return jsonify({'error': 'Volunteer already assigned to this task'}), 409
    cycle = db.execute("SELECT * FROM delivery_cycles WHERE id=?", (cycle_id,)).fetchone()
    slot_id = str(uuid.uuid4())
    db.execute(
        "INSERT INTO volunteer_slots (id,cycle_id,family_id,task_type,task_date,claimed_by,claimed_at,status,created_at) VALUES (?,?,?,?,?,?,?,?,?)",
        (slot_id, cycle_id, family_id, task_type,
         cycle['delivery_date_start'] if cycle else None,
         claimed_by, now(), 'claimed', now())
    )
    db.commit()
    row = db.execute(
        "SELECT vs.*, v.name as volunteer_name FROM volunteer_slots vs LEFT JOIN volunteers v ON vs.claimed_by=v.id WHERE vs.id=?",
        (slot_id,)
    ).fetchone()
    return jsonify(dict(row)), 201

@app.route('/api/volunteer-slots/<sid>', methods=['DELETE'])
@require_auth(roles=['admin'])
def delete_volunteer_slot(sid):
    db = get_db()
    if not db.execute("SELECT id FROM volunteer_slots WHERE id=?", (sid,)).fetchone():
        return jsonify({'error': 'Not found'}), 404
    db.execute("DELETE FROM volunteer_slots WHERE id=?", (sid,))
    db.commit()
    return jsonify({'ok': True})

# ── Portal: Family Sign-up ─────────────────────────────────────────────────────

@app.route('/api/portal/families/<cycle_id>')
@require_portal_auth()
def portal_get_families(cycle_id):
    """Families enrolled in a cycle with the current volunteer's per-task signup status."""
    db = get_db()
    cycle = db.execute("SELECT * FROM delivery_cycles WHERE id=?", (cycle_id,)).fetchone()
    if not cycle:
        return jsonify({'error': 'Cycle not found'}), 404
    vol_id = g.pv['volunteer_id']

    # Only join to orders that are genuinely confirmed (not pending, cancelled, or skipped)
    families = db.execute(
        '''SELECT f.id, f.name, f.family_size, f.family_code, f.address, f.city,
                  fr.status as order_status, fr.id as request_id
           FROM families f
           LEFT JOIN food_requests fr ON fr.family_id = f.id AND fr.cycle_id = ?
                                     AND fr.status IN ('confirmed','auto_confirmed','submitted','delivered')
           WHERE f.status = 'active'
           ORDER BY f.name''',
        (cycle_id,)
    ).fetchall()

    task_types = db.execute(
        "SELECT slug, label, is_family_slot FROM volunteer_task_types WHERE is_active=1 ORDER BY display_order, label"
    ).fetchall()

    result = []
    for fam in families:
        fam_dict = dict(fam)
        # Slots are pre-created on the write paths (cycle creation → _pre_create_slots_for_cycle,
        # family activation → _pre_create_slots_for_family, order confirmation → _ensure_volunteer_slots),
        # so this read endpoint no longer creates/commits them (audit P1.9: a GET must not mutate,
        # and the SELECT-then-INSERT here was the duplicate-slot race source).
        slots = db.execute(
            '''SELECT vs.id, vs.task_type, vs.claimed_by, vs.status,
                      v.name as claimed_by_name
               FROM volunteer_slots vs
               LEFT JOIN volunteers v ON vs.claimed_by = v.id
               WHERE vs.cycle_id=? AND vs.family_id=? AND vs.status!='cancelled'
               ORDER BY vs.claimed_at''',
            (cycle_id, fam['id'])
        ).fetchall()
        my_slots    = {}   # {task_type: slot_id}  — slots I claimed/confirmed
        my_status   = {}   # {task_type: 'claimed'|'confirmed'}
        taken_by    = {}   # {task_type: volunteer_name}  — slots taken by someone else
        vol_counts  = {}   # {task_type: total_count}
        for s in slots:
            tt = s['task_type']
            vol_counts[tt] = vol_counts.get(tt, 0) + 1
            if s['claimed_by'] == vol_id and s['status'] in ('claimed', 'confirmed'):
                my_slots[tt]  = s['id']
                my_status[tt] = s['status']
            elif s['claimed_by'] and s['status'] in ('claimed', 'confirmed'):
                if tt not in taken_by:
                    taken_by[tt] = s['claimed_by_name'] or 'A volunteer'
        fam_dict['my_slots']         = my_slots
        fam_dict['my_status']        = my_status   # 'claimed' = pending order, 'confirmed' = order placed
        fam_dict['taken_by']         = taken_by
        fam_dict['volunteer_counts'] = vol_counts
        # Address only for volunteers signed up for delivery on a confirmed order
        if 'delivery' not in my_slots or not fam_dict.get('order_status'):
            fam_dict['address'] = None
            fam_dict['city']    = None
        result.append(fam_dict)

    # No db.commit() — this endpoint is now read-only (slots pre-created on write paths).

    return jsonify({
        'cycle':      dict(cycle),
        'families':   result,
        'task_types': [dict(t) for t in task_types]
    })

@app.route('/api/portal/signup', methods=['POST'])
@require_portal_auth()
def portal_signup():
    """Volunteer claims the open slot for a family+task in a cycle.
    One volunteer per slot — claims the existing open slot rather than creating a new row."""
    d = request.json or {}
    cycle_id   = d.get('cycle_id')
    family_id  = d.get('family_id')
    task_types = d.get('task_types', [])
    if not cycle_id or not family_id or not task_types:
        return jsonify({'error': 'cycle_id, family_id, task_types required'}), 422
    if (not isinstance(task_types, list) or len(task_types) > 10
            or any(not isinstance(t, str) or not t.strip() for t in task_types)):
        return jsonify({'error': 'task_types must be a list of valid task names'}), 422
    task_types = list(dict.fromkeys(t.strip() for t in task_types))
    vol_id = g.pv['volunteer_id']
    db = get_db()
    cycle  = db.execute("SELECT * FROM delivery_cycles WHERE id=?", (cycle_id,)).fetchone()
    family = db.execute("SELECT * FROM families WHERE id=?", (family_id,)).fetchone()
    if not cycle:
        return jsonify({'error': 'Cycle not found'}), 404
    if cycle['status'] not in ('upcoming', 'open', 'shopping') or cycle['delivery_date_start'] < datetime.utcnow().strftime('%Y-%m-%d'):
        return jsonify({'error': 'This delivery cycle is not open for volunteer sign-up'}), 409
    if not family or family['status'] != 'active':
        return jsonify({'error': 'Family not found or inactive'}), 404

    placeholders = ','.join('?' for _ in task_types)
    valid_tasks = {
        r['slug'] for r in db.execute(
            f'''SELECT slug FROM volunteer_task_types
                WHERE is_active=1 AND is_family_slot=1 AND slug IN ({placeholders})''',
            task_types
        ).fetchall()
    }
    invalid_tasks = [t for t in task_types if t not in valid_tasks]
    if invalid_tasks:
        return jsonify({'error': f'Invalid family task type: {invalid_tasks[0]}'}), 422

    order_confirmed = bool(db.execute(
        '''SELECT 1 FROM food_requests
           WHERE cycle_id=? AND family_id=?
             AND status IN ('confirmed','auto_confirmed','submitted','delivered')
           LIMIT 1''',
        (cycle_id, family_id)
    ).fetchone())
    claim_status = 'confirmed' if order_confirmed else 'claimed'

    claimed = []
    ts = now()
    for task_type in task_types:
        # Already mine — skip silently
        if db.execute(
            "SELECT id FROM volunteer_slots WHERE cycle_id=? AND family_id=? AND task_type=? AND claimed_by=? AND status IN ('claimed','confirmed')",
            (cycle_id, family_id, task_type, vol_id)
        ).fetchone():
            continue

        # Already taken by someone else?  Check BEFORE touching the open slot.
        taken = db.execute(
            '''SELECT v.name FROM volunteer_slots vs
               JOIN volunteers v ON vs.claimed_by = v.id
               WHERE vs.cycle_id=? AND vs.family_id=? AND vs.task_type=?
                 AND vs.status IN ('claimed','confirmed') AND vs.claimed_by != ?''',
            (cycle_id, family_id, task_type, vol_id)
        ).fetchone()
        if taken:
            return jsonify({'error': f'{task_type.capitalize()} is already assigned to {taken["name"]}'}), 409

        open_slot = db.execute(
            "SELECT id FROM volunteer_slots WHERE cycle_id=? AND family_id=? AND task_type=? AND status='open'",
            (cycle_id, family_id, task_type)
        ).fetchone()

        if open_slot:
            # Atomic claim: AND status='open' guard + rowcount check closes the race
            # where two volunteers pass the "taken?" check simultaneously — without
            # this, both got 201 + confirmation emails and the last write silently won.
            cur = db.execute(
                "UPDATE volunteer_slots SET claimed_by=?, claimed_at=?, status=?, updated_at=? WHERE id=? AND status='open'",
                (vol_id, ts, claim_status, ts, open_slot['id'])
            )
            if cur.rowcount == 0:
                # Returning without commit rolls back any earlier task_type claims
                # in this request — preserves the original all-or-nothing behavior.
                holder = db.execute(
                    "SELECT v.name FROM volunteer_slots vs JOIN volunteers v ON vs.claimed_by=v.id WHERE vs.id=?",
                    (open_slot['id'],)
                ).fetchone()
                who = holder['name'] if holder else 'another volunteer'
                return jsonify({'error': f'{task_type.capitalize()} was just claimed by {who}'}), 409
        else:
            active_slot = db.execute(
                '''SELECT status FROM volunteer_slots
                   WHERE cycle_id=? AND family_id=? AND task_type=? AND status!='cancelled' ''',
                (cycle_id, family_id, task_type)
            ).fetchone()
            if active_slot:
                return jsonify({'error': f'{task_type.capitalize()} is not available'}), 409
            try:
                db.execute(
                    "INSERT INTO volunteer_slots (id,cycle_id,family_id,task_type,task_date,claimed_by,claimed_at,status,created_at) VALUES (?,?,?,?,?,?,?,?,?)",
                    (str(uuid.uuid4()), cycle_id, family_id, task_type,
                     cycle['delivery_date_start'] if task_type == 'delivery' else None,
                     vol_id, ts, claim_status, ts)
                )
            except sqlite3.IntegrityError:
                return jsonify({'error': f'{task_type.capitalize()} was just claimed by another volunteer'}), 409
        claimed.append(task_type)

    db.commit()

    # Email confirmation to volunteer
    if claimed:
        vol_row = db.execute("SELECT * FROM volunteers WHERE id=?", (vol_id,)).fetchone()
        vol = dict(vol_row) if vol_row else {}
        fam = dict(family) if family else {}
        vol_email = vol.get('email')
        if vol_email:
            fcode      = fam.get('family_code', '')
            task_label = ', '.join(t.capitalize() for t in claimed)
            assignment_word = 'confirmed' if order_confirmed else 'reserved'
            body = (f"Assalamu Alaikum {vol.get('name', '')},\n\n"
                    f"Your volunteer slot has been {assignment_word} for: {task_label}\n"
                    f"Family: {fcode} - Size: {fam.get('family_size', '?')}\n"
                    f"Delivery: {cycle['delivery_date_start']}\n"
                    f"JazakAllah Khair!\n\n— Sihha Food Program")
            if order_confirmed and 'delivery' in claimed and fam.get('address'):
                body += f"\nAddress: {fam['address']}, {fam.get('city', '')}"
            subject = f"Sihha {assignment_word.capitalize()}: {task_label}"
            _email_send(vol_email, subject, body)

    return jsonify({'ok': True, 'claimed': claimed}), 201

@app.route('/api/portal/cancel/<slot_id>', methods=['DELETE'])
@require_portal_auth()
def portal_cancel_slot(slot_id):
    """Volunteer releases their slot back to open so another volunteer can claim it."""
    vol_id = g.pv['volunteer_id']
    db = get_db()
    slot = db.execute(
        "SELECT * FROM volunteer_slots WHERE id=? AND claimed_by=?", (slot_id, vol_id)
    ).fetchone()
    if not slot:
        return jsonify({'error': 'Not found or not yours'}), 404
    if slot['status'] not in ('claimed', 'confirmed'):
        return jsonify({'error': 'Only active assignments can be cancelled'}), 409
    # Release back to open — preserve prev_claimed_by for portal history
    db.execute(
        "UPDATE volunteer_slots SET prev_claimed_by=claimed_by, claimed_by=NULL, claimed_at=NULL, status='open', updated_at=? WHERE id=?",
        (now(), slot_id)
    )
    db.commit()
    return jsonify({'ok': True})

# ── Email Reminders ────────────────────────────────────────────────────────────

def _send_reminders_job():
    """Core reminder logic. Called by scheduler and by the admin trigger endpoint.
    Uses a direct DB connection (no Flask request context needed).
    DB-idempotent: reminder_log UNIQUE(slot_id, sent_to) prevents double-sends
    even when both gunicorn workers run the job simultaneously."""
    conn = make_conn()  # FK enforcement + busy_timeout (audit 2.5)
    try:
        # Delivery dates are Central-time calendar dates. UTC can already be on
        # the next day during the Chicago evening and would skip the real target.
        target_date = (_today_central() + timedelta(days=2)).isoformat()
        slots = conn.execute(
            '''SELECT vs.*, v.name as vol_name, v.email as vol_email,
                      f.name as family_name, f.family_code, f.address, f.city
               FROM volunteer_slots vs
               JOIN volunteers v ON vs.claimed_by = v.id
               JOIN families f ON vs.family_id = f.id
               WHERE vs.status='confirmed' AND vs.task_date=?
               AND v.email IS NOT NULL AND TRIM(v.email) != '' ''',
            (target_date,)
        ).fetchall()
        sent = 0
        for s in slots:
            vol_email = (s['vol_email'] or '').strip()
            if not vol_email:
                continue
            try:
                conn.execute(
                    "INSERT INTO reminder_log (id,slot_id,sent_to,sent_at) VALUES (?,?,?,?)",
                    (str(uuid.uuid4()), s['id'], vol_email, datetime.utcnow().isoformat())
                )
                conn.commit()
            except sqlite3.IntegrityError:
                continue  # Already sent to this volunteer for this slot
            fcode = s['family_code'] or ''
            if s['task_type'] == 'delivery':
                body = (f"Assalamu Alaikum {s['vol_name']},\n\n"
                        f"Reminder: You have a delivery in 2 days!\n\n"
                        f"Family ID: {fcode}\n"
                        f"Address: {s['address']}, {s['city']}\n"
                        f"Please deliver by 5pm.\n\n"
                        f"JazakAllah Khair!\n\n— Sihha Food Program")
                subject = f"Sihha Reminder: Delivery on {target_date}"
            else:
                body = (f"Assalamu Alaikum {s['vol_name']},\n\n"
                        f"Reminder: You have a shopping task in 2 days!\n\n"
                        f"Family ID: {fcode}\n"
                        f"Drop off at Abu Baqr by Sunday 2pm.\n"
                        f"Send receipt to treasurer after shopping.\n\n"
                        f"JazakAllah Khair!\n\n— Sihha Food Program")
                subject = f"Sihha Reminder: Shopping on {target_date}"
            try:
                delivered = _email_send(vol_email, subject, body)
            except Exception as exc:
                log.warning(f'Email reminder failed for slot {s["id"]}: {exc}')
                delivered = False
            if delivered:
                sent += 1
            else:
                # The unique row is a successful-send guard, not an attempt log.
                # Remove it after failure so the next scheduler run can retry.
                conn.execute(
                    "DELETE FROM reminder_log WHERE slot_id=? AND sent_to=?",
                    (s['id'], vol_email)
                )
                conn.commit()
        log.info(f'Email Reminders: {sent} sent for target date {target_date}')
        return sent, target_date
    finally:
        conn.close()

def _destructive_guard(expected_confirm):
    """Two-key safety for destructive admin endpoints (audit P2). Returns an
    (response, status) error tuple to return immediately, or None to proceed.
    Requires BOTH: an ops-level opt-in env var (ALLOW_DESTRUCTIVE_OPS truthy) AND
    an exact confirmation phrase in the JSON body — so neither a stray POST nor a
    forgotten env flag alone can wipe live data. Backups are only once-daily."""
    enabled = os.environ.get('ALLOW_DESTRUCTIVE_OPS', '').strip().lower() in ('1', 'true', 'yes', 'on')
    if not enabled:
        return jsonify({'error': 'Destructive operations are disabled on this deployment. '
                                 'Set ALLOW_DESTRUCTIVE_OPS=1 in the environment to enable them.'}), 403
    confirm = (request.json or {}).get('confirm', '')
    if confirm != expected_confirm:
        return jsonify({'error': f'Confirmation required. Re-send this request with '
                                 f'body {{"confirm": "{expected_confirm}"}} to proceed.'}), 400
    return None


@app.route('/api/admin/wipe-test-data', methods=['POST'])
@require_auth(roles=['admin'])
def wipe_test_data():
    """Wipe all operational data. Preserves: users, food catalog, donations, sessions.
    Gated by _destructive_guard — requires ALLOW_DESTRUCTIVE_OPS + confirm phrase."""
    _blocked = _destructive_guard('WIPE-ALL-DATA')
    if _blocked:
        return _blocked
    db = get_db()
    db.execute('PRAGMA foreign_keys=OFF')
    counts = {}
    for t in ['reminder_log','portal_sessions','food_request_items','food_requests',
              'volunteer_slots','receipts','reimbursements','cycle_assignments',
              'delivery_cycles','volunteers','families']:
        try:
            n = db.execute(f'SELECT COUNT(*) FROM {t}').fetchone()[0]
            db.execute(f'DELETE FROM {t}')
            counts[t] = n
        except Exception:
            counts[t] = 0
    db.execute('PRAGMA foreign_keys=ON')
    db.commit()
    log.info(f'wipe-test-data: {counts}')
    return jsonify({'ok': True, 'wiped': counts})


@app.route('/api/admin/import-historical', methods=['POST'])
@require_auth(roles=['admin'])
def import_historical():
    """Import real families, volunteers and historical cycles in one shot.
    Bypasses auto-enrollment — handles delivery matrix directly."""
    data = request.json or {}
    db   = get_db()
    n    = now()

    results = {'families': 0, 'volunteers': 0, 'cycles': 0, 'food_requests': 0, 'errors': []}

    # ── Families ─────────────────────────────────────────────────────────────
    fam_id_map = {}  # name → id
    for f in data.get('families', []):
        fid = str(uuid.uuid4())
        # Derive family_size from bundle_size
        bsize = f.get('bundle_size', 'M')
        size_map = {'S': 2, 'M': 4, 'L': 6}
        family_size = size_map.get(bsize, 4)
        # Generate family_code
        phone = f.get('phone', '')
        digits = ''.join(c for c in phone if c.isdigit())
        code = digits[-4:] + bsize if digits else fid[:4].upper()
        try:
            db.execute(
                '''INSERT INTO families
                   (id,name,phone,address,city,family_size,dietary_notes,status,
                    source,family_code,created_at)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
                (fid, f['name'], phone, f.get('address',''),
                 f.get('city','Rochester'), family_size,
                 f.get('dietary_notes',''), 'active', 'import', code, n)
            )
            fam_id_map[f['name']] = fid
            results['families'] += 1
        except Exception as e:
            results['errors'].append(f"Family {f['name']}: {e}")

    # ── Volunteers ───────────────────────────────────────────────────────────
    vol_id_map = {}  # name → id
    for v in data.get('volunteers', []):
        vid = str(uuid.uuid4())
        try:
            db.execute(
                '''INSERT INTO volunteers
                   (id,name,phone,email,status,source,created_at)
                   VALUES (?,?,?,?,?,?,?)''',
                (vid, v['name'], v.get('phone',''), v.get('email',''),
                 'active', 'import', n)
            )
            vol_id_map[v['name']] = vid
            results['volunteers'] += 1
        except Exception as e:
            results['errors'].append(f"Volunteer {v['name']}: {e}")

    db.commit()

    # ── Food items for pre-populating requests ───────────────────────────────
    all_items = [r['id'] for r in db.execute("SELECT id FROM food_items WHERE is_active=1").fetchall()]

    # ── Historical cycles ────────────────────────────────────────────────────
    for cyc in data.get('cycles', []):
        cid = str(uuid.uuid4())
        try:
            db.execute(
                '''INSERT INTO delivery_cycles
                   (id,title,delivery_date_start,delivery_date_end,
                    request_open_at,request_close_at,status,notes,created_at)
                   VALUES (?,?,?,?,?,?,?,?,?)''',
                (cid, cyc['title'], cyc['delivery_date_start'], cyc['delivery_date_end'],
                 '', '', 'delivered', 'Imported from historical records', n)
            )
            results['cycles'] += 1
        except Exception as e:
            results['errors'].append(f"Cycle {cyc['title']}: {e}")
            continue

        # Create food_request for every family — confirmed if they received, skipped otherwise
        delivered_names = set(cyc.get('delivered_families', []))
        for fname, fid in fam_id_map.items():
            status = 'confirmed' if fname in delivered_names else 'skipped'
            token  = str(uuid.uuid4())
            rid    = str(uuid.uuid4())
            try:
                db.execute(
                    '''INSERT INTO food_requests
                       (id,cycle_id,family_id,bundle_size,submitted_at,status,
                        confirmation_token,confirmed_at)
                       VALUES (?,?,?,?,?,?,?,?)''',
                    (rid, cid, fid, 'M', cyc['delivery_date_start'],
                     status, token,
                     cyc['delivery_date_end'] if status == 'confirmed' else None)
                )
                if status == 'confirmed':
                    for item_id in all_items:
                        db.execute(
                            'INSERT OR IGNORE INTO food_request_items (id,request_id,food_item_id,selected) VALUES (?,?,?,1)',
                            (str(uuid.uuid4()), rid, item_id)
                        )
                results['food_requests'] += 1
            except Exception as e:
                results['errors'].append(f"food_request {fname}/{cyc['title']}: {e}")

    db.commit()
    log.info(f'import-historical: {results}')
    return jsonify({'ok': True, 'results': results})


@app.route('/api/admin/seed-cycles-2026', methods=['POST'])
@require_auth(roles=['admin'])
def seed_cycles_2026():
    """Create all bi-weekly 2026 delivery cycles (May–Dec). DESTRUCTIVE: deletes all
    existing 2026 cycles first, so it's gated by _destructive_guard (ALLOW_DESTRUCTIVE_OPS
    + confirm phrase) to prevent wiping live 2026 cycles on a stray POST."""
    _blocked = _destructive_guard('SEED-CYCLES-2026')
    if _blocked:
        return _blocked
    from datetime import date, timedelta

    def build_cycles():
        cycles = []
        d = date(2026, 5, 9)   # first Saturday after April 25-26 delivery
        while d <= date(2026, 12, 31):
            sat = d
            sun = d + timedelta(days=1)
            order_close = sat - timedelta(days=7)
            order_open  = sat - timedelta(days=13)
            month_name  = sat.strftime('%B')
            cycles.append({
                'title':               f'{month_name} {sat.day}–{sun.day}, 2026',
                'delivery_date_start': sat.isoformat(),
                'delivery_date_end':   sun.isoformat(),
                'request_open_at':     f'{order_open.isoformat()}T08:00:00',
                'request_close_at':    f'{order_close.isoformat()}T23:59:00',
                'status':              'upcoming',
                'notes':               'Auto-seeded by admin endpoint',
            })
            d += timedelta(days=14)
        return cycles

    db = get_db()
    cycles_to_seed = build_cycles()

    # Fix schema before any inserts
    try:
        _fix_delivery_cycles_schema(db)
    except Exception as _e:
        log.error(f'seed_cycles_2026: schema fix failed: {_e}')
        return jsonify({'error': f'Schema migration failed: {_e}'}), 500

    # Wipe any existing 2026 cycles (regardless of status) and reseed cleanly
    deleted = db.execute(
        "DELETE FROM delivery_cycles WHERE delivery_date_start >= '2026-01-01'"
    ).rowcount
    db.commit()

    created = 0
    cycle_ids = []
    for c in cycles_to_seed:
        cid = str(uuid.uuid4())
        db.execute(
            '''INSERT INTO delivery_cycles
               (id, title, delivery_date_start, delivery_date_end,
                request_open_at, request_close_at, status, notes, created_by, created_at)
               VALUES (?,?,?,?,?,?,?,?,?,?)''',
            (cid, c['title'], c['delivery_date_start'], c['delivery_date_end'],
             c['request_open_at'], c['request_close_at'], 'upcoming', c['notes'],
             g.user['user_id'], now())
        )
        created += 1
        cycle_ids.append(cid)
    db.commit()
    # Pre-create volunteer slots for all active families across all new cycles
    total_slots = 0
    try:
        for cid in cycle_ids:
            total_slots += _pre_create_slots_for_cycle(db, cid)
        db.commit()
    except Exception as _e:
        log.warning(f'seed-cycles-2026: slot pre-creation failed: {_e}')
    log.info(f'seed-cycles-2026: deleted={deleted}, created={created}, slots_created={total_slots}')
    return jsonify({'ok': True, 'created': created, 'deleted': deleted, 'slots_created': total_slots})


@app.route('/api/reminders/trigger', methods=['POST'])
@require_auth(roles=['admin'])
def trigger_reminders():
    """Admin manual trigger — also used if Railway Cron is configured."""
    sent, target_date = _send_reminders_job()
    return jsonify({'ok': True, 'reminders_sent': sent, 'target_date': target_date})


@app.route('/api/admin/db-debug', methods=['GET'])
@require_auth(roles=['admin'])
def db_debug():
    """Diagnostic: show delivery_cycles schema + all rows."""
    try:
        db = get_db()
        schema_row = db.execute(
            "SELECT sql FROM sqlite_master WHERE type='table' AND name='delivery_cycles'"
        ).fetchone()
        schema_str = schema_row['sql'] if schema_row else None

        cycles_raw = db.execute(
            "SELECT id, title, delivery_date_start, status FROM delivery_cycles ORDER BY delivery_date_start"
        ).fetchall()
        cycles = [{'id': r['id'], 'title': r['title'],
                   'date': r['delivery_date_start'], 'status': r['status']}
                  for r in cycles_raw]

        test_error = None
        try:
            test_id = '__dbdebug_test__'
            db.execute(
                "INSERT INTO delivery_cycles (id,title,delivery_date_start,delivery_date_end,request_open_at,request_close_at,status,created_by,created_at) VALUES (?,?,?,?,?,?,?,?,?)",
                (test_id,'Test','2099-01-01','2099-01-02','','','upcoming','diag','2026-01-01')
            )
            db.execute("DELETE FROM delivery_cycles WHERE id=?", (test_id,))
            db.commit()
        except Exception as e:
            test_error = str(e)

        return jsonify({
            'ok': True,
            'schema_has_upcoming': 'upcoming' in (schema_str or ''),
            'schema': schema_str,
            'cycle_count': len(cycles),
            'cycles': cycles,
            'upcoming_insert_test': 'OK' if test_error is None else f'FAILED: {test_error}'
        })
    except Exception as e:
        log.exception('db-debug failed')  # traceback goes to logs, never to the client
        return jsonify({'ok': False, 'error': str(e)}), 500

def _send_family_confirmation_reminders():
    """7 days before delivery: email all active families with a link to /family portal.
    Does NOT create food_request rows — families place orders via the portal (single creation path).
    Idempotent via reminder_log (slot_id='opt_in_{cycle_id}', sent_to=family_id)."""
    conn = make_conn()  # FK enforcement + busy_timeout (audit 2.5)
    try:
        target      = (datetime.utcnow() + timedelta(days=7)).strftime('%Y-%m-%d')
        base_url    = os.environ.get('APP_URL', 'https://sihha-ops-hub-production.up.railway.app')
        portal_link = f"{base_url}/family"

        cycles = conn.execute(
            "SELECT * FROM delivery_cycles WHERE delivery_date_start = ? AND status IN ('upcoming','open')",
            (target,)
        ).fetchall()

        sent = 0
        for cycle in cycles:
            cycle_id = cycle['id']
            log_key  = f'opt_in_{cycle_id}'

            families = conn.execute(
                "SELECT id, name, email FROM families WHERE status='active' AND email IS NOT NULL AND TRIM(email) != ''"
            ).fetchall()

            for fam in families:
                fam_email = (fam['email'] or '').strip()
                if not fam_email:
                    continue
                # Reserve the guard row BEFORE sending (audit 2.6). UNIQUE(slot_id,
                # sent_to) makes the INSERT the atomic arbiter across both workers —
                # previously the send happened first, so two workers firing at 09:00
                # could both pass the check and double-email the family.
                cur = conn.execute(
                    "INSERT OR IGNORE INTO reminder_log (id, slot_id, sent_to, sent_at) VALUES (?,?,?,?)",
                    (str(uuid.uuid4()), log_key, fam['id'], datetime.utcnow().isoformat())
                )
                if cur.rowcount == 0:
                    continue  # other worker or earlier run already handled this family
                conn.commit()  # publish the claim before the slow network call
                body = (
                    f"Assalamu Alaikum {fam['name']},\n\n"
                    f"Sihha has a food delivery on {cycle['delivery_date_start']}.\n"
                    f"Please log in to place or manage your order:\n{portal_link}\n\n"
                    f"JazakAllah Khair!\n\n— Sihha Food Program"
                )
                if _email_send(fam_email, f'Sihha Food Delivery — {cycle["delivery_date_start"]}', body):
                    sent += 1
                else:
                    # Send failed — release the guard so tomorrow's run retries
                    conn.execute("DELETE FROM reminder_log WHERE slot_id=? AND sent_to=?",
                                 (log_key, fam['id']))
                    conn.commit()
            conn.commit()

        log.info(f'Family opt-in notifications: {sent} emails sent for delivery {target}')
        return sent
    finally:
        conn.close()

def _skip_nonresponding_families():
    """5 days before delivery (cutoff): mark non-responding families as skipped.
    Handles two cases:
    1. Legacy pending_confirmation rows (created by old scheduler) — UPDATE to skipped.
    2. Active families with NO food_request for this cycle — INSERT a skipped row (tracking only)."""
    import json as _json
    conn = make_conn()  # FK enforcement + busy_timeout (audit 2.5)
    try:
        target  = (datetime.utcnow() + timedelta(days=5)).strftime('%Y-%m-%d')
        ts_skip = datetime.utcnow().isoformat()
        cycles  = conn.execute(
            "SELECT * FROM delivery_cycles WHERE delivery_date_start = ? AND status IN ('upcoming','open','shopping')",
            (target,)
        ).fetchall()

        total_skipped = 0
        for cycle in cycles:
            cycle_id = cycle['id']

            # 1. Legacy: mark any pending_confirmation rows as skipped
            legacy = conn.execute(
                "SELECT fr.id FROM food_requests fr WHERE fr.cycle_id=? AND fr.status='pending_confirmation'",
                (cycle_id,)
            ).fetchall()
            if legacy:
                conn.execute(
                    'UPDATE food_requests SET status=\'skipped\', confirmed_at=? WHERE id IN ({})'.format(
                        ','.join('?' * len(legacy))),
                    [ts_skip] + [r['id'] for r in legacy]
                )
                for _r in legacy:
                    conn.execute(
                        "INSERT INTO food_request_events (id,request_id,event_type,actor,payload,created_at) VALUES (?,?,?,?,?,?)",
                        (str(uuid.uuid4()), _r['id'], 'auto_skipped', 'scheduler',
                         _json.dumps({'note': 'no response by cutoff (legacy row)'}), ts_skip)
                    )
                total_skipped += len(legacy)

            # 2. Active families with no food_request row for this cycle → INSERT skipped
            active_fams = conn.execute(
                "SELECT f.id, f.family_size FROM families f WHERE f.status='active'"
            ).fetchall()
            for fam in active_fams:
                exists = conn.execute(
                    "SELECT id FROM food_requests WHERE cycle_id=? AND family_id=?",
                    (cycle_id, fam['id'])
                ).fetchone()
                if exists:
                    continue
                brow = conn.execute(
                    "SELECT bundle_size FROM bundle_size_rules WHERE min_household<=? AND (max_household IS NULL OR max_household>=?) ORDER BY min_household DESC LIMIT 1",
                    (fam['family_size'] or 1, fam['family_size'] or 1)
                ).fetchone()
                bsize = brow['bundle_size'] if brow else 'M'
                rid   = str(uuid.uuid4())
                conn.execute(
                    "INSERT OR IGNORE INTO food_requests (id,cycle_id,family_id,bundle_size,submitted_at,status) VALUES (?,?,?,?,?,?)",
                    (rid, cycle_id, fam['id'], bsize, ts_skip, 'skipped')
                )
                inserted = conn.execute("SELECT id FROM food_requests WHERE id=?", (rid,)).fetchone()
                if inserted:
                    conn.execute(
                        "INSERT INTO food_request_events (id,request_id,event_type,actor,payload,created_at) VALUES (?,?,?,?,?,?)",
                        (str(uuid.uuid4()), rid, 'auto_skipped', 'scheduler',
                         _json.dumps({'note': 'no order placed by cutoff'}), ts_skip)
                    )
                    total_skipped += 1

            conn.commit()

        log.info(f'Cutoff: {total_skipped} families marked skipped for delivery {target}')
        return total_skipped
    finally:
        conn.close()

def _release_unconfirmed_slots_job():
    """Daily job: release claimed slots where delivery is ≤3 days away and family has no confirmed order.
    Sends WA to each released volunteer. Idempotent via reminder_log (key: slot_id + 'autorelease').
    """
    from datetime import date as _date, timedelta as _td
    conn = make_conn()  # FK enforcement + busy_timeout (audit 2.5)
    try:
        cutoff_date = (_date.today() + _td(days=3)).isoformat()
        today_str   = _date.today().isoformat()

        # Find claimed slots within 3 days where the family's order is NOT confirmed
        slots = conn.execute(
            '''SELECT vs.id, vs.task_type, vs.claimed_by,
                      dc.delivery_date_start, dc.title as cycle_title,
                      f.name as family_name, f.id as family_id,
                      v.email as vol_email, v.name as vol_name
               FROM volunteer_slots vs
               JOIN delivery_cycles dc ON vs.cycle_id = dc.id
               JOIN families f ON vs.family_id = f.id
               JOIN volunteers v ON vs.claimed_by = v.id
               WHERE vs.status IN ('claimed','confirmed')
                 AND dc.delivery_date_start >= ?
                 AND dc.delivery_date_start <= ?
                 AND NOT EXISTS (
                     SELECT 1 FROM food_requests fr
                     WHERE fr.cycle_id = vs.cycle_id
                       AND fr.family_id = vs.family_id
                       AND fr.status IN ('confirmed','auto_confirmed','submitted')
                 )''',
            (today_str, cutoff_date)
        ).fetchall()

        released = 0
        for slot in slots:
            # Idempotency: skip if already sent auto-release for this slot
            already = conn.execute(
                "SELECT id FROM reminder_log WHERE slot_id=? AND sent_to='autorelease'",
                (slot['id'],)
            ).fetchone()
            if already:
                continue

            # Release the slot
            conn.execute(
                "UPDATE volunteer_slots SET prev_claimed_by=claimed_by, claimed_by=NULL, "
                "claimed_at=NULL, status='open', updated_at=? WHERE id=?",
                (datetime.utcnow().isoformat(), slot['id'])
            )
            # Log idempotency guard — rowcount gates the email below (audit 2.6):
            # if the other worker's INSERT landed first, it owns the notification.
            guard = conn.execute(
                "INSERT OR IGNORE INTO reminder_log (id, slot_id, sent_to, sent_at) VALUES (?,?,?,?)",
                (str(uuid.uuid4()), slot['id'], 'autorelease', datetime.utcnow().isoformat())
            )
            conn.commit()
            if guard.rowcount == 0:
                continue  # other worker already released + notified for this slot
            released += 1

            # Email volunteer
            vol_email = (slot['vol_email'] or '').strip() if slot['vol_email'] else ''
            if vol_email:
                try:
                    body = (f"Assalamu Alaikum {slot['vol_name']},\n\n"
                            f"{slot['family_name']} has not placed an order for {slot['cycle_title']} "
                            f"(delivery {slot['delivery_date_start']}).\n"
                            f"Your {slot['task_type']} slot has been released — no action needed.\n\n"
                            f"JazakAllah Khair for signing up!\n\n— Sihha Food Program")
                    _email_send(vol_email, f'Sihha Slot Released — {slot["cycle_title"]}', body)
                except Exception:
                    pass

        log.info(f'_release_unconfirmed_slots_job: released {released} slots (checked {len(slots)})')
        return released
    finally:
        conn.close()


# ── Daily DB backup (Phase 0.1 — audit 2026-06-09) ───────────────────────────
# The entire dataset is one SQLite file on one Railway volume. This job writes a
# date-stamped snapshot to data/backups/ via the SQLite online-backup API (safe
# under WAL while the app is serving traffic) and rotates old copies.
# Idempotent across the 2 gunicorn workers: skips if today's backup exists.
# NOTE: still on the same volume — off-site copy is backlog item 0.2.

BACKUP_DIR  = os.environ.get('BACKUP_DIR', os.path.join(os.path.dirname(DB_PATH) or '.', 'backups'))
BACKUP_KEEP = int(os.environ.get('BACKUP_KEEP', 14))

def _offsite_backup(dest):
    """Phase 0.2: email the gzipped daily snapshot off-Railway.
    Opt-in via BACKUP_EMAIL env var (recipient inbox). Skips silently if unset.
    A volume failure then costs at most one day of data instead of everything."""
    recipient = os.environ.get('BACKUP_EMAIL', '').strip()
    if not recipient:
        return
    try:
        import gzip
        with open(dest, 'rb') as f:
            gz = gzip.compress(f.read(), compresslevel=9)
        stamp = os.path.basename(dest)
        if len(gz) > EMAIL_ATTACHMENT_MAX_BYTES:
            _email_send(recipient, f'Sihha backup TOO LARGE to email — {stamp}',
                        f'Compressed DB snapshot is {len(gz)//1024//1024} MB '
                        f'(>{EMAIL_ATTACHMENT_MAX_BYTES//1_000_000} MB email cap). '
                        f'Time to move off-site backups to object storage (S3/B2) — see MEMORY.md backlog 0.2.')
            log.warning(f'Off-site backup skipped — gz size {len(gz)} bytes exceeds email cap')
            return
        ok = _email_send(recipient, f'Sihha daily DB backup — {stamp}',
                         f'Attached: gzipped SQLite snapshot {stamp} ({len(gz)//1024} KB compressed). '
                         f'Restore: gunzip, then replace data/sihaa.db on the Railway volume.',
                         attachment=(f'{stamp}.gz', gz))
        log.info(f'Off-site backup email: {"sent" if ok else "FAILED"} ({len(gz)//1024} KB)')
    except Exception as e:
        log.error(f'Off-site backup FAILED: {e}')


def _daily_backup_job():
    try:
        os.makedirs(BACKUP_DIR, exist_ok=True)
        stamp = datetime.utcnow().strftime('%Y%m%d')
        dest = os.path.join(BACKUP_DIR, f'sihaa-{stamp}.db')
        if os.path.exists(dest):
            return
        tmp = f'{dest}.tmp{os.getpid()}'
        src = sqlite3.connect(DB_PATH)
        try:
            dst = sqlite3.connect(tmp)
            try:
                src.backup(dst)
            finally:
                dst.close()
        finally:
            src.close()
        os.replace(tmp, dest)  # atomic; worker race just overwrites identical snapshot
        # Rotate: keep newest BACKUP_KEEP daily snapshots
        snaps = sorted(f for f in os.listdir(BACKUP_DIR)
                       if f.startswith('sihaa-') and f.endswith('.db'))
        for old in snaps[:-BACKUP_KEEP]:
            try:
                os.remove(os.path.join(BACKUP_DIR, old))
            except OSError:
                pass
        log.info(f'Daily DB backup OK: {dest} ({os.path.getsize(dest)//1024} KB), '
                 f'{min(len(snaps), BACKUP_KEEP)} snapshots retained')
        _offsite_backup(dest)  # Phase 0.2 — no-op unless BACKUP_EMAIL env var set
    except Exception as e:
        log.error(f'Daily DB backup FAILED: {e}')


def _purge_expired_sessions_job():
    """Audit 3.6: delete expired admin/portal sessions nightly — previously they
    accumulated forever (only per-token DELETE on explicit logout)."""
    conn = make_conn()
    try:
        cutoff = datetime.utcnow().isoformat()
        n1 = conn.execute("DELETE FROM sessions WHERE expires_at < ?", (cutoff,)).rowcount
        n2 = conn.execute("DELETE FROM portal_sessions WHERE expires_at < ?", (cutoff,)).rowcount
        conn.commit()
        if n1 or n2:
            log.info(f'Session purge: removed {n1} admin + {n2} portal expired sessions')
    except Exception as e:
        log.error(f'Session purge FAILED: {e}')
    finally:
        conn.close()


def _daily_heartbeat_job():
    """Phase 0.4: daily ops digest to active admin users. Its arrival proves the
    scheduler and app are alive; its content surfaces silent failures (missing
    backup, stuck cycle, growing pending queues) before families notice."""
    conn = make_conn()
    try:
        today = datetime.utcnow().strftime('%Y-%m-%d')
        try:
            snaps = sorted(f for f in os.listdir(BACKUP_DIR)
                           if f.startswith('sihaa-') and f.endswith('.db'))
            if snaps:
                latest = snaps[-1]
                size_kb = os.path.getsize(os.path.join(BACKUP_DIR, latest)) // 1024
                fresh = latest == f'sihaa-{datetime.utcnow().strftime("%Y%m%d")}.db'
                backup_line = f'Backup: {latest} ({size_kb} KB)' + ('' if fresh else ' — ⚠️ NOT from today')
            else:
                backup_line = '⚠️ NO BACKUPS FOUND on volume'
        except Exception:
            backup_line = '⚠️ Backup directory unreadable'

        cyc = conn.execute(
            "SELECT * FROM delivery_cycles WHERE status IN ('open','shopping','upcoming') "
            "ORDER BY delivery_date_start LIMIT 1").fetchone()
        if cyc:
            orders = conn.execute(
                "SELECT COUNT(*) FROM food_requests WHERE cycle_id=? AND status IN ('confirmed','submitted')",
                (cyc['id'],)).fetchone()[0]
            open_slots = conn.execute(
                "SELECT COUNT(*) FROM volunteer_slots WHERE cycle_id=? AND status='open'",
                (cyc['id'],)).fetchone()[0]
            cycle_line = (f"Active cycle: {cyc['title']} [{cyc['status']}] — "
                          f"{orders} orders, {open_slots} open volunteer slots")
        else:
            cycle_line = 'No active cycle'

        emails_24h = conn.execute(
            "SELECT COUNT(*) FROM reminder_log WHERE sent_at >= ?",
            ((datetime.utcnow() - timedelta(days=1)).isoformat(),)).fetchone()[0]
        pend_fam = conn.execute("SELECT COUNT(*) FROM families   WHERE status='pending'").fetchone()[0]
        pend_vol = conn.execute("SELECT COUNT(*) FROM volunteers WHERE status='pending'").fetchone()[0]
        pend_rec = conn.execute("SELECT COUNT(*) FROM receipts   WHERE status='pending'").fetchone()[0]

        body = (f"Sihha Ops Hub — daily heartbeat {today} (UTC)\n\n"
                f"{backup_line}\n"
                f"{cycle_line}\n"
                f"Notification emails sent (24h): {emails_24h}\n"
                f"Pending review: {pend_fam} families · {pend_vol} volunteers · {pend_rec} receipts\n\n"
                f"If this email stops arriving, the app or its scheduler is down — check Railway.")

        # One heartbeat per day across both workers (same guard pattern as reminders)
        guard = conn.execute(
            "INSERT OR IGNORE INTO reminder_log (id, slot_id, sent_to, sent_at) VALUES (?,?,?,?)",
            (str(uuid.uuid4()), f'heartbeat_{today}', 'admins', datetime.utcnow().isoformat()))
        conn.commit()
        if guard.rowcount == 0:
            return
        admins = conn.execute(
            "SELECT email FROM users WHERE role='admin' AND active=1 "
            "AND email IS NOT NULL AND TRIM(email)!=''").fetchall()
        sent = sum(1 for a in admins
                   if _email_send(a['email'], f'Sihha daily heartbeat — {today}', body))
        log.info(f'Heartbeat digest sent to {sent}/{len(admins)} admins')
    except Exception as e:
        log.error(f'Heartbeat job FAILED: {e}')
    finally:
        conn.close()


# ── Bootstrap on startup (runs under both gunicorn and direct execution) ──────

bootstrap_db()

# ── APScheduler: daily 8am UTC reminder job ───────────────────────────────────
# Runs ONCE in the gunicorn master (started with --preload), NOT per worker — the
# jobs are registered at import time and --preload imports the module a single time
# in the master before forking. reminder_log idempotency remains as defense-in-depth.
# job_defaults (audit P1.8): APScheduler's default misfire_grace_time is 1 second, so
# any delay (a redeploy landing on a cron time, GC/IO stall) would SILENTLY skip that
# run entirely — dropping a day's reminders/backup. 3600s grace lets a delayed job
# still fire within the hour; coalesce collapses multiple missed runs into one.
try:
    from apscheduler.schedulers.background import BackgroundScheduler
    _scheduler = BackgroundScheduler(
        timezone='UTC',
        job_defaults={'misfire_grace_time': 3600, 'coalesce': True},
    )
    _scheduler.add_job(_send_reminders_job, 'cron', hour=8, minute=0,
                       id='daily_reminders', replace_existing=True)
    _scheduler.add_job(_send_family_confirmation_reminders, 'cron', hour=9, minute=0,
                       id='family_opt_in_notifications', replace_existing=True)
    _scheduler.add_job(_skip_nonresponding_families, 'cron', hour=9, minute=30,
                       id='family_cutoff_skip', replace_existing=True)
    _scheduler.add_job(_release_unconfirmed_slots_job, 'cron', hour=10, minute=0,
                       id='auto_release_unconfirmed_slots', replace_existing=True)
    _scheduler.add_job(_daily_backup_job, 'cron', hour=7, minute=30,
                       id='daily_db_backup', replace_existing=True)
    _scheduler.add_job(_daily_heartbeat_job, 'cron', hour=11, minute=0,
                       id='daily_heartbeat', replace_existing=True)
    _scheduler.add_job(_purge_expired_sessions_job, 'cron', hour=6, minute=45,
                       id='session_purge', replace_existing=True)
    _scheduler.add_job(_sync_wix_donations_job, 'cron', minute=0,
                       id='wix_donation_sync', replace_existing=True)
    _scheduler.start()
    log.info('APScheduler started — Wix donation sync hourly (:00); DB backup 07:30, email reminders 08:00, family opt-in 09:00, cutoff 09:30, auto-release slots 10:00, heartbeat 11:00 UTC')
    # Take a backup immediately on deploy too (idempotent — skips if today's exists)
    _daily_backup_job()
except ImportError:
    log.warning('APScheduler not installed. Run: pip install apscheduler')
except Exception as _e:
    log.warning(f'APScheduler failed to start: {_e}')

# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    log.info(f'Sihha Ops Hub starting on port {PORT}')
    app.run(host='0.0.0.0', port=PORT, debug=False)
